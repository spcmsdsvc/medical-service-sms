"""
===============================================================================
MEDICAL SERVICE MANAGEMENT SYSTEM - ENTERPRISE MASTER BACKEND
===============================================================================
Version: 5.4.5 (Activity Log Tuning & PH Time Integration)
Framework: Flask / SQLAlchemy / Flask-Login
Branding: MEDICAL SERVICE

-------------------------------------------------------------------------------
CORE ARCHITECTURAL MODULES AND BUSINESS LOGIC:
-------------------------------------------------------------------------------

1.  IDENTITY & ACCESS MANAGEMENT (IAM):
    - Role-based security architecture supporting SuperAdmin, RegionalAdmin, and Engineer.
    - Password encryption using Werkzeug PBKDF2 hashing algorithms.
    - Individual security preference module for credential rotation.
    - v5.4: Hard link between User accounts and Engineer profiles via user_id.

2.  ACTIVITY LOGGING (v5.4.5):
    - Background audit trail recording all significant system actions.
    - UPDATED: Disabled login/logout logging to reduce noise.
    - Captures Calendar and Database changes for Admin oversight.
    - Timestamps strictly follow Manila Time (UTC+8).

3.  CLIENT & CENTER DATABASE:
    - Expanded 9-contact matrix to track diverse hospital department leads.
    - Intelligent acronym parsing (e.g. SLMC vs St Lukes).

4.  TECHNICAL WORKFLOW & COMPLIANCE:
    - Status lifecycle management: In Progress -> For Continuation -> Completed.
    - Strict Technical Service Report (TSR) file validation for task closure.

5.  SCHEDULING & CHAIN SYNCHRONIZATION (REPAIRED v5.2.2):
    - Multi-day visit grouping for high-density calendar rendering.
    - HIERO-SYNC: Regional Admin (Kevin) restricted to Cebu/Davao branches.

6.  REPORTING & ANALYTICS:
    - Overhauled Service Dashboard focusing on Open Technical Tasks.
    - Live-refreshing Activity Log for Admins.
===============================================================================
"""

from flask import (
    Flask, 
    render_template, 
    request, 
    jsonify, 
    Response, 
    redirect, 
    url_for, 
    flash,
    send_file
, session)

from flask_sqlalchemy import SQLAlchemy

from flask_login import (
    LoginManager, 
    UserMixin, 
    login_user, 
    logout_user, 
    login_required, 
    current_user
)

from werkzeug.security import (
    generate_password_hash, 
    check_password_hash
)

from werkzeug.utils import secure_filename

from sqlalchemy import (
    func, 
    and_, 
    or_
)

from sqlalchemy.orm import joinedload, selectinload

from datetime import (
    datetime, 
    timedelta,
    timezone
)

# --- NEW SECURITY IMPORTS ---
from flask_wtf.csrf import CSRFProtect
from dotenv import load_dotenv
import os
import re
import io
import csv
import secrets
import string
import smtplib
import html
import json
import threading
import requests
import base64
import zipfile
import zlib
import tempfile
from email.message import EmailMessage
from email.utils import formataddr

# --- APPLICATION CORE INITIALIZATION ---

# Load local environment variables from .env when developing locally.
# The real .env file must stay private and must not be included in zip/git.
load_dotenv()

app = Flask(__name__)

# Master Branding Identity: MEDICAL SERVICE
# SECURITY: SECRET_KEY must come from the environment. Do not hardcode or fallback to a real key.
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
if not app.config['SECRET_KEY']:
    raise RuntimeError("SECRET_KEY is not set. Add it to your local .env file.")

# --- EMAIL NOTIFICATION CONFIGURATION ---
# Email notification settings. Brevo API is default; SMTP remains as fallback.
app.config['SMTP_HOST'] = os.environ.get('SMTP_HOST', 'smtp.office365.com')
app.config['SMTP_PORT'] = int(os.environ.get('SMTP_PORT', '587'))
app.config['SMTP_TIMEOUT'] = int(os.environ.get('SMTP_TIMEOUT', '8'))
app.config['SMTP_USERNAME'] = os.environ.get('SMTP_USERNAME', 'jonamar@shimadzu.com.ph')
app.config['SMTP_PASSWORD'] = os.environ.get('SMTP_PASSWORD')

# Sender identity used by Brevo/SMTP.
# Railway currently uses MAIL_SENDER_EMAIL / MAIL_SENDER_NAME, while older code
# used SMTP_SENDER_EMAIL / SMTP_SENDER_NAME. Support both, with SMTP_* taking
# priority when present.
app.config['SMTP_SENDER_EMAIL'] = (
    os.environ.get('SMTP_SENDER_EMAIL') or
    os.environ.get('MAIL_SENDER_EMAIL') or
    app.config['SMTP_USERNAME']
)
app.config['SMTP_SENDER_NAME'] = (
    os.environ.get('SMTP_SENDER_NAME') or
    os.environ.get('MAIL_SENDER_NAME') or
    'SPC-Medical Service Scheduler'
)
app.config['EMAIL_NOTIFICATIONS_ENABLED'] = os.environ.get('EMAIL_NOTIFICATIONS_ENABLED', 'true').strip().lower() in {'1', 'true', 'yes', 'on'}
app.config['EMAIL_PROVIDER'] = os.environ.get('EMAIL_PROVIDER', 'brevo').strip().lower()
app.config['BREVO_API_KEY'] = os.environ.get('BREVO_API_KEY')
app.config['BREVO_API_URL'] = os.environ.get('BREVO_API_URL', 'https://api.brevo.com/v3/smtp/email')
app.config['BREVO_TIMEOUT'] = int(os.environ.get('BREVO_TIMEOUT', '12'))

# Static internal CC recipients for TSR emails sent to clients.
# Keep this backend-controlled so schedulers cannot accidentally modify CC.
STATIC_TSR_CLIENT_CC_EMAILS = [
    'r_aretano@shimadzu.com.ph',
    'rnrio@shimadzu.com.ph',
    'kevin@shimadzu.com.ph',
    'diary@shimadzu.com.ph',
    'hanna@shimadzu.com.ph'
]


# Approved email-safe font stacks for TSR client emails.
# Tenorite is tried first, with safe fallbacks for Outlook/Gmail/mobile clients.
TSR_EMAIL_FONT_STACKS = {
    'tenorite': "'Tenorite', 'Aptos', 'Segoe UI', Arial, sans-serif",
    'arial': "Arial, sans-serif",
    'verdana': "Verdana, Geneva, sans-serif",
    'tahoma': "Tahoma, Geneva, sans-serif",
    'georgia': "Georgia, serif"
}
DEFAULT_TSR_EMAIL_FONT_KEY = 'tenorite'


def get_tsr_email_font_stack(font_key=None):
    """Return an approved TSR email font stack."""
    requested = (font_key or DEFAULT_TSR_EMAIL_FONT_KEY or '').strip().lower()
    return TSR_EMAIL_FONT_STACKS.get(requested, TSR_EMAIL_FONT_STACKS[DEFAULT_TSR_EMAIL_FONT_KEY])




# Enable Global CSRF Protection
csrf = CSRFProtect(app)

# --- DATABASE ENGINE CONFIGURATION ---

basedir = os.path.abspath(os.path.dirname(__file__))

# Railway persistent SQLite database support.
# IMPORTANT: Railway container storage under /app is temporary and is wiped on redeploy.
# The database must live inside the attached Railway volume at /data.
if os.environ.get('RAILWAY_ENVIRONMENT'):
    os.makedirs('/data', exist_ok=True)

    legacy_db_path = os.path.join(basedir, 'scheduler.db')
    railway_db_path = os.path.join('/data', 'scheduler.db')

    # One-time safe restore:
    # If Railway volume DB is missing/empty/fresh-bootstrap-sized, copy the bundled
    # legacy scheduler.db from the deployed app folder into the persistent volume.
    # If FORCE_RESTORE_SCHEDULER_DB=true is set temporarily in Railway Variables,
    # overwrite /data/scheduler.db once and keep a timestamped backup first.
    try:
        force_restore_scheduler_db = (
            os.environ.get('FORCE_RESTORE_SCHEDULER_DB', '').strip().lower()
            in {'1', 'true', 'yes', 'on'}
        )

        should_restore_legacy_db = (
            os.path.exists(legacy_db_path)
            and (
                force_restore_scheduler_db
                or not os.path.exists(railway_db_path)
                or os.path.getsize(railway_db_path) < 10240
            )
        )

        if should_restore_legacy_db:
            import shutil
            if os.path.exists(railway_db_path):
                backup_existing_db_path = (
                    railway_db_path + '.before_restore_' +
                    datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
                )
                shutil.copy2(railway_db_path, backup_existing_db_path)
                print(f'[DATABASE] Existing Railway DB backed up to {backup_existing_db_path}.', flush=True)

            shutil.copy2(legacy_db_path, railway_db_path)
            print('[DATABASE] Restored bundled scheduler.db into Railway volume /data/scheduler.db.', flush=True)
    except Exception as db_restore_error:
        print(f'[DATABASE] Legacy scheduler.db restore skipped: {db_restore_error}', flush=True)

    DATABASE_PATH = railway_db_path
else:
    DATABASE_PATH = os.path.join(basedir, 'scheduler.db')

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + DATABASE_PATH
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# --- FILE UPLOAD SYSTEM CONFIGURATION ---

# Railway persistent upload storage support
if os.environ.get('RAILWAY_ENVIRONMENT'):
    UPLOAD_FOLDER = '/data/uploads/reports'
else:
    UPLOAD_FOLDER = os.path.join(basedir, 'static/uploads/reports')

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_UPLOAD_MB', '25')) * 1024 * 1024

# SECURITY UPDATE: Allowed file extensions for Technical Service Reports
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'docx', 'xlsx', 'csv'}

def allowed_file(filename):
    """ Helper: Validates if the file extension is permitted """
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def is_tsr_filename(filename):
    """A valid completion TSR must contain TSR in the original/display filename."""
    return bool(filename and 'TSR' in filename.upper())


def validate_uploaded_report_files():
    """Validate uploaded report files before any database mutation."""
    invalid = []
    for file_obj in request.files.getlist('report_file'):
        if not file_obj or not file_obj.filename:
            continue
        if not allowed_file(file_obj.filename):
            invalid.append(file_obj.filename)
    if invalid:
        return False, f"Unsupported report file type: {', '.join(invalid)}"
    return True, None


def uploaded_files_have_tsr():
    """Returns True if any newly uploaded file has TSR in its filename."""
    return any(
        file_obj and file_obj.filename and is_tsr_filename(file_obj.filename)
        for file_obj in request.files.getlist('report_file')
    )


def existing_files_have_tsr(filenames):
    """Returns True if any preserved/stored file has TSR in its filename."""
    return any(is_tsr_filename(filename) for filename in filenames or [])

if not os.path.exists(UPLOAD_FOLDER):
    try:
        os.makedirs(UPLOAD_FOLDER)
    except OSError as folder_err:
        print(f"CRITICAL ERROR: Directory initialization failure: {folder_err}")

db = SQLAlchemy(app)


_shift_file_original_filename_ready = False


def derive_original_filename_from_stored_filename(filename):
    """Return clean client-facing filename from randomized stored filename."""
    safe_name = os.path.basename(clean_str(filename))
    if not safe_name:
        return ''

    match = re.match(r'^shift_\d+_[0-9a-fA-F]{16}_(.+)$', safe_name)
    if match:
        return match.group(1)

    return safe_name


def get_shift_file_display_name(file_rec_or_filename):
    """Return original upload filename for email/UI while keeping disk filename safe."""
    if hasattr(file_rec_or_filename, 'filename'):
        original = clean_str(getattr(file_rec_or_filename, 'original_filename', ''))
        if original:
            return os.path.basename(original)
        return derive_original_filename_from_stored_filename(getattr(file_rec_or_filename, 'filename', ''))

    return derive_original_filename_from_stored_filename(file_rec_or_filename)


def get_shift_file_disk_name(file_rec_or_filename):
    """Return randomized disk filename used for storage/delete lookup."""
    if hasattr(file_rec_or_filename, 'filename'):
        return os.path.basename(clean_str(getattr(file_rec_or_filename, 'filename', '')))
    return os.path.basename(clean_str(file_rec_or_filename))

def get_unique_upload_filename(preferred_filename):
    """Return a clean unique upload filename without random prefixes.

    Used for generated Online TSR PDFs so browser/static downloads keep the
    same clean filename as the Offline TSR Download PDF button.
    """
    safe_name = secure_filename(os.path.basename(clean_str(preferred_filename) or ''))
    if not safe_name:
        safe_name = 'Online_TSR.pdf'

    stem, ext = os.path.splitext(safe_name)
    ext = ext or '.pdf'
    candidate = f"{stem}{ext}"
    counter = 2

    while os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], candidate)):
        candidate = f"{stem}_{counter}{ext}"
        counter += 1

    return candidate


# --- LOGIN MANAGER CONFIGURATION ---

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# --- GLOBAL SYSTEM CONSTANTS ---

LEAVE_CATEGORIES = [
    "Sick Leave", 
    "Vacation Leave", 
    "Emergency Leave", 
    "Paternity Leave", 
    "Maternity Leave"
]

# --- v5.4.4 TIMEZONE HELPER (PH TIME UTC+8) ---

def get_manila_time():
    """ Returns current time in Philippines (UTC+8) """
    return datetime.now(timezone.utc) + timedelta(hours=8)

def get_manila_today():
    """Return current Manila calendar date."""
    return get_manila_time().date()


def should_notify_schedule(shift):
    """Only notify engineers for today/future schedules.

    Backfilling historical schedules should stay silent because those jobs are
    already in the past and do not require engineer email alerts.
    """
    if not shift or not getattr(shift, 'start_time', None):
        return False
    return shift.start_time.date() >= get_manila_today()


def should_notify_snapshot(snapshot):
    """Only notify delete events for today/future schedule snapshots."""
    if not snapshot:
        return False

    raw_date = snapshot.get('date_iso') or snapshot.get('date')
    if raw_date:
        parsed = parse_date(raw_date)
        if parsed:
            return parsed >= get_manila_today()

    return True


# --- DATABASE MODELS ---

class User(UserMixin, db.Model):
    must_change_password = db.Column(db.Boolean, default=False)
    """
    Identity and Access Management (IAM) Model.
    Permission Hierarchy:
    - superadmin: Full authority over personnel accounts and system resets.
    - regional_admin: Oversight of Cebu and Davao branches only.
    - admin: Standard management (Legacy/General).
    - engineer: Technical staff (Self-access only).
    """
    id = db.Column(db.Integer, primary_key=True)
    
    username = db.Column(db.String(100), unique=True, nullable=False)
    
    password = db.Column(db.String(200), nullable=False)
    
    # Permission Scope Definition (Mapped to UI visibility)
    role = db.Column(db.String(20), default='engineer') 

    # v5.4.0: Direct Relationship to Engineer profile
    engineer_rel = db.relationship('Engineer', backref='user_account', uselist=False)

    @property
    def engineer_profile(self):
        """Personalization v5.4: Links user account to the correct engineer profile.

        Important:
        - Admin Kevin must resolve by employee ID, not by first-name matching.
        - This prevents another engineer named Kevin from being treated as the regional admin profile.
        """
        if self.engineer_rel:
            return self.engineer_rel

        username = (self.username or '').strip().lower()
        if username == REGIONAL_ADMIN_USERNAME:
            return Engineer.query.filter_by(employee_id=REGIONAL_ADMIN_EMPLOYEE_ID).first()

        # Fallback for older engineer accounts without user_id linkage.
        return Engineer.query.filter(Engineer.name.ilike(f"{self.username}%")).first()


class ActivityLog(db.Model):
    """
    NEW MODULE v5.2: Audit Trail.
    Records significant actions for Admin oversight.
    """
    id = db.Column(db.Integer, primary_key=True)
    
    user = db.Column(db.String(100), nullable=False)
    
    action = db.Column(db.String(255), nullable=False)
    
    timestamp = db.Column(db.DateTime, default=get_manila_time)


class Engineer(db.Model):
    """
    Technical Personnel Model.
    Stores names and initials of medical equipment technicians.
    Includes manual Employee ID enforcement for Version 5.2.1.
    """
    id = db.Column(db.Integer, primary_key=True)
    
    # v5.4.0: Foreign Key to User Account
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    # NEW v5.2.1: Manual employee identifier
    employee_id = db.Column(db.String(50), unique=True, nullable=False)
    
    name = db.Column(db.String(100), nullable=False)
    
    initials = db.Column(db.String(10), nullable=False)
    
    phone = db.Column(db.String(20))
    
    email = db.Column(db.String(100))
    
    branch = db.Column(db.String(50))
    
    # Linked schedules with cascade deletion
    shifts = db.relationship(
        'Shift',
        backref=db.backref('engineer', foreign_keys='Shift.engineer_id'),
        lazy=True,
        cascade="all, delete-orphan",
        foreign_keys='Shift.engineer_id'
    )


class Client(db.Model):
    """
    Medical Center / Customer Model.
    Features a 9-field contact matrix to support complex 
    hospital administrative hierarchies.
    """
    id = db.Column(db.Integer, primary_key=True)
    
    name = db.Column(db.String(100), nullable=False)
    
    address = db.Column(db.String(200))
    
    # --- Departmental Contact Matrix ---
    
    # Technical Contact Group Alpha
    contact_person_1 = db.Column(db.String(100))
    contact_number_1 = db.Column(db.String(50))
    email_address_1 = db.Column(db.String(100))
    
    # Technical Contact Group Beta
    contact_person_2 = db.Column(db.String(100))
    contact_number_2 = db.Column(db.String(50))
    email_address_2 = db.Column(db.String(100))
    
    # Technical Contact Group Gamma
    contact_person_3 = db.Column(db.String(100))
    contact_number_3 = db.Column(db.String(50))
    email_address_3 = db.Column(db.String(100))
    
    # Logic Relationships
    products = db.relationship('Product', backref='owner', lazy=True)
    
    shifts = db.relationship('Shift', backref='client', lazy=True)



class Contact(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    name = db.Column(db.String(100))
    phone = db.Column(db.String(50))
    email = db.Column(db.String(100))


class Product(db.Model):
    """
    Medical Equipment Inventory Model.
    Identified by unique Serial Number.
    """
    serial_number = db.Column(db.String(100), primary_key=True)
    
    name = db.Column(db.String(100), nullable=False)
    
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'))
    
    start_warranty_date = db.Column(db.Date)
    
    end_warranty_date = db.Column(db.Date)
    
    # History of technical visits for this asset
    shifts = db.relationship('Shift', backref='product', lazy=True)


class ShiftFile(db.Model):
    """
    Multi-Attachment Repository.
    Allows engineers to upload multiple Service Reports.

    filename = randomized/safe server filename.
    original_filename = clean client-facing upload filename.
    """
    id = db.Column(db.Integer, primary_key=True)

    shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=False)

    filename = db.Column(db.String(200), nullable=False)

    original_filename = db.Column(db.String(200), nullable=True)

    uploaded_at = db.Column(db.DateTime, default=get_manila_time)




class ShiftEngineer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'))
    engineer_id = db.Column(db.Integer, db.ForeignKey('engineer.id'))


class Shift(db.Model):
    """
    The Core Service Record Model.
    Tracks technical site visits and progress.
    """
    id = db.Column(db.Integer, primary_key=True)
    
    title = db.Column(db.String(100), nullable=False)
    
    start_time = db.Column(db.DateTime, nullable=False)
    
    end_time = db.Column(db.DateTime, nullable=False)
    
    engineer_id = db.Column(db.Integer, db.ForeignKey('engineer.id'), nullable=False)
    
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'))
    
    product_id = db.Column(db.String(100), db.ForeignKey('product.serial_number'))
    
    status = db.Column(db.String(50), default='In Progress')
    
    # Link to the NEW multi-file table
    files = db.relationship(
        'ShiftFile', 
        backref='shift', 
        lazy=True, 
        cascade="all, delete-orphan"
    )
    
    created_at = db.Column(db.DateTime, default=get_manila_time)
    group_id = db.Column(db.String(50), index=True)

    # Linked time override support:
    # A child override keeps its own date/time/engineer assignment, but remains linked
    # to the original parent schedule for shared job details such as client/task/product/status/files.
    parent_shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=True, index=True)
    override_engineer_id = db.Column(db.Integer, db.ForeignKey('engineer.id'), nullable=True, index=True)
    override_kind = db.Column(db.String(30), nullable=True, index=True)


class TsrKnowledgeEntry(db.Model):
    """Reusable TSR complaint/action knowledge captured from engineer TSR drafts."""
    id = db.Column(db.Integer, primary_key=True)
    complaint = db.Column(db.Text, nullable=False)
    actions_taken = db.Column(db.Text, nullable=False)
    client_name = db.Column(db.String(200), nullable=True, index=True)
    product_name = db.Column(db.String(200), nullable=True, index=True)
    serial_number = db.Column(db.String(120), nullable=True, index=True)
    task = db.Column(db.String(200), nullable=True, index=True)
    service_category = db.Column(db.String(100), nullable=True)
    engineer_name = db.Column(db.String(120), nullable=True)
    source_shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=True, index=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True, index=True)
    created_at = db.Column(db.DateTime, default=get_manila_time, index=True)


class OnlineTsrSubmission(db.Model):
    """Phase 3 online TSR save intake record before PDF generation/attachment."""
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=False, index=True)
    tsr_number = db.Column(db.String(120), nullable=True, index=True)
    client_name = db.Column(db.String(200), nullable=True)
    product_name = db.Column(db.String(200), nullable=True)
    serial_number = db.Column(db.String(120), nullable=True, index=True)
    submitted_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True, index=True)
    submitted_by_name = db.Column(db.String(120), nullable=True)
    status = db.Column(db.String(40), default='received', index=True)
    payload_json = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=get_manila_time, index=True)


_tsr_knowledge_entry_table_ready = False
_online_tsr_submission_table_ready = False


def ensure_tsr_knowledge_entry_table():
    """Create the TSR knowledge table on existing live SQLite databases if missing."""
    global _tsr_knowledge_entry_table_ready
    if _tsr_knowledge_entry_table_ready:
        return
    try:
        TsrKnowledgeEntry.__table__.create(db.engine, checkfirst=True)
        _tsr_knowledge_entry_table_ready = True
    except Exception as table_error:
        print(f"[TSR-KB] Unable to ensure tsr_knowledge_entry table: {table_error}", flush=True)
        raise


def ensure_online_tsr_submission_table():
    """Create the Phase 3 online TSR submission table on live SQLite if missing."""
    global _online_tsr_submission_table_ready
    if _online_tsr_submission_table_ready:
        return
    try:
        OnlineTsrSubmission.__table__.create(db.engine, checkfirst=True)
        _online_tsr_submission_table_ready = True
    except Exception as table_error:
        print(f"[ONLINE-TSR] Unable to ensure online_tsr_submission table: {table_error}", flush=True)
        raise


# --- ADVANCED DUPLICATE ENGINE ---
ADDRESS_ALIASES = {
    "bgc": ["bgc", "bgc city", "bonifacio global city", "taguig bgc"],
    "makati": ["makati", "makati city"],
    "qc": ["quezon city", "qc"],
}

def normalize_address(addr):
    if not addr:
        return ""
    addr = addr.lower().strip()
    for k, vals in ADDRESS_ALIASES.items():
        for v in vals:
            if v in addr:
                return k
    return addr

def similarity(a,b):
    from difflib import SequenceMatcher
    return SequenceMatcher(None,a,b).ratio()

# --- SYSTEM HELPERS & DATA SANITIZATION ---

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def clean_int(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str in ["", "null", "None", "undefined", "NaN"]:
        return None
    try:
        return int(val_str)
    except (ValueError, TypeError):
        return None


def clean_str(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str in ["", "null", "None", "undefined"]:
        return None
    return val_str


def parse_date(date_str):
    if not date_str:
        return None
    val_str = str(date_str).strip()
    if val_str in ["", "null", "None"]:
        return None
    try:
        return datetime.strptime(val_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def generate_acronym(text):
    if not text:
        return ""
    clean = re.sub(r'[^a-zA-Z\s]', '', text).strip().upper()
    words = clean.split()
    if not words:
        return ""
    if len(words) == 1:
        return words[0]
    return "".join([w[0] for w in words])


def check_for_duplicate_client(new_name, new_addr):
    name_up = new_name.strip().upper()
    acr_incoming = generate_acronym(new_name)
    addr_norm = normalize_address(new_addr)

    for c in Client.query.all():
        existing_name = c.name.strip().upper()
        existing_acr = generate_acronym(c.name)
        existing_addr = normalize_address(c.address)

        if addr_norm != existing_addr:
            continue

        if name_up == existing_name or name_up in existing_name or existing_name in name_up:
            return c

        if acr_incoming == existing_acr:
            return c

        if similarity(name_up, existing_name) > 0.85:
            return c

    return None


def log_activity(action):
    """ Helper v5.4.3: Records live system events with auto-commit fix """
    if current_user and current_user.is_authenticated:
        db.session.add(ActivityLog(user=current_user.username.capitalize(), action=action))
        db.session.commit() # FIXED: Ensure log is saved immediately


def generate_temp_password(length=14):
    """Generates a one-time temporary password for first-run/local testing bootstrap."""
    alphabet = string.ascii_letters + string.digits + "!@#$%&*"
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def bootstrap_user(username, role, plain_password=None):
    """
    Creates a missing bootstrap user safely.
    - Uses SEED_PASSWORD_<USERNAME> from .env when provided.
    - Otherwise generates a random temporary password for local first-run testing.
    - Always forces password change on first login.
    """
    existing = User.query.filter_by(username=username).first()
    if existing:
        return existing, None, False

    temp_password = plain_password or generate_temp_password()
    user = User(
        username=username,
        password=generate_password_hash(temp_password),
        role=role,
        must_change_password=True
    )
    db.session.add(user)
    db.session.flush()
    return user, temp_password, True


_emergency_superadmin_bootstrap_ready = False


def ensure_emergency_superadmin_from_env():
    """Create/reset Jonamar superadmin only when SEED_PASSWORD_JONAMAR is present.

    This is for fresh Railway volume recovery when /data/scheduler.db is empty.
    Remove SEED_PASSWORD_JONAMAR from Railway Variables after login works.
    """
    global _emergency_superadmin_bootstrap_ready

    if _emergency_superadmin_bootstrap_ready:
        return

    seed_password = os.environ.get('SEED_PASSWORD_JONAMAR')
    if not seed_password:
        _emergency_superadmin_bootstrap_ready = True
        return

    username = 'jonamar'
    user = User.query.filter_by(username=username).first()

    if user:
        user.password = generate_password_hash(seed_password)
        user.role = 'superadmin'
        user.must_change_password = True
        print("[BOOTSTRAP] Reset jonamar superadmin password from SEED_PASSWORD_JONAMAR.", flush=True)
    else:
        db.session.add(User(
            username=username,
            password=generate_password_hash(seed_password),
            role='superadmin',
            must_change_password=True
        ))
        print("[BOOTSTRAP] Created jonamar superadmin from SEED_PASSWORD_JONAMAR.", flush=True)

    db.session.commit()
    _emergency_superadmin_bootstrap_ready = True


# --- AUTHORIZATION POLICY HELPERS ---

DEVELOPER_SUPERADMIN_USERNAME = 'jonamar'
MANAGER_USERNAMES = {'rodito', 'robert'}
SCHEDULER_USERNAMES = {'diary', 'hanna'}
SUPERADMIN_DISPLAY_USERNAMES = {'jonamar', 'robert'}
SUPERADMIN_USERNAMES = {DEVELOPER_SUPERADMIN_USERNAME} | MANAGER_USERNAMES | SCHEDULER_USERNAMES
PROTECTED_PASSWORD_USERNAMES = {DEVELOPER_SUPERADMIN_USERNAME}
REGIONAL_ADMIN_USERNAME = 'kevin'
REGIONAL_ADMIN_EMPLOYEE_ID = '15-148'
REGIONAL_ADMIN_BRANCHES = {'Cebu', 'Davao'}


def _username_of(user=None):
    target = user or current_user
    return (getattr(target, 'username', '') or '').strip().lower()


def get_display_role(user):
    """Business-facing role label for Settings/UI without weakening backend authority."""
    username = _username_of(user)
    raw_role = (getattr(user, 'role', '') or '').strip().lower()

    if username == DEVELOPER_SUPERADMIN_USERNAME:
        return 'Superadmin'
    if username == 'robert':
        return 'Superadmin'
    if username == 'rodito':
        return 'Manager'
    if username in SCHEDULER_USERNAMES:
        return 'Scheduler'
    if username == REGIONAL_ADMIN_USERNAME:
        return 'Administrator'
    if raw_role == 'engineer':
        return 'Engineer'
    if raw_role == 'regional_admin':
        return 'Administrator'
    if raw_role == 'superadmin':
        return 'Superadmin'
    return raw_role.capitalize() if raw_role else 'User'


def is_scheduler_user(user=None):
    target = user or current_user
    return bool(
        target and
        getattr(target, 'is_authenticated', False) and
        getattr(target, 'role', None) == 'superadmin' and
        _username_of(target) in SCHEDULER_USERNAMES
    )


def is_superadmin_user(user=None):
    target = user or current_user
    return bool(
        target and
        getattr(target, 'is_authenticated', False) and
        getattr(target, 'role', None) == 'superadmin' and
        _username_of(target) in SUPERADMIN_USERNAMES
    )


def is_regional_admin_user(user=None):
    target = user or current_user
    if not (
        target and
        getattr(target, 'is_authenticated', False) and
        getattr(target, 'role', None) == 'regional_admin' and
        _username_of(target) == REGIONAL_ADMIN_USERNAME
    ):
        return False

    # Unique identifier for admin Kevin.
    # If the profile exists, it must be the intended Kevin Bautista employee record.
    profile = getattr(target, 'engineer_profile', None)
    if profile:
        return getattr(profile, 'employee_id', None) == REGIONAL_ADMIN_EMPLOYEE_ID

    # Allow the account during first-run/bootstrap before the Engineer row is created.
    return True


def is_admin_authorized(user=None):
    """System-management authorization. The old generic admin role is intentionally not accepted."""
    return is_superadmin_user(user) or is_regional_admin_user(user)


def can_reset_password_for_user(target_user, actor=None):
    """Password reset policy for the new hierarchy."""
    actor = actor or current_user
    actor_username = _username_of(actor)
    target_username = _username_of(target_user)
    target_role = (getattr(target_user, 'role', '') or '').strip().lower()

    if not target_user:
        return False, 'No account'

    # Nobody can change Jonamar/developer-superadmin password except Jonamar through own-password flow.
    if target_username in PROTECTED_PASSWORD_USERNAMES and actor_username != target_username:
        return False, 'This protected account password cannot be changed.'

    if not is_admin_authorized(actor):
        return False, 'Denied'

    # Kevin keeps current admin scope but should not reset protected core accounts.
    if is_regional_admin_user(actor):
        if target_role in {'superadmin', 'regional_admin'}:
            return False, 'You are not allowed to change this password.'
        return True, None

    # Schedulers cannot change passwords of superadmins or engineers.
    if is_scheduler_user(actor):
        if target_role in {'superadmin', 'engineer'}:
            return False, 'Schedulers cannot change passwords of superadmins or engineers.'
        return True, None

    # Managers/Robert have full authority except Jonamar handled above.
    if is_superadmin_user(actor):
        return True, None

    return False, 'Denied'


def denied(message='Denied'):
    return jsonify({'message': message}), 403


def get_shift_assigned_engineer_ids(shift):
    if not shift:
        return []
    linked_ids = [se.engineer_id for se in ShiftEngineer.query.filter_by(shift_id=shift.id).all()]
    if linked_ids:
        return linked_ids
    return [shift.engineer_id] if shift.engineer_id else []


def can_modify_schedule_for_engineer_ids(engineer_ids):
    """
    Calendar write policy:
    - named superadmins: all branches
    - Kevin/regional_admin: Cebu and Davao only, never Manila
    - engineer: own linked engineer profile only
    """
    cleaned_ids = [eid for eid in (engineer_ids or []) if eid]

    if is_superadmin_user():
        return True

    if is_regional_admin_user():
        if not cleaned_ids:
            return False
        engineers = [db.session.get(Engineer, eid) for eid in cleaned_ids]
        return all(e and e.branch in REGIONAL_ADMIN_BRANCHES for e in engineers)

    if getattr(current_user, 'role', None) == 'engineer':
        my_profile = current_user.engineer_profile
        return bool(my_profile and cleaned_ids and all(eid == my_profile.id for eid in cleaned_ids))

    return False


def can_create_schedule_for_engineer_ids(engineer_ids):
    """Schedule creation policy.

    Admins keep their normal branch authority.

    Engineers may create a multi-engineer schedule only when their own linked
    Engineer profile is included in the selected team. Teammates may be from
    any branch. This matches the timeline rule where engineers can start only
    from their own row, then add any teammate inside the modal.
    """
    cleaned_ids = [eid for eid in (engineer_ids or []) if eid]

    if is_superadmin_user():
        return True

    if is_regional_admin_user():
        if not cleaned_ids:
            return False
        engineers = [db.session.get(Engineer, eid) for eid in cleaned_ids]
        return all(e and e.branch in REGIONAL_ADMIN_BRANCHES for e in engineers)

    if getattr(current_user, 'role', None) == 'engineer':
        my_profile = current_user.engineer_profile
        return bool(my_profile and cleaned_ids and my_profile.id in cleaned_ids)

    return False


def can_modify_schedule_shift(shift):
    return can_modify_schedule_for_engineer_ids(get_shift_assigned_engineer_ids(shift))


def get_current_user_engineer_id():
    """Return the logged-in engineer profile ID when the user is an engineer."""
    if not (
        current_user and
        getattr(current_user, 'is_authenticated', False) and
        getattr(current_user, 'role', None) == 'engineer'
    ):
        return None

    profile = getattr(current_user, 'engineer_profile', None)
    return profile.id if profile else None


def is_current_engineer_assigned_to_shift(shift):
    """Allow engineers to work on schedules where they are one of the assigned staff."""
    my_engineer_id = get_current_user_engineer_id()
    if not shift or not my_engineer_id:
        return False

    assigned_ids = get_shift_assigned_engineer_ids(shift)
    if my_engineer_id in assigned_ids:
        return True

    return bool(
        getattr(shift, 'override_engineer_id', None) == my_engineer_id or
        getattr(shift, 'engineer_id', None) == my_engineer_id
    )


def can_work_on_existing_schedule_shift(shift):
    """Permission for non-destructive schedule work such as TSR upload/status/time edit.

    Admins keep normal branch authority. Engineers may update a schedule only
    when their linked Engineer profile is assigned to that schedule.
    """
    if not shift:
        return False

    if is_superadmin_user() or is_regional_admin_user():
        return can_modify_schedule_shift(shift)

    return is_current_engineer_assigned_to_shift(shift)


def can_submit_update_engineer_ids_for_scope(master_shift, requested_engineer_ids, edit_scope, override_engineer_id=None):
    """Validate requested engineer IDs for update_shift.

    Engineers can save/upload/complete schedules assigned to them without being
    blocked by other assigned engineers on the same job. They still cannot add,
    remove, or reassign other engineers. For custom time override, they can only
    modify their own engineer row.
    """
    if is_superadmin_user() or is_regional_admin_user():
        return can_modify_schedule_for_engineer_ids(requested_engineer_ids)

    my_engineer_id = get_current_user_engineer_id()
    if not my_engineer_id or not master_shift:
        return False

    requested_ids = [eid for eid in (requested_engineer_ids or []) if eid]
    if edit_scope == 'engineer_day_time_override':
        target_id = override_engineer_id or (requested_ids[0] if requested_ids else None)
        return bool(target_id == my_engineer_id and set(requested_ids or [target_id]) == {my_engineer_id})

    if not is_current_engineer_assigned_to_shift(master_shift):
        return False

    existing_ids = get_shift_assigned_engineer_ids(master_shift)
    return set(map(int, requested_ids)) == set(map(int, existing_ids))


def get_schedule_email_recipients(engineer_ids):
    """Return unique assigned engineers with valid email addresses."""
    recipients = []
    seen_emails = set()

    for engineer_id in engineer_ids or []:
        engineer = db.session.get(Engineer, engineer_id)
        email_addr = clean_str(getattr(engineer, 'email', None)) if engineer else None

        if not engineer or not email_addr:
            continue

        normalized = email_addr.lower()
        if normalized in seen_emails:
            continue

        seen_emails.add(normalized)
        recipients.append(engineer)

    return recipients


def format_schedule_email_text(shift, assigned_engineers, created_by):
    """Plain-text email body for newly created schedules."""
    client_name = shift.client.name if shift and shift.client else "N/A"
    product_name = shift.product.name if shift and shift.product else "N/A"
    product_serial = shift.product.serial_number if shift and shift.product else ""
    product_label = product_name
    if product_serial:
        product_label = f"{product_serial} / {product_name}"

    assigned_names = ", ".join([engineer.name for engineer in assigned_engineers]) or "N/A"

    return (
        "New schedule assigned\n\n"
        f"Date: {shift.start_time.strftime('%B %d, %Y')}\n"
        f"Time: {shift.start_time.strftime('%I:%M %p')} - {shift.end_time.strftime('%I:%M %p')}\n"
        f"Task: {shift.title}\n"
        f"Client: {client_name}\n"
        f"Product: {product_label}\n"
        f"Status: {shift.status or 'In Progress'}\n"
        f"Assigned engineer(s): {assigned_names}\n"
        f"Assigned by: {created_by}\n\n"
        "Please check the Medical Service Scheduler for full details."
    )


def format_schedule_email_html(shift, assigned_engineers, created_by):
    """HTML email body for newly created schedules."""
    client_name = shift.client.name if shift and shift.client else "N/A"
    product_name = shift.product.name if shift and shift.product else "N/A"
    product_serial = shift.product.serial_number if shift and shift.product else ""
    product_label = product_name
    if product_serial:
        product_label = f"{product_serial} / {product_name}"

    assigned_names = ", ".join([engineer.name for engineer in assigned_engineers]) or "N/A"

    rows = [
        ("Date", shift.start_time.strftime('%B %d, %Y')),
        ("Time", f"{shift.start_time.strftime('%I:%M %p')} - {shift.end_time.strftime('%I:%M %p')}"),
        ("Task", shift.title),
        ("Client", client_name),
        ("Product", product_label),
        ("Status", shift.status or 'In Progress'),
        ("Assigned engineer(s)", assigned_names),
        ("Assigned by", created_by)
    ]

    detail_rows = "".join(
        f"<tr><td style='padding:6px 10px;font-weight:700;border-bottom:1px solid #e5e7eb;'>{html.escape(label)}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #e5e7eb;'>{html.escape(str(value or ''))}</td></tr>"
        for label, value in rows
    )

    return f"""
    <div style="font-family:Arial,sans-serif;color:#111827;line-height:1.5;">
        <h2 style="margin:0 0 12px;color:#0f172a;">New schedule assigned</h2>
        <table style="border-collapse:collapse;border:1px solid #e5e7eb;min-width:420px;">
            {detail_rows}
        </table>
        <p style="margin-top:16px;color:#374151;">Please check the Medical Service Scheduler for full details.</p>
    </div>
    """




def format_schedule_event_email_text(action_label, shift, assigned_engineers, actor_name, extra_note=None):
    """Plain-text email body for schedule create/update/move/delete notifications."""
    client_name = shift.client.name if shift and shift.client else "N/A"
    product_name = shift.product.name if shift and shift.product else "N/A"
    product_serial = shift.product.serial_number if shift and shift.product else ""
    product_label = product_name
    if product_serial:
        product_label = f"{product_serial} / {product_name}"

    assigned_names = ", ".join([engineer.name for engineer in assigned_engineers]) or "N/A"

    body = (
        f"Schedule {action_label}\n\n"
        f"Date: {shift.start_time.strftime('%B %d, %Y')}\n"
        f"Time: {shift.start_time.strftime('%I:%M %p')} - {shift.end_time.strftime('%I:%M %p')}\n"
        f"Task: {shift.title}\n"
        f"Client: {client_name}\n"
        f"Product: {product_label}\n"
        f"Status: {shift.status or 'In Progress'}\n"
        f"Assigned engineer(s): {assigned_names}\n"
        f"Updated by: {actor_name}\n"
    )

    if extra_note:
        body += f"Note: {extra_note}\n"

    body += "\nPlease check the Medical Service Scheduler for full details."
    return body


def format_schedule_event_email_html(action_label, shift, assigned_engineers, actor_name, extra_note=None):
    """HTML email body for schedule create/update/move/delete notifications."""
    client_name = shift.client.name if shift and shift.client else "N/A"
    product_name = shift.product.name if shift and shift.product else "N/A"
    product_serial = shift.product.serial_number if shift and shift.product else ""
    product_label = product_name
    if product_serial:
        product_label = f"{product_serial} / {product_name}"

    assigned_names = ", ".join([engineer.name for engineer in assigned_engineers]) or "N/A"

    rows = [
        ("Date", shift.start_time.strftime('%B %d, %Y')),
        ("Time", f"{shift.start_time.strftime('%I:%M %p')} - {shift.end_time.strftime('%I:%M %p')}"),
        ("Task", shift.title),
        ("Client", client_name),
        ("Product", product_label),
        ("Status", shift.status or 'In Progress'),
        ("Assigned engineer(s)", assigned_names),
        ("Updated by", actor_name)
    ]

    if extra_note:
        rows.append(("Note", extra_note))

    detail_rows = "".join(
        f"<tr><td style='padding:6px 10px;font-weight:700;border-bottom:1px solid #e5e7eb;'>{html.escape(label)}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #e5e7eb;'>{html.escape(str(value or ''))}</td></tr>"
        for label, value in rows
    )

    return f"""
    <div style="font-family:Arial,sans-serif;color:#111827;line-height:1.5;">
        <h2 style="margin:0 0 12px;color:#0f172a;">Schedule {html.escape(action_label)}</h2>
        <table style="border-collapse:collapse;border:1px solid #e5e7eb;min-width:420px;">
            {detail_rows}
        </table>
        <p style="margin-top:16px;color:#374151;">Please check the Medical Service Scheduler for full details.</p>
    </div>
    """


def send_schedule_event_notification_async(app_obj, shift_id, engineer_ids, action_label, actor_username, extra_note=None):
    """Send schedule event notifications in the background for create/edit/move/delete actions."""
    def worker():
        with app_obj.app_context():
            try:
                shift = db.session.get(Shift, shift_id)
                if not shift:
                    print(f"[EMAIL] {action_label} skipped: shift #{shift_id} not found.", flush=True)
                    return

                if not should_notify_schedule(shift):
                    print(
                        f"[EMAIL] {action_label} skipped: shift #{shift_id} is in the past ({shift.start_time.date()}).",
                        flush=True
                    )
                    return

                assigned_engineers = get_schedule_email_recipients(engineer_ids)
                recipient_emails = [engineer.email for engineer in assigned_engineers if engineer.email]

                print(f"[EMAIL] {action_label} notification started for shift #{shift_id}", flush=True)
                print(f"[EMAIL] Preparing notification for engineer IDs: {engineer_ids}", flush=True)

                if not recipient_emails:
                    print(f"[EMAIL] {action_label} skipped: selected engineer(s) have no email address.", flush=True)
                    return

                actor_name = (actor_username or 'Scheduler').capitalize()
                subject_date = shift.start_time.strftime('%b %d, %Y')
                subject = f"Schedule {action_label} - {subject_date} - {shift.title}"

                text_body = format_schedule_event_email_text(action_label, shift, assigned_engineers, actor_name, extra_note)
                html_body = format_schedule_event_email_html(action_label, shift, assigned_engineers, actor_name, extra_note)

                email_sent, email_message = send_email_notification(recipient_emails, subject, text_body, html_body)
                print(f"[EMAIL] {action_label} final result: sent={email_sent} | {email_message}", flush=True)

                if email_sent:
                    db.session.add(ActivityLog(
                        user=actor_name,
                        action=f"Sent schedule {action_label.lower()} email notification: {shift.title}"
                    ))
                    db.session.commit()
            except Exception as worker_error:
                print(f"[EMAIL] {action_label} worker failed: {worker_error}", flush=True)

    email_thread = threading.Thread(target=worker, daemon=True)
    email_thread.start()
    print(f"[EMAIL] {action_label} background email thread queued.", flush=True)


def send_schedule_deleted_notification_async(app_obj, snapshot, recipient_engineer_ids, actor_username):
    """Send delete notification from a snapshot because the Shift row is removed before email sends."""
    def worker():
        with app_obj.app_context():
            try:
                if not should_notify_snapshot(snapshot):
                    print(
                        f"[EMAIL] Deleted skipped: snapshot #{snapshot.get('id')} is in the past ({snapshot.get('date_iso') or snapshot.get('date_label')}).",
                        flush=True
                    )
                    return

                assigned_engineers = get_schedule_email_recipients(recipient_engineer_ids)
                recipient_emails = [engineer.email for engineer in assigned_engineers if engineer.email]

                print(f"[EMAIL] Deleted notification started for shift snapshot #{snapshot.get('id')}", flush=True)

                if not recipient_emails:
                    print("[EMAIL] Deleted skipped: selected engineer(s) have no email address.", flush=True)
                    return

                actor_name = (actor_username or 'Scheduler').capitalize()
                assigned_names = ", ".join([engineer.name for engineer in assigned_engineers]) or "N/A"

                subject = f"Schedule Deleted - {snapshot.get('date_label')} - {snapshot.get('title')}"

                text_body = (
                    "Schedule Deleted\n\n"
                    f"Date: {snapshot.get('date_label')}\n"
                    f"Time: {snapshot.get('time_label')}\n"
                    f"Task: {snapshot.get('title')}\n"
                    f"Client: {snapshot.get('client_name')}\n"
                    f"Product: {snapshot.get('product_name')}\n"
                    f"Assigned engineer(s): {assigned_names}\n"
                    f"Deleted by: {actor_name}\n\n"
                    "This schedule was removed from the Medical Service Scheduler."
                )

                rows = [
                    ("Date", snapshot.get('date_label')),
                    ("Time", snapshot.get('time_label')),
                    ("Task", snapshot.get('title')),
                    ("Client", snapshot.get('client_name')),
                    ("Product", snapshot.get('product_name')),
                    ("Assigned engineer(s)", assigned_names),
                    ("Deleted by", actor_name)
                ]

                detail_rows = "".join(
                    f"<tr><td style='padding:6px 10px;font-weight:700;border-bottom:1px solid #e5e7eb;'>{html.escape(str(label))}</td>"
                    f"<td style='padding:6px 10px;border-bottom:1px solid #e5e7eb;'>{html.escape(str(value or ''))}</td></tr>"
                    for label, value in rows
                )

                html_body = f"""
                <div style="font-family:Arial,sans-serif;color:#111827;line-height:1.5;">
                    <h2 style="margin:0 0 12px;color:#991b1b;">Schedule Deleted</h2>
                    <table style="border-collapse:collapse;border:1px solid #e5e7eb;min-width:420px;">
                        {detail_rows}
                    </table>
                    <p style="margin-top:16px;color:#374151;">This schedule was removed from the Medical Service Scheduler.</p>
                </div>
                """

                email_sent, email_message = send_email_notification(recipient_emails, subject, text_body, html_body)
                print(f"[EMAIL] Deleted final result: sent={email_sent} | {email_message}", flush=True)

                if email_sent:
                    db.session.add(ActivityLog(
                        user=actor_name,
                        action=f"Sent schedule deleted email notification: {snapshot.get('title')}"
                    ))
                    db.session.commit()
            except Exception as worker_error:
                print(f"[EMAIL] Deleted worker failed: {worker_error}", flush=True)

    email_thread = threading.Thread(target=worker, daemon=True)
    email_thread.start()
    print("[EMAIL] Deleted background email thread queued.", flush=True)



def normalize_email_list(email_list):
    """Return unique cleaned email addresses while preserving order."""
    clean_emails = []
    seen = set()

    for email_addr in email_list or []:
        email_addr = clean_str(email_addr)
        if not email_addr:
            continue

        normalized = email_addr.lower()
        if normalized in seen:
            continue

        seen.add(normalized)
        clean_emails.append(email_addr)

    return clean_emails


def get_static_tsr_client_cc_emails():
    """Return static backend-controlled CC recipients for TSR client emails."""
    return normalize_email_list(STATIC_TSR_CLIENT_CC_EMAILS)


def get_current_user_email_for_tsr_cc():
    """Return the logged-in sender email for TSR client CC.

    Engineers use their linked Engineer profile email.
    Scheduler accounts use their known company email.
    """
    if not current_user or not getattr(current_user, 'is_authenticated', False):
        return None

    profile = getattr(current_user, 'engineer_profile', None)
    email_addr = clean_str(getattr(profile, 'email', None)) if profile else None

    if not email_addr:
        scheduler_email_map = {
            'diary': 'diary@shimadzu.com.ph',
            'hanna': 'hanna@shimadzu.com.ph',
            'kevin': 'kevin@shimadzu.com.ph'
        }
        email_addr = scheduler_email_map.get(_username_of(current_user))

    if not email_addr:
        return None

    return email_addr.lower()


def get_tsr_client_cc_emails_for_current_sender():
    """Return static TSR CC recipients plus the current sender email."""
    cc_emails = get_static_tsr_client_cc_emails()
    sender_email = get_current_user_email_for_tsr_cc()

    if sender_email:
        cc_emails.append(sender_email)

    return normalize_email_list(cc_emails)


def send_brevo_email_notification(to_emails, subject, text_body, html_body=None):
    """Send email through Brevo API over HTTPS so it works when SMTP ports are blocked."""
    print("[EMAIL] Brevo API notification requested.", flush=True)

    if not app.config.get('EMAIL_NOTIFICATIONS_ENABLED'):
        print("[EMAIL] Skipped: EMAIL_NOTIFICATIONS_ENABLED=false", flush=True)
        return False, 'Email notifications disabled.'

    api_key = app.config.get('BREVO_API_KEY')
    if not api_key:
        print("[EMAIL] Skipped: BREVO_API_KEY is not configured in .env", flush=True)
        return False, 'BREVO_API_KEY is not configured.'

    clean_recipients = []
    seen = set()
    for email_addr in to_emails or []:
        email_addr = clean_str(email_addr)
        if not email_addr:
            continue
        normalized = email_addr.lower()
        if normalized not in seen:
            seen.add(normalized)
            clean_recipients.append(email_addr)

    if not clean_recipients:
        return False, 'No recipient email addresses.'

    sender_email = app.config.get('SMTP_SENDER_EMAIL') or app.config.get('SMTP_USERNAME')
    sender_name = app.config.get('SMTP_SENDER_NAME') or 'SPC-Medical Service Scheduler'

    payload = {
        'sender': {
            'name': sender_name,
            'email': sender_email
        },
        'to': [{'email': email_addr} for email_addr in clean_recipients],
        'subject': subject,
        'textContent': text_body
    }

    if html_body:
        payload['htmlContent'] = html_body

    headers = {
        'accept': 'application/json',
        'api-key': api_key,
        'content-type': 'application/json'
    }

    try:
        print(f"[EMAIL] Brevo sending via HTTPS API to {', '.join(clean_recipients)}", flush=True)
        response = requests.post(
            app.config.get('BREVO_API_URL'),
            headers=headers,
            json=payload,
            timeout=app.config.get('BREVO_TIMEOUT')
        )

        if 200 <= response.status_code < 300:
            print(f"[EMAIL] Brevo success: {response.status_code} {response.text}", flush=True)
            return True, f"Brevo email sent to {len(clean_recipients)} recipient(s)."

        print(f"[EMAIL] Brevo failed: {response.status_code} {response.text}", flush=True)
        return False, f"Brevo failed: {response.status_code} {response.text}"

    except Exception as brevo_error:
        print(f"[EMAIL] Brevo exception: {brevo_error}", flush=True)
        return False, str(brevo_error)



def send_brevo_email_with_attachments(to_emails, subject, text_body, html_body=None, attachments=None, cc_emails=None):
    """Send email through Brevo API with optional file attachments."""
    print("[EMAIL-CLIENT] Brevo API client email with attachments requested.", flush=True)

    if not app.config.get('EMAIL_NOTIFICATIONS_ENABLED'):
        return False, 'Email notifications disabled.'

    api_key = app.config.get('BREVO_API_KEY')
    if not api_key:
        return False, 'BREVO_API_KEY is not configured.'

    clean_recipients = []
    seen = set()
    for email_addr in to_emails or []:
        email_addr = clean_str(email_addr)
        if not email_addr:
            continue
        normalized = email_addr.lower()
        if normalized not in seen:
            seen.add(normalized)
            clean_recipients.append(email_addr)

    if not clean_recipients:
        return False, 'No recipient email addresses.'

    sender_email = app.config.get('SMTP_SENDER_EMAIL') or app.config.get('SMTP_USERNAME')
    sender_name = app.config.get('SMTP_SENDER_NAME') or 'SPC-Medical Service Scheduler'

    clean_cc_recipients = normalize_email_list(cc_emails)

    payload = {
        'sender': {
            'name': sender_name,
            'email': sender_email
        },
        'to': [{'email': email_addr} for email_addr in clean_recipients],
        'subject': subject,
        'textContent': text_body
    }

    if clean_cc_recipients:
        payload['cc'] = [{'email': email_addr} for email_addr in clean_cc_recipients]

    if html_body:
        payload['htmlContent'] = html_body

    attachment_payload = []
    for attachment in attachments or []:
        filename = attachment.get('display_name') or attachment.get('original_filename') or attachment.get('filename')
        file_path = attachment.get('path')
        if not filename or not file_path or not os.path.exists(file_path):
            continue

        with open(file_path, 'rb') as file_obj:
            attachment_payload.append({
                'name': filename,
                'content': base64.b64encode(file_obj.read()).decode('ascii')
            })

    if attachment_payload:
        payload['attachment'] = attachment_payload

    headers = {
        'accept': 'application/json',
        'api-key': api_key,
        'content-type': 'application/json'
    }

    try:
        response = requests.post(
            app.config.get('BREVO_API_URL'),
            headers=headers,
            json=payload,
            timeout=app.config.get('BREVO_TIMEOUT')
        )

        if 200 <= response.status_code < 300:
            return True, f"Brevo client email sent to {len(clean_recipients)} recipient(s)."

        return False, f"Brevo failed: {response.status_code} {response.text}"

    except Exception as brevo_error:
        return False, str(brevo_error)


def send_email_with_attachments(to_emails, subject, text_body, html_body=None, attachments=None, cc_emails=None):
    """Send email with optional attachments through configured provider."""
    provider = app.config.get('EMAIL_PROVIDER', 'brevo')

    if provider == 'brevo':
        return send_brevo_email_with_attachments(to_emails, subject, text_body, html_body, attachments, cc_emails=cc_emails)

    if not app.config.get('EMAIL_NOTIFICATIONS_ENABLED'):
        return False, 'Email notifications disabled.'

    smtp_password = app.config.get('SMTP_PASSWORD')
    smtp_username = app.config.get('SMTP_USERNAME')

    if not smtp_password or not smtp_username:
        return False, 'SMTP_USERNAME or SMTP_PASSWORD is not configured.'

    clean_recipients = []
    seen = set()
    for email_addr in to_emails or []:
        email_addr = clean_str(email_addr)
        if not email_addr:
            continue
        normalized = email_addr.lower()
        if normalized not in seen:
            seen.add(normalized)
            clean_recipients.append(email_addr)

    if not clean_recipients:
        return False, 'No recipient email addresses.'

    clean_cc_recipients = normalize_email_list(cc_emails)

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = formataddr((app.config.get('SMTP_SENDER_NAME'), app.config.get('SMTP_SENDER_EMAIL')))
    msg['To'] = ", ".join(clean_recipients)
    if clean_cc_recipients:
        msg['Cc'] = ", ".join(clean_cc_recipients)
    msg.set_content(text_body)

    if html_body:
        msg.add_alternative(html_body, subtype='html')

    for attachment in attachments or []:
        filename = attachment.get('display_name') or attachment.get('original_filename') or attachment.get('filename')
        file_path = attachment.get('path')
        if not filename or not file_path or not os.path.exists(file_path):
            continue

        with open(file_path, 'rb') as file_obj:
            msg.add_attachment(
                file_obj.read(),
                maintype='application',
                subtype='octet-stream',
                filename=filename
            )

    try:
        with smtplib.SMTP(app.config.get('SMTP_HOST'), app.config.get('SMTP_PORT'), timeout=app.config.get('SMTP_TIMEOUT')) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(smtp_username, smtp_password)
            smtp.send_message(msg)
        return True, f"Email sent to {len(clean_recipients)} recipient(s)."
    except Exception as email_error:
        return False, str(email_error)


def send_email_notification(to_emails, subject, text_body, html_body=None):
    """Send schedule email through configured provider. Defaults to Brevo API."""
    provider = app.config.get('EMAIL_PROVIDER', 'brevo')

    if provider == 'brevo':
        return send_brevo_email_notification(to_emails, subject, text_body, html_body)

    print("[EMAIL] SMTP schedule email notification requested.", flush=True)

    if not app.config.get('EMAIL_NOTIFICATIONS_ENABLED'):
        print("[EMAIL] Skipped: EMAIL_NOTIFICATIONS_ENABLED=false", flush=True)
        return False, 'Email notifications disabled.'

    smtp_password = app.config.get('SMTP_PASSWORD')
    smtp_username = app.config.get('SMTP_USERNAME')

    if not smtp_password or not smtp_username:
        print("[EMAIL] Skipped: SMTP_USERNAME or SMTP_PASSWORD is not configured in .env", flush=True)
        return False, 'SMTP_USERNAME or SMTP_PASSWORD is not configured.'

    clean_recipients = []
    seen = set()
    for email_addr in to_emails or []:
        email_addr = clean_str(email_addr)
        if not email_addr:
            continue
        normalized = email_addr.lower()
        if normalized not in seen:
            seen.add(normalized)
            clean_recipients.append(email_addr)

    if not clean_recipients:
        return False, 'No recipient email addresses.'

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = formataddr((app.config.get('SMTP_SENDER_NAME'), app.config.get('SMTP_SENDER_EMAIL')))
    msg['To'] = ", ".join(clean_recipients)
    msg.set_content(text_body)

    if html_body:
        msg.add_alternative(html_body, subtype='html')

    try:
        with smtplib.SMTP(app.config.get('SMTP_HOST'), app.config.get('SMTP_PORT'), timeout=app.config.get('SMTP_TIMEOUT')) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            print(f"[EMAIL] Connecting to {app.config.get('SMTP_HOST')}:{app.config.get('SMTP_PORT')} as {smtp_username}", flush=True)
            smtp.login(smtp_username, smtp_password)
            smtp.send_message(msg)
        print(f"[EMAIL] Success: sent to {', '.join(clean_recipients)}", flush=True)
        return True, f"Email sent to {len(clean_recipients)} recipient(s)."
    except Exception as email_error:
        print(f"[EMAIL] Failed: {email_error}", flush=True)
        return False, str(email_error)


def notify_engineers_for_new_schedule(first_shift, engineer_ids):
    """Notify assigned engineers after a new schedule has been committed."""
    print(f"[EMAIL] Preparing notification for engineer IDs: {engineer_ids}", flush=True)

    if not first_shift:
        print("[EMAIL] Skipped: no shift to notify.", flush=True)
        return False, 'No shift to notify.'

    if not should_notify_schedule(first_shift):
        print(f"[EMAIL] Skipped: new schedule is in the past ({first_shift.start_time.date()}).", flush=True)
        return False, 'Past schedule notification skipped.'

    assigned_engineers = get_schedule_email_recipients(engineer_ids)
    recipient_emails = [engineer.email for engineer in assigned_engineers if engineer.email]

    if not recipient_emails:
        print("[EMAIL] Skipped: selected engineer(s) have no email address.", flush=True)
        return False, 'No assigned engineer email addresses.'

    print(f"[EMAIL] Recipients: {', '.join(recipient_emails)}", flush=True)

    created_by = current_user.username.capitalize() if current_user and current_user.is_authenticated else 'Scheduler'
    subject_date = first_shift.start_time.strftime('%b %d, %Y')
    subject = f"New Schedule Assigned - {subject_date} - {first_shift.title}"

    text_body = format_schedule_email_text(first_shift, assigned_engineers, created_by)
    html_body = format_schedule_email_html(first_shift, assigned_engineers, created_by)

    return send_email_notification(recipient_emails, subject, text_body, html_body)




def notify_engineers_for_new_schedule_async(app_obj, first_shift_id, engineer_ids, shift_title, date_label, actor_username):
    """Run schedule email notification in a background thread so schedule save never waits for SMTP."""
    def worker():
        with app_obj.app_context():
            try:
                shift = db.session.get(Shift, first_shift_id)
                if not shift:
                    print("[EMAIL] Background skipped: shift not found.", flush=True)
                    return

                if not should_notify_schedule(shift):
                    print(
                        f"[EMAIL] Background skipped: new schedule is in the past ({shift.start_time.date()}).",
                        flush=True
                    )
                    return

                # Build/sent email without relying on request/current_user context.
                assigned_engineers = get_schedule_email_recipients(engineer_ids)
                recipient_emails = [engineer.email for engineer in assigned_engineers if engineer.email]

                print(f"[EMAIL] Background started for shift #{first_shift_id}", flush=True)
                print(f"[EMAIL] Preparing notification for engineer IDs: {engineer_ids}", flush=True)

                if not recipient_emails:
                    print("[EMAIL] Background skipped: selected engineer(s) have no email address.", flush=True)
                    return

                print(f"[EMAIL] Recipients: {', '.join(recipient_emails)}", flush=True)

                created_by = (actor_username or 'Scheduler').capitalize()
                subject_date = shift.start_time.strftime('%b %d, %Y')
                subject = f"New Schedule Assigned - {subject_date} - {shift.title}"

                text_body = format_schedule_email_text(shift, assigned_engineers, created_by)
                html_body = format_schedule_email_html(shift, assigned_engineers, created_by)

                email_sent, email_message = send_email_notification(recipient_emails, subject, text_body, html_body)
                print(f"[EMAIL] Background final result: sent={email_sent} | {email_message}", flush=True)

                if email_sent:
                    db.session.add(ActivityLog(
                        user=created_by,
                        action=f"Sent schedule email notification: {shift_title} on {date_label}"
                    ))
                    db.session.commit()
            except Exception as worker_error:
                print(f"[EMAIL] Background worker failed: {worker_error}", flush=True)

    email_thread = threading.Thread(target=worker, daemon=True)
    email_thread.start()
    print("[EMAIL] Background email thread queued.", flush=True)


def is_mobile_request():
    """Best-effort mobile browser detection for engineer login routing."""
    user_agent = (request.headers.get('User-Agent') or '').lower()
    mobile_markers = (
        'iphone',
        'ipod',
        'android',
        'blackberry',
        'windows phone',
        'mobile',
        'opera mini',
        'opera mobi'
    )
    return any(marker in user_agent for marker in mobile_markers)

# --- AUTHENTICATION MODULE ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_rec = User.query.filter_by(username=username).first()
        if user_rec and check_password_hash(user_rec.password, password):
            login_user(user_rec)

            if user_rec.must_change_password:
                session['force_pw_change'] = True
            else:
                session.pop('force_pw_change', None)

            if getattr(user_rec, 'role', None) == 'engineer' and is_mobile_request():
                return redirect(url_for('timeline_page'))

            return redirect(url_for('dashboard_page'))
        flash('Invalid Credentials - Access Denied')
    return render_template('login.html')


@app.route('/logout')
def logout():
    # log_activity removed v5.4.5 per request
    logout_user()
    return redirect(url_for('login'))




# --- PWA / OFFLINE FIELD OPERATIONS FOUNDATION ---
# Step P1 + P2: Install shell, offline fallback, and improved critical-page caching.
# These routes are static-safe. Authenticated business data still requires login.

@app.route('/manifest.json')
def pwa_manifest():
    """PWA manifest for installable field-service experience."""
    manifest = {
        "name": "MEDICAL SERVICE Scheduler",
        "short_name": "MED SERVICE",
        "description": "Medical Service scheduling and offline field operations platform.",
        "start_url": "/timeline?source=pwa",
        "scope": "/",
        "display": "standalone",
        "background_color": "#f4f7f6",
        "theme_color": "#2c3e50",
        "orientation": "portrait",
        "categories": ["business", "productivity", "medical"],
        "icons": [
            {
                "src": "/pwa-icon.svg",
                "sizes": "any",
                "type": "image/svg+xml",
                "purpose": "any maskable"
            }
        ]
    }
    response = jsonify(manifest)
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.route('/pwa-icon.svg')
def pwa_icon():
    """Lightweight inline SVG app icon to avoid requiring static image files in early PWA phase."""
    svg = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 512 512">
<rect width="512" height="512" rx="112" fill="#2c3e50"/>
<circle cx="256" cy="256" r="150" fill="#ffffff"/>
<path d="M236 142h40v94h94v40h-94v94h-40v-94h-94v-40h94z" fill="#d63384"/>
<path d="M122 396h268" stroke="#ffffff" stroke-width="28" stroke-linecap="round" opacity=".92"/>
</svg>"""
    return Response(svg, mimetype="image/svg+xml", headers={"Cache-Control": "public, max-age=86400"})


@app.route('/offline')
def offline_page():
    """Offline fallback page shown by the service worker when navigation is unavailable."""
    html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Offline - MEDICAL SERVICE</title>
    <style>
        :root { color-scheme: light; }
        body {
            margin: 0;
            min-height: 100vh;
            display: grid;
            place-items: center;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
            background: linear-gradient(135deg, #2c3e50, #0f172a);
            color: #0f172a;
            padding: 18px;
        }
        .card {
            width: min(100%, 440px);
            background: #fff;
            border-radius: 24px;
            padding: 28px;
            box-shadow: 0 24px 70px rgba(0,0,0,.35);
            text-align: center;
        }
        .icon {
            width: 72px;
            height: 72px;
            margin: 0 auto 18px;
            border-radius: 22px;
            display: grid;
            place-items: center;
            background: #e7f1ff;
            color: #0d6efd;
            font-size: 34px;
            font-weight: 900;
        }
        h1 { font-size: 1.35rem; margin: 0 0 8px; }
        p { color: #64748b; line-height: 1.45; margin: 0 0 16px; }
        .hint {
            background: #f8fafc;
            border: 1px solid #e2e8f0;
            border-radius: 16px;
            padding: 14px;
            font-size: .88rem;
            font-weight: 700;
            color: #334155;
        }
        button {
            width: 100%;
            min-height: 48px;
            margin-top: 16px;
            border: 0;
            border-radius: 14px;
            background: #0d6efd;
            color: #fff;
            font-weight: 900;
            font-size: .95rem;
        }
    </style>
</head>
<body>
    <main class="card">
        <div class="icon">+</div>
        <h1>You are offline</h1>
        <p>MEDICAL SERVICE could not reach the server. Cached pages and saved offline TSR drafts remain available if you opened the app once while online.</p>
        <div class="hint">For field use, open Timeline or Offline TSR while online before going to a remote area.</div>
        <button onclick="location.reload()">Try Again</button>
    </main>
</body>
</html>"""
    return Response(html, mimetype="text/html", headers={"Cache-Control": "no-cache"})



def build_shift_datetime_bounds(start_date, end_date):
    """Return inclusive-exclusive datetime bounds for indexed Shift.start_time filtering.

    Avoids wrapping Shift.start_time in func.date(), which slows SQLite on Railway
    because indexes on start_time cannot be used efficiently.
    """
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())
    return start_dt, end_dt



@app.route('/offline-tsr')
@login_required
def offline_tsr_page():
    """Standalone offline-capable TSR form shell for field engineers. Precached by the service worker for mobile/desktop field use.

    Offline Storage Health is server-side gated to authorized admins only so
    engineers do not receive the diagnostic panel markup or controls.
    """
    return render_template(
        'offline_tsr.html',
        offline_storage_health_admin=is_admin_authorized()
    )


@app.route('/get_offline_tsr_schedule_options')
@login_required
def get_offline_tsr_schedule_options():
    """Return current user's field schedules for the standalone Offline TSR picker.

    The standalone TSR page stores these options locally so engineers can select
    a previously synced schedule even when the device is offline.
    """
    ensure_shift_file_original_filename_column()

    today = get_manila_today()
    start_window = today - timedelta(days=7)
    end_window = today + timedelta(days=45)
    start_dt, end_dt = build_shift_datetime_bounds(start_window, end_window)

    query = (
        Shift.query
        .options(
            joinedload(Shift.client),
            joinedload(Shift.product),
            joinedload(Shift.engineer),
            selectinload(Shift.files)
        )
        .filter(Shift.start_time >= start_dt)
        .filter(Shift.start_time < end_dt)
        .order_by(Shift.start_time.asc())
    )

    shifts = query.all()
    options = []

    for shift in shifts:
        # Standalone TSR is for real customer/equipment service work only.
        if not shift.client_id:
            continue

        assigned_ids = get_shift_assigned_engineer_ids(shift)

        if getattr(current_user, 'role', None) == 'engineer' and not is_current_engineer_assigned_to_shift(shift):
            continue

        if is_regional_admin_user():
            assigned_engineers = [db.session.get(Engineer, eid) for eid in assigned_ids if eid]
            assigned_engineers = [eng for eng in assigned_engineers if eng]
            if assigned_engineers and not any(eng.branch in REGIONAL_ADMIN_BRANCHES for eng in assigned_engineers):
                continue

        assigned_engineers = [db.session.get(Engineer, eid) for eid in assigned_ids if eid]
        assigned_engineers = [eng for eng in assigned_engineers if eng]

        if getattr(current_user, 'role', None) == 'engineer':
            serviced_engineer = getattr(current_user, 'engineer_profile', None) or shift.engineer
        else:
            serviced_engineer = shift.engineer or (assigned_engineers[0] if assigned_engineers else None)

        client = shift.client
        product = shift.product

        options.append({
            'id': shift.id,
            'date_iso': shift.start_time.date().isoformat() if shift.start_time else '',
            'date_label': shift.start_time.strftime('%b %d, %Y') if shift.start_time else '',
            'time_start': shift.start_time.strftime('%H:%M') if shift.start_time else '',
            'time_end': shift.end_time.strftime('%H:%M') if shift.end_time else '',
            'task': shift.title or '',
            'status': shift.status or '',
            'client_name': client.name if client else '',
            'client_address': client.address if client else '',
            'client_id': shift.client_id,
            'product_name': product.name if product else '',
            'product_id': shift.product_id or '',
            'engineers': assigned_ids,
            'engineer_names': [eng.name for eng in assigned_engineers],
            'serviced_by': serviced_engineer.name if serviced_engineer else '',
            'serviced_by_initials': serviced_engineer.initials if serviced_engineer else '',
            'branch': getattr(serviced_engineer, 'branch', '') if serviced_engineer else ''
        })

    return jsonify({
        'schedules': options,
        'generated_at': get_manila_time().isoformat(),
        'current_engineer': {
            'id': getattr(getattr(current_user, 'engineer_profile', None), 'id', None),
            'name': getattr(getattr(current_user, 'engineer_profile', None), 'name', '') or getattr(current_user, 'username', ''),
            'initials': getattr(getattr(current_user, 'engineer_profile', None), 'initials', '') or ''
        }
    })



# --- TSR KNOWLEDGE BASE (Phase 6C.1 Step 1) ---

def tsr_knowledge_current_engineer_name():
    profile = getattr(current_user, 'engineer_profile', None)
    return clean_str(getattr(profile, 'name', None)) or clean_str(getattr(current_user, 'username', None)) or ''


def tsr_knowledge_entry_to_dict(entry):
    return {
        'id': entry.id,
        'complaint': entry.complaint or '',
        'actions_taken': entry.actions_taken or '',
        'client_name': entry.client_name or '',
        'product_name': entry.product_name or '',
        'serial_number': entry.serial_number or '',
        'task': entry.task or '',
        'service_category': entry.service_category or '',
        'engineer_name': entry.engineer_name or '',
        'created_at': entry.created_at.strftime('%Y-%m-%d') if entry.created_at else ''
    }


def tsr_knowledge_rank_score(entry, q='', search_type='all', client='', product='', serial='', task=''):
    """Return a relevance score so TSR knowledge results show best matches first.

    Ranking priority:
    1. Exact complaint/action keyword match
    2. Same serial number
    3. Same product/model
    4. Same client/task context
    5. Most recent entry
    """
    def norm(value):
        return (value or '').strip().lower()

    q_norm = norm(q)
    client_norm = norm(client)
    product_norm = norm(product)
    serial_norm = norm(serial)
    task_norm = norm(task)

    complaint_norm = norm(getattr(entry, 'complaint', ''))
    actions_norm = norm(getattr(entry, 'actions_taken', ''))
    entry_client_norm = norm(getattr(entry, 'client_name', ''))
    entry_product_norm = norm(getattr(entry, 'product_name', ''))
    entry_serial_norm = norm(getattr(entry, 'serial_number', ''))
    entry_task_norm = norm(getattr(entry, 'task', ''))

    score = 0

    if search_type == 'complaint':
        searchable_texts = [complaint_norm]
    elif search_type == 'action':
        searchable_texts = [actions_norm]
    else:
        searchable_texts = [complaint_norm, actions_norm, entry_client_norm, entry_product_norm, entry_serial_norm, entry_task_norm]

    if q_norm:
        for text_value in searchable_texts:
            if not text_value:
                continue
            if text_value == q_norm:
                score += 120
            elif text_value.startswith(q_norm):
                score += 90
            elif q_norm in text_value:
                score += 70

        q_words = [word for word in re.split(r'\s+', q_norm) if len(word) >= 3]
        if q_words:
            matched_words = sum(1 for word in q_words if any(word in text_value for text_value in searchable_texts))
            score += matched_words * 8

    if serial_norm and entry_serial_norm:
        if entry_serial_norm == serial_norm:
            score += 60
        elif serial_norm in entry_serial_norm or entry_serial_norm in serial_norm:
            score += 40

    if product_norm and entry_product_norm:
        if entry_product_norm == product_norm:
            score += 35
        elif product_norm in entry_product_norm or entry_product_norm in product_norm:
            score += 25

    if client_norm and entry_client_norm:
        if entry_client_norm == client_norm:
            score += 25
        elif client_norm in entry_client_norm or entry_client_norm in client_norm:
            score += 15

    if task_norm and entry_task_norm:
        if entry_task_norm == task_norm:
            score += 15
        elif task_norm in entry_task_norm or entry_task_norm in task_norm:
            score += 8

    if getattr(entry, 'created_at', None):
        age_days = max((get_manila_time().replace(tzinfo=None) - entry.created_at).days, 0)
        if age_days <= 30:
            score += 10
        elif age_days <= 90:
            score += 6
        elif age_days <= 365:
            score += 3

    return score


@app.route('/search_tsr_knowledge')
@login_required
def search_tsr_knowledge():
    """Search reusable complaint/action history for Offline TSR reference.

    Phase 6C.1B:
    - type=complaint searches the complaint field directly.
    - type=action searches actions taken directly.
    - Context values are fallback helpers, not hard blockers, so keyword search
      still works even when product/serial/client data is incomplete.
    """
    if not (is_admin_authorized() or getattr(current_user, 'role', None) == 'engineer'):
        return denied()

    ensure_tsr_knowledge_entry_table()

    q = clean_str(request.args.get('q')) or ''
    client = clean_str(request.args.get('client')) or ''
    product = clean_str(request.args.get('product')) or ''
    serial = clean_str(request.args.get('serial')) or ''
    task = clean_str(request.args.get('task')) or ''
    search_type = (clean_str(request.args.get('type')) or 'all').lower()
    if search_type not in {'all', 'complaint', 'action'}:
        search_type = 'all'

    query = TsrKnowledgeEntry.query
    filters = []

    if q:
        like = f"%{q}%"
        if search_type == 'complaint':
            filters.append(TsrKnowledgeEntry.complaint.ilike(like))
        elif search_type == 'action':
            filters.append(TsrKnowledgeEntry.actions_taken.ilike(like))
        else:
            filters.append(or_(
                TsrKnowledgeEntry.complaint.ilike(like),
                TsrKnowledgeEntry.actions_taken.ilike(like),
                TsrKnowledgeEntry.client_name.ilike(like),
                TsrKnowledgeEntry.product_name.ilike(like),
                TsrKnowledgeEntry.serial_number.ilike(like),
                TsrKnowledgeEntry.task.ilike(like)
            ))
    else:
        context_filters = []
        if serial:
            context_filters.append(TsrKnowledgeEntry.serial_number.ilike(f"%{serial}%"))
        if product:
            context_filters.append(TsrKnowledgeEntry.product_name.ilike(f"%{product}%"))
        if client:
            context_filters.append(TsrKnowledgeEntry.client_name.ilike(f"%{client}%"))
        if task:
            context_filters.append(TsrKnowledgeEntry.task.ilike(f"%{task}%"))
        if context_filters:
            filters.append(or_(*context_filters))

    if search_type == 'complaint':
        filters.append(TsrKnowledgeEntry.complaint.isnot(None))
        filters.append(TsrKnowledgeEntry.complaint != '')
    elif search_type == 'action':
        filters.append(TsrKnowledgeEntry.actions_taken.isnot(None))
        filters.append(TsrKnowledgeEntry.actions_taken != '')

    if filters:
        query = query.filter(and_(*filters))

    # Fetch a wider recent pool, then sort by relevance in Python so live SQLite
    # keeps compatibility without requiring full-text search or schema changes.
    entries = query.order_by(TsrKnowledgeEntry.created_at.desc()).limit(100).all()
    ranked_entries = sorted(
        entries,
        key=lambda entry: (
            tsr_knowledge_rank_score(entry, q=q, search_type=search_type, client=client, product=product, serial=serial, task=task),
            entry.created_at or datetime.min
        ),
        reverse=True
    )[:25]

    return jsonify({
        'status': 'success',
        'type': search_type,
        'count': len(ranked_entries),
        'entries': [tsr_knowledge_entry_to_dict(entry) for entry in ranked_entries]
    })



@app.route('/get_tsr_knowledge_quick_picks')
@login_required
def get_tsr_knowledge_quick_picks():
    """Return equipment-tied recent/common TSR knowledge quick picks.

    Phase 2 Step 4:
    - Quick picks are intentionally tied to the selected equipment context.
    - Same serial number is preferred; same product/model is allowed as fallback.
    - No global fallback is returned, so engineers do not see unrelated machines.
    """
    if not (is_admin_authorized() or getattr(current_user, 'role', None) == 'engineer'):
        return denied()

    ensure_tsr_knowledge_entry_table()

    product = clean_str(request.args.get('product')) or ''
    serial = clean_str(request.args.get('serial')) or ''
    client = clean_str(request.args.get('client')) or ''
    task = clean_str(request.args.get('task')) or ''

    if not serial and not product:
        return jsonify({
            'status': 'success',
            'equipment_scoped': False,
            'message': 'Select a schedule/equipment first to load equipment-specific quick picks.',
            'recent_complaints': [],
            'common_complaints': [],
            'recent_actions': [],
            'common_actions': []
        })

    context_filters = []
    if serial:
        context_filters.append(TsrKnowledgeEntry.serial_number.ilike(f"%{serial}%"))
    if product:
        context_filters.append(TsrKnowledgeEntry.product_name.ilike(f"%{product}%"))

    entries = (
        TsrKnowledgeEntry.query
        .filter(or_(*context_filters))
        .order_by(TsrKnowledgeEntry.created_at.desc())
        .limit(250)
        .all()
    )

    ranked_entries = sorted(
        entries,
        key=lambda entry: (
            tsr_knowledge_rank_score(entry, q='', search_type='all', client=client, product=product, serial=serial, task=task),
            entry.created_at or datetime.min
        ),
        reverse=True
    )

    def text_key(value):
        return re.sub(r'\s+', ' ', (value or '').strip().lower())

    def build_recent(field_name, limit=6):
        recent = []
        seen = set()
        for entry in ranked_entries:
            value = clean_str(getattr(entry, field_name, None)) or ''
            key = text_key(value)
            if not key or key in seen:
                continue
            seen.add(key)
            recent.append(tsr_knowledge_entry_to_dict(entry))
            if len(recent) >= limit:
                break
        return recent

    def build_common(field_name, limit=6):
        grouped = {}
        for entry in ranked_entries:
            value = clean_str(getattr(entry, field_name, None)) or ''
            key = text_key(value)
            if not key:
                continue
            if key not in grouped:
                grouped[key] = {
                    'count': 0,
                    'entry': entry,
                    'latest': entry.created_at or datetime.min,
                    'score': tsr_knowledge_rank_score(entry, q='', search_type='all', client=client, product=product, serial=serial, task=task)
                }
            grouped[key]['count'] += 1
            if (entry.created_at or datetime.min) > grouped[key]['latest']:
                grouped[key]['latest'] = entry.created_at or datetime.min
                grouped[key]['entry'] = entry
            grouped[key]['score'] = max(grouped[key]['score'], tsr_knowledge_rank_score(entry, q='', search_type='all', client=client, product=product, serial=serial, task=task))

        common = sorted(
            grouped.values(),
            key=lambda item: (item['count'], item['score'], item['latest']),
            reverse=True
        )[:limit]

        results = []
        for item in common:
            entry_dict = tsr_knowledge_entry_to_dict(item['entry'])
            entry_dict['usage_count'] = item['count']
            results.append(entry_dict)
        return results

    return jsonify({
        'status': 'success',
        'equipment_scoped': True,
        'context': {
            'product_name': product,
            'serial_number': serial,
            'client_name': client,
            'task': task
        },
        'recent_complaints': build_recent('complaint'),
        'common_complaints': build_common('complaint'),
        'recent_actions': build_recent('actions_taken'),
        'common_actions': build_common('actions_taken')
    })




def build_online_tsr_pdf_filename(submission, shift, tsr_number='', payload=None):
    """Build the visible Online TSR PDF filename using the Offline TSR download rule.

    Visible/original filename must match the browser Download PDF filename:
    - NCS_TSR_Shimadzu_Client_Product(Serial)_Task_MMDDYYYY.pdf
    - NCS_TSR_B_Shimadzu_Client_Product(Serial)_Task_MMDDYYYY.pdf when billing/PO applies

    The actual stored disk filename is still randomized later by ShiftFile.filename.
    """
    payload = payload if isinstance(payload, dict) else {}
    selected_schedule = payload.get('selectedSchedule') if isinstance(payload.get('selectedSchedule'), dict) else {}

    # If the frontend explicitly supplies its download filename in a future patch,
    # trust it first so both sides remain perfectly aligned.
    frontend_filename = (
        clean_str(payload.get('pdf_filename')) or
        clean_str(payload.get('download_filename')) or
        clean_str(payload.get('tsr_pdf_filename')) or
        clean_str(payload.get('_download_filename'))
    )

    if frontend_filename:
        safe_name = secure_filename(os.path.basename(frontend_filename))
        if safe_name:
            if not safe_name.lower().endswith('.pdf'):
                safe_name = f"{safe_name}.pdf"
            if 'TSR' not in safe_name.upper():
                safe_name = f"TSR_{safe_name}"
            return safe_name

    def first_value(*keys):
        for key in keys:
            value = clean_str(payload.get(key))
            if value:
                return value
            value = clean_str(selected_schedule.get(key))
            if value:
                return value
        return ''

    def clean_filename_part(value, fallback):
        value = clean_str(value) or fallback
        value = re.sub(r'[^A-Za-z0-9()\-]+', '_', value)
        value = re.sub(r'_+', '_', value).strip('_')
        return value or fallback

    def format_tsr_date(value):
        raw_value = clean_str(value) or ''
        parsed = None
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y', '%b %d, %Y'):
            try:
                parsed = datetime.strptime(raw_value, fmt).date()
                break
            except Exception:
                pass
        if not parsed and getattr(shift, 'start_time', None):
            parsed = shift.start_time.date()
        if not parsed:
            parsed = get_manila_today()
        return parsed.strftime('%m%d%Y')

    client_name = first_value('tsr-customer-name', 'client_name') or clean_str(getattr(submission, 'client_name', '')) or 'Client'
    product_name = first_value('tsr-equipment-model', 'product_name') or clean_str(getattr(submission, 'product_name', '')) or 'Product'
    serial_number = first_value('tsr-serial-no', 'product_id', 'serial_number') or clean_str(getattr(submission, 'serial_number', ''))
    task_name = first_value('task') or clean_str(getattr(shift, 'title', '')) or 'Service'
    service_date = first_value('tsr-service-date', 'date_iso', 'date_label')

    product_label = clean_filename_part(product_name, 'Product')
    serial_clean = clean_filename_part(serial_number, '') if serial_number else ''
    if serial_clean:
        product_label = f"{product_label}({serial_clean})"

    # Match frontend billing filename condition: add B when billing/PO fields are present.
    billing_fields = [
        payload.get('tsr-billing-statement'),
        payload.get('tsr-purchase-order'),
        payload.get('billing_statement'),
        payload.get('purchase_order'),
    ]
    billing_flag = any(clean_str(value) for value in billing_fields)

    prefix = 'NCS_TSR_B_Shimadzu' if billing_flag else 'NCS_TSR_Shimadzu'
    raw_name = "_".join([
        prefix,
        clean_filename_part(client_name, 'Client'),
        product_label,
        clean_filename_part(task_name, 'Service'),
        format_tsr_date(service_date)
    ]) + '.pdf'

    safe_name = secure_filename(raw_name)
    if not safe_name.lower().endswith('.pdf'):
        safe_name = f"{safe_name}.pdf"
    if 'TSR' not in safe_name.upper():
        safe_name = f"TSR_{safe_name}"
    return safe_name

def generate_online_tsr_submission_pdf(submission, shift, payload):
    """Generate a basic official TSR PDF from the online Offline TSR payload.

    Phase 3 Step 2 only creates the server-side PDF file.
    Attaching it to the schedule happens in Phase 3 Step 3.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    selected_schedule = payload.get('selectedSchedule') if isinstance(payload.get('selectedSchedule'), dict) else {}

    def payload_value(*keys):
        for key in keys:
            value = clean_str(payload.get(key))
            if value:
                return value
        return ''

    def schedule_value(*keys):
        for key in keys:
            value = clean_str(selected_schedule.get(key))
            if value:
                return value
        return ''

    def safe_paragraph(value):
        return Paragraph(html.escape(clean_str(value) or ''), body_style)

    tsr_number = payload_value('tsr-number', 'tsr_number') or clean_str(submission.tsr_number) or f"ONLINE-{submission.id}"
    pdf_filename = build_online_tsr_pdf_filename(submission, shift, tsr_number, payload=payload)
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TSRTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        alignment=1,
        spaceAfter=8
    )
    section_style = ParagraphStyle(
        'TSRSection',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=12,
        textColor=colors.HexColor('#0f172a'),
        spaceBefore=8,
        spaceAfter=4
    )
    body_style = ParagraphStyle(
        'TSRBody',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        wordWrap='CJK'
    )

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm
    )

    story = [
        Paragraph('TECHNICAL SERVICE REPORT', title_style),
        Paragraph('Generated by MEDICAL SERVICE Online TSR Save', body_style),
        Spacer(1, 6)
    ]

    def detail_table(rows):
        table_rows = [[Paragraph(f'<b>{html.escape(label)}</b>', body_style), safe_paragraph(value)] for label, value in rows]
        table = Table(table_rows, colWidths=[44 * mm, 130 * mm])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8fafc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        return table

    story.append(Paragraph('Customer / Equipment', section_style))
    story.append(detail_table([
        ('TSR No.', tsr_number),
        ('Customer Name', payload_value('tsr-customer-name') or schedule_value('client_name')),
        ('Address', payload_value('tsr-address') or schedule_value('client_address')),
        ('Requested By', payload_value('tsr-requested-by')),
        ('Department', payload_value('tsr-department')),
        ('Equipment / Model', payload_value('tsr-equipment-model') or schedule_value('product_name')),
        ('Serial No.', payload_value('tsr-serial-no') or schedule_value('product_id')),
        ('Service Category', payload_value('tsr-service-category') or payload_value('tsr-service-category-other')),
    ]))

    story.append(Paragraph('Service Details', section_style))
    story.append(detail_table([
        ('Schedule ID', getattr(shift, 'id', '')),
        ('Task', schedule_value('task') or clean_str(getattr(shift, 'title', ''))),
        ('Date of Service', payload_value('tsr-service-date') or schedule_value('date_iso', 'date_label')),
        ('Time Started', payload_value('tsr-time-started') or schedule_value('time_start')),
        ('Time Finished', payload_value('tsr-time-finished') or schedule_value('time_end')),
        ('Serviced By', payload_value('tsr-serviced-by') or clean_str(submission.submitted_by_name)),
        ('Acknowledged By', payload_value('tsr-acknowledged-by')),
        ('Contact No.', payload_value('tsr-contact-no')),
        ('Email Add.', payload_value('tsr-email-add')),
    ]))

    story.append(Paragraph('Complaint', section_style))
    story.append(detail_table([('Complaint', payload_value('tsr-complaint'))]))

    story.append(Paragraph('Actions Taken', section_style))
    story.append(detail_table([('Actions Taken', payload_value('tsr-actions-taken'))]))

    story.append(Paragraph('Remarks / Recommendations', section_style))
    story.append(detail_table([('Remarks', payload_value('tsr-remarks'))]))

    story.append(Paragraph('Documents / Billing', section_style))
    story.append(detail_table([
        ('Submitted Documents', payload_value('tsr-documents')),
        ('Billing Statement', payload_value('tsr-billing-statement')),
        ('Sales Invoice', payload_value('tsr-sales-invoice')),
        ('Purchase Order', payload_value('tsr-purchase-order')),
    ]))

    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"Online submission #{submission.id} generated on {get_manila_time().strftime('%Y-%m-%d %H:%M')} PH time.",
        body_style
    ))

    doc.build(story)
    return pdf_filename


def get_tsr_completion_linked_shifts(shift):
    """Return schedules that should complete together after online TSR Save.

    Covers multi-day chains via group_id and linked engineer/day time overrides
    via parent_shift_id. For single multi-engineer schedules, updating the one
    shared Shift row is enough because every engineer card reads that same row.
    """
    if not shift:
        return []

    linked_by_id = {shift.id: shift}

    if shift.group_id:
        group_shifts = Shift.query.filter_by(group_id=shift.group_id).all()
        for group_shift in group_shifts:
            linked_by_id[group_shift.id] = group_shift

        group_ids = list(linked_by_id.keys())
        if group_ids:
            overrides = Shift.query.filter(
                Shift.parent_shift_id.in_(group_ids),
                Shift.override_kind == 'time_override'
            ).all()
            for override in overrides:
                linked_by_id[override.id] = override

    if shift.parent_shift_id:
        parent_shift = db.session.get(Shift, shift.parent_shift_id)
        if parent_shift:
            linked_by_id[parent_shift.id] = parent_shift
            if parent_shift.group_id:
                for group_shift in Shift.query.filter_by(group_id=parent_shift.group_id).all():
                    linked_by_id[group_shift.id] = group_shift

    return list(linked_by_id.values())


def complete_linked_schedules_for_online_tsr(shift):
    """Mark the selected schedule and its linked multi-day/override rows completed."""
    completed = []
    for linked_shift in get_tsr_completion_linked_shifts(shift):
        if not linked_shift or not linked_shift.client_id:
            continue
        if (linked_shift.status or '') != 'Completed':
            linked_shift.status = 'Completed'
        completed.append(linked_shift.id)
    return sorted(set(completed))


def attach_online_tsr_pdf_to_shift(shift, pdf_filename):
    """Attach generated online TSR PDF to the selected schedule using ShiftFile.

    Phase 3 Step 3 keeps disk filenames randomized while preserving the
    readable generated TSR filename in original_filename for UI/email/report use.
    Phase 3 Step 4 explicitly stamps uploaded_at so generated TSRs sort and
    display consistently in Reports/Archive just like manual TSR uploads.
    """
    ensure_shift_file_original_filename_column()

    safe_original = secure_filename(os.path.basename(clean_str(pdf_filename) or ''))
    if not safe_original:
        raise ValueError('Generated TSR PDF filename is empty.')

    source_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_original)
    if not os.path.exists(source_path):
        raise FileNotFoundError(f'Generated TSR PDF was not found: {safe_original}')

    disk_filename = get_unique_upload_filename(safe_original)
    disk_path = os.path.join(app.config['UPLOAD_FOLDER'], disk_filename)

    os.replace(source_path, disk_path)

    file_rec = ShiftFile(
        shift_id=shift.id,
        filename=disk_filename,
        original_filename=safe_original,
        uploaded_at=get_manila_time()
    )
    db.session.add(file_rec)
    return file_rec


def attach_uploaded_online_tsr_pdf_to_shift(shift, uploaded_pdf, display_filename=None):
    """Attach the exact browser-generated Offline TSR PDF blob to a schedule.

    This keeps Save Online visually identical to the Download PDF button because
    the frontend uploads the same PDF blob it would download locally.
    Disk storage is still randomized; original_filename remains clean/readable.
    """
    ensure_shift_file_original_filename_column()

    if not uploaded_pdf or not getattr(uploaded_pdf, 'filename', None):
        raise ValueError('No generated TSR PDF was uploaded.')

    requested_name = (
        clean_str(display_filename) or
        clean_str(getattr(uploaded_pdf, 'filename', None)) or
        'Online_TSR.pdf'
    )
    safe_original = secure_filename(os.path.basename(requested_name))
    if not safe_original:
        safe_original = 'Online_TSR.pdf'
    if not safe_original.lower().endswith('.pdf'):
        safe_original = f"{safe_original}.pdf"
    if 'TSR' not in safe_original.upper():
        safe_original = f"TSR_{safe_original}"

    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    disk_filename = get_unique_upload_filename(safe_original)
    disk_path = os.path.join(app.config['UPLOAD_FOLDER'], disk_filename)

    uploaded_pdf.save(disk_path)

    if not os.path.exists(disk_path) or os.path.getsize(disk_path) <= 0:
        raise ValueError('Uploaded TSR PDF was empty or could not be saved.')

    file_rec = ShiftFile(
        shift_id=shift.id,
        filename=disk_filename,
        original_filename=safe_original,
        uploaded_at=get_manila_time()
    )
    db.session.add(file_rec)
    return file_rec


def attach_uploaded_online_tsr_extra_files_to_shift(shift, excluded_file_ids=None):
    """Attach extra offline TSR images/signatures/documents uploaded with a queued TSR.

    Phase 4 Step 7 backend support:
    - accepts multipart files aside from the generated TSR PDF
    - preserves the original filenames
    - uses the same ShiftFile archive table
    - skips unsupported extensions and empty files safely
    """
    ensure_shift_file_original_filename_column()

    excluded_file_ids = {id(file_obj) for file_obj in (excluded_file_ids or []) if file_obj}
    saved_files = []
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

    attachment_field_names = {
        'tsr_attachment',
        'tsr_attachments',
        'tsr_signature',
        'tsr_signatures',
        'tsr_image',
        'tsr_images',
        'attachment',
        'attachments',
        'image',
        'images',
        'signature',
        'signatures'
    }

    for field_name in request.files:
        if field_name in {'tsr_pdf', 'pdf', 'file'}:
            continue
        if field_name == 'report_file' and request.files.get(field_name) and id(request.files.get(field_name)) in excluded_file_ids:
            continue

        if not (
            field_name in attachment_field_names or
            field_name.startswith('tsr_attachment') or
            field_name.startswith('tsr_image') or
            field_name.startswith('tsr_signature')
        ):
            continue

        for file_obj in request.files.getlist(field_name):
            if not file_obj or not getattr(file_obj, 'filename', None):
                continue
            if id(file_obj) in excluded_file_ids:
                continue
            if not allowed_file(file_obj.filename):
                print(f"[ONLINE-TSR] Skipped unsupported offline TSR attachment: {file_obj.filename}", flush=True)
                continue

            original_name = secure_filename(os.path.basename(file_obj.filename))
            if not original_name:
                continue

            disk_filename = get_unique_upload_filename(original_name)
            disk_path = os.path.join(app.config['UPLOAD_FOLDER'], disk_filename)
            file_obj.save(disk_path)

            if not os.path.exists(disk_path) or os.path.getsize(disk_path) <= 0:
                try:
                    os.remove(disk_path)
                except OSError:
                    pass
                continue

            file_rec = ShiftFile(
                shift_id=shift.id,
                filename=disk_filename,
                original_filename=original_name,
                uploaded_at=get_manila_time()
            )
            db.session.add(file_rec)
            saved_files.append(file_rec)

    return saved_files


@app.route('/save_offline_tsr_online', methods=['POST'])
@csrf.exempt
@login_required
def save_offline_tsr_online():
    """Phase 3 Step 3: receive online Offline TSR payload, generate PDF, and attach it to the selected schedule.

    This validates the selected schedule, stores the TSR payload, creates the
    server-side PDF, and saves it through the same ShiftFile attachment system
    used by manually uploaded TSR files.
    """
    if not (is_admin_authorized() or getattr(current_user, 'role', None) == 'engineer'):
        return denied()

    ensure_online_tsr_submission_table()

    if request.is_json:
        payload = request.get_json(silent=True) or {}
    else:
        raw_payload = (
            request.form.get('payload_json') or
            request.form.get('payload') or
            request.form.get('tsr_payload') or
            request.form.get('data') or
            request.form.get('json') or
            ''
        )
        if raw_payload:
            try:
                payload = json.loads(raw_payload)
            except Exception:
                payload = {}
        else:
            payload = dict(request.form.items())

    for filename_key in ('pdf_filename', 'download_filename', 'tsr_pdf_filename'):
        if request.form.get(filename_key) and not payload.get(filename_key):
            payload[filename_key] = request.form.get(filename_key)

    selected_schedule = payload.get('selectedSchedule') if isinstance(payload.get('selectedSchedule'), dict) else {}
    schedule_id = (
        clean_int(payload.get('schedule_id')) or
        clean_int(payload.get('selectedScheduleId')) or
        clean_int(selected_schedule.get('id'))
    )

    if not schedule_id:
        return jsonify({'status': 'error', 'message': 'Please select a schedule before saving online.'}), 400

    shift = db.session.get(Shift, schedule_id)
    if not shift:
        return jsonify({'status': 'error', 'message': 'Selected schedule was not found.'}), 404

    if not can_work_on_existing_schedule_shift(shift):
        return denied('You are not allowed to save a TSR for this schedule.')

    tsr_number = (clean_str(payload.get('tsr-number')) or clean_str(payload.get('tsr_number')) or '')[:120]
    client_name = (clean_str(payload.get('tsr-customer-name')) or clean_str(selected_schedule.get('client_name')) or '')[:200]
    product_name = (clean_str(payload.get('tsr-equipment-model')) or clean_str(selected_schedule.get('product_name')) or '')[:200]
    serial_number = (clean_str(payload.get('tsr-serial-no')) or clean_str(selected_schedule.get('product_id')) or '')[:120]
    submitted_by = tsr_knowledge_current_engineer_name()[:120]

    submission = OnlineTsrSubmission(
        shift_id=shift.id,
        tsr_number=tsr_number,
        client_name=client_name,
        product_name=product_name,
        serial_number=serial_number,
        submitted_by_user_id=getattr(current_user, 'id', None),
        submitted_by_name=submitted_by,
        status='received',
        payload_json=json.dumps(payload, ensure_ascii=False)
    )
    db.session.add(submission)
    db.session.flush()

    try:
        uploaded_pdf = None
        if request.files:
            uploaded_pdf = (
                request.files.get('tsr_pdf') or
                request.files.get('pdf') or
                request.files.get('file') or
                request.files.get('report_file')
            )

        if uploaded_pdf and getattr(uploaded_pdf, 'filename', ''):
            pdf_filename = build_online_tsr_pdf_filename(submission, shift, tsr_number, payload=payload)
            attached_file = attach_uploaded_online_tsr_pdf_to_shift(shift, uploaded_pdf, display_filename=pdf_filename)
            pdf_source = 'frontend_blob'
        else:
            pdf_filename = generate_online_tsr_submission_pdf(submission, shift, payload)
            attached_file = attach_online_tsr_pdf_to_shift(shift, pdf_filename)
            pdf_source = 'backend_reportlab_fallback'

        extra_attached_files = attach_uploaded_online_tsr_extra_files_to_shift(
            shift,
            excluded_file_ids=[uploaded_pdf, attached_file]
        )
        completed_shift_ids = complete_linked_schedules_for_online_tsr(shift)
        payload['_generated_pdf_filename'] = pdf_filename
        payload['_attached_file_id'] = attached_file.id
        payload['_attached_disk_filename'] = attached_file.filename
        payload['_attached_display_filename'] = get_shift_file_display_name(attached_file)
        payload['_extra_attachment_file_ids'] = [file_rec.id for file_rec in extra_attached_files]
        payload['_extra_attachment_filenames'] = [get_shift_file_display_name(file_rec) for file_rec in extra_attached_files]
        payload['_pdf_source'] = pdf_source
        payload['_completed_shift_ids'] = completed_shift_ids
        payload['_generated_pdf_at'] = get_manila_time().isoformat()
        submission.payload_json = json.dumps(payload, ensure_ascii=False)
        submission.status = 'completed'
        db.session.add(ActivityLog(
            user=(getattr(current_user, 'username', '') or submitted_by or 'System').capitalize(),
            action=f"Generated online TSR PDF and completed linked schedule(s): {pdf_filename} for schedule #{shift.id} (+{len(extra_attached_files)} extra attachment(s))"
        ))
        db.session.commit()
    except Exception as pdf_error:
        submission.status = 'pdf_error'
        db.session.commit()
        print(f"[ONLINE-TSR] PDF generation/attachment failed for submission #{submission.id}: {pdf_error}", flush=True)
        return jsonify({
            'status': 'error',
            'message': 'TSR was received but PDF generation or attachment failed. Please try again.',
            'submission_id': submission.id,
            'phase': '3-step-3'
        }), 500

    return jsonify({
        'status': 'success',
        'message': 'TSR saved online, PDF generated, attached to schedule, and linked schedule status completed.',
        'submission_id': submission.id,
        'schedule_id': shift.id,
        'completed_shift_ids': completed_shift_ids,
        'tsr_number': tsr_number,
        'pdf_filename': pdf_filename,
        'attached_file_id': attached_file.id,
        'attached_filename': get_shift_file_display_name(attached_file),
        'attached_disk_filename': attached_file.filename,
        'extra_attachment_count': len(extra_attached_files),
        'extra_attachment_file_ids': [file_rec.id for file_rec in extra_attached_files],
        'extra_attachment_filenames': [get_shift_file_display_name(file_rec) for file_rec in extra_attached_files],
        'pdf_source': pdf_source,
        'phase': '3-step-5-same-pdf'
    })


@app.route('/save_tsr_knowledge_entry', methods=['POST'])
@csrf.exempt
@login_required
def save_tsr_knowledge_entry():
    """Save current complaint/actions as reusable TSR knowledge."""
    if not (is_admin_authorized() or getattr(current_user, 'role', None) == 'engineer'):
        return denied()

    ensure_tsr_knowledge_entry_table()

    payload = request.get_json(silent=True) or {}
    complaint = clean_str(payload.get('complaint')) or ''
    actions_taken = clean_str(payload.get('actions_taken')) or ''

    if len(complaint) < 3 and len(actions_taken) < 3:
        return jsonify({'status': 'error', 'message': 'Please enter a complaint or action before saving.'}), 400

    client_name = (clean_str(payload.get('client_name')) or '')[:200]
    product_name = (clean_str(payload.get('product_name')) or '')[:200]
    serial_number = (clean_str(payload.get('serial_number')) or '')[:120]
    task = (clean_str(payload.get('task')) or '')[:200]
    service_category = (clean_str(payload.get('service_category')) or '')[:100]
    source_shift_id = clean_int(payload.get('schedule_id'))

    duplicate_query = TsrKnowledgeEntry.query.filter(
        func.lower(func.trim(TsrKnowledgeEntry.complaint)) == complaint.strip().lower(),
        func.lower(func.trim(TsrKnowledgeEntry.actions_taken)) == actions_taken.strip().lower()
    )

    if serial_number:
        duplicate_query = duplicate_query.filter(func.lower(func.trim(TsrKnowledgeEntry.serial_number)) == serial_number.strip().lower())
    elif product_name:
        duplicate_query = duplicate_query.filter(func.lower(func.trim(TsrKnowledgeEntry.product_name)) == product_name.strip().lower())

    existing_entry = duplicate_query.order_by(TsrKnowledgeEntry.created_at.desc()).first()
    if existing_entry:
        return jsonify({
            'status': 'success',
            'duplicate': True,
            'message': 'Knowledge base already has this complaint/actions pattern.',
            'entry': tsr_knowledge_entry_to_dict(existing_entry)
        })

    entry = TsrKnowledgeEntry(
        complaint=complaint[:5000],
        actions_taken=actions_taken[:7000],
        client_name=client_name,
        product_name=product_name,
        serial_number=serial_number,
        task=task,
        service_category=service_category,
        engineer_name=tsr_knowledge_current_engineer_name()[:120],
        source_shift_id=source_shift_id,
        created_by_user_id=getattr(current_user, 'id', None)
    )
    db.session.add(entry)
    db.session.commit()

    return jsonify({
        'status': 'success',
        'duplicate': False,
        'message': 'Complaint/actions saved to the TSR knowledge base.',
        'entry': tsr_knowledge_entry_to_dict(entry)
    })

@app.route('/service-worker.js')
def pwa_service_worker():
    """Service worker for PWA install shell, critical page caching, and offline fallback."""
    sw = r"""const CACHE_VERSION = 'medical-service-pwa-step6b-real-pdf-export';
const APP_SHELL_CACHE = `${CACHE_VERSION}-shell`;
const RUNTIME_CACHE = `${CACHE_VERSION}-runtime`;

const APP_SHELL = [
  '/',
  '/login',
  '/timeline',
  '/offline-tsr',
  '/offline',
  '/manifest.json',
  '/pwa-icon.svg',
  'https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css',
  'https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js',
  'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css'
];

const FIELD_SAFE_ROUTES = [
  '/timeline',
  '/offline-tsr',
  '/offline',
  '/login'
];

self.addEventListener('install', event => {
  self.skipWaiting();

  event.waitUntil(
    caches.open(APP_SHELL_CACHE).then(cache => {
      return Promise.allSettled(APP_SHELL.map(url => cache.add(url)));
    })
  );
});

self.addEventListener('activate', event => {
  event.waitUntil(
    caches.keys().then(keys => Promise.all(
      keys
        .filter(key => ![APP_SHELL_CACHE, RUNTIME_CACHE].includes(key))
        .map(key => caches.delete(key))
    )).then(() => self.clients.claim())
  );
});

async function networkFirst(request) {
  try {
    const response = await fetch(request);

    if (response && response.ok) {
      const cache = await caches.open(RUNTIME_CACHE);
      cache.put(request, response.clone());
    }

    return response;
  } catch (err) {
    const runtimeCache = await caches.open(RUNTIME_CACHE);
    const runtimeCached = await runtimeCache.match(request);
    if (runtimeCached) return runtimeCached;

    const shellCache = await caches.open(APP_SHELL_CACHE);
    const shellCached = await shellCache.match(request);
    if (shellCached) return shellCached;

    const cached = await caches.match(request);
    if (cached) return cached;

    return caches.match('/offline');
  }
}

async function cacheFirst(request) {
  const cached = await caches.match(request);
  if (cached) return cached;

  const response = await fetch(request);
  if (response && response.ok) {
    const cache = await caches.open(APP_SHELL_CACHE);
    cache.put(request, response.clone());
  }

  return response;
}

async function staleWhileRevalidate(request) {
  const cache = await caches.open(RUNTIME_CACHE);
  const cached = await cache.match(request);

  const networkPromise = fetch(request)
    .then(response => {
      if (response && response.ok) {
        cache.put(request, response.clone());
      }
      return response;
    })
    .catch(() => cached);

  return cached || networkPromise;
}

self.addEventListener('fetch', event => {
  const request = event.request;

  if (request.method !== 'GET') return;

  const url = new URL(request.url);
  const isSameOrigin = url.origin === self.location.origin;

  if (request.mode === 'navigate') {
    event.respondWith(networkFirst(request));
    return;
  }

  // PRODUCT-2 INSTANT REFRESH FIX:
  // All dynamic JSON/API reads must be network-first so recent product/schedule
  // mutations are visible immediately instead of returning stale PWA runtime cache.
  if (isSameOrigin && (
      url.pathname.startsWith('/get_') ||
      url.pathname.startsWith('/api/') ||
      url.pathname.startsWith('/preview_tsr_archive') ||
      url.pathname.startsWith('/download_tsr_archive') ||
      url.pathname === '/search_products' ||
      url.pathname === '/search_clients'
  )) {
    event.respondWith(networkFirst(request));
    return;
  }

  if (
    isSameOrigin &&
    (
      url.pathname.startsWith('/static/') ||
      url.pathname === '/manifest.json' ||
      url.pathname === '/pwa-icon.svg' ||
      url.pathname === '/offline'
    )
  ) {
    event.respondWith(cacheFirst(request));
    return;
  }

  if (!isSameOrigin) {
    event.respondWith(cacheFirst(request));
    return;
  }

  event.respondWith(staleWhileRevalidate(request));
});

self.addEventListener('message', event => {
  if (!event.data) return;

  if (event.data.type === 'SKIP_WAITING') {
    self.skipWaiting();
  }

  if (event.data.type === 'CACHE_FIELD_ROUTES') {
    event.waitUntil(
      caches.open(APP_SHELL_CACHE).then(cache => {
        return Promise.allSettled(FIELD_SAFE_ROUTES.map(async route => {
          const response = await fetch(route, { cache: 'reload' });
          if (response && response.ok) {
            await cache.put(route, response.clone());
          }
        }));
      })
    );
  }
});
"""
    return Response(
        sw,
        mimetype="application/javascript",
        headers={
            "Cache-Control": "no-cache",
            "Service-Worker-Allowed": "/"
        }
    )


# --- PAGE NAVIGATION ROUTES ---

@app.route('/')
@login_required
def dashboard_page():
    """ Overhauled Overview Hub - Personalization enabled for v5.1 """
    return render_template('dashboard.html')


@app.route('/activity_page')
@login_required
def activity_page():
    """Full audit trail viewer for admin-authorized users."""
    if not is_admin_authorized():
        return redirect(url_for('dashboard_page'))
    return render_template('activity.html')


@app.route('/analytics_page')
@login_required
def analytics_page():
    """Analytics dashboard for authorized management users."""
    if not is_admin_authorized():
        return redirect(url_for('dashboard_page'))
    return render_template('analytics.html')


@app.route('/reports_page')
@login_required
def reports_page():
    """Reports hub. Admins see full intelligence; engineers see TSR archive only."""
    if not (is_admin_authorized() or getattr(current_user, 'role', None) == 'engineer'):
        return redirect(url_for('dashboard_page'))

    reports_admin_view = is_admin_authorized()
    return render_template(
        'reports.html',
        # Existing names kept for backward compatibility with older reports.html versions.
        reports_admin=reports_admin_view,
        reports_engineer_only=(getattr(current_user, 'role', None) == 'engineer' and not reports_admin_view),
        # New decluttered/professional reports.html expects this name.
        reports_admin_view=reports_admin_view
    )


@app.route('/timeline')
@login_required
def timeline_page():
    """ Main technical scheduling matrix interface """
    engineer_id = None
    engineer_employee_id = ''

    profile = getattr(current_user, 'engineer_profile', None)
    if profile:
        engineer_id = profile.id
        engineer_employee_id = profile.employee_id or ''

    return render_template(
        'timeline.html',
        logged_in_engineer_id=engineer_id,
        logged_in_engineer_employee_id=engineer_employee_id
    )


@app.route('/engineers_page')
@login_required
def engineers_page():
    """ Personnel view - Restricted access handled in HTML templates """
    return render_template('engineers.html')


@app.route('/clients_page')
@login_required
def clients_page():
    """ Medical Center directory - Restricted access handled in HTML templates """
    return render_template('clients.html')


@app.route('/products_page')
@login_required
def products_page():
    """ Asset inventory - Modification permitted for all technical staff """
    return render_template('products.html')


@app.route('/settings')
@login_required
def settings_page():
    """ Profile hub for Credential Rotation and Admin Management """
    all_accounts = []
    backup_superadmin = is_superadmin_user()
    if is_admin_authorized():
        all_accounts = User.query.order_by(User.username).all()
        for account in all_accounts:
            account.display_role = get_display_role(account)
            can_reset, _ = can_reset_password_for_user(account)
            account.can_reset_password = can_reset
    return render_template(
        'settings.html',
        users=all_accounts,
        backup_superadmin=backup_superadmin
    )


def get_active_sqlite_database_path():
    """Return the active SQLite database file path for backup export."""
    database_path = getattr(db.engine.url, 'database', None)

    if database_path:
        if not os.path.isabs(database_path):
            database_path = os.path.join(basedir, database_path)
        return os.path.abspath(database_path)

    return os.path.join(basedir, 'scheduler.db')


def get_backup_upload_roots():
    """Return existing upload roots that may contain TSR/report files.

    Railway deployments may store uploads in /data/uploads/reports, while local
    or older deployments may still use static/uploads/reports. Backing up all
    existing candidates prevents missing PDFs when the active upload folder
    differs from the expected Railway path.
    """
    candidates = []

    configured_upload = clean_str(app.config.get('UPLOAD_FOLDER'))
    if configured_upload:
        candidates.append(configured_upload)

    candidates.extend([
        '/data/uploads',
        '/data/uploads/reports',
        os.path.join(basedir, 'static', 'uploads'),
        os.path.join(basedir, 'static', 'uploads', 'reports')
    ])

    upload_roots = []
    seen = set()

    for candidate in candidates:
        if not candidate:
            continue

        abs_candidate = os.path.abspath(candidate)
        normalized = abs_candidate.replace('\\', '/').rstrip('/')

        if normalized.endswith('/uploads/reports'):
            abs_candidate = os.path.dirname(abs_candidate)
            normalized = abs_candidate.replace('\\', '/').rstrip('/')

        if normalized in seen:
            continue

        if os.path.exists(abs_candidate):
            seen.add(normalized)
            upload_roots.append(abs_candidate)

    return upload_roots


def add_path_to_backup_zip(zip_handle, source_path, archive_prefix):
    """Add one file or directory to a backup ZIP, preserving relative paths."""
    source_path = os.path.abspath(source_path)

    if not os.path.exists(source_path):
        return 0

    file_count = 0

    archive_prefix = archive_prefix.replace('\\', '/').strip('/')

    if os.path.isfile(source_path):
        zip_handle.write(source_path, os.path.join(archive_prefix, os.path.basename(source_path)))
        return 1

    # Add a directory marker so uploads/reports is visible even when empty.
    zip_handle.writestr(f"{archive_prefix}/", "")

    reports_dir = os.path.join(source_path, 'reports')
    if os.path.isdir(reports_dir):
        zip_handle.writestr(f"{archive_prefix}/reports/", "")

    for root, _, files in os.walk(source_path):
        for filename in files:
            full_path = os.path.join(root, filename)
            if not os.path.isfile(full_path):
                continue

            # Never include old backup ZIPs inside a new backup.
            if filename.lower().endswith('.zip') and 'backup' in filename.lower():
                continue

            relative_path = os.path.relpath(full_path, source_path)
            archive_name = os.path.join(archive_prefix, relative_path)
            zip_handle.write(full_path, archive_name)
            file_count += 1

    return file_count


@app.route('/admin/download-backup')
@login_required
def download_system_backup():
    """Superadmin-only manual backup download for SQLite DB and uploads."""
    if not is_superadmin_user():
        return denied('Only superadmins can download system backups.')

    db.session.commit()

    timestamp = get_manila_time().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"medical_service_backup_{timestamp}.zip"
    temp_file = tempfile.NamedTemporaryFile(prefix='medical_service_backup_', suffix='.zip', delete=False)
    temp_file_path = temp_file.name
    temp_file.close()

    db_path = get_active_sqlite_database_path()
    upload_roots = get_backup_upload_roots()

    try:
        with zipfile.ZipFile(temp_file_path, 'w', zipfile.ZIP_DEFLATED) as backup_zip:
            db_count = add_path_to_backup_zip(backup_zip, db_path, 'database')
            upload_count = 0
            uploaded_roots_manifest = []

            if upload_roots:
                for upload_root in upload_roots:
                    root_name = os.path.basename(upload_root.rstrip(os.sep)) or 'uploads'
                    archive_prefix = 'uploads' if root_name == 'uploads' else f"uploads/{root_name}"
                    added_count = add_path_to_backup_zip(backup_zip, upload_root, archive_prefix)
                    upload_count += added_count
                    uploaded_roots_manifest.append({
                        'path': upload_root,
                        'archive_prefix': archive_prefix,
                        'file_count': added_count
                    })
            else:
                backup_zip.writestr('uploads/', '')
                backup_zip.writestr('uploads/reports/', '')

            manifest = {
                'generated_at_manila': get_manila_time().isoformat(),
                'generated_by': getattr(current_user, 'username', 'unknown'),
                'database_path': db_path,
                'database_included': bool(db_count),
                'upload_roots': uploaded_roots_manifest,
                'upload_file_count': upload_count,
                'app': 'MEDICAL SERVICE Scheduler'
            }
            backup_zip.writestr('backup_manifest.json', json.dumps(manifest, indent=2))

        db.session.add(ActivityLog(
            user=(getattr(current_user, 'username', '') or 'Superadmin').capitalize(),
            action=f"Downloaded system backup: {backup_filename}"
        ))
        db.session.commit()

        response = send_file(
            temp_file_path,
            as_attachment=True,
            download_name=backup_filename,
            mimetype='application/zip'
        )
        response.headers['Cache-Control'] = 'no-store'

        @response.call_on_close
        def cleanup_backup_file():
            try:
                os.remove(temp_file_path)
            except OSError:
                pass

        return response

    except Exception as backup_error:
        try:
            os.remove(temp_file_path)
        except OSError:
            pass
        print(f"[BACKUP] Manual backup failed: {backup_error}", flush=True)
        return jsonify({
            'status': 'error',
            'message': 'System backup could not be generated. Please check server logs.'
        }), 500




# --- DATA RETRIEVAL API MODULE ---


ACTIVITY_CATEGORY_META = {
    'Schedule': {'icon': 'fa-calendar-check', 'label': 'Schedule'},
    'Client': {'icon': 'fa-hospital', 'label': 'Client'},
    'Product': {'icon': 'fa-boxes-stacked', 'label': 'Product'},
    'Engineer': {'icon': 'fa-user-gear', 'label': 'Engineer'},
    'Export': {'icon': 'fa-file-export', 'label': 'Export'},
    'Security': {'icon': 'fa-shield-halved', 'label': 'Security'},
    'Email': {'icon': 'fa-envelope', 'label': 'Email'},
    'System': {'icon': 'fa-circle-info', 'label': 'System'}
}

ACTIVITY_ACTION_META = {
    'Create': {'icon': 'fa-circle-plus', 'label': 'Create'},
    'Update': {'icon': 'fa-pen-to-square', 'label': 'Update'},
    'Move': {'icon': 'fa-arrows-up-down-left-right', 'label': 'Move'},
    'Delete': {'icon': 'fa-trash-can', 'label': 'Delete'},
    'Export': {'icon': 'fa-file-export', 'label': 'Export'},
    'Import': {'icon': 'fa-file-import', 'label': 'Import'},
    'Email': {'icon': 'fa-envelope', 'label': 'Email'},
    'Security': {'icon': 'fa-shield-halved', 'label': 'Security'},
    'Other': {'icon': 'fa-circle-info', 'label': 'Other'}
}


def classify_activity_action(action):
    """Return a normalized audit category for dashboard badges and filters."""
    text = (action or '').lower()

    if 'email notification' in text or 'sent schedule' in text and 'email' in text:
        return 'Email'
    if 'password' in text or 'unauthorized' in text or 'denied' in text:
        return 'Security'
    if 'export' in text:
        return 'Export'
    if 'client' in text or 'medical center' in text:
        return 'Client'
    if 'product' in text or 'equipment' in text or 'inventory' in text:
        return 'Product'
    if 'engineer' in text or 'technical staff' in text or 'personnel' in text or 'profile' in text:
        return 'Engineer'
    if 'schedule' in text or 'calendar' in text or 'record' in text or 'bulk-purged' in text or 'wiped technical' in text:
        return 'Schedule'

    return 'System'


def classify_activity_verb(action):
    """Return the action subtype used by the advanced activity log UI."""
    text = (action or '').lower()

    if 'sent schedule' in text and 'email' in text:
        return 'Email'
    if 'password' in text or 'unauthorized' in text or 'denied' in text:
        return 'Security'
    if 'export' in text:
        return 'Export'
    if 'import' in text:
        return 'Import'
    if any(token in text for token in ['added ', 'created ', 'registered ', 'new client', 'new schedule']):
        return 'Create'
    if any(token in text for token in ['updated ', 'modified ', 'changed ', 'reset ', 'forced password']):
        return 'Update'
    if 'moved ' in text:
        return 'Move'
    if any(token in text for token in ['deleted ', 'removed ', 'purged ', 'wiped ', 'bulk-purged']):
        return 'Delete'

    return 'Other'


def infer_activity_branch(action):
    """Best-effort branch hint from activity text."""
    text = (action or '').lower()
    branches = []
    for branch in ['Manila', 'Cebu', 'Davao']:
        if branch.lower() in text:
            branches.append(branch)
    return ', '.join(branches)


def activity_log_to_dict(log):
    """Serialize an ActivityLog row for API responses."""
    category = classify_activity_action(log.action)
    action_type = classify_activity_verb(log.action)
    return {
        'id': log.id,
        'user': log.user,
        'action': log.action,
        'type': category,
        'action_type': action_type,
        'branch': infer_activity_branch(log.action),
        'timestamp': log.timestamp.strftime("%Y-%m-%d %H:%M"),
        'date': log.timestamp.strftime("%b %d, %Y"),
        'time': log.timestamp.strftime("%I:%M %p")
    }


def get_activity_users():
    """Return available users seen in activity logs for filter dropdowns."""
    return [
        user for (user,) in (
            db.session.query(ActivityLog.user)
            .filter(ActivityLog.user.isnot(None))
            .distinct()
            .order_by(ActivityLog.user.asc())
            .all()
        )
        if user
    ]


def activity_scope_query(query):
    """
    Activity visibility scope:
    - named superadmins/managers/schedulers: all logs
    - Kevin/regional_admin: Cebu/Davao-related logs plus his own actions
    - engineers: denied by route guard
    """
    if is_regional_admin_user() and not is_superadmin_user():
        regional_terms = [
            ActivityLog.user.ilike(REGIONAL_ADMIN_USERNAME),
            ActivityLog.action.ilike('%Cebu%'),
            ActivityLog.action.ilike('%Davao%'),
            ActivityLog.action.ilike('%Kevin%')
        ]
        return query.filter(or_(*regional_terms))

    return query


def parse_activity_date_bounds():
    """Resolve and normalize date range filter for activity logs."""
    start_date = parse_date(request.args.get('start_date'))
    end_date = parse_date(request.args.get('end_date'))

    if start_date and end_date and end_date < start_date:
        start_date, end_date = end_date, start_date

    return start_date, end_date



@app.route('/get_recent_activity')
@login_required
def get_recent_activity():
    """API v5.2: Fetches the latest 15 actions for the Admin Dashboard feed."""
    query = activity_scope_query(ActivityLog.query)
    logs = query.order_by(ActivityLog.timestamp.desc()).limit(15).all()
    return jsonify([activity_log_to_dict(log) for log in logs])



def build_activity_query():
    """Shared activity query builder for full log API and CSV export."""
    query = activity_scope_query(ActivityLog.query)

    search = clean_str(request.args.get('q'))
    user_filter = clean_str(request.args.get('user'))
    type_filter = clean_str(request.args.get('type'))
    action_filter = clean_str(request.args.get('action_type'))
    branch_filter = clean_str(request.args.get('branch'))
    start_date, end_date = parse_activity_date_bounds()

    if search:
        like = f"%{search}%"
        query = query.filter(or_(ActivityLog.action.ilike(like), ActivityLog.user.ilike(like)))

    if user_filter:
        query = query.filter(ActivityLog.user.ilike(user_filter))

    if start_date:
        query = query.filter(func.date(ActivityLog.timestamp) >= start_date)

    if end_date:
        query = query.filter(func.date(ActivityLog.timestamp) <= end_date)

    # Derived filters are applied after query because category/action type/branch are inferred from action text.
    return query, type_filter, action_filter, branch_filter


@app.route('/get_activity_filter_options')
@login_required
def get_activity_filter_options():
    """Filter dropdown data for the advanced Activity Log page."""
    if not is_admin_authorized():
        return denied()

    return jsonify({
        'users': get_activity_users(),
        'categories': list(ACTIVITY_CATEGORY_META.keys()),
        'action_types': list(ACTIVITY_ACTION_META.keys()),
        'branches': ['Manila', 'Cebu', 'Davao']
    })


@app.route('/get_activity_logs')
@login_required
def get_activity_logs():
    """Full activity log API with search, category/action/branch filters, date range, and pagination."""
    if not is_admin_authorized():
        return denied()

    page = clean_int(request.args.get('page')) or 1
    per_page = clean_int(request.args.get('per_page')) or 25
    page = max(page, 1)
    per_page = min(max(per_page, 10), 100)

    query, type_filter, action_filter, branch_filter = build_activity_query()
    logs = query.order_by(ActivityLog.timestamp.desc()).all()

    if type_filter:
        logs = [log for log in logs if classify_activity_action(log.action) == type_filter]

    if action_filter:
        logs = [log for log in logs if classify_activity_verb(log.action) == action_filter]

    if branch_filter:
        logs = [log for log in logs if branch_filter.lower() in (log.action or '').lower()]

    total = len(logs)
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    page_logs = logs[start_idx:end_idx]

    category_counts = {}
    action_counts = {}
    for log in logs:
        category = classify_activity_action(log.action)
        action_type = classify_activity_verb(log.action)
        category_counts[category] = category_counts.get(category, 0) + 1
        action_counts[action_type] = action_counts.get(action_type, 0) + 1

    return jsonify({
        'page': page,
        'per_page': per_page,
        'total': total,
        'pages': (total + per_page - 1) // per_page,
        'category_counts': dict(sorted(category_counts.items())),
        'action_counts': dict(sorted(action_counts.items())),
        'logs': [activity_log_to_dict(log) for log in page_logs]
    })


@app.route('/export_activity_logs')
@login_required
def export_activity_logs():
    """CSV export for filtered activity logs."""
    if not is_admin_authorized():
        return denied()

    query, type_filter, action_filter, branch_filter = build_activity_query()
    logs = query.order_by(ActivityLog.timestamp.desc()).all()

    if type_filter:
        logs = [log for log in logs if classify_activity_action(log.action) == type_filter]

    if action_filter:
        logs = [log for log in logs if classify_activity_verb(log.action) == action_filter]

    if branch_filter:
        logs = [log for log in logs if branch_filter.lower() in (log.action or '').lower()]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Timestamp', 'User', 'Category', 'Action Type', 'Branch Hint', 'Action'])

    for log in logs:
        writer.writerow([
            log.timestamp.strftime("%Y-%m-%d %H:%M"),
            log.user,
            classify_activity_action(log.action),
            classify_activity_verb(log.action),
            infer_activity_branch(log.action),
            log.action
        ])

    output.seek(0)
    log_activity("Exported activity log report")
    filename = f"activity_logs_{get_manila_time().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )


@app.route('/get_engineers')
@login_required
def get_engineers():
    """ Personnel Retrieval API providing contact info and account metadata """
    engineers = Engineer.query.order_by(Engineer.name).all()
    results = []
    
    for e in engineers:
        # Cross-reference engineer ID with user database (Hard Link v5.4)
        account = db.session.get(User, e.user_id) if e.user_id else None
        if not account:
            fname = e.name.split()[0].lower()
            account = User.query.filter_by(username=fname).first()

            # Do not let the special regional admin account attach by loose first-name match
            # to another technician named Kevin.
            if (
                account and
                account.username == REGIONAL_ADMIN_USERNAME and
                e.employee_id != REGIONAL_ADMIN_EMPLOYEE_ID
            ):
                account = None

        results.append({
            'id': e.id, 
            'employee_id': e.employee_id,
            'name': e.name, 
            'initials': e.initials, 
            'phone': e.phone or "",
            'email': e.email or "",
            'branch': e.branch or "",
            'account_role': account.role if account else None,
            'display_role': get_display_role(account) if account else None,
            'user_id': account.id if account else None
        })
        
    return jsonify(results)


@app.route('/get_clients')
@login_required
def get_clients():
    clients = Client.query.order_by(Client.name).all()
    results = []
    for c in clients:
        contacts = Contact.query.filter_by(client_id=c.id).all()
        entry = {'id': c.id, 'name': c.name, 'address': c.address}
        for idx, ct in enumerate(contacts, start=1):
            entry[f'cp{idx}'] = ct.name
            entry[f'cn{idx}'] = ct.phone
            entry[f'ce{idx}'] = ct.email
        results.append(entry)
    return jsonify(results)


@app.route('/get_products')
@login_required
def get_products():
    """ Inventory Retrieval API featuring machine ownership mapping """
    products = Product.query.all()
    results = []
    
    for p in products:
        results.append({
            'serial_number': p.serial_number, 
            'name': p.name, 
            'client_id': p.client_id, 
            'client_name': p.owner.name if p.owner else "N/A", 
            'start_warranty': p.start_warranty_date.isoformat() if p.start_warranty_date else "", 
            'end_warranty': p.end_warranty_date.isoformat() if p.end_warranty_date else ""
        })
        
    return jsonify(results)


@app.route('/get_open_tasks')
@login_required
def get_open_tasks():
    """ 
    DASHBOARD CORE LOGIC API:
    Retrieves all technical visits where status is NOT 'Completed'.
    """
    invalid_dashboard_categories = ['Completed', 'Training'] + LEAVE_CATEGORIES
    
    active_shifts = Shift.query.filter(
        and_(
            Shift.status.notin_(invalid_dashboard_categories),
            Shift.client_id.isnot(None)
        )
    ).all()
    
    results = []
    for s in active_shifts:

        eng_names = [
            db.session.get(Engineer, se.engineer_id).name
            for se in ShiftEngineer.query.filter_by(shift_id=s.id).all()
            if db.session.get(Engineer, se.engineer_id)
        ]

        results.append({
            'id': s.id,
            'task_date': s.start_time.strftime("%Y-%m-%d"),
            'created_at': s.created_at.strftime("%Y-%m-%d %H:%M"),
            'engineer': ", ".join(eng_names) if eng_names else "None",
            'client': s.client.name if s.client else "N/A",
            'task': s.title,
            'status': s.status
        })
        
    return jsonify(results)


@app.route('/get_timeline_data')
@login_required
def get_timeline_data():
    ensure_shift_file_original_filename_column()

    """
    Main calendar grid API. Maps engineers to weekly date slots.

    Step 11 safe optimization:
    - Applies optional branch filter on engineers server-side.
    - Loads all shifts for the visible week in one query.
    - Keeps legacy Shift.engineer_id fallback so older records still display.
    - Preloads client/product/files and precomputes assignment/group metadata.
    """
    offset = clean_int(request.args.get('offset', 0)) or 0
    branch_filter = clean_str(request.args.get('branch')) or 'ALL'

    # Regional admin view policy:
    # - may view ALL branches and Manila schedules
    # - REGIONAL filter means Cebu + Davao combined
    # - write permissions are still restricted by can_modify_schedule_for_engineer_ids()
    if is_regional_admin_user():
        allowed_view_filters = {'ALL', 'REGIONAL', 'Manila'} | REGIONAL_ADMIN_BRANCHES
        if branch_filter not in allowed_view_filters:
            branch_filter = 'ALL'

    target_dt = (get_manila_time()).date() + timedelta(weeks=offset)
    target_dt -= timedelta(days=target_dt.weekday())
    week_span = [(target_dt + timedelta(days=i)) for i in range(7)]
    week_start = week_span[0]
    week_end = week_span[-1]

    engineer_query = Engineer.query.order_by(Engineer.name)
    if branch_filter == 'REGIONAL':
        engineer_query = engineer_query.filter(Engineer.branch.in_(REGIONAL_ADMIN_BRANCHES))
    elif branch_filter != 'ALL':
        engineer_query = engineer_query.filter(Engineer.branch == branch_filter)

    engineers = engineer_query.all()
    visible_engineer_ids = {engineer.id for engineer in engineers}

    schedule_mapping = {
        technician.id: {day.isoformat(): [] for day in week_span}
        for technician in engineers
    }

    if visible_engineer_ids:
        week_start_dt, week_end_dt = build_shift_datetime_bounds(week_start, week_end)

        weekly_shifts = (
            Shift.query
            .options(
                joinedload(Shift.client),
                joinedload(Shift.product),
                selectinload(Shift.files)
            )
            .filter(Shift.start_time >= week_start_dt)
            .filter(Shift.start_time < week_end_dt)
            .order_by(Shift.start_time.asc())
            .all()
        )

        weekly_shift_ids = [shift.id for shift in weekly_shifts]

        shift_engineer_map = {shift_id: [] for shift_id in weekly_shift_ids}
        if weekly_shift_ids:
            assignment_rows = (
                db.session.query(ShiftEngineer.shift_id, ShiftEngineer.engineer_id)
                .filter(ShiftEngineer.shift_id.in_(weekly_shift_ids))
                .all()
            )
            for shift_id, engineer_id in assignment_rows:
                shift_engineer_map.setdefault(shift_id, []).append(engineer_id)

        group_ids = {shift.group_id for shift in weekly_shifts if shift.group_id}
        group_ranges = {}
        if group_ids:
            group_rows = (
                db.session.query(
                    Shift.group_id,
                    func.min(func.date(Shift.start_time)),
                    func.max(func.date(Shift.start_time))
                )
                .filter(Shift.group_id.in_(group_ids))
                .group_by(Shift.group_id)
                .all()
            )
            group_ranges = {
                group_id: (start_date, end_date)
                for group_id, start_date, end_date in group_rows
            }

        for shift in weekly_shifts:
            assigned_engineer_ids = shift_engineer_map.get(shift.id) or []

            # Legacy compatibility: some records may only have Shift.engineer_id.
            if not assigned_engineer_ids and shift.engineer_id:
                assigned_engineer_ids = [shift.engineer_id]

            visible_assigned_ids = [eid for eid in assigned_engineer_ids if eid in visible_engineer_ids]
            if not visible_assigned_ids:
                continue

            if shift.group_id and shift.group_id in group_ranges:
                start_date, end_date = group_ranges[shift.group_id]
            else:
                start_date = shift.start_time.date().isoformat()
                end_date = shift.end_time.date().isoformat()

            payload = {
                'id': shift.id,
                'client_name': shift.client.name if shift.client else "",
                'client_address': shift.client.address if shift.client else "",
                'product_name': shift.product.name if shift.product else "",
                'task': shift.title,
                'time_start': shift.start_time.strftime("%H:%M"),
                'time_end': shift.end_time.strftime("%H:%M"),
                'client_id': shift.client_id,
                'product_id': shift.product_id,
                'status': shift.status,
                'files': [get_shift_file_display_name(file_record) or file_record.filename for file_record in shift.files],
                'file_details': [
                    {
                        'id': file_record.id,
                        'filename': get_shift_file_display_name(file_record) or file_record.filename,
                        'disk_filename': file_record.filename,
                        'display_name': get_shift_file_display_name(file_record),
                        'download_url': url_for('download_tsr_archive_file', file_id=file_record.id, scope='all') if is_tsr_filename(get_shift_file_display_name(file_record) or file_record.filename) else '',
                        'uploaded_at': file_record.uploaded_at.isoformat() if file_record.uploaded_at else ''
                    }
                    for file_record in shift.files
                ],
                'engineers': assigned_engineer_ids,
                'day_owner_engineer_id': shift.engineer_id,
                'day_owner_engineer_name': shift.engineer.name if shift.engineer else '',
                'day_owner_engineer_initials': shift.engineer.initials if shift.engineer else '',
                'start_date': str(start_date),
                'end_date': str(end_date),
                'group_id': shift.group_id,
                'parent_shift_id': shift.parent_shift_id,
                'override_engineer_id': shift.override_engineer_id,
                'override_kind': shift.override_kind,
                'is_time_override': is_shift_time_override(shift)
            }

            day_key = shift.start_time.date().isoformat()
            for engineer_id in visible_assigned_ids:
                if day_key in schedule_mapping.get(engineer_id, {}):
                    schedule_mapping[engineer_id][day_key].append(payload)

    return jsonify({
        'days': [
            {
                'iso': day.isoformat(),
                'display_date': day.strftime("%b %d"),
                'display_day': day.strftime("%A")
            }
            for day in week_span
        ],
        'engineers': [
            {
                'id': engineer.id,
                'name': engineer.name,
                'branch': engineer.branch or "Manila",
                'initials': engineer.initials or ''
            }
            for engineer in engineers
        ],
        'schedule': schedule_mapping,
        'branch_filter': branch_filter,
        'current_range': f"{week_span[0].strftime('%B %d')} - {week_span[6].strftime('%B %d, %Y')}"
    })


# --- ANALYTICS MODULE ---

def analytics_date_bounds():
    """Resolve analytics date range from query string or default to current month."""
    today = get_manila_time().date()

    start_date = parse_date(request.args.get('start_date'))
    end_date = parse_date(request.args.get('end_date'))

    if not start_date or not end_date:
        start_date = today.replace(day=1)
        end_date = today

    if end_date < start_date:
        start_date, end_date = end_date, start_date

    return start_date, end_date


def classify_schedule_type(shift):
    """Normalize schedule type for analytics breakdown."""
    title = shift.title or ''

    if title in LEAVE_CATEGORIES:
        return 'Leave'
    if title == 'In Office' or title.startswith('In Office:'):
        return 'In Office'
    if title == 'Traveling to Client' or title.startswith('Traveling to Client:'):
        return 'Travel'
    if title == 'Pull-out Parts' or title.startswith('Pull-out Parts:'):
        return 'Pull-out'
    if title == 'Holiday' or title.startswith('Holiday:'):
        return 'Holiday'
    if shift.client_id:
        return 'Site Visit'
    return 'Training'


def get_shift_engineer_records(shift):
    """Return assigned Engineer records for a shift, falling back to primary engineer."""
    linked_engineers = [
        db.session.get(Engineer, se.engineer_id)
        for se in ShiftEngineer.query.filter_by(shift_id=shift.id).all()
    ]
    linked_engineers = [eng for eng in linked_engineers if eng]

    if linked_engineers:
        return linked_engineers

    fallback = db.session.get(Engineer, shift.engineer_id) if shift.engineer_id else None
    return [fallback] if fallback else []


def analytics_requested_branch():
    """Read requested analytics branch filter while respecting Kevin/regional scope."""
    branch = clean_str(request.args.get('branch')) or 'ALL'

    if is_regional_admin_user():
        allowed_view_filters = {'ALL', 'REGIONAL', 'Manila'} | REGIONAL_ADMIN_BRANCHES
        if branch in allowed_view_filters:
            return branch
        return 'ALL'

    if branch not in {'ALL', 'REGIONAL', 'Manila', 'Cebu', 'Davao'}:
        return 'ALL'

    return branch


def analytics_branch_label(branch):
    if branch == 'ALL':
        return 'All Branches'
    if branch == 'REGIONAL':
        return 'Cebu + Davao Branches'
    return f'{branch} Branch'


def analytics_visible_engineer_ids():
    """Restrict analytics scope by role and optional branch filter."""
    branch = analytics_requested_branch()

    engineer_query = db.session.query(Engineer.id)

    if is_regional_admin_user():
        # Kevin/regional_admin may view analytics for all branches, including Manila.
        # Mutation endpoints still restrict write access to Cebu/Davao only.
        pass
    elif not is_superadmin_user():
        return []

    if branch == 'REGIONAL':
        engineer_query = engineer_query.filter(Engineer.branch.in_(REGIONAL_ADMIN_BRANCHES))
    elif branch != 'ALL':
        engineer_query = engineer_query.filter(Engineer.branch == branch)

    return [eid for (eid,) in engineer_query.all()]


def analytics_scope_query(query):
    """Apply role scope to analytics queries."""
    visible_ids = analytics_visible_engineer_ids()

    if visible_ids is None:
        return query

    if not visible_ids:
        return query.filter(False)

    return (
        query
        .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
        .filter(ShiftEngineer.engineer_id.in_(visible_ids))
    )


@app.route('/get_analytics_summary')
@login_required
def get_analytics_summary():
    """Management analytics API: schedules, branch workload, engineer workload, type/status breakdown."""
    if not is_admin_authorized():
        return denied()

    start_date, end_date = analytics_date_bounds()
    today = get_manila_time().date()

    def count_between(start_d, end_d):
        q = Shift.query.filter(
            func.date(Shift.start_time) >= start_d,
            func.date(Shift.start_time) <= end_d
        )
        q = analytics_scope_query(q)
        return len({shift.id for shift in q.all()})

    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    month_start = today.replace(day=1)

    range_query = Shift.query.filter(
        func.date(Shift.start_time) >= start_date,
        func.date(Shift.start_time) <= end_date
    )
    range_query = analytics_scope_query(range_query)
    raw_shifts = range_query.order_by(Shift.start_time.asc()).all()
    shifts = list({shift.id: shift for shift in raw_shifts}.values())

    completed = 0
    active = 0
    branches = {}
    engineers = {}
    categories = {}
    statuses = {}
    open_statuses = {
        'In Progress': 0,
        'For Continuation': 0,
        'Waiting for P.O': 0,
        'Waiting for Parts': 0
    }

    total_assignment_count = 0

    for shift in shifts:
        schedule_type = classify_schedule_type(shift)
        categories[schedule_type] = categories.get(schedule_type, 0) + 1

        status = shift.status or 'In Progress'
        statuses[status] = statuses.get(status, 0) + 1

        if status == 'Completed':
            completed += 1
        else:
            active += 1

        if status in open_statuses and shift.client_id:
            open_statuses[status] += 1

        assigned_engineers = get_shift_engineer_records(shift)
        if not assigned_engineers:
            branches['Unassigned'] = branches.get('Unassigned', 0) + 1
            continue

        total_assignment_count += len(assigned_engineers)

        for eng in assigned_engineers:
            branch = eng.branch or 'Unassigned'
            branches[branch] = branches.get(branch, 0) + 1

            if eng.name not in engineers:
                engineers[eng.name] = {
                    'name': eng.name,
                    'branch': branch,
                    'count': 0,
                    'completed': 0,
                    'active': 0
                }

            engineers[eng.name]['count'] += 1
            if status == 'Completed':
                engineers[eng.name]['completed'] += 1
            else:
                engineers[eng.name]['active'] += 1

    busiest_engineers = sorted(
        engineers.values(),
        key=lambda x: (-x['count'], x['name'])
    )[:15]

    return jsonify({
        'status': 'success',
        'range': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        },
        'branch': analytics_requested_branch(),
        'branch_label': analytics_branch_label(analytics_requested_branch()),
        'today': count_between(today, today),
        'week': count_between(week_start, week_end),
        'month': count_between(month_start, today),
        'range_total': len(shifts),
        'completed': completed,
        'active': active,
        'assignment_total': total_assignment_count,
        'branches': dict(sorted(branches.items())),
        'engineers': busiest_engineers,
        'categories': dict(sorted(categories.items())),
        'statuses': dict(sorted(statuses.items())),
        'open_statuses': open_statuses
    })






def shift_has_tsr_file(shift):
    """Return True when a shift has at least one TSR attachment."""
    filenames = []
    for file_rec in getattr(shift, 'files', []) or []:
        filenames.append(get_shift_file_display_name(file_rec))
        filenames.append(get_shift_file_disk_name(file_rec))
    return existing_files_have_tsr(filenames)


def tsr_archive_requested_scope():
    """Return requested TSR archive scope.

    Engineers default to "my" TSRs, but may request "all" for read-only
    reference access to other TSRs.
    """
    scope = (request.args.get('scope') or 'my').strip().lower()
    return 'all' if scope == 'all' else 'my'


def user_can_view_shift_tsr_archive(shift, scope=None):
    """Reports archive visibility.

    - Admin-authorized users follow the normal branch/reporting scope.
    - Engineers default to their own TSRs, but may switch to all TSRs for
      read/download reference only.
    """
    if not shift:
        return False

    if is_admin_authorized():
        return True

    if getattr(current_user, 'role', None) == 'engineer':
        if (scope or 'my') == 'all':
            return True
        return is_current_engineer_assigned_to_shift(shift)

    return False


def tsr_archive_shift_to_dict(shift):
    """Serialize one shift with TSR files for the archive UI.

    Phase 3 Step 4 keeps generated online TSR PDFs compatible with the
    existing Reports archive by exposing the same preview/download URLs while
    adding lightweight metadata that the frontend can safely ignore.
    """
    files = []
    archive_scope = tsr_archive_requested_scope()
    sorted_files = sorted(
        getattr(shift, 'files', []) or [],
        key=lambda file_rec: getattr(file_rec, 'uploaded_at', None) or datetime.min,
        reverse=True
    )

    for file_rec in sorted_files:
        display_name = get_shift_file_display_name(file_rec)
        disk_name = get_shift_file_disk_name(file_rec)
        if not existing_files_have_tsr([display_name, disk_name]):
            continue

        filename_for_ext = display_name or disk_name or ''
        ext = filename_for_ext.rsplit('.', 1)[-1].lower() if '.' in filename_for_ext else ''
        is_generated_online_tsr = bool(
            ext == 'pdf' and
            (display_name or '').upper().startswith('TSR_') and
            (disk_name or '').startswith(f'shift_{shift.id}_')
        )

        files.append({
            'id': file_rec.id,
            'name': display_name or disk_name,
            'uploaded_at': file_rec.uploaded_at.strftime('%Y-%m-%d %H:%M') if getattr(file_rec, 'uploaded_at', None) else '',
            'file_type': ext,
            'is_pdf': ext == 'pdf',
            'source': 'Generated Online TSR' if is_generated_online_tsr else 'Uploaded TSR',
            'preview_url': url_for('preview_tsr_archive_file', file_id=file_rec.id, scope=archive_scope),
            'download_url': url_for('download_tsr_archive_file', file_id=file_rec.id, scope=archive_scope)
        })

    engineers = get_shift_engineer_records(shift)

    return {
        'id': shift.id,
        'date': shift.start_time.strftime('%Y-%m-%d') if shift.start_time else '',
        'client': shift.client.name if shift.client else 'N/A',
        'product': shift.product.name if shift.product else 'N/A',
        'serial': shift.product.serial_number if shift.product else '',
        'task': shift.title or '',
        'status': shift.status or '',
        'engineers': ', '.join([engineer.name for engineer in engineers]) or 'N/A',
        'files': files
    }


@app.route('/get_tsr_archive')
@login_required
def get_tsr_archive():
    """Searchable TSR archive with engineer/admin access control."""
    if not (is_admin_authorized() or getattr(current_user, 'role', None) == 'engineer'):
        return denied()

    start_date, end_date = analytics_date_bounds()
    search = clean_str(request.args.get('q')) or ''
    scope = tsr_archive_requested_scope()

    query = Shift.query.options(
        joinedload(Shift.client),
        joinedload(Shift.product),
        selectinload(Shift.files)
    ).filter(
        func.date(Shift.start_time) >= start_date,
        func.date(Shift.start_time) <= end_date,
        Shift.client_id.isnot(None)
    )

    if is_admin_authorized():
        query = analytics_scope_query(query)

    shifts = list({shift.id: shift for shift in query.order_by(Shift.start_time.desc()).limit(500).all()}.values())

    rows = []
    q = search.lower()
    for shift in shifts:
        if not user_can_view_shift_tsr_archive(shift, scope):
            continue
        if not shift_has_tsr_file(shift):
            continue

        row = tsr_archive_shift_to_dict(shift)

        if q:
            haystack = ' '.join([
                row.get('date', ''),
                row.get('client', ''),
                row.get('product', ''),
                row.get('serial', ''),
                row.get('task', ''),
                row.get('status', ''),
                row.get('engineers', ''),
                ' '.join([file_row.get('name', '') for file_row in row.get('files', [])])
            ]).lower()
            if q not in haystack:
                continue

        rows.append(row)
        if len(rows) >= 100:
            break

    return jsonify({
        'status': 'success',
        'rows': rows,
        'count': len(rows),
        'scope': scope,
        'admin': is_admin_authorized()
    })


def _tsr_archive_error_response(message, status_code=404):
    """Return a visible archive error instead of silently redirecting to Reports."""
    print(f"[TSR-ARCHIVE] {message}", flush=True)

    if (
        request.path.startswith('/preview_tsr_archive_pdf_meta') or
        request.path.startswith('/preview_tsr_archive_pdf_page') or
        request.path.startswith('/preview_tsr_archive_content') or
        'application/json' in (request.headers.get('Accept') or '')
    ):
        return jsonify({'status': 'error', 'message': message}), status_code

    safe_message = html.escape(message)
    reports_url = url_for('reports_page')
    return Response(
        f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>TSR Archive Notice</title>
  <style>
    body{{margin:0;min-height:100vh;display:grid;place-items:center;background:#f8fafc;color:#0f172a;font-family:Arial,sans-serif;padding:18px;}}
    .card{{max-width:680px;background:#fff;border-radius:18px;padding:24px;box-shadow:0 12px 34px rgba(15,23,42,.12);border:1px solid #e5e7eb;}}
    h1{{font-size:1.15rem;margin:0 0 8px;}}
    p{{color:#475569;line-height:1.5;}}
    code{{background:#f1f5f9;padding:2px 6px;border-radius:6px;}}
    a{{display:inline-block;margin-top:10px;background:#0d6efd;color:#fff;text-decoration:none;font-weight:900;padding:10px 14px;border-radius:10px;}}
  </style>
</head>
<body>
  <main class="card">
    <h1>Unable to open TSR file</h1>
    <p>{safe_message}</p>
    <p>This page is shown instead of silently reloading Reports so the problem can be identified.</p>
    <a href="{reports_url}">Back to Reports</a>
  </main>
</body>
</html>""",
        status=status_code,
        mimetype='text/html',
        headers={'Cache-Control': 'no-store'}
    )


def _unique_existing_paths(paths):
    """Return unique existing file paths from a candidate list."""
    found = []
    seen = set()
    for candidate in paths:
        if not candidate:
            continue
        normalized = os.path.abspath(candidate)
        key = normalized.replace('\\', '/').lower()
        if key in seen:
            continue
        seen.add(key)
        if os.path.isfile(normalized):
            found.append(normalized)
    return found


def get_tsr_archive_file_candidate_paths(disk_name, display_name=None):
    """Return possible storage paths for TSR archive files.

    Railway stores current files in /data/uploads/reports. Older local/live
    builds may have files under static/uploads/reports, and restored backups may
    contain either the randomized disk filename or the original display name.
    """
    safe_disk_name = os.path.basename(clean_str(disk_name) or '')
    safe_display_name = os.path.basename(clean_str(display_name) or '')
    names = []
    for name in (safe_disk_name, safe_display_name, derive_original_filename_from_stored_filename(safe_disk_name)):
        if name and name not in names:
            names.append(name)

    roots = []
    for root in (
        app.config.get('UPLOAD_FOLDER'),
        '/data/uploads/reports',
        '/data/uploads',
        os.path.join(basedir, 'static', 'uploads', 'reports'),
        os.path.join(basedir, 'static', 'uploads')
    ):
        root = clean_str(root)
        if root and root not in roots:
            roots.append(root)

    candidates = []
    for root in roots:
        for name in names:
            candidates.append(os.path.join(root, name))
            candidates.append(os.path.join(root, 'reports', name))

    return candidates


def _resolve_tsr_archive_file_for_preview(file_id):
    """Return validated TSR archive file info for preview/download routes."""
    print(f"[TSR-ARCHIVE] Resolving file_id={file_id} path={request.path}", flush=True)

    file_rec = db.session.get(ShiftFile, file_id)
    if not file_rec:
        return None, _tsr_archive_error_response(f'TSR file record #{file_id} was not found.', 404)

    shift = db.session.get(Shift, file_rec.shift_id)
    scope = tsr_archive_requested_scope()
    if not user_can_view_shift_tsr_archive(shift, scope):
        return None, _tsr_archive_error_response(
            f'You do not have permission to open TSR file #{file_id} with scope={scope}.',
            403
        )

    display_name = get_shift_file_display_name(file_rec)
    disk_name = get_shift_file_disk_name(file_rec)
    if not existing_files_have_tsr([display_name, disk_name]):
        return None, _tsr_archive_error_response(
            f'File #{file_id} is not recognized as a TSR file: {display_name or disk_name or "unnamed"}.',
            400
        )

    candidate_paths = get_tsr_archive_file_candidate_paths(disk_name, display_name)
    found_paths = _unique_existing_paths(candidate_paths)
    if not found_paths:
        print(
            f"[TSR-ARCHIVE] Missing physical file for file_id={file_id}; "
            f"disk_name={disk_name}; display_name={display_name}; "
            f"checked={candidate_paths[:8]}",
            flush=True
        )
        return None, _tsr_archive_error_response(
            'The TSR record exists in the database, but the physical file was not found in the active upload folders. '
            f'Disk filename: {disk_name or "blank"}. Display filename: {display_name or "blank"}.',
            404
        )

    file_path = found_paths[0]
    ext_source = display_name or disk_name or file_path
    ext = ext_source.rsplit('.', 1)[-1].lower() if '.' in ext_source else ''

    print(f"[TSR-ARCHIVE] Resolved file_id={file_id} to {file_path}", flush=True)
    return {
        'file_rec': file_rec,
        'shift': shift,
        'scope': scope,
        'display_name': display_name,
        'disk_name': disk_name,
        'file_path': file_path,
        'ext': ext
    }, None


@app.route('/preview_tsr_archive_file/<int:file_id>')
@login_required
def preview_tsr_archive_file(file_id):
    """Authenticated browser-safe TSR preview page.

    PDFs are rendered as server-generated PNG pages so Brave/mobile browsers do
    not auto-download or depend on browser PDF plugins. Images and CSV files
    are displayed inside the same viewer. DOCX/XLSX remain download-only.
    """
    import json

    resolved, error_response = _resolve_tsr_archive_file_for_preview(file_id)
    if error_response:
        return error_response

    file_rec = resolved['file_rec']
    scope = resolved['scope']
    ext = resolved['ext']
    safe_name = resolved['display_name'] or resolved['disk_name']

    content_url = url_for('preview_tsr_archive_content', file_id=file_rec.id, scope=scope)
    pdf_meta_url = url_for('preview_tsr_archive_pdf_meta', file_id=file_rec.id, scope=scope)
    pdf_page_url_template = url_for(
        'preview_tsr_archive_pdf_page',
        file_id=file_rec.id,
        page_number=0,
        scope=scope
    ).replace('/0?', '/__PAGE__?')
    download_url = url_for('download_tsr_archive_file', file_id=file_rec.id, scope=scope)

    viewer_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Preview - {html.escape(safe_name)}</title>
  <style>
    body {{ margin:0; background:#f8fafc; color:#0f172a; font-family:Arial,sans-serif; }}
    .viewer-header {{ position:sticky; top:0; z-index:10; display:flex; justify-content:space-between; align-items:center; gap:12px; padding:10px 14px; background:#111827; color:#fff; box-shadow:0 4px 18px rgba(15,23,42,.18); }}
    .viewer-title {{ font-weight:900; font-size:.9rem; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
    .viewer-actions {{ display:flex; gap:8px; flex-shrink:0; }}
    .viewer-actions a, .viewer-actions button {{ border:0; border-radius:10px; padding:8px 10px; font-size:.8rem; font-weight:900; text-decoration:none; cursor:pointer; }}
    .viewer-actions a {{ background:#0d6efd; color:#fff; }}
    .viewer-actions button {{ background:#e5e7eb; color:#111827; }}
    #viewer {{ padding:14px; }}
    .page-wrap {{ display:grid; gap:14px; justify-items:center; }}
    img.preview-page {{ display:block; max-width:100%; height:auto; background:#fff; border-radius:8px; box-shadow:0 8px 24px rgba(15,23,42,.12); }}
    img.preview-image {{ display:block; max-width:100%; height:auto; margin:0 auto; background:#fff; border-radius:8px; box-shadow:0 8px 24px rgba(15,23,42,.12); }}
    pre {{ white-space:pre-wrap; word-break:break-word; background:#fff; border-radius:14px; padding:14px; box-shadow:0 8px 24px rgba(15,23,42,.08); }}
    .message {{ max-width:720px; margin:40px auto; background:#fff; border-radius:18px; padding:22px; box-shadow:0 8px 28px rgba(15,23,42,.08); text-align:center; }}
    .message h2 {{ margin:0 0 8px; font-size:1.05rem; }}
    .message p {{ color:#64748b; line-height:1.45; }}
    .page-loading {{ max-width:980px; width:100%; margin:0 auto; background:#fff; border-radius:12px; padding:18px; color:#64748b; font-weight:900; text-align:center; box-shadow:0 8px 24px rgba(15,23,42,.08); }}
    @media(max-width:768px){{ .viewer-header {{ align-items:flex-start; flex-direction:column; }} .viewer-actions {{ width:100%; }} .viewer-actions a, .viewer-actions button {{ flex:1; text-align:center; min-height:42px; }} #viewer {{ padding:10px; }} }}
  </style>
</head>
<body>
  <div class="viewer-header">
    <div class="viewer-title">{html.escape(safe_name)}</div>
    <div class="viewer-actions">
      <button type="button" onclick="location.reload()">Reload</button>
      <a href="{html.escape(download_url)}">Download</a>
    </div>
  </div>
  <main id="viewer"><div class="message"><h2>Loading preview...</h2><p>Please wait while the TSR loads.</p></div></main>

  <script>
    const CONTENT_URL = {json.dumps(content_url)};
    const PDF_META_URL = {json.dumps(pdf_meta_url)};
    const PDF_PAGE_URL_TEMPLATE = {json.dumps(pdf_page_url_template)};
    const FILE_EXT = {json.dumps(ext)};
    const FILE_NAME = {json.dumps(safe_name)};
    const viewer = document.getElementById('viewer');

    function showMessage(title, message) {{
      viewer.innerHTML = `<div class="message"><h2>${{title}}</h2><p>${{message}}</p><p><a href="{html.escape(download_url)}">Download file</a></p></div>`;
    }}

    async function renderPdfServerPages() {{
      const metaResponse = await fetch(PDF_META_URL, {{ credentials: 'same-origin', cache: 'no-store' }});
      const meta = await metaResponse.json().catch(() => ({{}}));
      if (!metaResponse.ok || meta.status !== 'success' || !meta.page_count) {{
        throw new Error(meta.message || 'PDF preview metadata unavailable.');
      }}

      viewer.innerHTML = '<div class="page-wrap" id="pdf-pages"></div>';
      const pages = document.getElementById('pdf-pages');

      for (let pageNumber = 1; pageNumber <= meta.page_count; pageNumber++) {{
        const pageBox = document.createElement('div');
        pageBox.className = 'page-loading';
        pageBox.textContent = `Loading page ${{pageNumber}} of ${{meta.page_count}}...`;
        pages.appendChild(pageBox);

        const image = new Image();
        image.className = 'preview-page';
        image.alt = `${{FILE_NAME}} page ${{pageNumber}}`;
        image.loading = 'eager';
        image.decoding = 'async';
        const pageUrl = PDF_PAGE_URL_TEMPLATE.replace('__PAGE__', String(pageNumber)) + `&preview_ts=${{Date.now()}}`;

        await new Promise((resolve, reject) => {{
          const timeout = setTimeout(() => reject(new Error(`Page ${{pageNumber}} render timeout.`)), 25000);
          image.onload = () => {{ clearTimeout(timeout); resolve(); }};
          image.onerror = async () => {{
            clearTimeout(timeout);
            try {{
              const errorResponse = await fetch(pageUrl, {{ credentials: 'same-origin', cache: 'no-store' }});
              const errorData = await errorResponse.json().catch(() => ({{}}));
              reject(new Error(errorData.message || `Page ${{pageNumber}} failed to render.`));
            }} catch (innerErr) {{
              reject(new Error(`Page ${{pageNumber}} failed to render.`));
            }}
          }};
          image.src = pageUrl;
        }});

        pageBox.replaceWith(image);
      }}
    }}

    async function renderCsv() {{
      const response = await fetch(CONTENT_URL, {{ credentials: 'same-origin', cache: 'no-store' }});
      if (!response.ok) throw new Error('Unable to load text preview.');
      const text = await response.text();
      const pre = document.createElement('pre');
      pre.textContent = text;
      viewer.innerHTML = '';
      viewer.appendChild(pre);
    }}

    function renderImage() {{
      const image = document.createElement('img');
      image.className = 'preview-image';
      image.alt = FILE_NAME;
      image.src = CONTENT_URL;
      viewer.innerHTML = '';
      viewer.appendChild(image);
    }}

    (async function() {{
      try {{
        if (FILE_EXT === 'pdf') {{ await renderPdfServerPages(); return; }}
        if (['png','jpg','jpeg'].includes(FILE_EXT)) {{ renderImage(); return; }}
        if (FILE_EXT === 'csv') {{ await renderCsv(); return; }}
        showMessage('Preview not available for this file type', 'DOCX/XLSX previews are not supported inside the browser viewer. Please use Download.');
      }} catch (err) {{
        console.error(err);
        showMessage('Unable to preview this TSR', (err && err.message ? err.message + ' ' : '') + 'Please use Download.');
      }}
    }})();
  </script>
</body>
</html>"""

    return Response(viewer_html, mimetype='text/html', headers={'Cache-Control': 'no-store'})


@app.route('/preview_tsr_archive_pdf_meta/<int:file_id>')
@login_required
def preview_tsr_archive_pdf_meta(file_id):
    """Return PDF page count for server-rendered TSR preview."""
    resolved, error_response = _resolve_tsr_archive_file_for_preview(file_id)
    if error_response:
        return error_response

    if resolved['ext'] != 'pdf':
        return jsonify({'status': 'error', 'message': 'Not a PDF file.'}), 400

    try:
        import fitz  # PyMuPDF
        with fitz.open(resolved['file_path']) as pdf_doc:
            page_count = int(getattr(pdf_doc, 'page_count', 0) or len(pdf_doc))
        return jsonify({
            'status': 'success',
            'page_count': page_count,
            'filename': resolved['display_name'] or resolved['disk_name']
        })
    except Exception as preview_error:
        return jsonify({
            'status': 'error',
            'message': f'PDF preview renderer unavailable: {preview_error}'
        }), 500


@app.route('/preview_tsr_archive_pdf_page/<int:file_id>/<int:page_number>')
@login_required
def preview_tsr_archive_pdf_page(file_id, page_number):
    """Render one PDF page as PNG for Brave/mobile-safe TSR preview."""
    resolved, error_response = _resolve_tsr_archive_file_for_preview(file_id)
    if error_response:
        return error_response

    if resolved['ext'] != 'pdf':
        return jsonify({'status': 'error', 'message': 'Not a PDF file.'}), 400

    if page_number < 1:
        return jsonify({'status': 'error', 'message': 'Invalid page number.'}), 400

    try:
        import fitz  # PyMuPDF
        with fitz.open(resolved['file_path']) as pdf_doc:
            page_count = int(getattr(pdf_doc, 'page_count', 0) or len(pdf_doc))
            if page_number > page_count:
                return jsonify({'status': 'error', 'message': 'Page not found.'}), 404

            page = pdf_doc.load_page(page_number - 1)
            # Keep preview light enough for Brave/mobile and large scanned TSR PDFs.
            # A lower zoom prevents long renders that leave the viewer stuck on Loading.
            matrix = fitz.Matrix(1.35, 1.35)
            try:
                pix = page.get_pixmap(matrix=matrix, alpha=False, annots=True)
            except TypeError:
                pix = page.get_pixmap(matrix=matrix, alpha=False)
            image_bytes = pix.tobytes('jpeg', jpg_quality=82)

        return Response(
            image_bytes,
            mimetype='image/jpeg',
            headers={
                'Cache-Control': 'no-store',
                'Content-Disposition': f'inline; filename="tsr_preview_page_{page_number}.jpg"',
                'X-Content-Type-Options': 'nosniff'
            }
        )
    except Exception as preview_error:
        return jsonify({
            'status': 'error',
            'message': f'Unable to render PDF page: {preview_error}'
        }), 500


@app.route('/preview_tsr_archive_content/<int:file_id>')
@login_required
def preview_tsr_archive_content(file_id):
    """Authenticated raw TSR content endpoint used by the preview page for image/CSV rendering."""
    resolved, error_response = _resolve_tsr_archive_file_for_preview(file_id)
    if error_response:
        return error_response

    ext = resolved['ext']
    mimetype_map = {
        'png': 'image/png',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'csv': 'text/plain; charset=utf-8'
    }

    clean_preview_name = resolved['display_name'] or derive_original_filename_from_stored_filename(resolved['disk_name']) or resolved['disk_name']
    response = send_file(
        resolved['file_path'],
        as_attachment=False,
        download_name=clean_preview_name,
        mimetype=mimetype_map.get(ext, 'application/octet-stream')
    )
    response.headers['Content-Disposition'] = f'inline; filename="{clean_preview_name}"'
    response.headers['Cache-Control'] = 'no-store'
    return response


@app.route('/download_tsr_archive_file/<int:file_id>')
@login_required
def download_tsr_archive_file(file_id):
    """Authenticated TSR archive download with engineer/admin visibility guard."""
    resolved, error_response = _resolve_tsr_archive_file_for_preview(file_id)
    if error_response:
        return error_response

    disk_name = resolved['disk_name']
    clean_download_name = (
        resolved['display_name'] or
        derive_original_filename_from_stored_filename(disk_name) or
        disk_name or
        f'tsr_file_{file_id}'
    )
    response = send_file(resolved['file_path'], as_attachment=True, download_name=clean_download_name)
    response.headers['Content-Disposition'] = f'attachment; filename="{clean_download_name}"'
    response.headers['Cache-Control'] = 'no-store'
    return response


@app.route('/get_reports_summary')
@login_required
def get_reports_summary():
    """Reports dashboard API: TSR intelligence, product intelligence, and email report visibility."""
    if not is_admin_authorized():
        return denied()

    start_date, end_date = analytics_date_bounds()
    today = get_manila_time().date()

    range_query = Shift.query.options(
        joinedload(Shift.client),
        joinedload(Shift.product),
        selectinload(Shift.files)
    ).filter(
        func.date(Shift.start_time) >= start_date,
        func.date(Shift.start_time) <= end_date
    )
    range_query = analytics_scope_query(range_query)
    raw_shifts = range_query.order_by(Shift.start_time.desc()).all()
    shifts = list({shift.id: shift for shift in raw_shifts}.values())

    service_shifts = [shift for shift in shifts if shift.client_id]
    completed_service_shifts = [shift for shift in service_shifts if (shift.status or '') == 'Completed']

    def has_tsr_file(shift):
        filenames = []
        for file_rec in getattr(shift, 'files', []) or []:
            filenames.append(get_shift_file_display_name(file_rec))
            filenames.append(get_shift_file_disk_name(file_rec))
        return existing_files_have_tsr(filenames)

    tsr_attached = [shift for shift in service_shifts if has_tsr_file(shift)]
    missing_tsr = [shift for shift in completed_service_shifts if not has_tsr_file(shift)]

    missing_rows = []
    for shift in missing_tsr[:15]:
        engineers = get_shift_engineer_records(shift)
        missing_rows.append({
            'id': shift.id,
            'date': shift.start_time.strftime('%Y-%m-%d'),
            'client': shift.client.name if shift.client else 'N/A',
            'product': shift.product.name if shift.product else 'N/A',
            'serial': shift.product.serial_number if shift.product else '',
            'task': shift.title or '',
            'engineers': ', '.join([eng.name for eng in engineers]) or 'N/A'
        })

    tsr_by_client = {}
    tsr_by_engineer = {}
    tsr_by_product = {}

    for shift in tsr_attached:
        client_name = shift.client.name if shift.client else 'N/A'
        product_name = shift.product.name if shift.product else 'N/A'
        product_serial = shift.product.serial_number if shift.product else ''
        product_label = product_name if not product_serial else f"{product_name} ({product_serial})"
        tsr_by_client[client_name] = tsr_by_client.get(client_name, 0) + 1
        tsr_by_product[product_label] = tsr_by_product.get(product_label, 0) + 1
        for eng in get_shift_engineer_records(shift):
            tsr_by_engineer[eng.name] = tsr_by_engineer.get(eng.name, 0) + 1

    def top_items(counter, limit=8):
        return [
            {'label': label, 'count': count}
            for label, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))[:limit]
        ]

    all_products = Product.query.options(joinedload(Product.owner)).all()
    product_service_counts = {}
    product_last_service = {}

    for shift in service_shifts:
        if not shift.product_id:
            continue
        product_service_counts[shift.product_id] = product_service_counts.get(shift.product_id, 0) + 1
        current_last = product_last_service.get(shift.product_id)
        if not current_last or shift.start_time > current_last:
            product_last_service[shift.product_id] = shift.start_time

    high_maintenance = []
    for product in all_products:
        count = product_service_counts.get(product.serial_number, 0)
        if count <= 0:
            continue
        high_maintenance.append({
            'product': product.name,
            'serial': product.serial_number,
            'client': product.owner.name if product.owner else 'N/A',
            'count': count,
            'last_service': product_last_service.get(product.serial_number).strftime('%Y-%m-%d') if product_last_service.get(product.serial_number) else ''
        })
    high_maintenance.sort(key=lambda row: (-row['count'], row['client'], row['product']))

    repeat_service_products = [row for row in high_maintenance if row['count'] >= 2]

    warranty_active = 0
    warranty_expired = 0
    warranty_expiring = 0
    products_without_service = 0

    for product in all_products:
        if product.end_warranty_date:
            if product.end_warranty_date >= today:
                warranty_active += 1
                if product.end_warranty_date <= today + timedelta(days=90):
                    warranty_expiring += 1
            else:
                warranty_expired += 1
        if not product.shifts:
            products_without_service += 1

    email_query = ActivityLog.query.filter(
        func.date(ActivityLog.timestamp) >= start_date,
        func.date(ActivityLog.timestamp) <= end_date,
        ActivityLog.action.ilike('%Sent TSR to client%')
    ).order_by(ActivityLog.timestamp.desc())
    email_logs = email_query.limit(15).all()
    email_sent_count = email_query.count()

    recent_emails = []
    for log in email_logs:
        recent_emails.append({
            'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M'),
            'user': log.user,
            'action': log.action
        })

    return jsonify({
        'status': 'success',
        'range': {
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        },
        'branch': analytics_requested_branch(),
        'branch_label': analytics_branch_label(analytics_requested_branch()),
        'tsr': {
            'service_schedules': len(service_shifts),
            'completed_service_schedules': len(completed_service_shifts),
            'attached': len(tsr_attached),
            'missing': len(missing_tsr),
            'unsigned': 0,
            'pending_sync': 0,
            'by_client': top_items(tsr_by_client),
            'by_engineer': top_items(tsr_by_engineer),
            'by_product': top_items(tsr_by_product),
            'missing_rows': missing_rows
        },
        'product': {
            'total': len(all_products),
            'warranty_active': warranty_active,
            'warranty_expired': warranty_expired,
            'warranty_expiring': warranty_expiring,
            'without_service': products_without_service,
            'repeat_service_count': len(repeat_service_products),
            'high_maintenance': high_maintenance[:15]
        },
        'email': {
            'sent_count': email_sent_count,
            'failed_count': 0,
            'recent': recent_emails
        }
    })


@app.route('/export_reports_summary')
@login_required
def export_reports_summary():
    """CSV export for Reports page TSR/product/email intelligence."""
    if not is_admin_authorized():
        return denied()

    start_date, end_date = analytics_date_bounds()

    range_query = Shift.query.options(
        joinedload(Shift.client),
        joinedload(Shift.product),
        selectinload(Shift.files)
    ).filter(
        func.date(Shift.start_time) >= start_date,
        func.date(Shift.start_time) <= end_date
    )
    range_query = analytics_scope_query(range_query)
    shifts = list({shift.id: shift for shift in range_query.order_by(Shift.start_time.asc()).all()}.values())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Client', 'Product', 'Serial', 'Task', 'Status', 'TSR Attached', 'Engineers'])

    for shift in shifts:
        if not shift.client_id:
            continue
        filenames = []
        for file_rec in getattr(shift, 'files', []) or []:
            filenames.append(get_shift_file_display_name(file_rec))
            filenames.append(get_shift_file_disk_name(file_rec))
        engineers = get_shift_engineer_records(shift)
        writer.writerow([
            shift.start_time.strftime('%Y-%m-%d'),
            shift.client.name if shift.client else '',
            shift.product.name if shift.product else '',
            shift.product.serial_number if shift.product else '',
            shift.title or '',
            shift.status or '',
            'Yes' if existing_files_have_tsr(filenames) else 'No',
            ', '.join([eng.name for eng in engineers])
        ])

    output.seek(0)
    log_activity(f"Exported reports summary: {start_date.isoformat()} to {end_date.isoformat()}")
    filename = f"reports_summary_{start_date.isoformat()}_{end_date.isoformat()}.csv"
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@app.route('/export_analytics_summary')
@login_required
def export_analytics_summary():
    """CSV export for selected analytics range and branch."""
    if not is_admin_authorized():
        return denied()

    start_date, end_date = analytics_date_bounds()
    branch = analytics_requested_branch()

    range_query = Shift.query.filter(
        func.date(Shift.start_time) >= start_date,
        func.date(Shift.start_time) <= end_date
    )
    range_query = analytics_scope_query(range_query)
    raw_shifts = range_query.order_by(Shift.start_time.asc()).all()
    shifts = list({shift.id: shift for shift in raw_shifts}.values())

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Date', 'Start', 'End', 'Client', 'Task', 'Product', 'Status', 'Type', 'Engineers', 'Branches'])

    for shift in shifts:
        assigned_engineers = get_shift_engineer_records(shift)
        writer.writerow([
            shift.start_time.strftime('%Y-%m-%d'),
            shift.start_time.strftime('%H:%M'),
            shift.end_time.strftime('%H:%M'),
            shift.client.name if shift.client else '',
            shift.title,
            shift.product.name if shift.product else '',
            shift.status or '',
            classify_schedule_type(shift),
            ', '.join([eng.name for eng in assigned_engineers]),
            ', '.join(sorted({eng.branch or '' for eng in assigned_engineers}))
        ])

    output.seek(0)
    log_activity(f"Exported analytics summary: {start_date.isoformat()} to {end_date.isoformat()} ({analytics_branch_label(branch)})")
    filename = f"analytics_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}_{branch.lower()}.csv"
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-disposition': f'attachment; filename={filename}'}
    )


# --- CSV REPORT GENERATION MODULE ---

@app.route('/export_clients')
@login_required
def export_clients():
    """Full customer database CSV dump using dynamic Contact rows."""
    if not is_admin_authorized(): return denied()
    clients = Client.query.order_by(Client.name).all()

    max_contacts = 0
    client_contacts = {}
    for c in clients:
        contacts = Contact.query.filter_by(client_id=c.id).all()
        client_contacts[c.id] = contacts
        max_contacts = max(max_contacts, len(contacts))

    output = io.StringIO()
    writer = csv.writer(output)

    headers = ['Name', 'Address']
    for i in range(1, max_contacts + 1):
        headers.extend([f'Contact {i} Name', f'Contact {i} Phone', f'Contact {i} Email'])
    writer.writerow(headers)

    for c in clients:
        row = [c.name, c.address]
        for contact in client_contacts.get(c.id, []):
            row.extend([contact.name or '', contact.phone or '', contact.email or ''])
        for _ in range(max_contacts - len(client_contacts.get(c.id, []))):
            row.extend(['', '', ''])
        writer.writerow(row)

    output.seek(0)
    log_activity("Exported the Client Database")
    filename = f"medical_centers_{get_manila_time().strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-disposition": f"attachment; filename={filename}"})

@app.route('/export_engineers')
@login_required
def export_engineers():
    """ Technical personnel list CSV dump """
    if not is_admin_authorized(): return denied()
    engineers = Engineer.query.order_by(Engineer.name).all()
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Emp ID', 'Engineer Name', 'Shorthand Initials', 'Branch', 'Phone', 'Email'])
    for e in engineers:
        writer.writerow([e.employee_id, e.name, e.initials, e.branch, e.phone, e.email])
    output.seek(0)
    log_activity("Exported the Personnel Directory")
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-disposition": "attachment; filename=technicians_list.csv"})


@app.route('/export_products')
@login_required
def export_products():
    """ Machine Inventory CSV dump with warranty metadata """
    if not is_admin_authorized(): return denied()
    products = Product.query.all()
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Serial Number', 'Description', 'Current Owner', 'Start Date', 'Expiry Date'])
    for p in products:
        writer.writerow([p.serial_number, p.name, p.owner.name if p.owner else "N/A", p.start_warranty_date, p.end_warranty_date])
    output.seek(0)
    log_activity("Exported the Product Inventory")
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-disposition": "attachment; filename=inventory.csv"})


@app.route('/export_timeline')
@login_required
def export_timeline():
    """Weekly Grid schedule snapshot CSV dump aligned with multi-engineer assignments."""
    if not is_admin_authorized(): return denied()
    offset = clean_int(request.args.get('offset', 0)) or 0
    branch_filter = clean_str(request.args.get('branch')) or 'ALL'

    target_dt = (get_manila_time()).date() + timedelta(weeks=offset)
    target_dt -= timedelta(days=target_dt.weekday())
    days = [(target_dt + timedelta(days=i)) for i in range(7)]

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Technical Staff', 'Branch'] + [d.strftime('%b %d (%a)') for d in days])

    personnel_query = Engineer.query.order_by(Engineer.branch, Engineer.name)
    if branch_filter == 'REGIONAL':
        personnel_query = personnel_query.filter(Engineer.branch.in_(REGIONAL_ADMIN_BRANCHES))
    elif branch_filter != 'ALL':
        personnel_query = personnel_query.filter(Engineer.branch == branch_filter)

    personnel = personnel_query.all()

    for eng in personnel:
        row_data = [eng.name, eng.branch or '']
        for day in days:
            shifts = (
                db.session.query(Shift)
                .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
                .filter(
                    ShiftEngineer.engineer_id == eng.id,
                    func.date(Shift.start_time) == day
                )
                .order_by(Shift.start_time)
                .all()
            )

            if shifts:
                cell_items = []
                for s in shifts:
                    time_label = f"{s.start_time.strftime('%H:%M')}-{s.end_time.strftime('%H:%M')}"
                    client_label = s.client.name if s.client else ''
                    product_label = s.product.name if s.product else ''
                    status_label = s.status or ''
                    detail_parts = [f"[{time_label}] {s.title}"]
                    if client_label:
                        detail_parts.append(client_label)
                    if product_label:
                        detail_parts.append(f"Item: {product_label}")
                    if status_label and s.client_id:
                        detail_parts.append(status_label)
                    cell_items.append(" / ".join(detail_parts))
                row_data.append(" | ".join(cell_items))
            else:
                row_data.append("-")
        writer.writerow(row_data)

    output.seek(0)
    log_activity("Exported Weekly Schedule Snapshot")
    filename = f"weekly_schedule_{days[0].strftime('%Y%m%d')}_{days[-1].strftime('%Y%m%d')}.csv"
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-disposition": f"attachment; filename={filename}"})

@app.route('/check_client_duplicate')
@login_required
def check_client_duplicate():
    name = request.args.get('name')
    addr = request.args.get('address')

    if not name:
        return jsonify({'match': False})

    c = check_for_duplicate_client(name, addr)
    if c:
        return jsonify({
            'match': True,
            'name': c.name,
            'address': c.address,
            'id': c.id
        })
    return jsonify({'match': False})


@app.route('/search_clients')
@login_required
def search_clients():
    q = request.args.get('q', '').strip()
    addr = request.args.get('address', '').strip()

    if not q:
        return jsonify([])

    results = []
    for c in Client.query.all():
        score = similarity(q.upper(), c.name.upper())

        # boost if acronym match
        if generate_acronym(q) == generate_acronym(c.name):
            score += 0.2

        # boost if address similar
        if normalize_address(addr) == normalize_address(c.address):
            score += 0.1

        if score > 0.6:
            results.append({
                'id': c.id,
                'name': c.name,
                'address': c.address,
                'score': round(score, 2)
            })

    results = sorted(results, key=lambda x: x['score'], reverse=True)[:5]
    return jsonify(results)


# --- CLIENT EXCEL EXPORT ---
from openpyxl import Workbook

@app.route('/export_client_excel/<int:client_id>')
@login_required
def export_client_excel(client_id):
    if not is_admin_authorized(): return denied()
    client = db.session.get(Client, client_id)
    if not client:
        return "Not found", 404

    contacts = Contact.query.filter_by(client_id=client_id).all()
    products = Product.query.filter_by(client_id=client_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Report"

    row = 1

    ws.cell(row=row, column=1, value="Client Name")
    ws.cell(row=row, column=2, value=client.name)
    row += 1
    ws.cell(row=row, column=1, value="Address")
    ws.cell(row=row, column=2, value=client.address)
    row += 2

    ws.cell(row=row, column=1, value="Contacts")
    row += 1
    ws.cell(row=row, column=1, value="Name")
    ws.cell(row=row, column=2, value="Phone")
    ws.cell(row=row, column=3, value="Email")
    row += 1

    for c in contacts:
        ws.cell(row=row, column=1, value=c.name)
        ws.cell(row=row, column=2, value=c.phone)
        ws.cell(row=row, column=3, value=c.email)
        row += 1

    row += 1

    ws.cell(row=row, column=1, value="Products")
    row += 1
    ws.cell(row=row, column=1, value="Serial")
    ws.cell(row=row, column=2, value="Name")
    ws.cell(row=row, column=3, value="Warranty Start")
    ws.cell(row=row, column=4, value="Warranty End")
    ws.cell(row=row, column=5, value="Status")
    row += 1

    from datetime import datetime
    today = datetime.today().date()

    for p in products:
        status = "Active" if p.end_warranty_date and p.end_warranty_date >= today else "Expired"
        ws.cell(row=row, column=1, value=p.serial_number)
        ws.cell(row=row, column=2, value=p.name)
        ws.cell(row=row, column=3, value=str(p.start_warranty_date))
        ws.cell(row=row, column=4, value=str(p.end_warranty_date))
        ws.cell(row=row, column=5, value=status)
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{client.name}_report.xlsx".replace(" ", "_")

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

def can_engineer_manage_client_contacts():
    """Engineers may add/edit client contact rows but may not delete clients/contacts."""
    return bool(
        current_user and
        getattr(current_user, 'is_authenticated', False) and
        getattr(current_user, 'role', None) == 'engineer'
    )


def apply_client_contacts_without_deleting_existing(client_id, payload):
    """Add/edit client contacts while preserving existing rows.

    Existing contacts are matched by visible order because the legacy frontend
    submits cp1/cn1/ce1 style fields without contact IDs. Empty submitted rows
    are ignored instead of deleting existing records.
    """
    existing_contacts = Contact.query.filter_by(client_id=client_id).order_by(Contact.id.asc()).all()
    submitted_rows = []

    i = 1
    while f'cp{i}' in payload or f'cn{i}' in payload or f'ce{i}' in payload:
        name = clean_str(payload.get(f'cp{i}'))
        phone = clean_str(payload.get(f'cn{i}'))
        email = clean_str(payload.get(f'ce{i}'))

        if name or phone or email:
            submitted_rows.append((name, phone, email))
        i += 1

    for idx, (name, phone, email) in enumerate(submitted_rows):
        if idx < len(existing_contacts):
            contact = existing_contacts[idx]
            contact.name = name
            contact.phone = phone
            contact.email = email
        else:
            db.session.add(Contact(
                client_id=client_id,
                name=name,
                phone=phone,
                email=email
            ))

    # Preserve legacy first 3 contact columns from merged active contacts.
    merged_rows = submitted_rows[:]
    if len(existing_contacts) > len(submitted_rows):
        for contact in existing_contacts[len(submitted_rows):]:
            merged_rows.append((contact.name, contact.phone, contact.email))

    legacy_rows = merged_rows[:3]
    legacy_fields = [
        ('contact_person_1', 'contact_number_1', 'email_address_1'),
        ('contact_person_2', 'contact_number_2', 'email_address_2'),
        ('contact_person_3', 'contact_number_3', 'email_address_3')
    ]

    client_rec = db.session.get(Client, client_id)
    if client_rec:
        for idx, fields in enumerate(legacy_fields):
            values = legacy_rows[idx] if idx < len(legacy_rows) else (None, None, None)
            setattr(client_rec, fields[0], values[0])
            setattr(client_rec, fields[1], values[1])
            setattr(client_rec, fields[2], values[2])

    return len(submitted_rows)


# --- CLIENT MANAGEMENT CORE ACTIONS ---

@app.route('/add_client', methods=['POST'])
@login_required
def add_client():
    """ Hospital record entry with conflict detection. Access: Admin Levels. """
    if not is_admin_authorized(): return jsonify({'message': 'Denied'}), 403
    payload = request.get_json()
    name = clean_str(payload.get('name')); addr = clean_str(payload.get('address'))
    
    if not payload.get('force'):
        collision = check_for_duplicate_client(name, addr)
        if collision: 
            return jsonify({'status': 'conflict', 'message': f'Duplicate Found: "{collision.name}"', 'existing_id': collision.id}), 409
            
    new_hospital = Client(
        name=name, address=addr, 
        contact_person_1=clean_str(payload.get('cp1')), contact_number_1=clean_str(payload.get('cn1')), email_address_1=clean_str(payload.get('ce1')),
        contact_person_2=clean_str(payload.get('cp2')), contact_number_2=clean_str(payload.get('cn2')), email_address_2=clean_str(payload.get('ce2')),
        contact_person_3=clean_str(payload.get('cp3')), contact_number_3=clean_str(payload.get('cn3')), email_address_3=clean_str(payload.get('ce3'))
    )
    db.session.add(new_hospital)
    db.session.flush()
    i = 1
    while f'cp{i}' in payload:
        db.session.add(Contact(
            client_id=new_hospital.id,
            name=clean_str(payload.get(f'cp{i}')),
            phone=clean_str(payload.get(f'cn{i}')),
            email=clean_str(payload.get(f'ce{i}'))
        ))
        i += 1
    db.session.commit()
    log_activity(f"Added new client: {name}")
    return jsonify({'status': 'success'})


@app.route('/update_client/<int:id>', methods=['PUT'])
@login_required
def update_client(id):
    """Hospital record modification logic.

    Admins can edit full client details and contact rows.
    Engineers can add/edit contact rows only; they cannot delete contacts or
    modify the client name/address.
    """
    payload = request.get_json(silent=True) or {}
    client_rec = db.session.get(Client, id)
    if not client_rec:
        return jsonify({'message': 'Not Found'}), 404

    if can_engineer_manage_client_contacts() and not is_admin_authorized():
        submitted_count = apply_client_contacts_without_deleting_existing(id, payload)
        db.session.commit()
        log_activity(f"Updated client contacts: {client_rec.name}")
        return jsonify({
            'status': 'success',
            'contacts_updated': submitted_count,
            'contact_only_update': True
        })

    if not is_admin_authorized():
        return jsonify({'message': 'Denied'}), 403

    client_rec.name, client_rec.address = clean_str(payload.get('name')), clean_str(payload.get('address'))
    client_rec.contact_person_1, client_rec.contact_number_1, client_rec.email_address_1 = clean_str(payload.get('cp1')), clean_str(payload.get('cn1')), clean_str(payload.get('ce1'))
    client_rec.contact_person_2, client_rec.contact_number_2, client_rec.email_address_2 = clean_str(payload.get('cp2')), clean_str(payload.get('cn2')), clean_str(payload.get('ce2'))
    client_rec.contact_person_3, client_rec.contact_number_3, client_rec.email_address_3 = None, None, None

    Contact.query.filter_by(client_id=id).delete()
    i = 1
    while f'cp{i}' in payload:
        db.session.add(Contact(
            client_id=id,
            name=clean_str(payload.get(f'cp{i}')),
            phone=clean_str(payload.get(f'cn{i}')),
            email=clean_str(payload.get(f'ce{i}'))
        ))
        i += 1

    db.session.commit()
    log_activity(f"Modified details for client: {client_rec.name}")
    return jsonify({'status': 'success'})


@app.route('/delete_client/<int:id>', methods=['DELETE'])
@login_required
def delete_client(id):
    """ Hospital removal logic. Access: Admin Levels. """
    if not is_admin_authorized(): return jsonify({'message': 'Denied'}), 403
    target = db.session.get(Client, id)
    if target: 
        name = target.name
        db.session.delete(target); db.session.commit()
        log_activity(f"Permanently removed Client: {name}")
    return jsonify({'status': 'success'})


# --- SCHEDULE MANAGEMENT CORE ACTIONS ---


def build_conflict_response(collision, engineer=None, message_prefix='Schedule conflict detected'):
    """Return detailed conflict information for add/edit/drag-drop scheduling UX."""
    conflict_engineer = engineer
    if not conflict_engineer and collision:
        assigned = get_shift_engineer_records(collision)
        conflict_engineer = assigned[0] if assigned else None

    conflict_date = collision.start_time.strftime('%Y-%m-%d') if collision and collision.start_time else ''
    conflict_time = ''
    if collision and collision.start_time and collision.end_time:
        conflict_time = f"{collision.start_time.strftime('%I:%M %p')} - {collision.end_time.strftime('%I:%M %p')}"

    conflict_client = collision.client.name if collision and collision.client else ''
    conflict_product = collision.product.name if collision and collision.product else ''
    conflict_task = collision.title if collision else ''

    parts = [message_prefix]
    if conflict_engineer:
        parts.append(f"Engineer: {conflict_engineer.name}")
    if conflict_date:
        parts.append(f"Date: {conflict_date}")
    if conflict_time:
        parts.append(f"Time: {conflict_time}")
    if conflict_task:
        parts.append(f"Existing task: {conflict_task}")
    if conflict_client:
        parts.append(f"Client: {conflict_client}")

    return jsonify({
        'status': 'conflict',
        'message': " | ".join(parts),
        'conflict': {
            'engineer': conflict_engineer.name if conflict_engineer else 'Engineer',
            'engineer_id': conflict_engineer.id if conflict_engineer else None,
            'date': conflict_date,
            'time': conflict_time,
            'task': conflict_task,
            'client': conflict_client,
            'product': conflict_product,
            'status': collision.status if collision else ''
        }
    }), 409



def find_add_schedule_collision(engineer_id, iter_date, start_dt, end_dt):
    """Return a visible same-day collision for add_shift.

    The calendar stores current multi-day schedules as one Shift row per date.
    Some older live rows may still span multiple calendar dates in a single row.
    Those legacy rows are not displayed on the middle/end dates, but the old raw
    datetime overlap query still treated them as collisions. This helper keeps
    conflict checks aligned with what users see on the calendar: only schedules
    whose Shift.start_time calendar date is the date being added can block that
    date.
    """
    if not engineer_id or not iter_date or not start_dt or not end_dt:
        return None

    return (
        db.session.query(Shift)
        .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
        .filter(
            ShiftEngineer.engineer_id == engineer_id,
            func.date(Shift.start_time) == iter_date,
            Shift.start_time < end_dt,
            Shift.end_time > start_dt
        )
        .order_by(Shift.start_time.asc(), Shift.id.asc())
        .first()
    )

def get_shift_payload():
    """Return request data from either JSON or multipart/form-data."""
    if request.is_json:
        return request.get_json(silent=True) or {}
    return request.form


def normalize_shift_title(payload):
    """Return a safe non-empty shift title for all schedule categories."""
    title = clean_str(payload.get('title'))
    if not title:
        return None
    return title


def parse_engineer_ids(raw_value, fallback=None):
    """Parse engineer IDs from JSON array string, list, or single fallback value."""
    import json

    engineers = raw_value
    if isinstance(engineers, str):
        try:
            engineers = json.loads(engineers)
        except Exception:
            engineers = [engineers] if engineers.strip() else []

    if not engineers and fallback:
        engineers = [fallback]

    cleaned = []
    for item in engineers or []:
        eng_id = clean_int(item)
        if eng_id and eng_id not in cleaned:
            cleaned.append(eng_id)
    return cleaned


def parse_bool_flag(raw_value, default=False):
    """Parse browser form/json boolean values safely."""
    if raw_value is None:
        return default

    if isinstance(raw_value, bool):
        return raw_value

    return str(raw_value).strip().lower() in {'1', 'true', 'yes', 'on'}


def get_schedule_dates_between(start_date, end_date, include_weekends=True):
    """Return schedule dates between start/end, optionally skipping weekends."""
    dates = []
    current_date = start_date

    while current_date <= end_date:
        if include_weekends or current_date.weekday() < 5:
            dates.append(current_date)
        current_date += timedelta(days=1)

    return dates


def has_weekend_between(start_date, end_date):
    """Return True if a date range includes Saturday or Sunday."""
    current_date = start_date

    while current_date <= end_date:
        if current_date.weekday() >= 5:
            return True
        current_date += timedelta(days=1)

    return False


def should_preserve_existing_schedule_dates(existing_chain, requested_start, requested_end):
    """Return True when an edit targets the same outer date range.

    This is critical for schedules previously extended with weekends skipped:
    the modal still shows the outer range, but the database chain contains only
    actual scheduled dates. Status/detail-only updates must preserve that date
    set and must not recreate missing Saturday/Sunday rows.
    """
    if not existing_chain or not requested_start or not requested_end:
        return False

    existing_dates = sorted({shift.start_time.date() for shift in existing_chain if shift.start_time})
    if not existing_dates:
        return False

    return existing_dates[0] == requested_start and existing_dates[-1] == requested_end


def save_uploaded_shift_files(shift):
    """Persist uploaded report files and link them to a Shift record.

    Server storage remains randomized to avoid collisions/security issues.
    original_filename stores the clean client-facing upload filename.
    """
    ensure_shift_file_original_filename_column()

    uploaded_files = request.files.getlist('report_file')
    saved_filenames = []

    for file_obj in uploaded_files:
        if not file_obj or not file_obj.filename:
            continue
        if not allowed_file(file_obj.filename):
            # Uploads are validated before database changes; this is a final safety guard.
            continue

        original_name = secure_filename(file_obj.filename)
        if not original_name:
            continue

        unique_name = f"shift_{shift.id}_{secrets.token_hex(8)}_{original_name}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
        file_obj.save(file_path)

        db.session.add(ShiftFile(
            shift_id=shift.id,
            filename=unique_name,
            original_filename=original_name
        ))
        saved_filenames.append(unique_name)

    return saved_filenames


def delete_shift_engineer_links(shift_id):
    """Remove engineer assignment rows for a shift before deleting/reusing it."""
    ShiftEngineer.query.filter_by(shift_id=shift_id).delete(synchronize_session=False)


def repair_shift_engineer_links():
    """Remove orphan/duplicate ShiftEngineer rows left by older versions.

    This does not delete schedules. It only cleans assignment rows that point to
    missing shifts/engineers or exact duplicate links.
    """
    valid_shift_ids = {sid for (sid,) in db.session.query(Shift.id).all()}
    valid_engineer_ids = {eid for (eid,) in db.session.query(Engineer.id).all()}
    seen = set()
    changed = False

    for link in ShiftEngineer.query.all():
        key = (link.shift_id, link.engineer_id)
        if (link.shift_id not in valid_shift_ids or
                link.engineer_id not in valid_engineer_ids or
                key in seen):
            db.session.delete(link)
            changed = True
        else:
            seen.add(key)

    if changed:
        db.session.commit()



def get_shift_table_columns():
    """Return current SQLite columns for the shift table."""
    rows = db.session.execute(db.text("PRAGMA table_info(shift)")).fetchall()
    return {row[1] for row in rows}


def ensure_shift_override_columns():
    """Safe live migration for linked engineer/day time overrides.

    This adds nullable columns only. It does not delete, rewrite, or backfill
    existing schedules, so it is safe for the live SQLite database.
    """
    existing_columns = get_shift_table_columns()
    migrations = [
        ('parent_shift_id', "ALTER TABLE shift ADD COLUMN parent_shift_id INTEGER"),
        ('override_engineer_id', "ALTER TABLE shift ADD COLUMN override_engineer_id INTEGER"),
        ('override_kind', "ALTER TABLE shift ADD COLUMN override_kind VARCHAR(30)")
    ]

    changed = False
    for column_name, ddl in migrations:
        if column_name not in existing_columns:
            db.session.execute(db.text(ddl))
            changed = True
            print(f"[DB MIGRATION] Added shift.{column_name}", flush=True)

    if changed:
        db.session.commit()




def get_shift_file_table_columns():
    """Return current SQLite columns for the shift_file table."""
    rows = db.session.execute(db.text("PRAGMA table_info(shift_file)")).fetchall()
    return {row[1] for row in rows}


def ensure_shift_file_original_filename_column():
    """Safe live SQLite migration for original attachment names.

    This must be callable before timeline file loading because SQLAlchemy will
    select ShiftFile.original_filename as soon as Shift.files is loaded.
    """
    global _shift_file_original_filename_ready

    if _shift_file_original_filename_ready:
        return

    existing_columns = get_shift_file_table_columns()
    changed = False

    if 'original_filename' not in existing_columns:
        db.session.execute(db.text("ALTER TABLE shift_file ADD COLUMN original_filename VARCHAR(200)"))
        changed = True
        print("[DB MIGRATION] Added shift_file.original_filename", flush=True)

    rows = db.session.execute(
        db.text("SELECT id, filename, original_filename FROM shift_file")
    ).fetchall()

    for row in rows:
        row_id = row[0]
        stored_filename = row[1]
        current_original = row[2]

        if current_original:
            continue

        display_name = derive_original_filename_from_stored_filename(stored_filename)
        if not display_name:
            continue

        db.session.execute(
            db.text("UPDATE shift_file SET original_filename = :display_name WHERE id = :row_id"),
            {'display_name': display_name, 'row_id': row_id}
        )
        changed = True

    if changed:
        db.session.commit()

    _shift_file_original_filename_ready = True



def ensure_schedule_delete_indexes():
    """Safe live migration: add non-unique indexes used by preview/range delete.

    CREATE INDEX IF NOT EXISTS is safe for live SQLite and does not delete data.
    """
    index_ddls = [
        "CREATE INDEX IF NOT EXISTS idx_shift_group_start_delete ON shift (group_id, start_time)",
        "CREATE INDEX IF NOT EXISTS idx_shift_start_delete ON shift (start_time)",
        "CREATE INDEX IF NOT EXISTS idx_shiftengineer_engineer_shift_delete ON shift_engineer (engineer_id, shift_id)",
        "CREATE INDEX IF NOT EXISTS idx_shift_start_status_timeline ON shift (start_time, status)",
        "CREATE INDEX IF NOT EXISTS idx_shift_branch_owner_timeline ON shift (engineer_id, start_time)",
        "CREATE INDEX IF NOT EXISTS idx_shiftengineer_shift_engineer_timeline ON shift_engineer (shift_id, engineer_id)",
        "CREATE INDEX IF NOT EXISTS idx_shiftfile_shift_uploaded_timeline ON shift_file (shift_id, uploaded_at)"
    ]

    for ddl in index_ddls:
        db.session.execute(db.text(ddl))

    db.session.commit()



def normalize_edit_scope(scope_value):
    """Normalize edit-scope aliases from the timeline UI.

    Safety note:
    Same-day assigned-engineer edits must never fall through to the
    full linked-chain rebuild path, because that path deletes/recreates
    the whole base chain.
    """
    raw_scope = (scope_value or '').strip().lower()

    assigned_day_aliases = {
        'assigned_engineers_day_only',
        'assigned_engineers_day',
        'assigned_day_only',
        'all_assigned_engineers_day_only',
        'all_assigned_engineers_same_day',
        'assigned_engineers_same_day',
        'same_day_assigned_engineers',
        'this_day_assigned_engineers',
        'this_day_only',
        'day_only'
    }

    time_override_aliases = {
        'engineer_day_time_override',
        'this_engineer_day',
        'engineer_day',
        'this_engineer_day_only',
        'this_engineer_same_day',
        'custom_time',
        'custom_time_override'
    }

    if raw_scope in assigned_day_aliases:
        return 'assigned_engineers_day_only'

    if raw_scope in time_override_aliases:
        return 'engineer_day_time_override'

    return 'entire_schedule'

def is_shift_time_override(shift):
    """Return True when a shift is a linked custom-time child schedule."""
    return bool(
        shift and
        getattr(shift, 'parent_shift_id', None) and
        getattr(shift, 'override_kind', None) == 'time_override'
    )


def get_base_group_shifts(group_id, fallback_shift=None):
    """Return non-override shifts in a group, preserving legacy behavior."""
    if group_id:
        return (
            Shift.query
            .filter(
                Shift.group_id == group_id,
                or_(Shift.override_kind.is_(None), Shift.override_kind == '')
            )
            .order_by(Shift.start_time.asc())
            .all()
        )
    return [fallback_shift] if fallback_shift else []


def get_linked_time_overrides_for_shift_ids(shift_ids):
    """Return linked child time overrides for parent shift ids."""
    clean_ids = [sid for sid in (shift_ids or []) if sid]
    if not clean_ids:
        return []
    return (
        Shift.query
        .filter(
            Shift.parent_shift_id.in_(clean_ids),
            Shift.override_kind == 'time_override'
        )
        .order_by(Shift.start_time.asc())
        .all()
    )


def copy_shared_schedule_fields(target_shift, source_shift):
    """Sync shared job details while preserving target time/date/engineer override."""
    target_shift.title = source_shift.title
    target_shift.client_id = source_shift.client_id
    target_shift.product_id = source_shift.product_id
    target_shift.status = source_shift.status
    target_shift.group_id = source_shift.group_id


def replace_shift_files_from_names(target_shift, filenames):
    """Replace a shift's file rows using preserved filenames.

    Also preserves clean client-facing original filenames when possible.
    """
    ensure_shift_file_original_filename_column()

    clean_filenames = [
        os.path.basename(clean_str(filename))
        for filename in (filenames or [])
        if clean_str(filename)
    ]

    existing_display_by_filename = {}
    if clean_filenames:
        existing_rows = ShiftFile.query.filter(ShiftFile.filename.in_(clean_filenames)).all()
        existing_display_by_filename = {
            file_rec.filename: get_shift_file_display_name(file_rec)
            for file_rec in existing_rows
        }

    ShiftFile.query.filter_by(shift_id=target_shift.id).delete(synchronize_session=False)

    for filename in clean_filenames:
        db.session.add(ShiftFile(
            shift_id=target_shift.id,
            filename=filename,
            original_filename=(
                existing_display_by_filename.get(filename) or
                derive_original_filename_from_stored_filename(filename)
            )
        ))


def sync_override_files_from_parent(parent_shift, override_shift):
    """Keep linked override attachments aligned with the parent schedule."""
    filenames = [file_rec.filename for file_rec in parent_shift.files]
    replace_shift_files_from_names(override_shift, filenames)


def create_or_update_time_override(parent_shift, engineer_id, start_dt, end_dt):
    """Create/update a linked custom-time child schedule for one engineer/day.

    The child remains connected to the parent via parent_shift_id and group_id.
    Later shared-detail edits on the parent can sync to this child while keeping
    the custom time untouched.
    """
    if not parent_shift or not engineer_id:
        return None

    override = (
        Shift.query
        .filter_by(
            parent_shift_id=parent_shift.id,
            override_engineer_id=engineer_id,
            override_kind='time_override'
        )
        .first()
    )

    if not override:
        override = Shift(
            title=parent_shift.title,
            start_time=start_dt,
            end_time=end_dt,
            engineer_id=engineer_id,
            client_id=parent_shift.client_id,
            product_id=parent_shift.product_id,
            status=parent_shift.status,
            created_at=parent_shift.created_at or get_manila_time(),
            group_id=parent_shift.group_id,
            parent_shift_id=parent_shift.id,
            override_engineer_id=engineer_id,
            override_kind='time_override'
        )
        db.session.add(override)
        db.session.flush()
    else:
        copy_shared_schedule_fields(override, parent_shift)
        override.start_time = start_dt
        override.end_time = end_dt
        override.engineer_id = engineer_id
        override.override_engineer_id = engineer_id
        override.parent_shift_id = parent_shift.id
        override.override_kind = 'time_override'

    delete_shift_engineer_links(override.id)
    db.session.add(ShiftEngineer(shift_id=override.id, engineer_id=engineer_id))
    sync_override_files_from_parent(parent_shift, override)
    return override



def ensure_single_engineer_link(shift, engineer_id):
    """Ensure an override/solo shift has exactly one ShiftEngineer link."""
    if not shift or not engineer_id:
        return

    delete_shift_engineer_links(shift.id)
    db.session.add(ShiftEngineer(shift_id=shift.id, engineer_id=engineer_id))
    shift.engineer_id = engineer_id


def sync_override_shared_fields_preserve_time(override_shift, parent_shift):
    """Sync shared details from parent to override while preserving custom time and engineer."""
    if not override_shift or not parent_shift:
        return

    override_engineer_id = override_shift.override_engineer_id or override_shift.engineer_id
    copy_shared_schedule_fields(override_shift, parent_shift)
    override_shift.parent_shift_id = parent_shift.id
    override_shift.group_id = parent_shift.group_id
    override_shift.override_kind = 'time_override'
    override_shift.override_engineer_id = override_engineer_id

    if override_engineer_id:
        ensure_single_engineer_link(override_shift, override_engineer_id)

    sync_override_files_from_parent(parent_shift, override_shift)


def sync_linked_time_overrides(parent_shift):
    """Propagate shared parent job details to linked time overrides only."""
    if not parent_shift:
        return

    overrides = get_linked_time_overrides_for_shift_ids([parent_shift.id])
    for override in overrides:
        sync_override_shared_fields_preserve_time(override, parent_shift)


def detach_engineer_from_parent_shift_for_override(parent_shift, engineer_id):
    """Remove one engineer from the base shared shift when that engineer has a time override."""
    if not parent_shift or not engineer_id:
        return

    ShiftEngineer.query.filter_by(
        shift_id=parent_shift.id,
        engineer_id=engineer_id
    ).delete(synchronize_session=False)

    remaining_ids = [
        se.engineer_id
        for se in ShiftEngineer.query.filter_by(shift_id=parent_shift.id).all()
    ]

    if remaining_ids:
        parent_shift.engineer_id = remaining_ids[0]



CLIENT_EMAIL_REGEX = re.compile(r'(?<![A-Z0-9._%+\-])([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})(?![A-Z0-9._%+\-])', re.IGNORECASE)


NON_CLIENT_EMAIL_DOMAINS = {
    'shimadzu.com.ph',
    'verisign.com',
    'symantec.com',
    'digicert.com',
    'globalsign.com',
    'sectigo.com',
    'comodoca.com',
    'adobe.com',
    'microsoft.com'
}

NON_CLIENT_EMAIL_PREFIXES = {
    'cps-requests',
    'no-reply',
    'noreply',
    'support',
    'admin',
    'administrator',
    'webmaster',
    'postmaster'
}



def is_valid_manual_recipient_email(email_addr):
    """Validate manually typed recipients.

    Manual send recipients may include internal Shimadzu emails.
    Certificate/system metadata emails remain blocked.
    """
    email_addr = clean_str(email_addr)
    if not email_addr:
        return False

    lowered = email_addr.lower()
    if not CLIENT_EMAIL_REGEX.fullmatch(lowered):
        return False

    local_part, domain = lowered.rsplit('@', 1)

    manual_blocked_domains = {
        domain_name
        for domain_name in NON_CLIENT_EMAIL_DOMAINS
        if domain_name != 'shimadzu.com.ph'
    }

    if domain in manual_blocked_domains:
        return False

    if any(domain.endswith('.' + blocked_domain) for blocked_domain in manual_blocked_domains):
        return False

    if local_part in NON_CLIENT_EMAIL_PREFIXES:
        return False

    return True


def parse_manual_recipient_emails(raw_value):
    """Parse comma/semicolon/newline/space-separated manual recipient emails."""
    if raw_value is None:
        return []

    if isinstance(raw_value, (list, tuple, set)):
        raw_items = []
        for item in raw_value:
            raw_items.extend(parse_manual_recipient_emails(item))
        return normalize_email_list(raw_items)

    normalized_text = str(raw_value).replace(';', ',').replace('\n', ',').replace('\r', ',')
    candidates = []

    for chunk in normalized_text.split(','):
        chunk = chunk.strip()
        if not chunk:
            continue

        # Support accidental pasted strings like "A <a@test.com> B <b@test.com>".
        found_in_chunk = CLIENT_EMAIL_REGEX.findall(chunk)
        if found_in_chunk:
            candidates.extend(found_in_chunk)
        else:
            # Also support simple space-separated emails.
            candidates.extend([part.strip() for part in chunk.split() if part.strip()])

    clean_recipients = []
    seen = set()

    for email_addr in candidates:
        cleaned = str(email_addr).strip().strip('.,;:()[]<>').lower()
        if not cleaned or cleaned in seen:
            continue
        if not is_valid_manual_recipient_email(cleaned):
            continue

        seen.add(cleaned)
        clean_recipients.append(cleaned)

    return clean_recipients


def is_valid_client_email(email_addr):
    """Basic email validation and certificate/internal-domain suppression."""
    email_addr = clean_str(email_addr)
    if not email_addr:
        return False

    lowered = email_addr.lower()
    if not CLIENT_EMAIL_REGEX.fullmatch(lowered):
        return False

    local_part, domain = lowered.rsplit('@', 1)

    if domain in NON_CLIENT_EMAIL_DOMAINS:
        return False

    if any(domain.endswith('.' + blocked_domain) for blocked_domain in NON_CLIENT_EMAIL_DOMAINS):
        return False

    if local_part in NON_CLIENT_EMAIL_PREFIXES:
        return False

    return True


def score_email_candidate_from_context(raw_text, start_index, end_index):
    """Score email candidates so TSR form fields outrank PDF metadata/cert emails."""
    text_window = (raw_text or '')[max(0, start_index - 180):min(len(raw_text or ''), end_index + 180)].lower()

    score = 0

    # Strong signals from the TSR template/form.
    if 'email add' in text_window:
        score += 100
    if 'email address' in text_window:
        score += 80
    if 'client email' in text_window or 'customer email' in text_window:
        score += 80
    if 'customer' in text_window or 'client' in text_window:
        score += 20
    if 'contact' in text_window:
        score += 10

    # Negative signals from certificate/signature metadata.
    negative_terms = [
        'certificate', 'certification', 'certifying', 'authority',
        'verisign', 'digicert', 'symantec', 'globalsign', 'sectigo',
        'signature', 'adobe', 'timestamp', 'ocsp', 'crl', 'issuer',
        'subject', 'cps'
    ]
    for term in negative_terms:
        if term in text_window:
            score -= 80

    return score


def extract_emails_from_text(raw_text):
    """Extract unique client email addresses from text with metadata filtering.

    PDF files may contain certificate authority emails in hidden metadata.
    We prefer emails near TSR labels such as EMAIL ADD and suppress CA/system emails.
    """
    candidates = {}

    for match in CLIENT_EMAIL_REGEX.finditer(raw_text or ''):
        cleaned = match.group(1).strip().strip('.,;:()[]<>').lower()
        if not is_valid_client_email(cleaned):
            continue

        score = score_email_candidate_from_context(raw_text, match.start(), match.end())
        candidates[cleaned] = max(score, candidates.get(cleaned, -9999))

    # Prefer positive/contextual candidates. If none have positive context,
    # keep valid emails but still exclude certificate/system domains above.
    positive = [(email, score) for email, score in candidates.items() if score > 0]
    selected = positive if positive else list(candidates.items())

    selected.sort(key=lambda item: (-item[1], item[0]))

    return [email for email, score in selected]


def decode_pdf_string_value(raw_value):
    """Decode common PDF form/string value encodings.

    Handles:
    - literal strings: (client@example.com)
    - hex strings: <FEFF006A006F...>
    - escaped PDF string characters
    """
    if raw_value is None:
        return ''

    value = str(raw_value).strip()

    if value.startswith('<') and value.endswith('>') and not value.startswith('<<'):
        hex_value = re.sub(r'[^0-9A-Fa-f]', '', value[1:-1])
        try:
            raw_bytes = bytes.fromhex(hex_value)
            if raw_bytes.startswith(b'\xfe\xff'):
                return raw_bytes[2:].decode('utf-16-be', errors='ignore')
            if raw_bytes.startswith(b'\xff\xfe'):
                return raw_bytes[2:].decode('utf-16-le', errors='ignore')
            # Try UTF-16-BE even without BOM because many PDF form values use it.
            if len(raw_bytes) >= 2 and raw_bytes[0] == 0 and any(b != 0 for b in raw_bytes[1::2]):
                decoded = raw_bytes.decode('utf-16-be', errors='ignore')
                if decoded:
                    return decoded
            return raw_bytes.decode('utf-8', errors='ignore') or raw_bytes.decode('latin-1', errors='ignore')
        except Exception:
            return value

    if value.startswith('(') and value.endswith(')'):
        value = value[1:-1]

    replacements = {
        r'\(': '(',
        r'\)': ')',
        r'\\': '\\',
        r'\n': '\n',
        r'\r': '\r',
        r'\t': '\t',
        r'\b': '\b',
        r'\f': '\f'
    }
    for old, new in replacements.items():
        value = value.replace(old, new)

    def octal_repl(match):
        try:
            return chr(int(match.group(1), 8))
        except Exception:
            return match.group(0)

    return re.sub(r'\\([0-7]{1,3})', octal_repl, value)


def load_optional_pdf_reader():
    """Load an optional PDF reader without creating IDE/Pylance missing-import warnings."""
    try:
        import importlib
        pypdf_module = importlib.import_module('pypdf')
        return getattr(pypdf_module, 'PdfReader', None)
    except Exception:
        pass

    try:
        import importlib
        pypdf2_module = importlib.import_module('PyPDF2')
        return getattr(pypdf2_module, 'PdfReader', None)
    except Exception:
        return None


def extract_pdf_form_field_text_with_libraries(file_bytes):
    """Best-effort AcroForm/text extraction using optional PDF libraries.

    The app does not require pypdf/PyPDF2. If neither package exists, the
    system silently falls back to raw PDF form-field parsing below.
    """
    output = []
    PdfReader = load_optional_pdf_reader()

    if not PdfReader:
        return ''

    try:
        reader = PdfReader(io.BytesIO(file_bytes))

        for page in getattr(reader, 'pages', []) or []:
            try:
                output.append(page.extract_text() or '')
            except Exception:
                pass

        try:
            fields = reader.get_fields() or {}
            for field_name, field_data in fields.items():
                output.append(str(field_name or ''))
                if isinstance(field_data, dict):
                    for key in ('/T', '/TU', '/TM', '/V', '/DV'):
                        value = field_data.get(key)
                        if value is not None:
                            output.append(str(value))
                else:
                    output.append(str(field_data or ''))
        except Exception:
            pass

    except Exception:
        pass

    return '\n'.join([chunk for chunk in output if chunk])


def extract_pdf_stream_bytes(file_bytes):
    """Extract and inflate common PDF stream objects using stdlib only.

    Many form-based PDFs store visible field values inside compressed streams.
    This lets us detect emails even when no PDF library is installed.
    """
    stream_outputs = []

    for match in re.finditer(rb'stream\r?\n(.*?)\r?\nendstream', file_bytes or b'', flags=re.DOTALL):
        stream_data = match.group(1).strip(b'\r\n')

        # Raw stream text may already be readable.
        stream_outputs.append(stream_data)

        # Try zlib/FlateDecode. If not compressed, ignore.
        try:
            stream_outputs.append(zlib.decompress(stream_data))
            continue
        except Exception:
            pass

        # Sometimes there are extra bytes before a valid zlib header.
        for header in (b'\x78\x9c', b'\x78\xda', b'\x78\x01'):
            header_pos = stream_data.find(header)
            if header_pos > 0:
                try:
                    stream_outputs.append(zlib.decompress(stream_data[header_pos:]))
                    break
                except Exception:
                    pass

    return stream_outputs


def decode_pdf_bytes_to_text_chunks(byte_chunks):
    """Decode PDF byte chunks using encodings commonly seen in form PDFs."""
    output = []

    for chunk in byte_chunks or []:
        if not chunk:
            continue

        for encoding in ('utf-8', 'utf-16-be', 'utf-16-le', 'latin-1'):
            try:
                decoded = chunk.decode(encoding, errors='ignore')
                if decoded:
                    output.append(decoded)
            except Exception:
                pass

    return output


def extract_pdf_hex_encoded_emails(decoded_text):
    """Decode emails hidden as UTF-16/hex sequences in PDF content streams."""
    output = []

    for hex_match in re.finditer(r'<([0-9A-Fa-f]{12,})>', decoded_text or ''):
        hex_value = hex_match.group(1)
        try:
            raw_bytes = bytes.fromhex(hex_value)
        except Exception:
            continue

        for encoding in ('utf-16-be', 'utf-16-le', 'utf-8', 'latin-1'):
            try:
                decoded = raw_bytes.decode(encoding, errors='ignore')
            except Exception:
                decoded = ''
            if '@' in decoded:
                output.append(decoded)

    return output


def extract_pdf_form_field_text_raw(file_bytes):
    """Extract readable AcroForm/XFA values from raw PDF bytes without dependencies."""
    if not file_bytes:
        return ''

    output = []

    # Full raw file decodes.
    raw_chunks = [file_bytes]
    raw_chunks.extend(extract_pdf_stream_bytes(file_bytes))
    output.extend(decode_pdf_bytes_to_text_chunks(raw_chunks))

    combined_raw_text = '\n'.join(output)

    # PDF dictionary field values: /T (EMAIL ADD:) /V (client@example.com)
    token_pattern = re.compile(
        r'/(?:T|TU|TM|V|DV)\s*(\((?:\\.|[^\\)])*\)|<[^<>\r\n]+>)',
        re.IGNORECASE | re.DOTALL
    )
    for match in token_pattern.finditer(combined_raw_text):
        output.append(decode_pdf_string_value(match.group(1)))

    # XFA/XML field values can show up after stream inflation.
    for xml_match in re.finditer(r'<[^>]*(?:email|EMAIL|mail|MAIL)[^>]*>(.*?)</[^>]+>', combined_raw_text, flags=re.DOTALL):
        output.append(re.sub(r'<[^>]+>', ' ', xml_match.group(1)))

    # Some streams encode text as hex strings.
    output.extend(extract_pdf_hex_encoded_emails(combined_raw_text))

    return '\n'.join([chunk for chunk in output if chunk])


def extract_text_from_pdf_bytes(file_bytes):
    """Best-effort text extraction from PDFs.

    Phase 1 now supports:
    - PDF form fields / AcroForm values
    - compressed PDF streams via stdlib zlib
    - XFA/XML-like form content
    - normal text PDFs when optional pypdf/PyPDF2 is available
    - raw PDF fallback scanning

    This still does not OCR scanned image PDFs.
    """
    if not file_bytes:
        return ''

    extracted_parts = [
        extract_pdf_form_field_text_with_libraries(file_bytes),
        extract_pdf_form_field_text_raw(file_bytes)
    ]

    try:
        extracted_parts.append(file_bytes.decode('latin-1', errors='ignore'))
    except Exception:
        pass

    return '\n'.join([part for part in extracted_parts if part])


def extract_text_from_docx_bytes(file_bytes):
    """Extract text from docx XML safely using stdlib zipfile."""
    try:
        output = []
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as docx_zip:
            for name in docx_zip.namelist():
                if name.startswith('word/') and name.endswith('.xml'):
                    xml_text = docx_zip.read(name).decode('utf-8', errors='ignore')
                    xml_text = re.sub(r'<[^>]+>', ' ', xml_text)
                    output.append(xml_text)
        return ' '.join(output)
    except Exception:
        return ''


def extract_text_from_xlsx_bytes(file_bytes):
    """Extract text from xlsx shared strings/worksheets safely using stdlib zipfile."""
    try:
        output = []
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as xlsx_zip:
            for name in xlsx_zip.namelist():
                if name.endswith('.xml') and (
                    name.startswith('xl/sharedStrings') or
                    name.startswith('xl/worksheets/')
                ):
                    xml_text = xlsx_zip.read(name).decode('utf-8', errors='ignore')
                    xml_text = re.sub(r'<[^>]+>', ' ', xml_text)
                    output.append(xml_text)
        return ' '.join(output)
    except Exception:
        return ''


def extract_text_from_report_file(file_path, filename):
    """Best-effort text extraction from supported TSR attachment types."""
    if not file_path or not os.path.exists(file_path):
        return ''

    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    try:
        with open(file_path, 'rb') as file_obj:
            file_bytes = file_obj.read()
    except Exception:
        return ''

    if ext == 'pdf':
        return extract_text_from_pdf_bytes(file_bytes)
    if ext == 'docx':
        return extract_text_from_docx_bytes(file_bytes)
    if ext == 'xlsx':
        return extract_text_from_xlsx_bytes(file_bytes)
    if ext in {'csv', 'txt'}:
        return file_bytes.decode('utf-8', errors='ignore')
    if ext in {'png', 'jpg', 'jpeg'}:
        # Phase 1 does not OCR scanned/image TSRs.
        return ''

    return ''


def get_tsr_files_for_shift(shift):
    """Return stored TSR files attached to a shift.

    filename is the server/disk filename.
    display_name is the original/client-facing filename.
    """
    ensure_shift_file_original_filename_column()

    if not shift:
        return []

    files = []
    for file_rec in shift.files:
        display_name = get_shift_file_display_name(file_rec)
        if not is_tsr_filename(display_name) and not is_tsr_filename(file_rec.filename):
            continue

        safe_filename = get_shift_file_disk_name(file_rec)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
        if os.path.exists(file_path):
            files.append({
                'id': file_rec.id,
                'filename': safe_filename,
                'display_name': display_name or safe_filename,
                'path': file_path,
                'uploaded_at': file_rec.uploaded_at.isoformat() if file_rec.uploaded_at else ''
            })

    return files


def detect_client_emails_from_tsr_files(tsr_files):
    """Detect emails from TSR files and return candidates with source filenames."""
    email_sources = {}

    for tsr_file in tsr_files or []:
        extracted_text = extract_text_from_report_file(tsr_file.get('path'), tsr_file.get('filename'))
        for email_addr in extract_emails_from_text(extracted_text):
            email_sources.setdefault(email_addr, set()).add(tsr_file.get('filename'))

    return [
        {
            'email': email_addr,
            'sources': sorted(list(source_names))
        }
        for email_addr, source_names in sorted(email_sources.items())
    ]


def get_shift_client_email_fallbacks(shift):
    """Return client database emails as fallback choices when TSR detection is empty/uncertain."""
    if not shift or not shift.client_id:
        return []

    fallback_emails = []
    seen = set()

    contacts = Contact.query.filter_by(client_id=shift.client_id).all()
    for contact in contacts:
        email_addr = clean_str(contact.email)
        if email_addr and is_valid_client_email(email_addr) and email_addr.lower() not in seen:
            seen.add(email_addr.lower())
            fallback_emails.append({
                'email': email_addr.lower(),
                'name': contact.name or '',
                'source': 'client_contact'
            })

    client = db.session.get(Client, shift.client_id)
    for email_addr in [
        getattr(client, 'email_address_1', None),
        getattr(client, 'email_address_2', None),
        getattr(client, 'email_address_3', None)
    ] if client else []:
        email_addr = clean_str(email_addr)
        if email_addr and is_valid_client_email(email_addr) and email_addr.lower() not in seen:
            seen.add(email_addr.lower())
            fallback_emails.append({
                'email': email_addr.lower(),
                'name': '',
                'source': 'client_legacy_contact'
            })

    return fallback_emails


def build_tsr_client_email_subject(shift):
    """Build client-facing TSR email subject.

    Format:
        Task/Purpose - Equipment - Client - MM/DD/YY

    Missing fields are omitted instead of showing blank separators.
    """
    task = clean_str(getattr(shift, 'title', None)) if shift else None
    equipment = clean_str(getattr(shift.product, 'name', None)) if shift and shift.product else None
    client_name = clean_str(getattr(shift.client, 'name', None)) if shift and shift.client else None
    service_date = shift.start_time.strftime('%m/%d/%y') if shift and shift.start_time else None

    subject_parts = [
        task or 'Task',
        equipment or 'Equipment',
        client_name or 'Client',
        service_date or get_manila_time().strftime('%m/%d/%y')
    ]

    return " - ".join(subject_parts)


def build_tsr_client_email_bodies(shift, sender_name, font_key=None):
    client_name = shift.client.name if shift and shift.client else 'Valued Client'
    product_name = shift.product.name if shift and shift.product else ''
    product_serial = shift.product.serial_number if shift and shift.product else ''
    service_date = shift.start_time.strftime('%B %d, %Y') if shift and shift.start_time else ''
    task = shift.title if shift else ''

    product_label = product_name
    if product_serial and product_name:
        product_label = f"{product_serial} / {product_name}"
    elif product_serial:
        product_label = product_serial

    text_lines = [
        f"Dear {client_name},",
        "",
        "Good day.",
        "",
        "Attached is the Technical Service Report (TSR) for the completed service visit.",
        "",
        f"Service date: {service_date}",
        f"Task: {task}",
    ]

    if product_label:
        text_lines.append(f"Equipment: {product_label}")

    text_lines.extend([
        "",
        "Thank you.",
        "",
        "Medical Service Team"
    ])

    html_rows = [
        ("Service date", service_date),
        ("Task", task)
    ]
    if product_label:
        html_rows.append(("Equipment", product_label))

    detail_rows = ''.join(
        f"<tr><td style='padding:6px 10px;font-weight:700;border-bottom:1px solid #e5e7eb;'>{html.escape(label)}</td>"
        f"<td style='padding:6px 10px;border-bottom:1px solid #e5e7eb;'>{html.escape(str(value or ''))}</td></tr>"
        for label, value in html_rows
    )

    font_stack = get_tsr_email_font_stack(font_key)

    html_body = f"""
    <div style="font-family:{html.escape(font_stack)};color:#111827;line-height:1.5;">
        <p>Dear {html.escape(client_name)},</p>
        <p>Good day.</p>
        <p>Attached is the Technical Service Report (TSR) for the completed service visit.</p>
        <table style="border-collapse:collapse;border:1px solid #e5e7eb;min-width:360px;">
            {detail_rows}
        </table>
        <p style="margin-top:16px;">Thank you.</p>
        <p>Medical Service Team</p>
    </div>
    """

    return "\n".join(text_lines), html_body


@app.route('/preview_tsr_client_email/<int:shift_id>', methods=['GET'])
@login_required
def preview_tsr_client_email(shift_id):
    """Phase 1: Detect possible client email addresses from attached TSR files."""
    shift = db.session.get(Shift, shift_id)
    if not shift:
        return jsonify({'message': 'Schedule not found.'}), 404

    if not can_work_on_existing_schedule_shift(shift):
        return denied('You are not authorized to send TSR for this schedule.')

    tsr_files = get_tsr_files_for_shift(shift)
    detected_emails = detect_client_emails_from_tsr_files(tsr_files)
    fallback_emails = get_shift_client_email_fallbacks(shift)

    return jsonify({
        'status': 'success',
        'shift_id': shift.id,
        'schedule_status': shift.status or '',
        'is_completed': (shift.status or '') == 'Completed',
        'client_name': shift.client.name if shift.client else '',
        'task': shift.title,
        'service_date': shift.start_time.strftime('%Y-%m-%d') if shift.start_time else '',
        'tsr_files': [
            {
                'id': file_info.get('id'),
                'filename': file_info.get('filename'),
                'display_name': file_info.get('display_name') or file_info.get('filename'),
                'uploaded_at': file_info.get('uploaded_at')
            }
            for file_info in tsr_files
        ],
        'detected_emails': detected_emails,
        'fallback_emails': fallback_emails,
        'font_options': [
            {'key': key, 'label': key.title() if key != 'tenorite' else 'Tenorite'}
            for key in TSR_EMAIL_FONT_STACKS.keys()
        ],
        'default_font_key': DEFAULT_TSR_EMAIL_FONT_KEY,
        'can_send': bool(tsr_files),
        'note': 'Image/scanned TSR OCR is not enabled in Phase 1.'
    })


@app.route('/send_tsr_client_email/<int:shift_id>', methods=['POST'])
@login_required
def send_tsr_client_email(shift_id):
    """Phase 1: Send attached TSR to scheduler-confirmed client email."""
    shift = db.session.get(Shift, shift_id)
    if not shift:
        return jsonify({'message': 'Schedule not found.'}), 404

    if not can_work_on_existing_schedule_shift(shift):
        return denied('You are not authorized to send TSR for this schedule.')

    payload = request.get_json(silent=True) or {}
    recipient_emails = parse_manual_recipient_emails(payload.get('email'))
    if not recipient_emails:
        return jsonify({
            'message': 'Please provide at least one valid recipient email address.'
        }), 400

    tsr_files = get_tsr_files_for_shift(shift)
    if not tsr_files:
        return jsonify({'message': 'No attached TSR file found for this schedule.'}), 400

    subject = clean_str(payload.get('subject')) or build_tsr_client_email_subject(shift)
    sender_name = current_user.username.capitalize() if current_user and current_user.is_authenticated else 'Scheduler'
    font_key = clean_str(payload.get('font_key')) or DEFAULT_TSR_EMAIL_FONT_KEY
    text_body, html_body = build_tsr_client_email_bodies(shift, sender_name, font_key=font_key)

    static_cc_emails = get_tsr_client_cc_emails_for_current_sender()

    email_sent, email_message = send_email_with_attachments(
        recipient_emails,
        subject,
        text_body,
        html_body,
        attachments=tsr_files,
        cc_emails=static_cc_emails
    )

    if not email_sent:
        return jsonify({'message': email_message or 'Unable to send TSR email.'}), 500

    recipients_label = ', '.join(recipient_emails)
    cc_log = f" | CC: {', '.join(static_cc_emails)}" if static_cc_emails else ""
    db.session.add(ActivityLog(
        user=sender_name,
        action=f"Sent TSR to client: {recipients_label}{cc_log} | Font: {font_key} | {shift.title}"
    ))
    db.session.commit()

    recipients_label = ', '.join(recipient_emails)
    return jsonify({
        'status': 'success',
        'message': f'TSR sent to {recipients_label}.',
        'email': recipients_label,
        'emails': recipient_emails,
        'recipient_count': len(recipient_emails),
        'cc': static_cc_emails,
        'sender_copy_email': get_current_user_email_for_tsr_cc(),
        'font_key': font_key,
        'font_stack': get_tsr_email_font_stack(font_key),
        'attachments': [file_info.get('display_name') or file_info.get('filename') for file_info in tsr_files]
    })


@app.route('/add_shift', methods=['POST'])
@login_required
def add_shift():
    """Create a schedule chain. Supports JSON and multipart/form-data."""
    import uuid

    payload = get_shift_payload()

    files_ok, files_error = validate_uploaded_report_files()
    if not files_ok:
        return jsonify({'message': files_error}), 400

    start_d = parse_date(payload.get('start_date'))
    end_d = parse_date(payload.get('end_date'))
    start_time = payload.get('start_time')
    end_time = payload.get('end_time')

    if not start_d or not end_d or not start_time or not end_time:
        return jsonify({'message': 'Missing date or time fields'}), 400
    if end_d < start_d:
        return jsonify({'message': 'End date cannot be earlier than start date'}), 400

    include_weekends = parse_bool_flag(payload.get('include_weekends'), default=True)
    schedule_dates = get_schedule_dates_between(start_d, end_d, include_weekends=include_weekends)
    if not schedule_dates:
        return jsonify({'message': 'No schedule dates selected. The selected range only contains weekends.'}), 400

    engineers = parse_engineer_ids(payload.get('engineers'))
    if not engineers:
        fallback_id = clean_int(payload.get('engineer_id'))
        if fallback_id:
            engineers = [fallback_id]
    if not engineers:
        return jsonify({'message': 'No engineers selected'}), 400

    shift_title = normalize_shift_title(payload)
    if not shift_title:
        return jsonify({'message': 'Task/Purpose is required'}), 400

    if not can_create_schedule_for_engineer_ids(engineers):
        return denied('You are not authorized to add schedules for the selected engineer/branch.')

    requested_status = clean_str(payload.get('status')) or 'In Progress'
    is_work_schedule = bool(clean_int(payload.get('client_id')))
    if is_work_schedule and requested_status == 'Completed' and not uploaded_files_have_tsr():
        return jsonify({'message': 'A TSR file is required before marking this schedule as Completed.'}), 400

    group_id = str(uuid.uuid4())
    created_at = get_manila_time()

    # Check all requested dates/engineers before creating anything.
    for iter_date in schedule_dates:
        st_obj = datetime.combine(iter_date, datetime.strptime(start_time, '%H:%M').time())
        et_obj = datetime.combine(iter_date, datetime.strptime(end_time, '%H:%M').time())
        if et_obj <= st_obj:
            return jsonify({'message': 'End time must be later than start time'}), 400

        for e_id in engineers:
            collision = find_add_schedule_collision(e_id, iter_date, st_obj, et_obj)
            if collision:
                eng = db.session.get(Engineer, e_id)
                return build_conflict_response(collision, eng)

    first_shift = None
    for iter_date in schedule_dates:
        st_obj = datetime.combine(iter_date, datetime.strptime(start_time, '%H:%M').time())
        et_obj = datetime.combine(iter_date, datetime.strptime(end_time, '%H:%M').time())

        new_shift = Shift(
            title=shift_title,
            start_time=st_obj,
            end_time=et_obj,
            engineer_id=engineers[0],
            client_id=clean_int(payload.get('client_id')),
            product_id=clean_str(payload.get('product_id')),
            status=requested_status,
            created_at=created_at,
            group_id=group_id
        )

        db.session.add(new_shift)
        db.session.flush()
        delete_shift_engineer_links(new_shift.id)
        if first_shift is None:
            first_shift = new_shift

        for e_id in engineers:
            db.session.add(ShiftEngineer(shift_id=new_shift.id, engineer_id=e_id))

    saved_files = []
    if first_shift:
        saved_files = save_uploaded_shift_files(first_shift)

    date_label = start_d.isoformat() if start_d == end_d else f"{start_d.isoformat()} to {end_d.isoformat()}"
    if not include_weekends and has_weekend_between(start_d, end_d):
        date_label += " (weekends skipped)"
    engineer_names = [db.session.get(Engineer, e_id).name for e_id in engineers if db.session.get(Engineer, e_id)]
    log_action = f"Added calendar schedule: {shift_title} on {date_label}"
    if engineer_names:
        log_action += f" for {', '.join(engineer_names)}"
    if saved_files:
        log_action += f" with {len(saved_files)} report file(s)"

    db.session.add(ActivityLog(user=current_user.username.capitalize(), action=log_action))
    db.session.commit()

    if first_shift:
        notify_engineers_for_new_schedule_async(
            app,
            first_shift.id,
            engineers,
            shift_title,
            date_label,
            current_user.username if current_user and current_user.is_authenticated else 'Scheduler'
        )

    # Keep response simple and immediate; email runs in the background.
    return jsonify({'status': 'success'})


@app.route('/update_shift/<int:id>', methods=['POST'])
@login_required
def update_shift(id):
    """Update a schedule chain from the edit modal."""
    import uuid

    master_shift = db.session.get(Shift, id)
    if not master_shift:
        return jsonify({'message': 'Missing'}), 404

    payload = get_shift_payload()

    files_ok, files_error = validate_uploaded_report_files()
    if not files_ok:
        return jsonify({'message': files_error}), 400

    engineers = parse_engineer_ids(payload.get('engineers'))
    if not engineers:
        fallback_id = clean_int(payload.get('engineer_id'))
        if fallback_id:
            engineers = [fallback_id]
    if not engineers:
        return jsonify({'message': 'No engineers selected'}), 400

    shift_title = normalize_shift_title(payload)
    if not shift_title:
        return jsonify({'message': 'Task/Purpose is required'}), 400

    new_status = clean_str(payload.get('status')) or master_shift.status
    start_d = parse_date(payload.get('start_date'))
    end_d = parse_date(payload.get('end_date'))
    new_start_time = payload.get('start_time')
    new_end_time = payload.get('end_time')

    if not start_d or not end_d or not new_start_time or not new_end_time:
        return jsonify({'message': 'Missing date or time fields'}), 400
    if end_d < start_d:
        return jsonify({'message': 'End date cannot be earlier than start date'}), 400

    include_weekends = parse_bool_flag(payload.get('include_weekends'), default=True)
    requested_schedule_dates = get_schedule_dates_between(start_d, end_d, include_weekends=include_weekends)
    if not requested_schedule_dates:
        return jsonify({'message': 'No schedule dates selected. The selected range only contains weekends.'}), 400

    edit_scope = normalize_edit_scope(clean_str(payload.get('edit_scope')) or 'entire_schedule')
    override_engineer_id = clean_int(payload.get('override_engineer_id')) or clean_int(payload.get('engineer_id'))

    old_engineer_ids = get_shift_assigned_engineer_ids(master_shift)
    if not can_work_on_existing_schedule_shift(master_shift):
        return denied('You are not authorized to edit this existing schedule.')
    if not can_submit_update_engineer_ids_for_scope(master_shift, engineers, edit_scope, override_engineer_id):
        return denied('You are not authorized to assign schedules to the selected engineer/branch.')

    if edit_scope == 'engineer_day_time_override':
        if not override_engineer_id:
            return jsonify({'message': 'Missing engineer for time override.'}), 400
        if len(engineers) != 1:
            engineers = [override_engineer_id]

        if not can_modify_schedule_for_engineer_ids([override_engineer_id]):
            return denied('You are not authorized to edit this engineer schedule.')

        override_start = datetime.combine(start_d, datetime.strptime(new_start_time, '%H:%M').time())
        override_end = datetime.combine(start_d, datetime.strptime(new_end_time, '%H:%M').time())
        if override_end <= override_start:
            return jsonify({'message': 'End time must be later than start time'}), 400

        parent_shift = master_shift
        if is_shift_time_override(master_shift) and master_shift.parent_shift_id:
            parent_shift = db.session.get(Shift, master_shift.parent_shift_id) or master_shift

        if not can_work_on_existing_schedule_shift(parent_shift):
            return denied('You are not authorized to edit this existing schedule.')

        ignore_ids = [parent_shift.id]
        existing_override = (
            Shift.query
            .filter_by(
                parent_shift_id=parent_shift.id,
                override_engineer_id=override_engineer_id,
                override_kind='time_override'
            )
            .first()
        )
        if existing_override:
            ignore_ids.append(existing_override.id)

        collision = (
            db.session.query(Shift)
            .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
            .filter(
                ShiftEngineer.engineer_id == override_engineer_id,
                Shift.start_time < override_end,
                Shift.end_time > override_start,
                ~Shift.id.in_(ignore_ids)
            )
            .first()
        )
        if collision:
            eng = db.session.get(Engineer, override_engineer_id)
            return build_conflict_response(collision, eng)

        # Only time/date is customized here. Shared job fields stay tied to parent.
        # If the frontend submitted new shared details, update the parent first.
        parent_shift.title = shift_title
        parent_shift.client_id = clean_int(payload.get('client_id'))
        parent_shift.product_id = clean_str(payload.get('product_id'))
        parent_shift.status = new_status

        if not parent_shift.group_id:
            parent_shift.group_id = str(uuid.uuid4())

        detach_engineer_from_parent_shift_for_override(parent_shift, override_engineer_id)
        override_shift = create_or_update_time_override(parent_shift, override_engineer_id, override_start, override_end)

        saved_files = save_uploaded_shift_files(parent_shift)
        if saved_files:
            sync_override_files_from_parent(parent_shift, override_shift)

        eng = db.session.get(Engineer, override_engineer_id)
        log_activity(
            f"Added custom time override: {parent_shift.title} for {eng.name if eng else 'Engineer'} on {start_d.isoformat()}"
        )
        db.session.commit()

        send_schedule_event_notification_async(
            app,
            override_shift.id,
            [override_engineer_id],
            'Updated',
            current_user.username if current_user and current_user.is_authenticated else 'Scheduler',
            'Your schedule time was customized while remaining linked to the original job.'
        )

        return jsonify({
            'status': 'success',
            'override': True,
            'shift_id': override_shift.id,
            'parent_shift_id': parent_shift.id
        })

    if edit_scope == 'assigned_engineers_day_only':
        # New middle option:
        # Update the selected assigned engineers for THIS ONE DAY only.
        # This does not rebuild the whole linked multi-day group.
        # Existing custom-time children for this day stay linked and receive shared details.
        if start_d != end_d:
            end_d = start_d
        requested_schedule_dates = [start_d]

        day_start = datetime.combine(start_d, datetime.strptime(new_start_time, '%H:%M').time())
        day_end = datetime.combine(start_d, datetime.strptime(new_end_time, '%H:%M').time())
        if day_end <= day_start:
            return jsonify({'message': 'End time must be later than start time'}), 400

        day_shift = master_shift
        if is_shift_time_override(master_shift) and master_shift.parent_shift_id:
            day_shift = db.session.get(Shift, master_shift.parent_shift_id) or master_shift

        # If editing from another row in a linked group, locate the base parent for the requested day.
        if day_shift.group_id:
            matching_day_shift = (
                Shift.query
                .filter(
                    Shift.group_id == day_shift.group_id,
                    func.date(Shift.start_time) == start_d,
                    or_(Shift.override_kind.is_(None), Shift.override_kind == '')
                )
                .order_by(Shift.start_time.asc())
                .first()
            )
            if matching_day_shift:
                day_shift = matching_day_shift

        existing_day_engineer_ids = get_shift_assigned_engineer_ids(day_shift)
        linked_day_overrides = get_linked_time_overrides_for_shift_ids([day_shift.id])
        linked_day_override_ids = [override.id for override in linked_day_overrides]
        linked_day_override_engineer_ids = {
            (override.override_engineer_id or override.engineer_id)
            for override in linked_day_overrides
            if (override.override_engineer_id or override.engineer_id)
        }

        # Authorization must cover the original day assignment and the requested assignment.
        original_scope_ids = list(dict.fromkeys(existing_day_engineer_ids + list(linked_day_override_engineer_ids)))
        if is_superadmin_user() or is_regional_admin_user():
            if not can_modify_schedule_for_engineer_ids(original_scope_ids):
                return denied('You are not authorized to edit this existing schedule.')
            if not can_modify_schedule_for_engineer_ids(engineers):
                return denied('You are not authorized to assign schedules to the selected engineer/branch.')
        else:
            if not can_work_on_existing_schedule_shift(day_shift):
                return denied('You are not authorized to edit this existing schedule.')
            if set(map(int, engineers)) != set(map(int, existing_day_engineer_ids)):
                return denied('Engineers cannot reassign schedules to other engineers.')

        is_work_schedule = bool(clean_int(payload.get('client_id')))
        preserved_files = [file_rec.filename for file_rec in day_shift.files]
        if is_work_schedule and new_status == 'Completed':
            has_existing_tsr = existing_files_have_tsr(preserved_files)
            has_uploaded_tsr = uploaded_files_have_tsr()
            if not has_existing_tsr and not has_uploaded_tsr:
                return jsonify({'message': 'A TSR file is required before marking this schedule as Completed.'}), 400

        ignore_ids = [day_shift.id] + linked_day_override_ids
        base_engineers_for_day = [
            e_id for e_id in engineers
            if e_id not in linked_day_override_engineer_ids
        ]

        # Check conflict only for engineers who will be assigned to the base day shift.
        # Linked custom-time override engineers keep their own custom time and are ignored here.
        for e_id in base_engineers_for_day:
            collision = (
                db.session.query(Shift)
                .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
                .filter(
                    ShiftEngineer.engineer_id == e_id,
                    Shift.start_time < day_end,
                    Shift.end_time > day_start,
                    ~Shift.id.in_(ignore_ids)
                )
                .first()
            )
            if collision:
                eng = db.session.get(Engineer, e_id)
                return build_conflict_response(collision, eng)

        if not day_shift.group_id:
            day_shift.group_id = str(uuid.uuid4())

        day_shift.title = shift_title
        day_shift.start_time = day_start
        day_shift.end_time = day_end
        day_shift.client_id = clean_int(payload.get('client_id'))
        day_shift.product_id = clean_str(payload.get('product_id'))
        day_shift.status = new_status

        delete_shift_engineer_links(day_shift.id)
        for e_id in base_engineers_for_day:
            db.session.add(ShiftEngineer(shift_id=day_shift.id, engineer_id=e_id))

        if base_engineers_for_day:
            day_shift.engineer_id = base_engineers_for_day[0]
        elif linked_day_override_engineer_ids:
            # Keep a valid legacy engineer_id if all visible engineers for this date are override children.
            day_shift.engineer_id = list(linked_day_override_engineer_ids)[0]

        saved_files = save_uploaded_shift_files(day_shift)

        # Keep linked custom-time children for this date connected to this same day parent.
        # They keep custom time, but shared details/files follow the edited day schedule.
        for override in linked_day_overrides:
            sync_override_shared_fields_preserve_time(override, day_shift)

        engineer_names = [
            db.session.get(Engineer, e_id).name
            for e_id in engineers
            if db.session.get(Engineer, e_id)
        ]
        log_action = f"Updated assigned engineers for one day: {shift_title} on {start_d.isoformat()}"
        if engineer_names:
            log_action += f" for {', '.join(engineer_names)}"
        if saved_files:
            log_action += f" and uploaded {len(saved_files)} report file(s)"

        db.session.add(ActivityLog(user=current_user.username.capitalize(), action=log_action))
        db.session.commit()

        send_schedule_event_notification_async(
            app,
            day_shift.id,
            engineers,
            'Updated',
            current_user.username if current_user and current_user.is_authenticated else 'Scheduler',
            'This schedule was updated for this day only.'
        )

        return jsonify({
            'status': 'success',
            'day_only': True,
            'shift_id': day_shift.id,
            'date': start_d.isoformat()
        })

    # Safe in-place shared-detail/status update:
    # If the date range/time/team did not structurally change, do not delete/rebuild
    # the linked chain. This preserves custom-time override child rows.
    existing_group_id_for_fast_update = master_shift.group_id
    existing_chain_for_fast_update = get_base_group_shifts(existing_group_id_for_fast_update, master_shift) if existing_group_id_for_fast_update else [master_shift]
    existing_chain_for_fast_update = sorted(existing_chain_for_fast_update, key=lambda s: s.start_time)

    if edit_scope == 'entire_schedule' and existing_chain_for_fast_update:
        existing_dates = [s.start_time.date() for s in existing_chain_for_fast_update]
        requested_dates = requested_schedule_dates

        # If the outer start/end are unchanged, compare against the actual
        # existing chain dates. This preserves weekday-only chains created by
        # "skip weekends" even when only status/shared details are changed.
        if should_preserve_existing_schedule_dates(existing_chain_for_fast_update, start_d, end_d):
            requested_dates = existing_dates

        linked_overrides_for_fast_update = get_linked_time_overrides_for_shift_ids([s.id for s in existing_chain_for_fast_update])
        override_ids_by_parent_date_for_fast_update = {}
        parent_date_lookup_for_fast_update = {s.id: s.start_time.date() for s in existing_chain_for_fast_update}
        for override in linked_overrides_for_fast_update:
            parent_date = parent_date_lookup_for_fast_update.get(override.parent_shift_id)
            override_engineer_id = override.override_engineer_id or override.engineer_id
            if parent_date and override_engineer_id:
                override_ids_by_parent_date_for_fast_update.setdefault(parent_date, set()).add(override_engineer_id)

        # Structure means date range + assigned base engineers match.
        # Do NOT require every day to share the submitted modal time.
        # Some base days may intentionally have custom times from the
        # "assigned engineers + this day only" option.
        structure_matches = (existing_dates == requested_dates)

        if structure_matches:
            for existing_shift in existing_chain_for_fast_update:
                expected_base_engineers = [
                    e_id for e_id in engineers
                    if e_id not in override_ids_by_parent_date_for_fast_update.get(existing_shift.start_time.date(), set())
                ]
                current_base_engineers = get_shift_assigned_engineer_ids(existing_shift)

                if set(map(int, expected_base_engineers)) != set(map(int, current_base_engineers)):
                    structure_matches = False
                    break

        existing_time_pairs = {
            (
                s.start_time.strftime('%H:%M'),
                s.end_time.strftime('%H:%M')
            )
            for s in existing_chain_for_fast_update
        }
        submitted_time_pair = (new_start_time, new_end_time)

        # If all base days currently share one time, then a changed submitted
        # time is a real whole-chain time edit and may be applied to all.
        # If base days already have mixed times, preserve each day time during
        # shared-detail/status updates so day-only customizations do not revert.
        should_apply_whole_chain_time = (
            len(existing_time_pairs) == 1 and
            next(iter(existing_time_pairs)) != submitted_time_pair
        )

        if structure_matches:
            preserved_files = []
            for existing_shift in existing_chain_for_fast_update:
                for file_rec in existing_shift.files:
                    if file_rec.filename not in preserved_files:
                        preserved_files.append(file_rec.filename)

            is_work_schedule = bool(clean_int(payload.get('client_id')))
            if is_work_schedule and new_status == 'Completed':
                has_existing_tsr = existing_files_have_tsr(preserved_files)
                has_uploaded_tsr = uploaded_files_have_tsr()
                if not has_existing_tsr and not has_uploaded_tsr:
                    return jsonify({'message': 'A TSR file is required before marking this schedule as Completed.'}), 400

            first_updated_shift = existing_chain_for_fast_update[0]
            uploaded_count = 0

            for existing_shift in existing_chain_for_fast_update:
                existing_shift.title = shift_title
                existing_shift.client_id = clean_int(payload.get('client_id'))
                existing_shift.product_id = clean_str(payload.get('product_id'))
                existing_shift.status = new_status
                existing_shift.group_id = existing_group_id_for_fast_update or existing_shift.group_id or str(uuid.uuid4())

                if should_apply_whole_chain_time:
                    existing_shift.start_time = datetime.combine(
                        existing_shift.start_time.date(),
                        datetime.strptime(new_start_time, '%H:%M').time()
                    )
                    existing_shift.end_time = datetime.combine(
                        existing_shift.end_time.date(),
                        datetime.strptime(new_end_time, '%H:%M').time()
                    )

            uploaded_files = save_uploaded_shift_files(first_updated_shift)
            uploaded_count = len(uploaded_files)

            # Sync all linked custom-time overrides without touching their time/engineer assignment.
            for existing_shift in existing_chain_for_fast_update:
                sync_linked_time_overrides(existing_shift)

            date_label = start_d.isoformat() if start_d == end_d else f"{start_d.isoformat()} to {end_d.isoformat()}"
            if not include_weekends and has_weekend_between(start_d, end_d):
                date_label += " (weekends skipped)"
            engineer_names = [db.session.get(Engineer, e_id).name for e_id in engineers if db.session.get(Engineer, e_id)]
            log_action = f"Updated calendar schedule: {shift_title} on {date_label}"
            if engineer_names:
                log_action += f" for {', '.join(engineer_names)}"
            if uploaded_count:
                log_action += f" and uploaded {uploaded_count} report file(s)"

            db.session.add(ActivityLog(user=current_user.username.capitalize(), action=log_action))
            db.session.commit()

            send_schedule_event_notification_async(
                app,
                first_updated_shift.id,
                engineers,
                'Updated',
                current_user.username if current_user and current_user.is_authenticated else 'Scheduler',
                'Your assigned schedule was updated.'
            )

            return jsonify({
                'status': 'success',
                'in_place_update': True,
                'preserved_mixed_day_times': not should_apply_whole_chain_time and len(existing_time_pairs) > 1
            })

    # Destructive full-chain rebuild guard.
    # Only the explicit entire_schedule scope may continue into this path.
    # This prevents same-day assigned-engineer edits from accidentally deleting
    # other days in a multi-day linked schedule if a frontend alias is submitted.
    if edit_scope != 'entire_schedule':
        return jsonify({'message': f'Unsupported edit scope: {edit_scope}'}), 400

    if getattr(current_user, 'role', None) == 'engineer':
        return denied('Engineers can update assigned schedule details, status, TSR files, or their own time only. Full schedule rebuild is restricted to schedulers/admins.')

    group_id = master_shift.group_id or str(uuid.uuid4())
    old_chain = get_base_group_shifts(group_id, master_shift) if master_shift.group_id else [master_shift]

    # Preserve the existing actual scheduled dates when the modal's outer range
    # did not change. This prevents status/detail updates from re-adding
    # weekends that were intentionally skipped during a previous extension.
    if should_preserve_existing_schedule_dates(old_chain, start_d, end_d):
        requested_schedule_dates = sorted({shift.start_time.date() for shift in old_chain if shift.start_time})

    old_ids = [s.id for s in old_chain]
    original_created_at = master_shift.created_at

    # Preserve existing file links across the chain. Physical files remain on disk.
    preserved_files = []
    for old_shift in old_chain:
        for file_rec in old_shift.files:
            if file_rec.filename not in preserved_files:
                preserved_files.append(file_rec.filename)

    old_shift_date_map = {old_shift.start_time.date(): old_shift for old_shift in old_chain}
    old_parent_date_by_id = {old_shift.id: old_date for old_date, old_shift in old_shift_date_map.items()}
    linked_time_overrides = get_linked_time_overrides_for_shift_ids(old_ids)
    old_ids_with_overrides = old_ids + [override.id for override in linked_time_overrides]

    # Preserve engineer/day time overrides when rebuilding the parent chain.
    # If Engineer B has a custom-time child for Tuesday, B must NOT be re-added
    # to the rebuilt Tuesday base shift. The child remains linked and receives
    # shared detail updates below.
    override_engineer_ids_by_old_parent_date = {}
    for override in linked_time_overrides:
        parent_date = old_parent_date_by_id.get(override.parent_shift_id)
        override_engineer_id = override.override_engineer_id or override.engineer_id
        if parent_date and override_engineer_id:
            override_engineer_ids_by_old_parent_date.setdefault(parent_date, set()).add(override_engineer_id)

    is_work_schedule = bool(clean_int(payload.get('client_id')))
    if is_work_schedule and new_status == 'Completed':
        has_existing_tsr = existing_files_have_tsr(preserved_files)
        has_uploaded_tsr = uploaded_files_have_tsr()
        if not has_existing_tsr and not has_uploaded_tsr:
            return jsonify({'message': 'A TSR file is required before marking this schedule as Completed.'}), 400

    # Check conflicts before deleting/recreating the chain.
    for iter_date in requested_schedule_dates:
        st_obj = datetime.combine(iter_date, datetime.strptime(new_start_time, '%H:%M').time())
        et_obj = datetime.combine(iter_date, datetime.strptime(new_end_time, '%H:%M').time())
        if et_obj <= st_obj:
            return jsonify({'message': 'End time must be later than start time'}), 400

        base_engineers_for_date = [
            e_id for e_id in engineers
            if e_id not in override_engineer_ids_by_old_parent_date.get(iter_date, set())
        ]

        for e_id in base_engineers_for_date:
            collision = (
                db.session.query(Shift)
                .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
                .filter(
                    ShiftEngineer.engineer_id == e_id,
                    Shift.start_time < et_obj,
                    Shift.end_time > st_obj,
                    ~Shift.id.in_(old_ids_with_overrides)
                )
                .first()
            )
            if collision:
                eng = db.session.get(Engineer, e_id)
                return build_conflict_response(collision, eng)

    for old_shift in old_chain:
        delete_shift_engineer_links(old_shift.id)
        db.session.delete(old_shift)
    db.session.flush()

    first_new_shift = None
    new_shift_by_date = {}
    for iter_date in requested_schedule_dates:
        st_obj = datetime.combine(iter_date, datetime.strptime(new_start_time, '%H:%M').time())
        et_obj = datetime.combine(iter_date, datetime.strptime(new_end_time, '%H:%M').time())
        base_engineers_for_date = [
            e_id for e_id in engineers
            if e_id not in override_engineer_ids_by_old_parent_date.get(iter_date, set())
        ]

        if not base_engineers_for_date:
            continue

        new_shift = Shift(
            title=shift_title,
            start_time=st_obj,
            end_time=et_obj,
            engineer_id=base_engineers_for_date[0],
            client_id=clean_int(payload.get('client_id')),
            product_id=clean_str(payload.get('product_id')),
            status=new_status,
            group_id=group_id,
            created_at=original_created_at
        )

        db.session.add(new_shift)
        db.session.flush()
        delete_shift_engineer_links(new_shift.id)
        new_shift_by_date[iter_date] = new_shift
        if first_new_shift is None:
            first_new_shift = new_shift

        for e_id in base_engineers_for_date:
            db.session.add(ShiftEngineer(shift_id=new_shift.id, engineer_id=e_id))

    saved_files = []
    if first_new_shift:
        for filename in preserved_files:
            db.session.add(ShiftFile(shift_id=first_new_shift.id, filename=filename))
        saved_files = save_uploaded_shift_files(first_new_shift)

    # Re-link custom-time child schedules to the newly rebuilt parent shift for the same date.
    for override in linked_time_overrides:
        old_parent_date = old_parent_date_by_id.get(override.parent_shift_id)

        new_parent = new_shift_by_date.get(old_parent_date)
        if not new_parent:
            # No base schedule remains for this date. Keep the child schedule visible and
            # preserve its custom time/engineer assignment instead of deleting/orphaning it.
            override.title = shift_title
            override.client_id = clean_int(payload.get('client_id'))
            override.product_id = clean_str(payload.get('product_id'))
            override.status = new_status
            override.group_id = group_id
            override.override_kind = 'time_override'
            override_engineer_id = override.override_engineer_id or override.engineer_id
            override.override_engineer_id = override_engineer_id
            if override_engineer_id:
                ensure_single_engineer_link(override, override_engineer_id)
            continue

        sync_override_shared_fields_preserve_time(override, new_parent)

    date_label = start_d.isoformat() if start_d == end_d else f"{start_d.isoformat()} to {end_d.isoformat()}"
    if not include_weekends and has_weekend_between(start_d, end_d):
        date_label += " (weekends skipped)"
    engineer_names = [db.session.get(Engineer, e_id).name for e_id in engineers if db.session.get(Engineer, e_id)]
    log_action = f"Updated calendar schedule: {shift_title} on {date_label}"
    if engineer_names:
        log_action += f" for {', '.join(engineer_names)}"
    if saved_files:
        log_action += f" and uploaded {len(saved_files)} report file(s)"

    db.session.add(ActivityLog(user=current_user.username.capitalize(), action=log_action))
    db.session.commit()

    if first_new_shift:
        send_schedule_event_notification_async(
            app,
            first_new_shift.id,
            engineers,
            'Updated',
            current_user.username if current_user and current_user.is_authenticated else 'Scheduler',
            'Your assigned schedule was updated.'
        )

    return jsonify({'status': 'success'})


@app.route('/move_shift', methods=['POST'])
@login_required
def move_shift():
    """
    Smart drag-and-drop move endpoint.

    move_mode:
    - single: replace the dragged/source engineer with the target engineer.
    - team: keep the assigned engineer team and move the whole date chain only.
    """
    payload = request.get_json(silent=True) or {}

    shift_id = clean_int(payload.get('shift_id'))
    source_engineer_id = clean_int(payload.get('source_engineer_id'))
    new_engineer_id = clean_int(payload.get('new_engineer_id'))
    new_start_date = parse_date(payload.get('new_start_date'))
    move_mode = clean_str(payload.get('move_mode')) or 'single'

    if move_mode not in {'single', 'team'}:
        move_mode = 'single'

    if not shift_id or not new_engineer_id or not new_start_date:
        return jsonify({'message': 'Missing drag/drop schedule data.'}), 400

    master_shift = db.session.get(Shift, shift_id)
    target_engineer = db.session.get(Engineer, new_engineer_id)

    if not master_shift:
        return jsonify({'message': 'Schedule not found.'}), 404
    if not target_engineer:
        return jsonify({'message': 'Target engineer not found.'}), 404

    original_engineer_ids = get_shift_assigned_engineer_ids(master_shift)
    if not original_engineer_ids:
        original_engineer_ids = [master_shift.engineer_id]

    if not source_engineer_id:
        source_engineer_id = master_shift.engineer_id or (original_engineer_ids[0] if original_engineer_ids else None)

    original_engineer_names = [
        db.session.get(Engineer, engineer_id).name
        for engineer_id in original_engineer_ids
        if db.session.get(Engineer, engineer_id)
    ]

    if move_mode == 'team':
        target_engineer_ids = list(dict.fromkeys(original_engineer_ids))
    else:
        if new_engineer_id in original_engineer_ids:
            target_engineer_ids = list(dict.fromkeys(original_engineer_ids))
        else:
            target_engineer_ids = []
            replaced_source = False

            for engineer_id in original_engineer_ids:
                if source_engineer_id and engineer_id == source_engineer_id and not replaced_source:
                    target_engineer_ids.append(new_engineer_id)
                    replaced_source = True
                else:
                    target_engineer_ids.append(engineer_id)

            if not replaced_source:
                target_engineer_ids = [new_engineer_id]

            target_engineer_ids = list(dict.fromkeys(target_engineer_ids))

    if not can_modify_schedule_for_engineer_ids(original_engineer_ids):
        return denied('You are not authorized to move this schedule.')
    if not can_modify_schedule_for_engineer_ids(target_engineer_ids):
        return denied('You are not authorized to move schedules to the selected engineer/branch.')

    group_id = master_shift.group_id
    old_chain = get_base_group_shifts(group_id, master_shift) if group_id else [master_shift]
    if not old_chain:
        return jsonify({'message': 'Schedule chain not found.'}), 404

    old_chain = sorted(old_chain, key=lambda s: s.start_time)
    old_ids = [s.id for s in old_chain]
    old_start_date = min(s.start_time.date() for s in old_chain)
    old_end_date = max(s.start_time.date() for s in old_chain)
    day_offset = (new_start_date - old_start_date).days

    for old_shift in old_chain:
        new_start = old_shift.start_time + timedelta(days=day_offset)
        new_end = old_shift.end_time + timedelta(days=day_offset)

        if new_end <= new_start:
            return jsonify({'message': 'Invalid schedule time range.'}), 400

        for engineer_id in target_engineer_ids:
            collision = (
                db.session.query(Shift)
                .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
                .filter(
                    ShiftEngineer.engineer_id == engineer_id,
                    Shift.start_time < new_end,
                    Shift.end_time > new_start,
                    ~Shift.id.in_(old_ids)
                )
                .first()
            )
            if collision:
                conflict_engineer = db.session.get(Engineer, engineer_id)
                return build_conflict_response(collision, conflict_engineer, 'Move blocked by schedule conflict')

    primary_engineer_id = target_engineer_ids[0]

    for old_shift in old_chain:
        old_shift.start_time = old_shift.start_time + timedelta(days=day_offset)
        old_shift.end_time = old_shift.end_time + timedelta(days=day_offset)
        old_shift.engineer_id = primary_engineer_id

        delete_shift_engineer_links(old_shift.id)
        for engineer_id in target_engineer_ids:
            db.session.add(ShiftEngineer(shift_id=old_shift.id, engineer_id=engineer_id))

    new_end_date = max(s.start_time.date() for s in old_chain)
    old_date_label = old_start_date.isoformat() if old_start_date == old_end_date else f"{old_start_date.isoformat()} to {old_end_date.isoformat()}"
    new_date_label = new_start_date.isoformat() if new_start_date == new_end_date else f"{new_start_date.isoformat()} to {new_end_date.isoformat()}"
    target_engineer_names = [
        db.session.get(Engineer, engineer_id).name
        for engineer_id in target_engineer_ids
        if db.session.get(Engineer, engineer_id)
    ]

    mode_label = 'whole team' if move_mode == 'team' else 'single engineer'
    log_activity(
        f"Moved calendar schedule ({mode_label}): {master_shift.title} from {old_date_label} to {new_date_label}"
        + (f" | from: {', '.join(original_engineer_names)}" if original_engineer_names else "")
        + (f" | assigned: {', '.join(target_engineer_names)}" if target_engineer_names else "")
    )
    db.session.commit()

    send_schedule_event_notification_async(
        app,
        master_shift.id,
        target_engineer_ids,
        'Moved',
        current_user.username if current_user and current_user.is_authenticated else 'Scheduler',
        f'Moved from {old_date_label} to {new_date_label}.'
    )

    return jsonify({
        'status': 'success',
        'message': 'Schedule moved successfully.',
        'move_mode': move_mode,
        'engineers': target_engineer_ids,
        'new_start_date': new_start_date.isoformat(),
        'new_end_date': new_end_date.isoformat(),
        'old_start_date': old_start_date.isoformat(),
        'old_end_date': old_end_date.isoformat()
    })




def serialize_shift_delete_preview(shift):
    """Small, frontend-friendly schedule preview row for delete confirmation."""
    assigned_engineers = get_shift_engineer_records(shift)
    assigned_ids = [engineer.id for engineer in assigned_engineers]
    assigned_names = [engineer.name for engineer in assigned_engineers]

    return {
        'id': shift.id,
        'date': shift.start_time.date().isoformat() if shift.start_time else '',
        'date_label': shift.start_time.strftime('%a, %b %d, %Y') if shift.start_time else '',
        'time_label': f"{shift.start_time.strftime('%I:%M %p')} - {shift.end_time.strftime('%I:%M %p')}" if shift.start_time and shift.end_time else '',
        'task': shift.title,
        'client_name': shift.client.name if shift.client else '',
        'product_name': shift.product.name if shift.product else '',
        'status': shift.status or '',
        'engineer_ids': assigned_ids,
        'engineer_names': assigned_names,
        'engineers_label': ', '.join(assigned_names),
        'group_id': shift.group_id,
        'is_time_override': is_shift_time_override(shift),
        'parent_shift_id': shift.parent_shift_id,
        'override_engineer_id': shift.override_engineer_id
    }


def get_shift_delete_candidates(base_shift, delete_mode, start_date=None, end_date=None, engineer_id=None):
    """Resolve delete candidates without mutating the database.

    delete_mode values:
    - single: clicked schedule only
    - engineer_range: schedules for one engineer in selected date range
    - linked_chain: all base + linked override schedules in the clicked group
    - linked_range: base + linked override schedules in selected date range within clicked group
    """
    if not base_shift:
        return []

    delete_mode = clean_str(delete_mode) or 'single'
    if delete_mode not in {'single', 'engineer_range', 'linked_chain', 'linked_range'}:
        delete_mode = 'single'

    if delete_mode == 'single':
        return [base_shift]

    if delete_mode == 'engineer_range':
        if not engineer_id or not start_date or not end_date:
            return []

        return (
            db.session.query(Shift)
            .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
            .filter(
                ShiftEngineer.engineer_id == engineer_id,
                func.date(Shift.start_time) >= start_date,
                func.date(Shift.start_time) <= end_date
            )
            .order_by(Shift.start_time.asc(), Shift.id.asc())
            .all()
        )

    group_id = base_shift.group_id
    if not group_id:
        candidates = [base_shift]
    else:
        query = Shift.query.filter(Shift.group_id == group_id)

        if delete_mode == 'linked_range':
            if not start_date or not end_date:
                return []
            query = query.filter(
                func.date(Shift.start_time) >= start_date,
                func.date(Shift.start_time) <= end_date
            )

        candidates = query.order_by(Shift.start_time.asc(), Shift.id.asc()).all()

    # Unique by ID while preserving order.
    unique = {}
    for shift in candidates:
        unique[shift.id] = shift
    return list(unique.values())


def build_delete_preview_payload(candidates, delete_mode):
    """Build preview summary for delete modal."""
    rows = [serialize_shift_delete_preview(shift) for shift in candidates]
    affected_engineer_ids = sorted({
        engineer_id
        for row in rows
        for engineer_id in row.get('engineer_ids', [])
        if engineer_id
    })
    affected_engineer_names = []
    for engineer_id in affected_engineer_ids:
        engineer = db.session.get(Engineer, engineer_id)
        if engineer:
            affected_engineer_names.append(engineer.name)

    dates = [row['date'] for row in rows if row.get('date')]

    return {
        'status': 'success',
        'mode': delete_mode,
        'count': len(rows),
        'affected_engineer_count': len(affected_engineer_ids),
        'affected_engineers': affected_engineer_names,
        'start_date': min(dates) if dates else '',
        'end_date': max(dates) if dates else '',
        'items': rows
    }


def delete_shift_rows_with_cleanup(candidates):
    """Delete Shift rows and their assignment links.

    Physical report files are intentionally not removed here because the same
    stored filename may be referenced by linked parent/override rows.
    """
    deleted_ids = []
    notified_snapshots = []

    for shift in candidates:
        assigned_engineer_ids = get_shift_assigned_engineer_ids(shift)
        snapshot = {
            'id': shift.id,
            'title': shift.title,
            'date_iso': shift.start_time.date().isoformat() if shift.start_time else '',
            'date_label': shift.start_time.strftime('%B %d, %Y') if shift.start_time else '',
            'time_label': f"{shift.start_time.strftime('%I:%M %p')} - {shift.end_time.strftime('%I:%M %p')}" if shift.start_time and shift.end_time else '',
            'client_name': shift.client.name if shift.client else 'N/A',
            'product_name': shift.product.name if shift.product else 'N/A'
        }

        delete_shift_engineer_links(shift.id)
        db.session.delete(shift)
        deleted_ids.append(shift.id)
        notified_snapshots.append((snapshot, assigned_engineer_ids))

    return deleted_ids, notified_snapshots


@app.route('/preview_delete_shifts', methods=['POST'])
@login_required
def preview_delete_shifts():
    """Preview delete candidates before the frontend asks for final confirmation."""
    payload = request.get_json(silent=True) or {}

    shift_id = clean_int(payload.get('shift_id'))
    delete_mode = clean_str(payload.get('delete_mode')) or 'single'
    start_date = parse_date(payload.get('start_date'))
    end_date = parse_date(payload.get('end_date'))
    engineer_id = clean_int(payload.get('engineer_id'))

    if start_date and end_date and end_date < start_date:
        start_date, end_date = end_date, start_date

    base_shift = db.session.get(Shift, shift_id) if shift_id else None
    if not base_shift:
        return jsonify({'message': 'Schedule not found.'}), 404

    candidates = get_shift_delete_candidates(
        base_shift,
        delete_mode,
        start_date=start_date,
        end_date=end_date,
        engineer_id=engineer_id
    )

    if not candidates:
        return jsonify({
            'status': 'success',
            'mode': delete_mode,
            'count': 0,
            'affected_engineer_count': 0,
            'affected_engineers': [],
            'items': []
        })

    affected_ids = sorted({
        engineer_id
        for shift in candidates
        for engineer_id in get_shift_assigned_engineer_ids(shift)
        if engineer_id
    })

    if not can_modify_schedule_for_engineer_ids(affected_ids):
        return denied('You are not authorized to delete one or more schedules in this preview.')

    return jsonify(build_delete_preview_payload(candidates, delete_mode))


@app.route('/delete_shifts_previewed', methods=['POST'])
@login_required
def delete_shifts_previewed():
    """Delete the previewed set after frontend confirmation."""
    payload = request.get_json(silent=True) or {}

    shift_id = clean_int(payload.get('shift_id'))
    delete_mode = clean_str(payload.get('delete_mode')) or 'single'
    start_date = parse_date(payload.get('start_date'))
    end_date = parse_date(payload.get('end_date'))
    engineer_id = clean_int(payload.get('engineer_id'))

    if start_date and end_date and end_date < start_date:
        start_date, end_date = end_date, start_date

    base_shift = db.session.get(Shift, shift_id) if shift_id else None
    if not base_shift:
        return jsonify({'message': 'Schedule not found.'}), 404

    candidates = get_shift_delete_candidates(
        base_shift,
        delete_mode,
        start_date=start_date,
        end_date=end_date,
        engineer_id=engineer_id
    )

    if not candidates:
        return jsonify({'status': 'success', 'count': 0, 'deleted_ids': []})

    affected_ids = sorted({
        engineer_id
        for shift in candidates
        for engineer_id in get_shift_assigned_engineer_ids(shift)
        if engineer_id
    })

    if not can_modify_schedule_for_engineer_ids(affected_ids):
        return denied('You are not authorized to delete one or more selected schedules.')

    preview_payload = build_delete_preview_payload(candidates, delete_mode)
    deleted_ids, notified_snapshots = delete_shift_rows_with_cleanup(candidates)
    db.session.commit()

    actor_name = current_user.username.capitalize()
    mode_labels = {
        'single': 'single schedule',
        'engineer_range': 'engineer date range',
        'linked_chain': 'linked schedule chain',
        'linked_range': 'linked schedule selected dates'
    }
    log_activity(
        f"Deleted {len(deleted_ids)} schedule(s) via {mode_labels.get(delete_mode, delete_mode)}"
        + (f": {preview_payload.get('start_date')} to {preview_payload.get('end_date')}" if preview_payload.get('start_date') else "")
    )

    # Send individual delete notifications only for today/future schedules.
    for snapshot, recipient_engineer_ids in notified_snapshots:
        send_schedule_deleted_notification_async(
            app,
            snapshot,
            recipient_engineer_ids,
            current_user.username if current_user and current_user.is_authenticated else 'Scheduler'
        )

    return jsonify({
        'status': 'success',
        'count': len(deleted_ids),
        'deleted_ids': deleted_ids,
        'preview': preview_payload
    })



def enforce_completed_schedule_tsr_after_file_delete(linked_shifts, deleted_filename):
    """Prevent Completed schedules from remaining completed without a TSR.

    If the deleted file was the last TSR linked to a completed work schedule,
    downgrade that schedule to For Continuation immediately. This closes the
    refresh/F5 loophole after file deletion.
    """
    reverted_shift_ids = []

    for shift in linked_shifts or []:
        if not shift:
            continue

        is_work_schedule = bool(getattr(shift, 'client_id', None))
        current_status = (getattr(shift, 'status', None) or '').strip()

        if not is_work_schedule or current_status != 'Completed':
            continue

        remaining_filenames = [
            file_rec.filename
            for file_rec in ShiftFile.query.filter_by(shift_id=shift.id).all()
        ]

        if existing_files_have_tsr(remaining_filenames):
            continue

        shift.status = 'For Continuation'
        reverted_shift_ids.append(shift.id)

    if reverted_shift_ids:
        log_activity(
            f"Reverted completed schedule(s) to For Continuation after TSR deletion: "
            f"{', '.join(map(str, reverted_shift_ids))} | removed file: {deleted_filename}"
        )

    return reverted_shift_ids


@app.route('/delete_file/<string:filename>', methods=['DELETE'])
@login_required
def delete_file(filename):
    """Individual technical attachment removal.

    Permission model:
    - admins/schedulers can remove files according to normal schedule authority
    - engineers can remove files only from schedules they are assigned to/work on
    - shared file references are removed only when the user can work on every linked schedule row
    """
    safe_filename = os.path.basename(filename)
    if safe_filename != filename:
        return jsonify({'message': 'Invalid filename'}), 400

    file_recs = ShiftFile.query.filter_by(filename=safe_filename).all()
    if not file_recs:
        return jsonify({'message': 'File not found'}), 404

    linked_shifts = [
        db.session.get(Shift, file_rec.shift_id)
        for file_rec in file_recs
        if file_rec.shift_id
    ]
    linked_shifts = [shift for shift in linked_shifts if shift]

    if not linked_shifts:
        return jsonify({'message': 'Linked schedule not found for this file.'}), 404

    if not all(can_work_on_existing_schedule_shift(shift) for shift in linked_shifts):
        return denied('You are not authorized to delete this file.')

    try:
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], safe_filename))
    except OSError:
        pass

    for file_rec in file_recs:
        db.session.delete(file_rec)

    db.session.flush()
    reverted_shift_ids = enforce_completed_schedule_tsr_after_file_delete(linked_shifts, safe_filename)

    db.session.commit()
    log_activity(f"Deleted report file: {safe_filename}")
    return jsonify({
        'status': 'success',
        'filename': safe_filename,
        'removed_links': len(file_recs),
        'reverted_status_shift_ids': reverted_shift_ids,
        'reverted_status_count': len(reverted_shift_ids),
        'status_reverted_to': 'For Continuation' if reverted_shift_ids else None
    })


@app.route('/delete_shift/<int:id>', methods=['DELETE'])
@login_required
def delete_shift(id):
    """Individual technical record removal. Kept for existing frontend compatibility."""
    target = db.session.get(Shift, id)
    if target:
        if not can_modify_schedule_shift(target):
            return denied('You are not authorized to delete this schedule.')

        title = target.title
        deleted_ids, notified_snapshots = delete_shift_rows_with_cleanup([target])
        db.session.commit()
        log_activity(f"Wiped technical record: {title}")

        for snapshot, assigned_engineer_ids in notified_snapshots:
            send_schedule_deleted_notification_async(
                app,
                snapshot,
                assigned_engineer_ids,
                current_user.username if current_user and current_user.is_authenticated else 'Scheduler'
            )

        return jsonify({'status': 'success', 'count': len(deleted_ids), 'deleted_ids': deleted_ids})

    return jsonify({'status': 'success', 'count': 0, 'deleted_ids': []})


@app.route('/batch_delete_shifts', methods=['POST'])
@login_required
def batch_delete_shifts():
    """Technical personnel bulk range purge tool.

    Backward compatible with existing frontend.
    Optional payload:
    - preview_only: true returns preview without deleting.
    """
    payload = request.get_json(silent=True) or {}
    eng_id = clean_int(payload.get('engineer_id'))
    s_date = parse_date(payload.get('start_date'))
    e_date = parse_date(payload.get('end_date'))
    preview_only = parse_bool_flag(payload.get('preview_only'), default=False)

    if s_date and e_date and e_date < s_date:
        s_date, e_date = e_date, s_date

    engineer = db.session.get(Engineer, eng_id)

    if not engineer:
        return jsonify({'message': 'Engineer not found'}), 404
    if not s_date or not e_date:
        return jsonify({'message': 'Start and end date are required.'}), 400

    if not can_modify_schedule_for_engineer_ids([eng_id]):
        return denied('You are not authorized to wipe this schedule range.')

    purge_list = (
        db.session.query(Shift)
        .join(ShiftEngineer, Shift.id == ShiftEngineer.shift_id)
        .filter(
            ShiftEngineer.engineer_id == eng_id,
            func.date(Shift.start_time) >= s_date,
            func.date(Shift.start_time) <= e_date
        )
        .order_by(Shift.start_time.asc(), Shift.id.asc())
        .all()
    )

    preview_payload = build_delete_preview_payload(purge_list, 'engineer_range')
    if preview_only:
        return jsonify(preview_payload)

    deleted_ids, notified_snapshots = delete_shift_rows_with_cleanup(purge_list)
    db.session.commit()

    count = len(deleted_ids)
    log_activity(f"Bulk-purged {count} entries for {engineer.name}")

    for snapshot, assigned_engineer_ids in notified_snapshots:
        send_schedule_deleted_notification_async(
            app,
            snapshot,
            assigned_engineer_ids,
            current_user.username if current_user and current_user.is_authenticated else 'Scheduler'
        )

    return jsonify({'status': 'success', 'count': count, 'deleted_ids': deleted_ids})


# --- PERSONNEL HR ACTIONS ---

@app.route('/add_engineer', methods=['POST'])
@login_required
def add_engineer():
    """ Register new staff and auto-create User account """
    if not is_admin_authorized(): return jsonify({'message': 'Denied'}), 403
    p = request.get_json()
    emp_id = clean_str(p.get('employee_id'))
    name_val = clean_str(p['name'])

    if Engineer.query.filter_by(employee_id=emp_id).first():
        return jsonify({'message': f'Error: Employee ID {emp_id} is already taken.'}), 400

    first_name = name_val.split()[0].lower()
    if User.query.filter_by(username=first_name).first():
        first_name = name_val.replace(" ", "").lower()

    # Generate secure temp password
    temp_password = secrets.token_urlsafe(8)

    # Create linked system account
    new_user = User(
        username=first_name,
        password=generate_password_hash(temp_password),
        role='engineer',
        must_change_password=True
    )
    db.session.add(new_user)
    db.session.flush()

    new_eng = Engineer(
        user_id=new_user.id,
        employee_id=emp_id,
        name=name_val,
        initials=clean_str(p['initials']), 
        phone=clean_str(p.get('phone')),
        email=clean_str(p.get('email')), 
        branch=clean_str(p.get('branch'))
    )
    db.session.add(new_eng)

    db.session.commit()
    log_activity(f"Added technical staff: {name_val} (ID: {emp_id})")

    return jsonify({
        'status': 'success',
        'username': first_name,
        'temp_password': temp_password
    })


@app.route('/update_engineer/<int:id>', methods=['PUT'])
@login_required
def update_engineer(id):
    """Updates directory metadata.

    Admins can edit full personnel details.
    Engineers can edit only their own contact number/email.
    """
    p = request.get_json(silent=True) or {}
    eng = db.session.get(Engineer, id)
    if not eng:
        return jsonify({'message': 'Missing'}), 404

    if getattr(current_user, 'role', None) == 'engineer' and not is_admin_authorized():
        my_profile = getattr(current_user, 'engineer_profile', None)
        if not my_profile or my_profile.id != eng.id:
            return denied('Engineers can edit only their own contact information.')

        eng.phone = clean_str(p.get('phone'))
        eng.email = clean_str(p.get('email'))

        db.session.commit()
        log_activity(f"Updated own contact info: {eng.name}")
        return jsonify({'status': 'success', 'self_contact_update': True})

    if not is_admin_authorized():
        return jsonify({'message': 'Denied'}), 403

    new_eid = clean_str(p.get('employee_id'))
    if new_eid != eng.employee_id and Engineer.query.filter_by(employee_id=new_eid).first():
        return jsonify({'message': 'Employee ID already taken.'}), 400

    eng.employee_id = new_eid
    eng.name, eng.initials = clean_str(p['name']), clean_str(p['initials'])
    eng.phone, eng.email, eng.branch = clean_str(p.get('phone')), clean_str(p.get('email')), clean_str(p.get('branch'))

    db.session.commit()
    log_activity(f"Updated profile for: {eng.name}")
    return jsonify({'status': 'success'})


@app.route('/delete_engineer/<int:id>', methods=['DELETE'])
@login_required
def delete_engineer(id):
    """ Deletes engineer and associated system account. """
    if not is_admin_authorized(): return jsonify({'message': 'Denied'}), 403
    eng = db.session.get(Engineer, id)
    if eng:
        name = eng.name
        # Delete linked account
        user_acc = db.session.get(User, eng.user_id) if eng.user_id else None
        if user_acc: db.session.delete(user_acc)
        
        db.session.delete(eng); db.session.commit()
        log_activity(f"Permanently removed personnel: {name}")
        return jsonify({'status': 'success'})
    return jsonify({'message': 'Not Found'}), 404


@app.route('/make_user_admin/<int:user_id>', methods=['POST'])
@login_required
def make_user_admin(user_id):
    """Admin promotion has been removed by policy. Generic admin role is disabled."""
    return denied('Admin promotion has been disabled.')


@app.route('/reset_user_password/<int:user_id>', methods=['POST'])
@login_required
def reset_user_password(user_id):
    """ Master Admin password reset logic with hierarchy protection. """
    target_user = db.session.get(User, user_id)
    if not target_user:
        return jsonify({'message': 'No account'}), 404

    allowed, reason = can_reset_password_for_user(target_user)
    if not allowed:
        return denied(reason)

    new_pw = ''.join(secrets.choice(string.ascii_letters + string.digits) for i in range(8))
    target_user.password = generate_password_hash(new_pw)
    db.session.commit()
    log_activity(f"Reset password for user: {target_user.username}")
    return jsonify({'new_pw': new_pw})




@app.route('/change_password', methods=['POST'])
@login_required
def change_password():
    """ Standard and Admin target password rotation with hierarchy protection. """
    p = request.get_json()
    new_pw = p.get('new_password')
    target_id = p.get('target_user_id')

    if not new_pw or len(new_pw) < 4:
        return jsonify({'message': 'Error: 4+ chars required'}), 400

    if target_id:
        target_user = db.session.get(User, int(target_id))
        if not target_user:
            return jsonify({'message': 'No account'}), 404

        # Own password change remains allowed for the logged-in account.
        if int(target_id) != int(current_user.id):
            allowed, reason = can_reset_password_for_user(target_user)
            if not allowed:
                return denied(reason)

        target_user.password = generate_password_hash(new_pw)
        log_activity(f"Admin forced password change for: {target_user.username}")
    else:
        current_user.password = generate_password_hash(new_pw)
        log_activity("Changed their own login password")

    db.session.commit()
    return jsonify({'status': 'success'})





@app.route('/force_change_password_api', methods=['POST'])
@login_required
def force_change_password_api():
    data = request.get_json()
    new_pw = data.get('new_password')

    if not new_pw or len(new_pw) < 8:
        return jsonify({'message': 'Minimum 8 characters required'}), 400

    current_user.password = generate_password_hash(new_pw)
    current_user.must_change_password = False

    db.session.commit()
    session.pop('force_pw_change', None)

    return jsonify({'status': 'success'})

# --- EQUIPMENT INVENTORY ACTIONS ---

@app.route('/add_product', methods=['POST'])
@login_required
def add_product():
    """ Equipment entry logic. Access: Admin Levels & Engineers.

    Safety fix:
    - Prevents duplicate serial numbers from crashing SQLite UNIQUE constraint.
    - Returns a clean 409 response instead of a 500 server error.
    - Rolls back the session if a commit fails for any unexpected reason.
    """
    if not (is_admin_authorized() or current_user.role == 'engineer'):
        return jsonify({'message': 'Denied'}), 403

    d = request.get_json() or {}

    serial_number = clean_str(d.get('serial_number')).upper()
    product_name = clean_str(d.get('name'))

    if not serial_number:
        return jsonify({'message': 'Serial number is required.'}), 400

    if not product_name:
        return jsonify({'message': 'Product name is required.'}), 400

    existing_product = db.session.get(Product, serial_number)
    if existing_product:
        return jsonify({
            'status': 'duplicate',
            'message': f'Serial number {serial_number} already exists in inventory.',
            'serial_number': serial_number,
            'existing_name': existing_product.name or ''
        }), 409

    new_p = Product(
        serial_number=serial_number,
        name=product_name,
        client_id=clean_int(d.get('client_id')),
        start_warranty_date=parse_date(d.get('start_warranty')),
        end_warranty_date=parse_date(d.get('end_warranty'))
    )

    try:
        db.session.add(new_p)
        db.session.commit()
    except Exception as product_error:
        db.session.rollback()

        # A duplicate may still happen if two users submit the same serial at the same time.
        if 'UNIQUE constraint failed' in str(product_error) or 'product.serial_number' in str(product_error):
            return jsonify({
                'status': 'duplicate',
                'message': f'Serial number {serial_number} already exists in inventory.',
                'serial_number': serial_number
            }), 409

        print(f"[PRODUCT] Add product failed for {serial_number}: {product_error}", flush=True)
        return jsonify({'message': 'Unable to add product. Please try again.'}), 500

    log_activity(f"Added equipment: {new_p.serial_number}")
    return jsonify({'status': 'success'})


@app.route('/update_product/<serial_number>', methods=['PUT'])
@login_required
def update_product(serial_number):
    """Modify a medical asset, including safe serial-number replacement.

    Product.serial_number is the primary key, so changing it must also move
    linked schedules from the old serial to the new serial. This prevents the
    old delete/recreate workaround from breaking historical schedule links.
    """
    if not (is_admin_authorized() or current_user.role == 'engineer'):
        return jsonify({'message': 'Denied'}), 403

    d = request.get_json() or {}
    old_serial = clean_str(serial_number)
    p = db.session.get(Product, old_serial)
    if not p:
        return jsonify({'message': 'Missing'}), 404

    new_serial = clean_str(d.get('serial_number'))
    product_name = clean_str(d.get('name'))

    if not new_serial:
        return jsonify({'message': 'Serial number is required.'}), 400

    if not product_name:
        return jsonify({'message': 'Product name is required.'}), 400

    new_serial = new_serial.upper()
    old_serial = p.serial_number
    old_name = p.name or ''
    old_client_id = p.client_id
    old_start = p.start_warranty_date
    old_end = p.end_warranty_date

    new_client_id = clean_int(d.get('client_id'))
    new_start = parse_date(d.get('start_warranty'))
    new_end = parse_date(d.get('end_warranty'))

    try:
        if new_serial != old_serial:
            existing_product = db.session.get(Product, new_serial)
            if existing_product:
                return jsonify({
                    'status': 'duplicate',
                    'message': f'Serial number {new_serial} already exists in inventory.',
                    'serial_number': new_serial,
                    'existing_name': existing_product.name or ''
                }), 409

            linked_schedule_count = Shift.query.filter_by(product_id=old_serial).count()

            replacement = Product(
                serial_number=new_serial,
                name=product_name,
                client_id=new_client_id,
                start_warranty_date=new_start,
                end_warranty_date=new_end
            )
            db.session.add(replacement)
            db.session.flush()

            Shift.query.filter_by(product_id=old_serial).update(
                {'product_id': new_serial},
                synchronize_session=False
            )

            db.session.delete(p)
            db.session.commit()

            log_activity(
                f"Changed product serial number for {product_name} from {old_serial} to {new_serial} "
                f"({linked_schedule_count} linked schedule(s) updated)"
            )

            return jsonify({
                'status': 'success',
                'serial_changed': True,
                'old_serial': old_serial,
                'new_serial': new_serial,
                'linked_schedule_count': linked_schedule_count
            })

        p.name = product_name
        p.client_id = new_client_id
        p.start_warranty_date = new_start
        p.end_warranty_date = new_end
        db.session.commit()

        changed_fields = []
        if old_name != product_name:
            changed_fields.append('name')
        if old_client_id != new_client_id:
            changed_fields.append('owner')
        if old_start != new_start or old_end != new_end:
            changed_fields.append('warranty dates')

        field_note = f" ({', '.join(changed_fields)})" if changed_fields else ''
        log_activity(f"Updated product details: {p.serial_number}{field_note}")
        return jsonify({'status': 'success', 'serial_changed': False})

    except Exception as product_error:
        db.session.rollback()

        if 'UNIQUE constraint failed' in str(product_error) or 'product.serial_number' in str(product_error):
            return jsonify({
                'status': 'duplicate',
                'message': f'Serial number {new_serial} already exists in inventory.',
                'serial_number': new_serial
            }), 409

        print(f"[PRODUCT] Update product failed for {old_serial} -> {new_serial}: {product_error}", flush=True)
        return jsonify({'message': 'Unable to update product. Please try again.'}), 500


@app.route('/delete_product/<serial_number>', methods=['DELETE'])
@login_required
def delete_product(serial_number):
    """ Restricted to Admin Levels. """
    if not is_admin_authorized(): return jsonify({'message': 'Denied'}), 403
    target = db.session.get(Product, serial_number)
    if target: 
        name = target.name
        db.session.delete(target); db.session.commit()
        log_activity(f"Purged product record: {name}")
    return jsonify({'status': 'success'})



def csv_get(row, *names):
    """Case/spacing tolerant CSV row value getter."""
    if not row:
        return None

    normalized = {
        str(k or '').strip().lower().replace('_', ' ').replace('-', ' '): v
        for k, v in row.items()
    }

    for name in names:
        key = str(name or '').strip().lower().replace('_', ' ').replace('-', ' ')
        if key in normalized:
            return normalized[key]

    return None


def csv_find_client_id(owner_name):
    """Resolve imported client/owner names using exact, ilike, acronym, and similarity fallback."""
    owner_name = clean_str(owner_name)
    if not owner_name:
        return None

    target = Client.query.filter(Client.name.ilike(owner_name)).first()
    if target:
        return target.id

    owner_acronym = generate_acronym(owner_name)
    best_match = None
    best_score = 0

    for client in Client.query.all():
        score = similarity(owner_name.upper(), client.name.upper())
        if owner_acronym and owner_acronym == generate_acronym(client.name):
            score += 0.25
        if score > best_score:
            best_score = score
            best_match = client

    return best_match.id if best_match and best_score >= 0.75 else None



# --- ENTERPRISE BULK DATA IMPORT MODULE ---

@app.route('/import_clients', methods=['POST'])
@login_required
def import_clients():
    """CSV bulk ingestion for clients. Supports dynamic Contact rows and friendly counts."""
    if not is_admin_authorized():
        return jsonify({'message': 'Denied'}), 403

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'message': 'No CSV file selected.'}), 400
    if not file.filename.lower().endswith('.csv'):
        return jsonify({'message': 'Only .csv files are supported.'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
        reader = csv.DictReader(stream)

        if not reader.fieldnames:
            return jsonify({'message': 'CSV file is empty or missing headers.'}), 400

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in reader:
            name = clean_str(csv_get(row, 'Name', 'Client Name', 'Medical Center', 'Medical Center Name', 'Customer', 'Company Name'))
            address = clean_str(csv_get(row, 'Address', 'Office Address', 'Location'))

            if not name:
                skipped_count += 1
                continue

            existing = check_for_duplicate_client(name, address)

            if existing:
                client_rec = existing
                client_rec.name = name
                client_rec.address = address
                Contact.query.filter_by(client_id=client_rec.id).delete()
                updated_count += 1
            else:
                client_rec = Client(name=name, address=address)
                db.session.add(client_rec)
                db.session.flush()
                created_count += 1

            contact_rows = []

            # Flexible multi-contact import:
            # CP1/CN1/CE1, Contact 1 Name/Phone/Email, Contact Person 1, etc.
            for idx in range(1, 31):
                c_name = clean_str(csv_get(row, f'CP{idx}', f'Contact {idx} Name', f'Contact {idx}', f'Contact Person {idx}'))
                c_phone = clean_str(csv_get(row, f'CN{idx}', f'Contact {idx} Phone', f'Contact {idx} Number', f'Phone {idx}'))
                c_email = clean_str(csv_get(row, f'CE{idx}', f'Contact {idx} Email', f'Email {idx}'))

                if c_name or c_phone or c_email:
                    contact_rows.append((c_name, c_phone, c_email))

            # Simple one-contact CSV fallback.
            if not contact_rows:
                c_name = clean_str(csv_get(row, 'Contact Name', 'Contact Person', 'Main Contact'))
                c_phone = clean_str(csv_get(row, 'Phone', 'Contact Phone', 'Contact Number', 'Mobile'))
                c_email = clean_str(csv_get(row, 'Email', 'Contact Email'))
                if c_name or c_phone or c_email:
                    contact_rows.append((c_name, c_phone, c_email))

            for c_name, c_phone, c_email in contact_rows:
                db.session.add(Contact(
                    client_id=client_rec.id,
                    name=c_name,
                    phone=c_phone,
                    email=c_email
                ))

            # Keep legacy first 3 contact columns populated for backward compatibility.
            legacy = contact_rows[:3]
            fields = [
                ('contact_person_1', 'contact_number_1', 'email_address_1'),
                ('contact_person_2', 'contact_number_2', 'email_address_2'),
                ('contact_person_3', 'contact_number_3', 'email_address_3')
            ]
            for field_idx, field_names in enumerate(fields):
                values = legacy[field_idx] if field_idx < len(legacy) else (None, None, None)
                setattr(client_rec, field_names[0], values[0])
                setattr(client_rec, field_names[1], values[1])
                setattr(client_rec, field_names[2], values[2])

        db.session.commit()

        total = created_count + updated_count
        log_activity(f"Imported client CSV: {created_count} created, {updated_count} updated, {skipped_count} skipped")
        return jsonify({
            'status': 'success',
            'message': f'Client import complete: {created_count} created, {updated_count} updated, {skipped_count} skipped.',
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'total': total
        })

    except UnicodeDecodeError:
        return jsonify({'message': 'Unable to read CSV. Please save it as UTF-8 CSV and try again.'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Import failed: {str(e)}'}), 500


@app.route('/import_products', methods=['POST'])
@login_required
def import_products():
    """CSV bulk ingestion for product inventory with flexible headers and counts."""
    if not is_admin_authorized():
        return jsonify({'message': 'Denied'}), 403

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'message': 'No CSV file selected.'}), 400
    if not file.filename.lower().endswith('.csv'):
        return jsonify({'message': 'Only .csv files are supported.'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
        reader = csv.DictReader(stream)

        if not reader.fieldnames:
            return jsonify({'message': 'CSV file is empty or missing headers.'}), 400

        created_count = 0
        updated_count = 0
        skipped_count = 0
        unresolved_owner_count = 0

        for row in reader:
            serial = clean_str(csv_get(row, 'Serial Number', 'Serial', 'S/N', 'SN'))
            product_name = clean_str(csv_get(row, 'Description', 'Product Name', 'Name', 'Equipment Name', 'Item'))
            owner_name = clean_str(csv_get(row, 'Current Owner', 'Owner', 'Client', 'Client Name', 'Medical Center'))
            start_date = parse_date(csv_get(row, 'Start Date', 'Warranty Start', 'Start Warranty', 'Start Warranty Date'))
            end_date = parse_date(csv_get(row, 'Expiry Date', 'End Date', 'Warranty End', 'End Warranty', 'End Warranty Date'))

            if not serial or not product_name:
                skipped_count += 1
                continue

            client_id = csv_find_client_id(owner_name)
            if owner_name and not client_id:
                unresolved_owner_count += 1

            product = db.session.get(Product, serial)
            if product:
                product.name = product_name
                product.client_id = client_id
                product.start_warranty_date = start_date
                product.end_warranty_date = end_date
                updated_count += 1
            else:
                db.session.add(Product(
                    serial_number=serial,
                    name=product_name,
                    client_id=client_id,
                    start_warranty_date=start_date,
                    end_warranty_date=end_date
                ))
                created_count += 1

        db.session.commit()

        total = created_count + updated_count
        message = f'Product import complete: {created_count} created, {updated_count} updated, {skipped_count} skipped.'
        if unresolved_owner_count:
            message += f' {unresolved_owner_count} owner name(s) were not matched and were left blank.'

        log_activity(f"Imported product CSV: {created_count} created, {updated_count} updated, {skipped_count} skipped")
        return jsonify({
            'status': 'success',
            'message': message,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'unresolved_owners': unresolved_owner_count,
            'total': total
        })

    except UnicodeDecodeError:
        return jsonify({'message': 'Unable to read CSV. Please save it as UTF-8 CSV and try again.'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Import failed: {str(e)}'}), 500



@app.before_request
def ensure_runtime_sqlite_migrations_before_request():
    """Run lightweight SQLite compatibility migrations before normal routes.

    Needed when app is started by gunicorn/Railway instead of python app.py.
    Also ensures a fresh Railway volume database has all base tables before
    column/index migrations run.
    """
    if request.endpoint == 'static':
        return

    # Safe on existing SQLite databases; creates tables only when missing.
    db.create_all()

    # Fresh Railway volume emergency login bootstrap.
    ensure_emergency_superadmin_from_env()

    ensure_shift_file_original_filename_column()
    ensure_schedule_delete_indexes()


@app.before_request
def enforce_password_change():
    if current_user.is_authenticated and getattr(current_user, "must_change_password", False):
        if request.endpoint not in ['force_change_password_api','logout','static']:
            session['force_pw_change'] = True


@app.route('/get_shift_range/<int:shift_id>')
@login_required
def get_shift_range(shift_id):
    shift = db.session.get(Shift, shift_id)
    if not shift:
        return jsonify({'error': 'not found'}), 404

    group_id = shift.group_id
    chain = Shift.query.filter_by(group_id=group_id).all()
    dates = [s.start_time.date() for s in chain]

    return jsonify({
        'start_date': min(dates).isoformat(),
        'end_date': max(dates).isoformat()
    })

# --- APPLICATION STARTUP HELPERS ---


def migrate_hierarchy_accounts():
    """One-time safe alignment for the named hierarchy accounts.

    - Renames legacy 'hannah' account to 'hanna' when possible.
    - Aligns named authority accounts to their intended backend roles.
    - Does not change passwords.
    """
    legacy_hannah = User.query.filter_by(username='hannah').first()
    existing_hanna = User.query.filter_by(username='hanna').first()

    if legacy_hannah and not existing_hanna:
        legacy_hannah.username = 'hanna'

    intended_roles = {
        'jonamar': 'superadmin',
        'rodito': 'superadmin',
        'robert': 'superadmin',
        'diary': 'superadmin',
        'hanna': 'superadmin',
        'kevin': 'regional_admin'
    }

    for username, role in intended_roles.items():
        user = User.query.filter_by(username=username).first()
        if user:
            user.role = role

    # Unique Regional Admin Kevin binding.
    # The admin account must link only to employee_id 15-148.
    # Any other engineer named Kevin should not inherit the admin profile/role.
    kevin_user = User.query.filter_by(username=REGIONAL_ADMIN_USERNAME).first()
    if kevin_user:
        kevin_user.role = 'regional_admin'
        admin_kevin_engineer = Engineer.query.filter_by(employee_id=REGIONAL_ADMIN_EMPLOYEE_ID).first()

        if admin_kevin_engineer:
            admin_kevin_engineer.user_id = kevin_user.id

            for engineer in Engineer.query.filter(
                Engineer.user_id == kevin_user.id,
                Engineer.id != admin_kevin_engineer.id
            ).all():
                engineer.user_id = None

    db.session.commit()


def bootstrap_static_accounts():
    """
    Creates missing first-run admin accounts safely.

    Important behavior for live systems:
    - Existing users are never overwritten.
    - Existing passwords are never reset.
    - Existing engineer records are not duplicated.

    For fresh local/test databases:
    - Missing accounts are created with either SEED_PASSWORD_<USERNAME> from .env
      or a generated temporary password printed once in the terminal.
    - Bootstrapped accounts must change password after first login.
    """
    bootstrap_credentials = []

    # SuperAdmins
    # (username, is_engineer, emp_id, branch, full_name, initials, phone, email)
    super_users = [
        ('rodito', True, '99-055', 'Manila', 'Rodito Aretano Jr', 'RAJ', '09175961523', 'r_aretano@shimadzu.com.ph'),
        ('robert', True, '14-146', 'Manila', 'Robert Rio', 'RR', '09175075933', 'rnrio@shimadzu.com.ph'),
        ('jonamar', True, '18-185', 'Manila', 'Jonamar Paunil', 'JP', '09171323953', 'jonamar@shimadzu.com.ph'),
        ('diary', False, None, None, None, None, None, None),
        ('hanna', False, None, None, None, None, None, None)
    ]

    for u_name, is_eng, e_id, br, f_name, inits, ph, em in super_users:
        env_password = os.environ.get(f"SEED_PASSWORD_{u_name.upper()}")
        u_rec, temp_pw, created = bootstrap_user(u_name, 'superadmin', env_password)
        if created:
            bootstrap_credentials.append((u_name, temp_pw, 'superadmin'))

        if is_eng and e_id and not Engineer.query.filter_by(employee_id=e_id).first():
            db.session.add(Engineer(
                user_id=u_rec.id,
                employee_id=e_id,
                branch=br,
                name=f_name,
                initials=inits,
                phone=ph,
                email=em
            ))

    # Regional Admin (Kevin)
    kevin_env_password = os.environ.get("SEED_PASSWORD_KEVIN")
    kevin_rec, kevin_temp_pw, kevin_created = bootstrap_user('kevin', 'regional_admin', kevin_env_password)
    if kevin_created:
        bootstrap_credentials.append(('kevin', kevin_temp_pw, 'regional_admin'))

    if kevin_rec and not Engineer.query.filter_by(employee_id=REGIONAL_ADMIN_EMPLOYEE_ID).first():
        db.session.add(Engineer(
            user_id=kevin_rec.id,
            employee_id=REGIONAL_ADMIN_EMPLOYEE_ID,
            branch='Davao',
            name='Kevin Bautista',
            initials='KB',
            phone='09175117264',
            email='kevin@shimadzu.com.ph'
        ))

    db.session.commit()
    return bootstrap_credentials


def print_bootstrap_credentials(bootstrap_credentials):
    """Prints first-run temporary credentials only when new bootstrap users were created."""
    if not bootstrap_credentials:
        return

    print("\n" + "=" * 72)
    print("FIRST-RUN TEMPORARY LOGIN CREDENTIALS")
    print("Use these only for local testing. Change the password after login.")
    print("They are shown once because the database was missing these users.")
    print("=" * 72)
    for username, password, role in bootstrap_credentials:
        print(f"{role:15} username={username:<12} temporary_password={password}")
    print("=" * 72 + "\n")


def initialize_database():
    """
    Initializes database tables and first-run bootstrap records.

    This function is intentionally separate from app.run() so startup tasks are
    ordered, testable, and not mixed directly with the server entry point.
    """
    with app.app_context():
        db.create_all()
        ensure_shift_override_columns()
        ensure_shift_file_original_filename_column()
        ensure_schedule_delete_indexes()
        migrate_hierarchy_accounts()
        bootstrap_credentials = bootstrap_static_accounts()
        migrate_hierarchy_accounts()
        repair_shift_engineer_links()
        print_bootstrap_credentials(bootstrap_credentials)


def get_bool_env(name, default=False):
    """Reads common true/false environment values safely."""
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {'1', 'true', 'yes', 'on'}


def run_development_server():
    """
    Runs Flask's built-in server for local/ngrok development only.

    For ngrok/live testing, use:
    - FLASK_DEBUG=false
    - FLASK_HOST=127.0.0.1
    - ngrok http 5000
    """
    debug_mode = get_bool_env('FLASK_DEBUG', default=False)

    # Support both names. FLASK_HOST is what your project instructions use;
    # FLASK_RUN_HOST is also common in Flask environments.
    host = os.environ.get('FLASK_HOST') or os.environ.get('FLASK_RUN_HOST') or '127.0.0.1'
    port = int(os.environ.get('FLASK_PORT') or os.environ.get('FLASK_RUN_PORT') or '5000')

    app.run(debug=debug_mode, host=host, port=port)


# --- SYSTEM ENTRY POINT ---

if __name__ == '__main__':
    initialize_database()
    run_development_server()
