"""Behavioral coverage for the standalone Reimbursement Tracker register."""

import json
import os
import pathlib
import tempfile
import unittest
import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

os.environ.setdefault(
    'MEDICAL_SERVICE_TEST_DB',
    str(pathlib.Path(tempfile.gettempdir()) / 'medical_service_reimbursement_tracker_tests.db'),
)

import app as app_module  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


ROOT = pathlib.Path(__file__).resolve().parents[1]


class ReimbursementTrackerWorkflowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = app_module.app
        cls.app.config['TESTING'] = True
        cls.app.config['WTF_CSRF_ENABLED'] = False
        cls.suffix = uuid.uuid4().hex[:8]
        cls.created_user_ids = []
        cls.created_engineer_ids = []

        with cls.app.app_context():
            app_module.db.create_all()
            app_module.ensure_user_admin_capability_columns()
            app_module.ensure_reimbursement_tracker_schema()

            cls.superadmin = app_module.User.query.filter_by(username='jonamar').first()
            if not cls.superadmin:
                cls.superadmin = app_module.User(
                    username='jonamar',
                    password=app_module.generate_password_hash('test-password'),
                    role='superadmin',
                    is_active=True,
                )
                app_module.db.session.add(cls.superadmin)
                app_module.db.session.flush()
                cls.created_user_ids.append(cls.superadmin.id)
            cls.superadmin.reimbursement_tracker_access = False

            def create_user(username, **flags):
                user = app_module.User(
                    username=f'{username}_{cls.suffix}',
                    password=app_module.generate_password_hash('test-password'),
                    role='staff',
                    is_active=True,
                    **flags,
                )
                app_module.db.session.add(user)
                app_module.db.session.flush()
                cls.created_user_ids.append(user.id)
                return user

            cls.tracker_user = create_user('tracker_manager', reimbursement_tracker_access=True)
            cls.personnel_user = create_user('tracker_personnel', personnel_admin_access=True)
            cls.plain_user = create_user('tracker_plain')

            def create_engineer(name, initials, branch):
                engineer = app_module.Engineer(
                    employee_id=f'RT-{cls.suffix}-{len(cls.created_engineer_ids) + 1}',
                    name=name,
                    initials=initials,
                    branch=branch,
                )
                app_module.db.session.add(engineer)
                app_module.db.session.flush()
                cls.created_engineer_ids.append(engineer.id)
                return engineer

            cls.jfl_engineer = create_engineer('Jim Frederick Lim', 'JFL', 'Manila')
            cls.raj_engineer = create_engineer('Rodito Aretano Jr.', 'RAJ', 'Cebu')
            cls.jp_one = create_engineer('Jonamar Paunil', 'JP', 'Manila')
            cls.jp_two = create_engineer('Jocel Prudente', 'JP', 'Davao')
            app_module.db.session.commit()

            cls.superadmin_id = cls.superadmin.id
            cls.tracker_user_id = cls.tracker_user.id
            cls.personnel_user_id = cls.personnel_user.id
            cls.plain_user_id = cls.plain_user.id
            cls.jfl_id = cls.jfl_engineer.id
            cls.raj_id = cls.raj_engineer.id
            cls.jp_one_id = cls.jp_one.id
            cls.jp_two_id = cls.jp_two.id

    @classmethod
    def tearDownClass(cls):
        with cls.app.app_context():
            for entry in app_module.ReimbursementTrackerEntry.query.all():
                app_module.db.session.delete(entry)
            for engineer_id in reversed(cls.created_engineer_ids):
                engineer = app_module.db.session.get(app_module.Engineer, engineer_id)
                if engineer:
                    app_module.db.session.delete(engineer)
            for user_id in reversed(cls.created_user_ids):
                user = app_module.db.session.get(app_module.User, user_id)
                if user:
                    app_module.db.session.delete(user)
            app_module.db.session.commit()
            app_module.db.session.remove()

    def setUp(self):
        with self.app.app_context():
            for entry in app_module.ReimbursementTrackerEntry.query.all():
                app_module.db.session.delete(entry)
            state = app_module.db.session.get(app_module.ReimbursementTrackerBatchState, 1)
            if state is None:
                state = app_module.ReimbursementTrackerBatchState(
                    id=1,
                    current_batch_sequence=app_module.REIMBURSEMENT_TRACKER_FIRST_BATCH_NUMBER,
                )
                app_module.db.session.add(state)
            else:
                state.current_batch_sequence = app_module.REIMBURSEMENT_TRACKER_FIRST_BATCH_NUMBER
            app_module.db.session.commit()

    @classmethod
    def _client_for(cls, user_id):
        client = cls.app.test_client()
        with client.session_transaction() as session:
            session['_user_id'] = str(user_id)
            session['_fresh'] = True
        return client

    @classmethod
    def _payload(cls, engineer_id=None, reference='BATCH-032', **overrides):
        payload = {
            'reference': reference,
            'expected_batch_sequence': 32,
            'submission_date': '2026-08-11',
            'engineer_id': engineer_id or cls.jfl_id,
            'office': 'Manila',
            'description': 'Tracker test row',
            'paid_in_full': False,
            'paid_transfer_date': '',
            'remarks': '',
        }
        for field, _label in app_module.REIMBURSEMENT_TRACKER_EXPENSE_FIELDS:
            payload[field] = '0'
        payload.update(overrides)
        return payload

    def _add(self, client, **overrides):
        response = client.post('/add_reimbursement_tracker_entry', json=self._payload(**overrides))
        self.assertEqual(response.status_code, 201, response.get_data(as_text=True))
        return response.get_json()['entry']

    def test_authorization_and_escalation_boundary(self):
        personnel = self._client_for(self.personnel_user_id)
        self.assertEqual(personnel.get('/reimbursement_tracker').status_code, 302)
        self.assertEqual(personnel.get('/get_reimbursement_tracker_entries').status_code, 403)
        self.assertEqual(personnel.get('/export_reimbursement_tracker').status_code, 403)
        self.assertEqual(
            personnel.post('/add_reimbursement_tracker_entry', json=self._payload()).status_code,
            403,
        )
        self.assertEqual(
            personnel.put('/update_reimbursement_tracker_entry/999999', json={}).status_code,
            403,
        )
        self.assertEqual(personnel.delete('/delete_reimbursement_tracker_entry/999999').status_code, 403)
        self.assertEqual(self._client_for(self.plain_user_id).get('/reimbursement_tracker').status_code, 302)

        superadmin = self._client_for(self.superadmin_id)
        created = personnel.post('/add_engineer', json={
            'name': f'Unauthorized Tracker Grant {self.suffix}',
            'employee_id': f'RT-ESC-{self.suffix}',
            'initials': 'RTE',
            'reimbursement_tracker_access': True,
        })
        self.assertEqual(created.status_code, 403)

        self.assertEqual(superadmin.get('/reimbursement_tracker').status_code, 200)
        self.assertEqual(superadmin.get('/get_reimbursement_tracker_entries').status_code, 200)
        self.assertEqual(superadmin.get('/export_reimbursement_tracker').status_code, 200)

        granted = superadmin.post('/settings/update-approval-user', json={
            'user_id': self.plain_user_id,
            'reimbursement_tracker_access': True,
            'is_active': True,
        })
        self.assertEqual(granted.status_code, 200, granted.get_data(as_text=True))
        self.assertTrue(granted.get_json()['user']['reimbursement_tracker_access'])
        with self.app.app_context():
            plain = app_module.db.session.get(app_module.User, self.plain_user_id)
            self.assertTrue(plain.reimbursement_tracker_access)
            plain.reimbursement_tracker_access = False
            app_module.db.session.commit()

    def test_page_and_endpoints_flip_together(self):
        owner = self._client_for(self.tracker_user_id)
        row = self._add(owner)
        with self.app.app_context():
            user = app_module.db.session.get(app_module.User, self.tracker_user_id)
            user.reimbursement_tracker_access = False
            app_module.db.session.commit()

        denied_client = self._client_for(self.tracker_user_id)
        self.assertEqual(denied_client.get('/reimbursement_tracker').status_code, 302)
        self.assertEqual(denied_client.get('/get_reimbursement_tracker_entries').status_code, 403)
        self.assertEqual(denied_client.get('/export_reimbursement_tracker').status_code, 403)
        self.assertEqual(denied_client.post('/add_reimbursement_tracker_entry', json=self._payload()).status_code, 403)
        self.assertEqual(denied_client.put(f"/update_reimbursement_tracker_entry/{row['id']}", json={}).status_code, 403)
        self.assertEqual(denied_client.delete(f"/delete_reimbursement_tracker_entry/{row['id']}").status_code, 403)

        with self.app.app_context():
            user = app_module.db.session.get(app_module.User, self.tracker_user_id)
            user.reimbursement_tracker_access = True
            app_module.db.session.commit()
        self.assertEqual(owner.get('/reimbursement_tracker').status_code, 200)
        self.assertEqual(owner.get('/get_reimbursement_tracker_entries').status_code, 200)

    def test_settings_reports_stored_grant_and_unrelated_save_does_not_grant(self):
        with self.app.app_context():
            superadmin = app_module.db.session.get(app_module.User, self.superadmin_id)
            superadmin.reimbursement_tracker_access = False
            self.assertTrue(app_module.can_manage_reimbursement_tracker(superadmin))
            self.assertFalse(app_module.approval_user_to_dict(superadmin)['reimbursement_tracker_access'])

        client = self._client_for(self.superadmin_id)
        rendered = client.get('/settings/approval-routing-data')
        self.assertEqual(rendered.status_code, 200)
        row = next(item for item in rendered.get_json()['users'] if item['id'] == self.superadmin_id)
        self.assertFalse(row['reimbursement_tracker_access'])
        saved = client.post('/settings/update-approval-user', json={
            'user_id': self.superadmin_id,
            'is_active': True,
            'reimbursement_tracker_access': row['reimbursement_tracker_access'],
        })
        self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))
        with self.app.app_context():
            self.assertFalse(app_module.db.session.get(app_module.User, self.superadmin_id).reimbursement_tracker_access)

    def test_writes_are_refused_without_the_capability(self):
        owner = self._client_for(self.tracker_user_id)
        row = self._add(owner)
        with self.app.app_context():
            user = app_module.db.session.get(app_module.User, self.plain_user_id)
            user.reimbursement_tracker_access = False
            before = app_module.ReimbursementTrackerEntry.query.count()
            app_module.db.session.commit()

        intruder = self._client_for(self.plain_user_id)
        self.assertEqual(intruder.post('/add_reimbursement_tracker_entry', json=self._payload()).status_code, 403)
        self.assertEqual(
            intruder.post('/start_reimbursement_tracker_batch', json={'expected_batch_sequence': 32}).status_code,
            403,
        )
        self.assertEqual(intruder.put(f"/update_reimbursement_tracker_entry/{row['id']}", json={}).status_code, 403)
        self.assertEqual(intruder.delete(f"/delete_reimbursement_tracker_entry/{row['id']}").status_code, 403)
        with self.app.app_context():
            self.assertEqual(app_module.ReimbursementTrackerEntry.query.count(), before)

    def test_control_number_uses_engineer_initials_from_the_table(self):
        client = self._client_for(self.tracker_user_id)
        jfl = self._add(client, engineer_id=self.jfl_id, office='Manila')
        raj = self._add(client, engineer_id=self.raj_id, office='Cebu')
        self.assertTrue(jfl['control_number'].startswith('JFL-20260811-032'))
        self.assertTrue(raj['control_number'].startswith('RAJ-20260811-032'))
        self.assertNotIn('RB-', raj['control_number'])
        self.assertNotIn('#N/A', jfl['control_number'])

    def test_control_number_sequence_comes_from_batch_reference(self):
        client = self._client_for(self.tracker_user_id)
        first = self._add(client, engineer_id=self.jfl_id, office='Manila')
        second = self._add(client, engineer_id=self.jfl_id, office='Manila', description='Same batch again')
        third = self._add(client, engineer_id=self.raj_id, office='Cebu')
        self.assertTrue(first['control_number'].endswith('-032'))
        self.assertEqual(first['control_number'], second['control_number'])
        self.assertTrue(third['control_number'].endswith('-032'))

    def test_control_number_is_stable_after_engineer_initials_change(self):
        client = self._client_for(self.tracker_user_id)
        row = self._add(client, engineer_id=self.jfl_id, office='Manila')
        with self.app.app_context():
            engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
            engineer.initials = 'CHANGED'
            app_module.db.session.commit()
        updated = client.put(f"/update_reimbursement_tracker_entry/{row['id']}", json={'remarks': 'kept stable'})
        self.assertEqual(updated.status_code, 200, updated.get_data(as_text=True))
        self.assertEqual(updated.get_json()['entry']['control_number'], row['control_number'])
        self.assertEqual(updated.get_json()['entry']['engineer_initials'], 'JFL')
        with self.app.app_context():
            engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
            engineer.initials = 'JFL'
            app_module.db.session.commit()

    def test_total_is_recomputed_server_side(self):
        client = self._client_for(self.tracker_user_id)
        row = self._add(
            client,
            representation='1.25',
            per_diem='2.75',
            total='999999.99',
            paid_in_full=True,
            paid_transfer_date='2026-08-11',
        )
        self.assertEqual(row['total'], '4.00')
        self.assertEqual(row['paid_amount'], '4.00')
        updated = client.put(f"/update_reimbursement_tracker_entry/{row['id']}", json={'paid_in_full': False})
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.get_json()['entry']['paid_amount'], '')

    def test_paid_transition_notifies_only_existing_rows_and_retoggle_resends(self):
        client = self._client_for(self.tracker_user_id)
        paid_email = f'tracker-paid-{self.suffix}@example.test'
        with self.app.app_context():
            engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
            original_email = engineer.email
            engineer.email = paid_email
            app_module.db.session.commit()

        try:
            with patch.object(app_module, 'send_reimbursement_tracker_paid_email_async') as sender:
                already_paid = self._add(
                    client,
                    engineer_id=self.jfl_id,
                    office='Manila',
                    paid_in_full=True,
                    paid_transfer_date='2026-08-12',
                )
                sender.assert_not_called()

                pending = self._add(client, engineer_id=self.jfl_id, office='Manila')
                first_paid = client.put(
                    f"/update_reimbursement_tracker_entry/{pending['id']}",
                    json={'paid_in_full': True, 'paid_transfer_date': '2026-08-12'},
                )
                self.assertEqual(first_paid.status_code, 200, first_paid.get_data(as_text=True))
                sender.assert_called_once_with(self.app, pending['id'])

                still_paid = client.put(
                    f"/update_reimbursement_tracker_entry/{pending['id']}",
                    json={'remarks': 'Receipt rechecked'},
                )
                self.assertEqual(still_paid.status_code, 200, still_paid.get_data(as_text=True))
                self.assertEqual(sender.call_count, 1)

                unticked = client.put(
                    f"/update_reimbursement_tracker_entry/{pending['id']}",
                    json={'paid_in_full': False},
                )
                self.assertEqual(unticked.status_code, 200, unticked.get_data(as_text=True))

                retoggled = client.put(
                    f"/update_reimbursement_tracker_entry/{pending['id']}",
                    json={'paid_in_full': True, 'paid_transfer_date': '2026-08-13'},
                )
                self.assertEqual(retoggled.status_code, 200, retoggled.get_data(as_text=True))
                self.assertEqual(sender.call_count, 2)
                self.assertEqual(already_paid['paid_in_full'], True)
        finally:
            with self.app.app_context():
                engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
                engineer.email = original_email
                app_module.db.session.commit()

    def test_paid_transition_warns_without_engineer_address_and_does_not_cc_only(self):
        client = self._client_for(self.tracker_user_id)
        with self.app.app_context():
            engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
            original_email = engineer.email
            engineer.email = ''
            app_module.db.session.commit()

        try:
            with patch.object(app_module, 'send_reimbursement_tracker_paid_email_async') as sender:
                row = self._add(client, engineer_id=self.jfl_id, office='Manila')
                response = client.put(
                    f"/update_reimbursement_tracker_entry/{row['id']}",
                    json={'paid_in_full': True, 'paid_transfer_date': '2026-08-12'},
                )
                self.assertEqual(response.status_code, 200, response.get_data(as_text=True))
                self.assertIn('no email address', response.get_json().get('warning', '').lower())
                sender.assert_not_called()
        finally:
            with self.app.app_context():
                engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
                engineer.email = original_email
                app_module.db.session.commit()

    def test_paid_email_formatter_and_cc_group_are_provider_independent(self):
        client = self._client_for(self.tracker_user_id)
        engineer_email = f'tracker-engineer-{self.suffix}@example.test'
        cc_email = f'tracker-cc-{self.suffix}@example.test'
        with self.app.app_context():
            engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
            original_email = engineer.email
            engineer.email = engineer_email
            app_module.ensure_email_recipient_setting_table()
            app_module.db.session.add(app_module.EmailRecipientSetting(
                group_key='reimbursement_tracker_paid_cc',
                email=cc_email,
                display_name='Tracker CC Test',
                is_active=True,
                sort_order=1,
            ))
            app_module.db.session.commit()

        try:
            row = self._add(
                client,
                engineer_id=self.jfl_id,
                office='Manila',
                representation='10.25',
            )
            with self.app.app_context():
                entry = app_module.db.session.get(app_module.ReimbursementTrackerEntry, row['id'])
                subject, text_body, html_body = app_module.format_reimbursement_tracker_paid_email(entry)
                self.assertIn(row['control_number'], subject)
                self.assertIn(row['reference'], text_body)
                self.assertIn(row['control_number'], html_body)
                self.assertIn('PHP 10.25', text_body)
                self.assertEqual(
                    app_module.get_active_email_recipients_by_group('reimbursement_tracker_paid_cc'),
                    [cc_email],
                )

            with patch.object(app_module, 'send_email_with_attachments', return_value=(True, 'test send')) as dispatcher:
                with patch.object(app_module.threading, 'Thread') as thread_class:
                    queued = app_module.send_reimbursement_tracker_paid_email_async(self.app, row['id'])
                    self.assertTrue(queued)
                    thread_class.assert_called_once()
                    thread_class.call_args.kwargs['target']()
                dispatcher.assert_called_once()
                send_args, send_kwargs = dispatcher.call_args
                self.assertEqual(send_args[0], [engineer_email])
                self.assertEqual(send_kwargs['cc_emails'], [cc_email])
        finally:
            with self.app.app_context():
                engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
                engineer.email = original_email
                recipient = app_module.EmailRecipientSetting.query.filter_by(email=cc_email).first()
                if recipient:
                    app_module.db.session.delete(recipient)
                app_module.db.session.commit()

    def test_list_returns_offices_engineers_suggestion_and_duplicate_initial_warning(self):
        client = self._client_for(self.tracker_user_id)
        response = client.get('/get_reimbursement_tracker_entries')
        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertIn('Manila', body['offices'])
        self.assertIn('Cebu', body['offices'])
        self.assertIn('JP', body['duplicate_initials'])
        self.assertEqual(body['current_batch'], {'sequence': 32, 'reference': 'BATCH-032'})
        self.assertTrue(body['suggested_reference'].startswith('BATCH-'))

    def test_export_reproduces_the_workbook_layout(self):
        client = self._client_for(self.tracker_user_id)
        self._add(client, engineer_id=self.jfl_id, office='Manila', representation='12.50')
        self._add(client, engineer_id=self.raj_id, office='Cebu')
        response = client.get('/export_reimbursement_tracker?sort=submission_date&direction=asc')
        self.assertEqual(response.status_code, 200, response.status)
        workbook = load_workbook(BytesIO(response.data), data_only=False)
        worksheet = workbook['Sheet']
        self.assertEqual(
            [worksheet.cell(1, column).value for column in range(1, 8)],
            ['Reference', 'Submission Date', 'Control #', 'Reimbursement', 'Engineer', 'Office', 'Total'],
        )
        self.assertEqual(worksheet.freeze_panes, 'A2')
        self.assertRegex(worksheet.auto_filter.ref, r'^A1:G\d+$')
        self.assertEqual(worksheet.max_column, 7)
        self.assertIsNone(worksheet['H1'].value)
        self.assertIsInstance(worksheet['G2'].value, (int, float, Decimal))
        self.assertFalse(worksheet.merged_cells.ranges)
        self.assertFalse(workbook.calculation.iterate)
        formulas = [
            cell.value for row in worksheet.iter_rows()
            for cell in row if isinstance(cell.value, str) and cell.value.startswith('=')
        ]
        self.assertEqual(formulas, [])
        exported_text = '\n'.join(
            str(cell.value) for row in worksheet.iter_rows() for cell in row if cell.value is not None
        )
        for removed_text in (
            'Current Input', 'by: Diary Dizon', 'Accounting', 'Reimbursements',
            'Summary', 'Trasnportation', 'Payment Status to Engineers',
        ):
            self.assertNotIn(removed_text, exported_text)

    def test_tracker_labels_keep_the_register_spelling_outside_the_export(self):
        template = (ROOT / 'templates' / 'reimbursement_tracker.html').read_text(encoding='utf-8')
        labels = [label for _field, label in app_module.REIMBURSEMENT_TRACKER_EXPENSE_FIELDS]
        self.assertIn('Trasnportation', labels)
        self.assertIn('Hotel Accomodation', labels)
        self.assertIn('Trasnportation', template)
        self.assertIn('Hotel Accomodation', template)

    def test_export_route_is_network_only_and_cache_version_is_bumped(self):
        client = self._client_for(self.tracker_user_id)
        worker = client.get('/service-worker.js')
        self.assertEqual(worker.status_code, 200)
        body = worker.get_data(as_text=True)
        self.assertIn("'/export_'", body)
        prefixes = body.split('const NETWORK_ONLY_DOWNLOAD_PREFIXES = [', 1)[1].split('];', 1)[0]
        self.assertTrue(any('/export_reimbursement_tracker'.startswith(prefix.strip().strip("',\"")) for prefix in prefixes.splitlines() if "'/" in prefix))
        version = body.split("const CACHE_VERSION = '", 1)[1].split("'", 1)[0]
        self.assertGreaterEqual(int(version.split('-v', 1)[1].split('-', 1)[0]), 86)

    def test_release_manifest_has_tracker_item(self):
        manifest = json.loads((ROOT / 'static' / 'changelog' / 'releases.json').read_text(encoding='utf-8'))
        items = [item for release in manifest['releases'] for item in release.get('items', [])]
        self.assertTrue(any(item['item_key'] == '2026-08-11-reimbursement-tracker' for item in items))

    def test_tracker_item_sits_under_its_own_release_not_the_backup_one(self):
        """The item first shipped inside the release titled 'Backup Center Behaves Offline'.

        The coverage test passed because only the date is checked, but What's New would
        have shown a Reimbursement Tracker entry under a backup headline.
        """
        manifest = json.loads((ROOT / 'static' / 'changelog' / 'releases.json').read_text(encoding='utf-8'))
        owning = [
            release for release in manifest['releases']
            if any(item['item_key'] == '2026-08-11-reimbursement-tracker'
                   for item in release.get('items', []))
        ]
        self.assertEqual(len(owning), 1, 'the tracker item must belong to exactly one release')
        release = owning[0]
        self.assertEqual(release['release_date'], '2026-08-11')
        self.assertIn('Reimbursement Tracker', release['title'])
        self.assertNotIn('Backup', release['title'])

    # The theme-token guard that used to live here was scoped to this one template while
    # its docstring claimed it asserted the class. It therefore could not catch the same
    # misspelling when it reappeared on P.O. Details. It is now repo-wide, covering this
    # page along with every other, in tests/test_appearance_themes.py:
    # AppearanceThemeSourceTests.test_every_theme_token_used_anywhere_is_actually_defined.

    def test_first_batch_migration_normalizes_rows_and_is_idempotent(self):
        """The first state row is an atomic marker for the live BATCH-032 conversion."""
        with self.app.app_context():
            for entry in app_module.ReimbursementTrackerEntry.query.all():
                app_module.db.session.delete(entry)
            state = app_module.db.session.get(app_module.ReimbursementTrackerBatchState, 1)
            if state:
                app_module.db.session.delete(state)
            app_module.db.session.commit()

            first = app_module.ReimbursementTrackerEntry(
                reference='BATCH-031',
                submission_date=date(2026, 8, 11),
                control_number='legacy-control-1',
                batch_sequence=31,
                description='Keep this description',
                engineer_id=self.jfl_id,
                engineer_name_snapshot='Jim Frederick Lim',
                engineer_initials_snapshot='JFL',
                office='Manila',
                total=Decimal('12.50'),
                representation=Decimal('12.50'),
                paid_in_full=True,
                paid_amount=Decimal('12.50'),
            )
            second = app_module.ReimbursementTrackerEntry(
                reference='BATCH-040',
                submission_date=date(2026, 8, 12),
                control_number='legacy-control-2',
                batch_sequence=40,
                description='Keep this second description',
                engineer_id=self.raj_id,
                engineer_name_snapshot='Rodito Aretano Jr.',
                engineer_initials_snapshot='RAJ',
                office='Cebu',
                total=Decimal('8.75'),
                toll=Decimal('8.75'),
            )
            app_module.db.session.add_all([first, second])
            app_module.db.session.commit()
            before = {
                first.id: (first.description, str(first.total), bool(first.paid_in_full), str(first.paid_amount)),
                second.id: (second.description, str(second.total), bool(second.paid_in_full), str(second.paid_amount)),
            }

            app_module._reimbursement_tracker_schema_ready = False
            app_module.ensure_reimbursement_tracker_schema()
            rows = app_module.ReimbursementTrackerEntry.query.order_by(app_module.ReimbursementTrackerEntry.id).all()
            self.assertEqual(len(rows), 2)
            self.assertEqual([row.reference for row in rows], ['BATCH-032', 'BATCH-032'])
            self.assertEqual([row.batch_sequence for row in rows], [32, 32])
            self.assertEqual(rows[0].control_number, 'JFL-20260811-032')
            self.assertEqual(rows[1].control_number, 'RAJ-20260812-032')
            self.assertEqual(
                {row.id: (row.description, str(row.total), bool(row.paid_in_full), str(row.paid_amount)) for row in rows},
                before,
            )
            self.assertEqual(
                app_module.db.session.get(app_module.ReimbursementTrackerBatchState, 1).current_batch_sequence,
                32,
            )

            snapshot = [(row.id, row.reference, row.control_number) for row in rows]
            app_module._reimbursement_tracker_schema_ready = False
            app_module.ensure_reimbursement_tracker_schema()
            self.assertEqual(
                [(row.id, row.reference, row.control_number)
                 for row in app_module.ReimbursementTrackerEntry.query.order_by(app_module.ReimbursementTrackerEntry.id).all()],
                snapshot,
            )

    def test_shared_batch_reuses_until_explicit_transition(self):
        client = self._client_for(self.tracker_user_id)
        first = self._add(client, engineer_id=self.jfl_id, office='Manila')
        second = self._add(client, engineer_id=self.raj_id, office='Cebu')
        self.assertEqual(first['reference'], 'BATCH-032')
        self.assertEqual(second['reference'], 'BATCH-032')
        self.assertTrue(first['control_number'].endswith('-032'))
        self.assertTrue(second['control_number'].endswith('-032'))

        stale_start = client.post(
            '/start_reimbursement_tracker_batch',
            json={'expected_batch_sequence': 31},
        )
        self.assertEqual(stale_start.status_code, 409)
        self.assertEqual(stale_start.get_json()['current_batch']['reference'], 'BATCH-032')

        started = client.post(
            '/start_reimbursement_tracker_batch',
            json={'expected_batch_sequence': 32},
        )
        self.assertEqual(started.status_code, 200, started.get_data(as_text=True))
        self.assertEqual(started.get_json()['current_batch'], {'sequence': 33, 'reference': 'BATCH-033'})

        stale_add = client.post(
            '/add_reimbursement_tracker_entry',
            json=self._payload(reference='BATCH-032', expected_batch_sequence=32),
        )
        self.assertEqual(stale_add.status_code, 409)
        self.assertEqual(client.get('/get_reimbursement_tracker_entries').get_json()['entries'].__len__(), 2)

        third = self._add(
            client,
            reference='BATCH-033',
            expected_batch_sequence=33,
            engineer_id=self.jfl_id,
            office='Manila',
        )
        self.assertEqual(third['reference'], 'BATCH-033')
        self.assertTrue(third['control_number'].endswith('-033'))

    def test_existing_row_keeps_batch_when_edit_tries_to_move_it(self):
        client = self._client_for(self.tracker_user_id)
        row = self._add(client, engineer_id=self.jfl_id, office='Manila')
        started = client.post('/start_reimbursement_tracker_batch', json={'expected_batch_sequence': 32})
        self.assertEqual(started.status_code, 200)
        edit = client.put(
            f"/update_reimbursement_tracker_entry/{row['id']}",
            json=self._payload(
                reference='BATCH-033',
                expected_batch_sequence=33,
                description='Edited without moving the historical batch',
            ),
        )
        self.assertEqual(edit.status_code, 200, edit.get_data(as_text=True))
        self.assertEqual(edit.get_json()['entry']['reference'], 'BATCH-032')
        self.assertEqual(edit.get_json()['entry']['batch_sequence'], 32)
        self.assertTrue(edit.get_json()['entry']['control_number'].endswith('-032'))

    def test_batch_999_cannot_advance(self):
        client = self._client_for(self.tracker_user_id)
        with self.app.app_context():
            state = app_module.db.session.get(app_module.ReimbursementTrackerBatchState, 1)
            state.current_batch_sequence = 999
            app_module.db.session.commit()
        response = client.post('/start_reimbursement_tracker_batch', json={'expected_batch_sequence': 999})
        self.assertEqual(response.status_code, 409)
        self.assertIn('BATCH-999', response.get_json()['error'])

    def test_the_form_actually_consumes_the_batch_suggestion(self):
        """The Add form consumes the server-owned current batch and exposes the transition."""
        template = (ROOT / 'templates' / 'reimbursement_tracker.html').read_text(encoding='utf-8')
        self.assertIn('current_batch', template,
                      'the page must consume the server-owned current batch')
        self.assertIn('data.suggested_reference', template,
                      'the compatibility alias must remain readable by the page')
        self.assertRegex(
            template,
            r"byId\('rt-reference'\)\.value\s*=\s*state\.currentBatch\?\.reference",
            'the Add form must prefill the reference from the current batch',
        )
        self.assertIn('id="rt-start-batch-button"', template)
        self.assertIn("/start_reimbursement_tracker_batch", template)
        self.assertIn('readonly', template)
        self.assertNotRegex(
            template, r'placeholder="BATCH-001"',
            'a hard-coded BATCH-001 placeholder would invite a restart at the wrong batch',
        )

    def test_modal_body_can_scroll_independently_of_the_form_wrapper(self):
        """The form sits between .modal-content and .modal-body, which broke Bootstrap.

        .modal-dialog-scrollable caps .modal-body only when it is a direct flex child of
        .modal-content. With the <form> in between, the body grew to full height, the
        content clipped it, and everything past Others/Misc -- including Save -- was
        unreachable with no way to scroll to it.
        """
        template = (ROOT / 'templates' / 'reimbursement_tracker.html').read_text(encoding='utf-8')
        self.assertIn('modal-dialog-scrollable', template)
        # The form must carry the flex column that .modal-content would otherwise provide.
        self.assertRegex(
            template,
            r'\.rt-modal-content\s*>\s*form\s*\{[^}]*display:\s*flex[^}]*flex-direction:\s*column',
        )
        self.assertRegex(
            template,
            r'\.rt-modal-content\s*>\s*form\s*>\s*\.modal-body\s*\{[^}]*overflow-y:\s*auto',
        )

    def test_amount_suggestion_chips_keep_their_three_load_bearing_guards(self):
        """The chips shipped with no regression guard at all.

        Their behaviour was proved in a browser, but nothing stops a later edit removing
        one of the three properties that make them safe. There is no JS runner here, so
        this is the documented inline-template exception -- each assertion targets an
        outcome a user would feel, not a formatting choice:

        1. type="button" -- a bare <button> inside <form> submits it, so tapping a
           suggestion would save the row instead of filling a field.
        2. the editing row is excluded -- otherwise editing a row offers its own amount
           back as a "recent" value, which reads as corroboration of a figure being changed.
        3. zero and negative values are dropped -- most categories are 0 in any given row,
           so without the filter all ten fields grow a row of PHP 0.00 chips.
        """
        template = (ROOT / 'templates' / 'reimbursement_tracker.html').read_text(encoding='utf-8')

        chip_markup = [line for line in template.splitlines() if 'rt-suggest-chip' in line and '<button' in line]
        self.assertTrue(chip_markup, 'no suggestion chip markup found')
        for line in chip_markup:
            self.assertIn('type="button"', line,
                          'a suggestion chip without type="button" submits the form')

        self.assertRegex(
            template,
            r'recentRowsForEngineer\(\s*engineerId\s*,\s*state\.editingId\s*\)',
            'the renderer must exclude the row being edited',
        )
        self.assertRegex(
            template,
            r'if\s*\(\s*!numeric\s*\|\|\s*numeric\s*<=\s*0\s*\)',
            'zero and negative amounts must not become chips',
        )

    def test_control_date_format_constant_actually_drives_the_control_number(self):
        """The constant shipped unreferenced, so editing it changed nothing.

        It exists so the ambiguous unpadded day (Jan 12 and Nov 2 both render 2026112)
        is a one-line change. Swap in a padded format and the built number must follow.
        """
        self.assertEqual(
            app_module.reimbursement_tracker_control_number('JFL', date(2026, 8, 7), 31),
            'JFL-2026087-031',
        )
        original = app_module.REIMBURSEMENT_TRACKER_CONTROL_DATE_FORMAT
        try:
            app_module.REIMBURSEMENT_TRACKER_CONTROL_DATE_FORMAT = '{d.year}{d.month:02d}{d.day:02d}'
            self.assertEqual(
                app_module.reimbursement_tracker_control_number('JFL', date(2026, 8, 7), 31),
                'JFL-20260807-031',
                'the constant is not wired into the builder',
            )
        finally:
            app_module.REIMBURSEMENT_TRACKER_CONTROL_DATE_FORMAT = original

    def test_office_must_match_the_engineers_branch(self):
        """The workbook made this impossible: its engineer list was an INDIRECT on office.

        The page reproduces that, but the endpoint accepted any string at all -- so a
        Davao engineer could be filed under Manila, or under an office that exists nowhere.
        """
        client = self._client_for(self.tracker_user_id)
        before = client.get('/get_reimbursement_tracker_entries').get_json()['entries']

        mismatch = client.post('/add_reimbursement_tracker_entry',
                               json=self._payload(engineer_id=self.raj_id, office='Manila'))
        self.assertEqual(mismatch.status_code, 400)
        self.assertIn('Cebu', mismatch.get_json()['error'])

        unknown = client.post('/add_reimbursement_tracker_entry',
                              json=self._payload(engineer_id=self.jfl_id, office='Atlantis'))
        self.assertEqual(unknown.status_code, 400)

        # Positive control: the matching office still saves, so the rule refuses the
        # mismatch rather than refusing everything.
        ok = client.post('/add_reimbursement_tracker_entry',
                         json=self._payload(engineer_id=self.raj_id, office='Cebu'))
        self.assertEqual(ok.status_code, 201, ok.get_data(as_text=True))
        self.assertEqual(ok.get_json()['entry']['office'], 'Cebu')

        after = client.get('/get_reimbursement_tracker_entries').get_json()['entries']
        self.assertEqual(len(after), len(before) + 1, 'only the valid row may be stored')

    def test_a_transferred_engineer_does_not_make_old_rows_unsaveable(self):
        """Office is a snapshot of where the row was filed, deliberately denormalised.

        So the branch check must not fire when an existing row's stored office is
        resubmitted unchanged -- otherwise a transfer would freeze every historical row.
        """
        client = self._client_for(self.tracker_user_id)
        row = self._add(client, engineer_id=self.jfl_id, office='Manila')
        with self.app.app_context():
            engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
            engineer.branch = 'Davao'
            app_module.db.session.commit()
        try:
            edit = client.post(f'/update_reimbursement_tracker_entry/{row["id"]}',
                               json=self._payload(engineer_id=self.jfl_id, office='Manila',
                                                  description='Edited after the transfer'))
            self.assertEqual(edit.status_code, 200, edit.get_data(as_text=True))
            self.assertEqual(edit.get_json()['entry']['office'], 'Manila')
        finally:
            with self.app.app_context():
                engineer = app_module.db.session.get(app_module.Engineer, self.jfl_id)
                engineer.branch = 'Manila'
                app_module.db.session.commit()


if __name__ == '__main__':
    unittest.main()
