"""Behavioural coverage for the P.O. Details register and capability boundary."""

import pathlib
import unittest
import uuid
from datetime import date, timedelta
from io import BytesIO
from urllib.parse import quote

import app as app_module  # noqa: E402
from openpyxl import load_workbook
from sqlalchemy import event
from sqlalchemy.exc import IntegrityError


ROOT = pathlib.Path(__file__).resolve().parents[1]


class ServiceContractAllocationHelperTests(unittest.TestCase):
    """Pure schedule/fiscal controls for the Service Contract P.O. register."""

    def test_supported_and_legacy_frequency_boundaries(self):
        self.assertEqual(
            app_module.PO_CURRENT_TYPES,
            ('single', 'semi_annual', 'quarterly'),
        )
        self.assertEqual(app_module.normalize_purchase_order_type('Contract'), 'contract')
        self.assertNotIn('contract', app_module.PO_CURRENT_TYPES)

    def test_single_schedule_allocates_remainder_on_final_occurrence(self):
        schedule = app_module.purchase_order_schedule(
            date(2026, 8, 25), date(2028, 8, 25), 'single', '1000000'
        )
        self.assertEqual(
            [item['date_string'] for item in schedule],
            ['2026-08-25', '2027-08-25', '2028-08-25'],
        )
        self.assertEqual(
            [item['amount_string'] for item in schedule],
            ['333333.33', '333333.33', '333333.34'],
        )

    def test_fiscal_quarters_follow_april_start_and_january_rollover(self):
        self.assertEqual(
            app_module.purchase_order_fiscal_period(2026, 1),
            (date(2026, 4, 1), date(2026, 6, 30)),
        )
        self.assertEqual(
            app_module.purchase_order_fiscal_period(2026, 4),
            (date(2027, 1, 1), date(2027, 3, 31)),
        )

    def test_fiscal_quarter_requires_a_fiscal_year(self):
        with self.assertRaisesRegex(ValueError, 'Fiscal year and fiscal quarter'):
            app_module.purchase_order_filter_window({'fiscal_quarter': '2'})

    def test_month_end_clamps_when_anchored_schedule_crosses_short_month(self):
        schedule = app_module.purchase_order_schedule(
            date(2024, 2, 29), date(2025, 2, 28), 'single', '2'
        )
        self.assertEqual(
            [item['date_string'] for item in schedule],
            ['2024-02-29', '2025-02-28'],
        )

    def test_small_total_never_creates_negative_allocation(self):
        schedule = app_module.purchase_order_schedule(
            date(2026, 4, 1), date(2027, 1, 2), 'quarterly', '0.02'
        )
        self.assertEqual([item['amount_string'] for item in schedule], [
            '0.00', '0.00', '0.00', '0.02',
        ])
        self.assertEqual(sum(item['amount'] for item in schedule), app_module.Decimal('0.02'))

    def test_quarterly_end_anniversary_is_coverage_expiry_not_an_extra_visit(self):
        schedule = app_module.purchase_order_schedule(
            date(2026, 8, 1), date(2027, 8, 1), 'quarterly', '100000'
        )
        self.assertEqual(
            [item['date_string'] for item in schedule],
            ['2026-08-01', '2026-11-01', '2027-02-01', '2027-05-01'],
        )
        self.assertEqual(
            [item['amount_string'] for item in schedule],
            ['25000.00', '25000.00', '25000.00', '25000.00'],
        )

    def test_template_has_new_heading_filters_and_status_confirmation_contract(self):
        template = (ROOT / 'templates' / 'po_details.html').read_text(encoding='utf-8')
        layout = (ROOT / 'templates' / 'layout.html').read_text(encoding='utf-8')
        self.assertIn('Service Contract P.O. Details', template)
        self.assertIn("nav_link('/po_details', 'fa-file-invoice', 'P.O. Details')", layout)
        self.assertIn('1st Fiscal (Apr–Jun)', template)
        self.assertIn('id="po-fiscal-quarter" class="form-select" disabled', template)
        self.assertIn('id="po-semi-annual-total"', template)
        self.assertIn('async function addMachineSelection(product)', template)
        self.assertIn('Pending Under Contract', template)
        self.assertIn('Single is annual for contract-backed machines and one-time for machines marked No Expiry Set - No Contract.', template)
        self.assertIn('Select a frequency first', template)
        self.assertIn('schedule_mode_label', template)
        self.assertIn('This is a legacy P.O. type. Select Single, Semi Annual, or Quarterly before saving.', template)
        self.assertIn('under_contract_confirmation_required', template)
        self.assertIn('Computed Amount', template)


class ServiceContractPurchaseOrderEndpointTests(unittest.TestCase):
    """Endpoint controls use the repository's isolated test database only."""

    @classmethod
    def setUpClass(cls):
        cls.app = app_module.app
        cls.app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        cls.suffix = uuid.uuid4().hex[:8]
        with cls.app.app_context():
            app_module.db.create_all()
            app_module.ensure_user_admin_capability_columns()
            app_module.ensure_purchase_order_schema()
            cls.user = app_module.User(
                username=f'po_service_{cls.suffix}',
                password=app_module.generate_password_hash('test-password'),
                role='staff', is_active=True, po_admin_access=True,
            )
            cls.client_record = app_module.Client(
                name=f'Service Contract Client {cls.suffix}', address='Service address',
            )
            app_module.db.session.add_all([cls.user, cls.client_record])
            app_module.db.session.flush()
            cls.product = app_module.Product(
                serial_number=f'PO-SERVICE-{cls.suffix}', name='Service Machine',
                client_id=cls.client_record.id,
                start_warranty_date=date(2026, 8, 25),
                end_warranty_date=date(2028, 8, 25),
                under_contract=False,
            )
            app_module.db.session.add(cls.product)
            app_module.db.session.commit()
            cls.user_id = cls.user.id
            cls.client_id = cls.client_record.id
            cls.serial = cls.product.serial_number

    @classmethod
    def tearDownClass(cls):
        with cls.app.app_context():
            for po in app_module.PurchaseOrder.query.filter_by(client_id=cls.client_id).all():
                app_module.db.session.delete(po)
            product = app_module.db.session.get(app_module.Product, cls.serial)
            if product:
                app_module.db.session.delete(product)
            client = app_module.db.session.get(app_module.Client, cls.client_id)
            if client:
                app_module.db.session.delete(client)
            user = app_module.db.session.get(app_module.User, cls.user_id)
            if user:
                app_module.db.session.delete(user)
            app_module.db.session.commit()

    def setUp(self):
        self.client = self.app.test_client()
        response = self.client.post('/login', data={
            'username': f'po_service_{self.suffix}',
            'password': 'test-password',
        })
        self.assertIn(response.status_code, (302, 303))

    def test_status_confirmation_dates_and_schedule_are_server_owned(self):
        with self.app.app_context():
            product = app_module.db.session.get(app_module.Product, self.serial)
            product.under_contract = False
            product.start_warranty_date = date(2026, 8, 25)
            product.end_warranty_date = date(2028, 8, 25)
            app_module.db.session.commit()
        payload = {
            'client_id': self.client_id,
            'product_serials': [self.serial],
            'start_date': '1900-01-01',
            'end_date': '1900-01-02',
            'po_number': f'SERVICE-{self.suffix}',
            'po_type': 'single',
            'amount': '1000000',
        }
        pending = self.client.post('/add_purchase_order', json=payload)
        self.assertEqual(pending.status_code, 409)
        self.assertEqual(pending.get_json()['code'], 'under_contract_confirmation_required')
        payload['under_contract_confirmed_serials'] = [self.serial]
        created = self.client.post('/add_purchase_order', json=payload)
        self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
        body = created.get_json()['purchase_order']
        self.assertEqual(body['start_date'], '2026-08-25')
        self.assertEqual(body['end_date'], '2028-08-25')
        self.assertEqual(body['computed_amount'], '333333.33')
        self.assertEqual([item['amount_string'] for item in body['allocation_schedule']], [
            '333333.33', '333333.33', '333333.34',
        ])
        with self.app.app_context():
            self.assertTrue(app_module.db.session.get(app_module.Product, self.serial).under_contract)

    def test_same_machine_edit_preserves_saved_snapshot(self):
        with self.app.app_context():
            product = app_module.db.session.get(app_module.Product, self.serial)
            product.under_contract = True
            app_module.db.session.commit()
        created = self.client.post('/add_purchase_order', json={
            'client_id': self.client_id, 'product_serials': [self.serial],
            'po_number': f'SNAPSHOT-{self.suffix}', 'po_type': 'quarterly', 'amount': '1200',
        })
        self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
        created_body = created.get_json()['purchase_order']
        self.assertEqual(created_body['computed_amount'], '150.00')
        self.assertEqual(len(created_body['allocation_schedule']), 8)
        record_id = created.get_json()['purchase_order']['id']
        with self.app.app_context():
            product = app_module.db.session.get(app_module.Product, self.serial)
            product.start_warranty_date = date(2030, 1, 1)
            product.end_warranty_date = date(2031, 1, 1)
            app_module.db.session.commit()
        updated = self.client.put(f'/update_purchase_order/{record_id}', json={
            'client_id': self.client_id, 'product_serials': [self.serial],
            'po_number': f'SNAPSHOT-EDITED-{self.suffix}', 'po_type': 'quarterly', 'amount': '1200',
        })
        self.assertEqual(updated.status_code, 200, updated.get_data(as_text=True))
        self.assertEqual(updated.get_json()['purchase_order']['start_date'], '2026-08-25')
        self.assertEqual(updated.get_json()['purchase_order']['end_date'], '2028-08-25')

    def test_one_time_single_saves_without_contract_confirmation_or_product_mutation(self):
        serial = f'PO-ONE-TIME-{self.suffix}'
        with self.app.app_context():
            product = app_module.Product(
                serial_number=serial, name='One-time Service Machine',
                client_id=self.client_id, start_warranty_date=date(2026, 8, 25),
                end_warranty_date=None, under_contract=False,
            )
            app_module.db.session.add(product)
            app_module.db.session.commit()
        try:
            created = self.client.post('/add_purchase_order', json={
                'client_id': self.client_id, 'product_serials': [serial],
                'po_number': f'ONE-TIME-{self.suffix}', 'po_type': 'single',
                'amount': '1000000',
            })
            self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
            body = created.get_json()['purchase_order']
            self.assertEqual(body['start_date'], '2026-08-25')
            self.assertEqual(body['end_date'], '')
            self.assertEqual(body['schedule_mode'], 'one_time')
            self.assertEqual(body['schedule_mode_label'], 'One-time')
            self.assertEqual([item['date_string'] for item in body['allocation_schedule']], ['2026-08-25'])
            self.assertEqual([item['amount_string'] for item in body['allocation_schedule']], ['1000000.00'])
            self.assertEqual(body['computed_amount'], '1000000.00')
            with self.app.app_context():
                saved = app_module.db.session.get(app_module.Product, serial)
                self.assertFalse(saved.under_contract)
                listing = self.client.get('/get_purchase_orders').get_json()
                product_payload = next(item for item in listing['products'] if item['serial_number'] == serial)
                self.assertEqual(product_payload['contract_status'], 'No Expiry Set - No Contract')
                self.assertTrue(product_payload['one_time_eligible'])

            filtered = self.client.get(
                f'/export_purchase_orders?number={quote("ONE-TIME-")}&from=2026-08-25&to=2026-08-25'
            )
            self.assertEqual(filtered.status_code, 200)
            worksheet = load_workbook(BytesIO(filtered.data), data_only=False)['P.O. Register']
            self.assertEqual(worksheet['B2'].value.date(), date(2026, 8, 25))
            self.assertIsNone(worksheet['C2'].value)
            self.assertEqual(worksheet['K2'].value, 1000000)
            self.assertEqual(worksheet['L2'].value, '2026-08-25')
        finally:
            with self.app.app_context():
                for po in app_module.PurchaseOrder.query.join(app_module.PurchaseOrderMachine).filter(
                    app_module.PurchaseOrderMachine.product_serial == serial
                ).all():
                    app_module.db.session.delete(po)
                product = app_module.db.session.get(app_module.Product, serial)
                if product:
                    app_module.db.session.delete(product)
                app_module.db.session.commit()

    def test_one_time_single_requires_start_and_rejects_ranged_frequencies(self):
        serial = f'PO-ONE-TIME-MISSING-START-{self.suffix}'
        ranged_serial = f'PO-NO-END-RANGED-{self.suffix}'
        with self.app.app_context():
            app_module.db.session.add_all([
                app_module.Product(
                    serial_number=serial, name='Missing Start', client_id=self.client_id,
                    start_warranty_date=None, end_warranty_date=None, under_contract=False,
                ),
                app_module.Product(
                    serial_number=ranged_serial, name='No End Ranged', client_id=self.client_id,
                    start_warranty_date=date(2026, 8, 25), end_warranty_date=None,
                    under_contract=False,
                ),
            ])
            app_module.db.session.commit()
        try:
            missing_start = self.client.post('/add_purchase_order', json={
                'client_id': self.client_id, 'product_serial': serial,
                'po_number': f'ONE-TIME-MISSING-{self.suffix}', 'po_type': 'single',
            })
            self.assertEqual(missing_start.status_code, 400)
            self.assertIn('Start Date', missing_start.get_json()['error'])
            for po_type in ('semi_annual', 'quarterly'):
                rejected = self.client.post('/add_purchase_order', json={
                    'client_id': self.client_id, 'product_serial': ranged_serial,
                    'po_number': f'NO-END-{po_type}-{self.suffix}', 'po_type': po_type,
                })
                self.assertEqual(rejected.status_code, 400)
                self.assertIn('Product End Date', rejected.get_json()['error'])
        finally:
            with self.app.app_context():
                for serial_value in (serial, ranged_serial):
                    for po in app_module.PurchaseOrder.query.join(app_module.PurchaseOrderMachine).filter(
                        app_module.PurchaseOrderMachine.product_serial == serial_value
                    ).all():
                        app_module.db.session.delete(po)
                    product = app_module.db.session.get(app_module.Product, serial_value)
                    if product:
                        app_module.db.session.delete(product)
                app_module.db.session.commit()

    def test_one_time_machines_must_match_and_preserve_blank_end_snapshot(self):
        serials = [f'PO-ONE-TIME-MATCH-{self.suffix}-{index}' for index in (1, 2)]
        with self.app.app_context():
            app_module.db.session.add_all([
                app_module.Product(
                    serial_number=serials[0], name='One-time Match One', client_id=self.client_id,
                    start_warranty_date=date(2026, 8, 25), end_warranty_date=None, under_contract=False,
                ),
                app_module.Product(
                    serial_number=serials[1], name='One-time Match Two', client_id=self.client_id,
                    start_warranty_date=date(2026, 8, 26), end_warranty_date=None, under_contract=False,
                ),
            ])
            app_module.db.session.commit()
        try:
            mismatch = self.client.post('/add_purchase_order', json={
                'client_id': self.client_id, 'product_serials': serials,
                'po_number': f'ONE-TIME-MISMATCH-{self.suffix}', 'po_type': 'single',
            })
            self.assertEqual(mismatch.status_code, 400)
            self.assertIn('matching Product Start Dates', mismatch.get_json()['error'])

            created = self.client.post('/add_purchase_order', json={
                'client_id': self.client_id, 'product_serial': serials[0],
                'po_number': f'ONE-TIME-SNAPSHOT-{self.suffix}', 'po_type': 'single',
                'amount': '25.00',
            })
            self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
            record_id = created.get_json()['purchase_order']['id']
            with self.app.app_context():
                product = app_module.db.session.get(app_module.Product, serials[0])
                product.start_warranty_date = None
                product.end_warranty_date = date(2030, 1, 1)
                app_module.db.session.commit()
            updated = self.client.put(f'/update_purchase_order/{record_id}', json={
                'client_id': self.client_id, 'product_serial': serials[0],
                'po_number': f'ONE-TIME-SNAPSHOT-EDITED-{self.suffix}', 'po_type': 'single',
                'amount': '25.00',
            })
            self.assertEqual(updated.status_code, 200, updated.get_data(as_text=True))
            self.assertEqual(updated.get_json()['purchase_order']['start_date'], '2026-08-25')
            self.assertEqual(updated.get_json()['purchase_order']['end_date'], '')
        finally:
            with self.app.app_context():
                for serial_value in serials:
                    for po in app_module.PurchaseOrder.query.join(app_module.PurchaseOrderMachine).filter(
                        app_module.PurchaseOrderMachine.product_serial == serial_value
                    ).all():
                        app_module.db.session.delete(po)
                    product = app_module.db.session.get(app_module.Product, serial_value)
                    if product:
                        app_module.db.session.delete(product)
                app_module.db.session.commit()

    def test_one_time_and_contract_backed_machines_cannot_be_mixed(self):
        one_time_serial = f'PO-ONE-TIME-MIX-{self.suffix}'
        contract_serial = f'PO-CONTRACT-MIX-{self.suffix}'
        with self.app.app_context():
            app_module.db.session.add_all([
                app_module.Product(
                    serial_number=one_time_serial, name='One-time Mix Machine', client_id=self.client_id,
                    start_warranty_date=date(2026, 8, 25), end_warranty_date=None, under_contract=False,
                ),
                app_module.Product(
                    serial_number=contract_serial, name='Contract Mix Machine', client_id=self.client_id,
                    start_warranty_date=date(2026, 8, 25), end_warranty_date=date(2028, 8, 25), under_contract=True,
                ),
            ])
            app_module.db.session.commit()
        try:
            response = self.client.post('/add_purchase_order', json={
                'client_id': self.client_id, 'product_serials': [one_time_serial, contract_serial],
                'po_number': f'ONE-TIME-MIX-{self.suffix}', 'po_type': 'single',
            })
            self.assertEqual(response.status_code, 400)
            self.assertIn('cannot be mixed', response.get_json()['error'])
        finally:
            with self.app.app_context():
                for serial_value in (one_time_serial, contract_serial):
                    product = app_module.db.session.get(app_module.Product, serial_value)
                    if product:
                        app_module.db.session.delete(product)
                app_module.db.session.commit()


class PurchaseOrderWorkflowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = app_module.app
        cls.app.config['TESTING'] = True
        cls.app.config['WTF_CSRF_ENABLED'] = False
        cls.suffix = uuid.uuid4().hex[:8]
        cls.created_user_ids = []
        cls.created_client_ids = []
        cls.created_product_serials = []
        cls.created_engineer_ids = []

        with cls.app.app_context():
            app_module.db.create_all()
            app_module.ensure_user_admin_capability_columns()
            app_module.ensure_purchase_order_schema()

            cls.superadmin = app_module.User.query.filter_by(username='jonamar').first()
            if not cls.superadmin:
                cls.superadmin = app_module.User(
                    username='jonamar',
                    password=app_module.generate_password_hash('test-password'),
                    role='superadmin',
                    is_active=True,
                )
                app_module.db.session.add(cls.superadmin)
                app_module.db.session.flush()
                cls.created_user_ids.append(cls.superadmin.id)

            def create_user(username, **flags):
                user = app_module.User(
                    username=f'{username}_{cls.suffix}',
                    password=app_module.generate_password_hash('test-password'),
                    role='staff',
                    is_active=True,
                    **flags,
                )
                app_module.db.session.add(user)
                app_module.db.session.flush()
                cls.created_user_ids.append(user.id)
                return user

            cls.po_user = create_user('po_manager', po_admin_access=True)
            cls.personnel_user = create_user('po_personnel', personnel_admin_access=True)
            cls.plain_user = create_user('po_plain')

            cls.client_one = app_module.Client(
                name=f'P.O. Test Medical Center One {cls.suffix}',
                address='Test address one',
            )
            cls.client_two = app_module.Client(
                name=f'P.O. Test Medical Center Two {cls.suffix}',
                address='Test address two',
            )
            app_module.db.session.add_all([cls.client_one, cls.client_two])
            app_module.db.session.flush()
            cls.created_client_ids.extend([cls.client_one.id, cls.client_two.id])
            cls.product_one = app_module.Product(
                serial_number=f'PO-SN-ONE-{cls.suffix}',
                name=f'CT One {cls.suffix}',
                client_id=cls.client_one.id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            )
            cls.product_two = app_module.Product(
                serial_number=f'PO-SN-TWO-{cls.suffix}',
                name=f'CT Two {cls.suffix}',
                client_id=cls.client_two.id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            )
            app_module.db.session.add_all([cls.product_one, cls.product_two])
            app_module.db.session.commit()

            cls.superadmin_id = cls.superadmin.id
            cls.po_user_id = cls.po_user.id
            cls.personnel_user_id = cls.personnel_user.id
            cls.plain_user_id = cls.plain_user.id
            cls.client_one_id = cls.client_one.id
            cls.client_two_id = cls.client_two.id
            cls.product_one_serial = cls.product_one.serial_number
            cls.product_two_serial = cls.product_two.serial_number
            cls.created_product_serials.extend([
                cls.product_one_serial,
                cls.product_two_serial,
            ])

    @classmethod
    def tearDownClass(cls):
        with cls.app.app_context():
            for purchase_order in app_module.PurchaseOrder.query.all():
                if purchase_order.client_id in cls.created_client_ids:
                    app_module.db.session.delete(purchase_order)
            for serial_number in reversed(cls.created_product_serials):
                product = app_module.db.session.get(app_module.Product, serial_number)
                if product:
                    app_module.db.session.delete(product)
            for client_id in reversed(cls.created_client_ids):
                client = app_module.db.session.get(app_module.Client, client_id)
                if client:
                    app_module.db.session.delete(client)
            for user_id in reversed(cls.created_user_ids):
                user = app_module.db.session.get(app_module.User, user_id)
                if user:
                    app_module.db.session.delete(user)
            app_module.db.session.commit()
            app_module.db.session.remove()

    @classmethod
    def _client_for(cls, user_id):
        client = cls.app.test_client()
        with client.session_transaction() as session:
            session['_user_id'] = str(user_id)
            session['_fresh'] = True
        return client

    def test_authorization_and_escalation_boundary(self):
        personnel = self._client_for(self.personnel_user_id)
        self.assertEqual(personnel.get('/po_details').status_code, 302)
        self.assertEqual(personnel.get('/get_purchase_orders').status_code, 403)
        self.assertEqual(
            personnel.post('/add_engineer', json={
                'name': f'Unauthorized P.O. Grant {self.suffix}',
                'employee_id': f'PO-ESC-{self.suffix}',
                'initials': 'PEG',
                'po_admin_access': True,
            }).status_code,
            403,
        )

        plain = self._client_for(self.plain_user_id)
        self.assertEqual(plain.get('/po_details').status_code, 302)
        self.assertEqual(plain.get('/get_purchase_orders').status_code, 403)

        superadmin = self._client_for(self.superadmin_id)
        response = superadmin.post('/settings/update-approval-user', json={
            'user_id': self.plain_user_id,
            'po_admin_access': True,
            'is_active': True,
        })
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()['user']['po_admin_access'])
        with self.app.app_context():
            self.assertTrue(app_module.db.session.get(app_module.User, self.plain_user_id).po_admin_access)
            app_module.db.session.get(app_module.User, self.plain_user_id).po_admin_access = False
            app_module.db.session.commit()

        created = superadmin.post('/add_engineer', json={
            'name': f'Purchase Grant {self.suffix}',
            'employee_id': f'PO-GRANT-{self.suffix}',
            'initials': 'PGR',
            'po_admin_access': True,
        })
        self.assertEqual(created.status_code, 200)
        username = created.get_json()['username']
        with self.app.app_context():
            granted = app_module.User.query.filter_by(username=username).first()
            self.assertIsNotNone(granted)
            self.assertTrue(granted.po_admin_access)
            self.created_user_ids.append(granted.id)
            engineer = app_module.Engineer.query.filter_by(user_id=granted.id).first()
            if engineer:
                self.created_engineer_ids.append(engineer.id)

    def test_register_update_soft_duplicate_and_delete(self):
        client = self._client_for(self.po_user_id)
        self.assertEqual(client.get('/po_details').status_code, 200)

        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': self.product_one_serial,
            'po_date': '2026-08-06',
            'end_date': '2026-12-31',
            'po_number': ' PO-001  ',
            'po_type': 'quarterly',
            'amount': '1250.50',
        })
        self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
        record = created.get_json()['purchase_order']
        self.assertEqual(record['po_number'], 'PO-001')
        self.assertEqual(record['po_type'], 'quarterly')
        self.assertEqual(record['start_date'], '2026-08-01')
        self.assertEqual(record['end_date'], '2027-08-01')
        self.assertEqual(record['amount'], '1250.50')

        duplicate = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': self.product_one_serial,
            'po_date': '2026-08-07',
            'po_number': 'po-001',
            'po_type': 'single',
        })
        self.assertEqual(duplicate.status_code, 409)
        self.assertIn('duplicate', duplicate.get_json())

        forced = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': self.product_one_serial,
            'po_date': '2026-08-07',
            'po_number': 'po-001',
            'po_type': 'single',
            'force': True,
        })
        self.assertEqual(forced.status_code, 201)

        updated = client.put(f"/update_purchase_order/{record['id']}", json={
            'client_id': self.client_one_id,
            'product_serial': self.product_one_serial,
            'po_date': '2026-08-08',
            'po_number': 'PO-001-UPDATED',
            'po_type': 'semi_annual',
        })
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.get_json()['purchase_order']['po_type'], 'semi_annual')

        listing = client.get('/get_purchase_orders')
        self.assertEqual(listing.status_code, 200)
        self.assertGreaterEqual(len(listing.get_json()['purchase_orders']), 2)

        self.assertEqual(client.delete(f"/delete_purchase_order/{record['id']}").status_code, 200)

    def test_summary_cards_include_all_current_frequency_amount_totals(self):
        client = self._client_for(self.po_user_id)
        response = client.get('/po_details')
        self.assertEqual(response.status_code, 200)
        page = response.get_data(as_text=True)
        self.assertIn('id="po-semi-annual-total"', page)
        self.assertIn('id="po-single-total"', page)
        self.assertIn('Total amount: ₱0.00', page)
        self.assertIn("const formatPHP = (value) =>", page)
        self.assertIn("row.po_type === 'semi_annual'", page)
        self.assertIn("row.po_type === 'quarterly'", page)

    def test_client_delete_cascades_purchase_orders_without_touching_other_client(self):
        client = self._client_for(self.po_user_id)
        with self.app.app_context():
            delete_target = app_module.Client(
                name=f'P.O. Cascade Target {self.suffix}',
                address='Temporary cascade target',
            )
            app_module.db.session.add(delete_target)
            app_module.db.session.flush()
            delete_target_id = delete_target.id
            cascade_serial = f'PO-SN-CASCADE-{self.suffix}'
            app_module.db.session.add(app_module.Product(
                serial_number=cascade_serial,
                name=f'Cascade CT {self.suffix}',
                client_id=delete_target_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ))
            app_module.db.session.commit()
        self.created_product_serials.append(cascade_serial)
        first = client.post('/add_purchase_order', json={
            'client_id': delete_target_id,
            'product_serial': cascade_serial,
            'po_date': '2026-08-06',
            'po_number': f'CASCADE-ONE-{self.suffix}',
            'po_type': 'single',
        })
        second = client.post('/add_purchase_order', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'po_date': '2026-08-06',
            'end_date': '2026-08-20',
            'po_number': f'CASCADE-TWO-{self.suffix}',
            'po_type': 'quarterly',
        })
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)

        superadmin = self._client_for(self.superadmin_id)
        deleted = superadmin.delete(f'/delete_client/{delete_target_id}')
        self.assertEqual(deleted.status_code, 200)

        with self.app.app_context():
            self.assertIsNone(app_module.db.session.get(app_module.Client, delete_target_id))
            self.assertFalse(app_module.PurchaseOrder.query.filter_by(client_id=delete_target_id).first())
            survivor = app_module.PurchaseOrder.query.filter_by(client_id=self.client_two_id).first()
            self.assertIsNotNone(survivor)

    def test_model_has_machine_association_but_no_work_linkage(self):
        """A P.O. covers ordered machines, but never becomes a work-record relationship.

        The legacy product_serial column remains a write-only rollback mirror. The
        association table is the application read path and keeps Shift/TSR linkage
        forbidden: those fields would turn this financial register into work history.
        """
        columns = set(app_module.PurchaseOrder.__table__.columns.keys())
        self.assertTrue({'amount', 'end_date', 'product_serial'} <= columns)
        self.assertFalse({'shift_id', 'tsr_id', 'product_id'} & columns)
        self.assertEqual(
            list(app_module.PurchaseOrder.__table__.c.product_serial.foreign_keys)[0].target_fullname,
            'product.serial_number',
        )
        self.assertEqual(
            app_module.Client.purchase_orders.property.cascade.delete_orphan,
            True,
        )
        machine_columns = set(app_module.PurchaseOrderMachine.__table__.columns.keys())
        self.assertTrue({'purchase_order_id', 'product_serial', 'position'} <= machine_columns)
        self.assertFalse({'shift_id', 'tsr_id', 'product_id'} & machine_columns)
        self.assertEqual(
            list(app_module.PurchaseOrderMachine.__table__.c.purchase_order_id.foreign_keys)[0].target_fullname,
            'purchase_order.id',
        )
        self.assertEqual(
            list(app_module.PurchaseOrderMachine.__table__.c.product_serial.foreign_keys)[0].target_fullname,
            'product.serial_number',
        )
        self.assertTrue(app_module.PurchaseOrder.machines.property.cascade.delete_orphan)
        self.assertTrue(any(
            constraint.name == 'uq_purchase_order_machine_pair'
            for constraint in app_module.PurchaseOrderMachine.__table__.constraints
        ))

    def test_new_purchase_order_requires_a_registered_machine(self):
        client = self._client_for(self.po_user_id)
        response = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'po_date': '2026-08-06',
            'po_number': f'NO-MACHINE-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()['error'],
            'Select the equipment/machine for this P.O.',
        )

        unknown = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': f'NOT-REGISTERED-{self.suffix}',
            'po_date': '2026-08-06',
            'po_number': f'UNKNOWN-MACHINE-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(unknown.status_code, 400)
        self.assertEqual(
            unknown.get_json()['error'],
            f'Equipment not found. Add it in Products first: NOT-REGISTERED-{self.suffix}.',
        )

    def test_purchase_order_rejects_a_machine_owned_by_another_client(self):
        client = self._client_for(self.po_user_id)
        response = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': self.product_two_serial,
            'po_date': '2026-08-06',
            'po_number': f'WRONG-MACHINE-OWNER-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()['error'],
            f'That machine is not registered to the selected medical center: {self.product_two_serial}.',
        )

    def test_purchase_order_payload_includes_client_scoped_products(self):
        client = self._client_for(self.po_user_id)
        response = client.get('/get_purchase_orders')
        self.assertEqual(response.status_code, 200)
        products = response.get_json()['products']
        by_serial = {product['serial_number']: product for product in products}
        self.assertIn(self.product_one_serial, by_serial)
        self.assertEqual(by_serial[self.product_one_serial]['client_id'], self.client_one_id)
        self.assertTrue(by_serial[self.product_one_serial]['under_contract'])
        self.assertIn('end_warranty', by_serial[self.product_one_serial])

    def test_machine_filter_and_sort_reach_excel_export(self):
        client = self._client_for(self.po_user_id)
        sort_products = [
            app_module.Product(
                serial_number=f'PO-SORT-A-{self.suffix}',
                name=f'Alpha Sort Machine {self.suffix}',
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ),
            app_module.Product(
                serial_number=f'PO-SORT-B-{self.suffix}',
                name=f'Zulu Sort Machine {self.suffix}',
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ),
        ]
        sort_serials = [product.serial_number for product in sort_products]
        with self.app.app_context():
            app_module.db.session.add_all(sort_products)
            app_module.db.session.commit()
        self.created_product_serials.extend(sort_serials)

        created_ids = []
        for serial, number in (
            (sort_serials[1], f'MACHINE-SORT-B-{self.suffix}'),
            (sort_serials[0], f'MACHINE-SORT-A-{self.suffix}'),
        ):
            response = client.post('/add_purchase_order', json={
                'client_id': self.client_one_id,
                'product_serial': serial,
                'po_date': '2026-08-06',
                'po_number': number,
                'po_type': 'single',
                'amount': '10.00',
            })
            self.assertEqual(response.status_code, 201, response.get_data(as_text=True))
            created_ids.append(response.get_json()['purchase_order']['id'])

        response = client.get(
            '/export_purchase_orders'
            f'?machine={quote("Sort Machine")}&sort=machine&direction=asc'
        )
        self.assertEqual(response.status_code, 200)
        worksheet = load_workbook(BytesIO(response.data), data_only=False)['P.O. Register']
        self.assertEqual(worksheet.max_row, 5)
        self.assertEqual(worksheet['G2'].value, sort_serials[0])
        self.assertEqual(worksheet['G3'].value, sort_serials[1])
        self.assertEqual(worksheet['J5'].value, '=SUM(J2:J3)')

        for record_id in created_ids:
            self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)

    def test_deleting_referenced_product_is_blocked_and_preserves_the_po(self):
        client = self._client_for(self.po_user_id)
        serial = f'PO-DELETE-GUARD-{self.suffix}'
        with self.app.app_context():
            app_module.db.session.add(app_module.Product(
                serial_number=serial,
                name=f'Delete Guard Machine {self.suffix}',
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ))
            app_module.db.session.commit()
        self.created_product_serials.append(serial)

        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': serial,
            'po_date': '2026-08-06',
            'po_number': f'DELETE-GUARD-PO-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(created.status_code, 201)
        record_id = created.get_json()['purchase_order']['id']

        superadmin = self._client_for(self.superadmin_id)
        blocked = superadmin.delete(f'/delete_product/{quote(serial, safe="")}')
        self.assertEqual(blocked.status_code, 409)
        self.assertIn('P.O. record', blocked.get_json()['message'])
        with self.app.app_context():
            survivor = app_module.db.session.get(app_module.PurchaseOrder, record_id)
            self.assertIsNotNone(survivor)
            self.assertEqual(survivor.product_serial, serial)
            self.assertEqual(
                [machine.product_serial for machine in survivor.machines],
                [serial],
            )

        self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)
        self.assertEqual(superadmin.delete(f'/delete_product/{quote(serial, safe="")}').status_code, 200)

    def test_renaming_product_repoints_purchase_orders(self):
        client = self._client_for(self.po_user_id)
        old_serial = f'PO-RENAME-OLD-{self.suffix}'
        new_serial = f'PO-RENAME-NEW-{self.suffix}'.upper()
        with self.app.app_context():
            app_module.db.session.add(app_module.Product(
                serial_number=old_serial,
                name=f'Rename Machine {self.suffix}',
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ))
            app_module.db.session.commit()
        self.created_product_serials.extend([old_serial, new_serial])

        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': old_serial,
            'po_date': '2026-08-06',
            'po_number': f'RENAME-PO-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(created.status_code, 201)
        record_id = created.get_json()['purchase_order']['id']

        superadmin = self._client_for(self.superadmin_id)
        renamed = superadmin.put(f'/update_product/{quote(old_serial, safe="")}', json={
            'serial_number': new_serial,
            'name': f'Rename Machine {self.suffix}',
            'client_id': self.client_one_id,
            'under_contract': False,
        })
        self.assertEqual(renamed.status_code, 200, renamed.get_data(as_text=True))
        self.assertEqual(renamed.get_json()['linked_purchase_order_count'], 1)
        with self.app.app_context():
            repointed = app_module.db.session.get(app_module.PurchaseOrder, record_id)
            self.assertEqual(repointed.product_serial, new_serial)
            self.assertEqual(
                [machine.product_serial for machine in repointed.machines],
                [new_serial],
            )

        self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)

    def test_purchase_order_modal_has_scoped_machine_autocomplete(self):
        client = self._client_for(self.po_user_id)
        page = client.get('/po_details')
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        for expected in (
            'id="po-client-input"',
            'id="po-client"',
            'id="po-client-results"',
            'id="po-machine-input"',
            'id="po-machine"',
            'id="po-machine-results"',
            'id="po-machine-chips"',
            'No equipment registered for this client — add it in Products first.',
        ):
            self.assertIn(expected, html)

    def test_purchase_order_machine_picker_matches_product_contract_flag(self):
        client = self._client_for(self.po_user_id)
        page = client.get('/po_details')
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("function productCoverageDisplay(product)", html)
        self.assertIn("if (Boolean(product?.under_contract)) return 'Under Contract';", html)
        self.assertIn("return product?.end_warranty ?", html)
        self.assertIn("productCoverageDisplay(product),", html)

    def test_purchase_order_machine_picker_loads_all_client_equipment(self):
        client = self._client_for(self.po_user_id)
        page = client.get('/po_details')
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        machine_matches = html.split(
            "matches = state.products.filter((product) => {",
            1,
        )[1].split("if (!matches.length)", 1)[0]
        self.assertNotIn('.slice(0, 10)', machine_matches)
        self.assertIn('max-height: 13rem;', html)
        self.assertIn('overflow-y: auto;', html)

    def test_multi_machine_add_edit_replaces_links_and_returns_full_response(self):
        client = self._client_for(self.po_user_id)
        second_serial = f'PO-MULTI-SECOND-{self.suffix}'
        with self.app.app_context():
            app_module.db.session.add(app_module.Product(
                serial_number=second_serial,
                name=f'Multi Machine Two {self.suffix}',
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ))
            app_module.db.session.commit()
        self.created_product_serials.append(second_serial)

        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serials': [self.product_one_serial, second_serial.lower(), self.product_one_serial],
            'start_date': '2026-08-10',
            'po_number': f'MULTI-ADD-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
        record = created.get_json()['purchase_order']
        record_id = record['id']
        self.assertEqual(record['machine_serials'], [self.product_one_serial, second_serial])
        self.assertEqual(record['machine_count'], 2)
        self.assertEqual([machine['serial_number'] for machine in record['machines']], record['machine_serials'])
        self.assertEqual(record['product_serial'], self.product_one_serial)

        updated = client.put(f'/update_purchase_order/{record_id}', json={
            'client_id': self.client_one_id,
            'product_serials': [second_serial],
            'start_date': '2026-08-10',
            'po_number': f'MULTI-REPLACED-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(updated.status_code, 200, updated.get_data(as_text=True))
        updated_record = updated.get_json()['purchase_order']
        self.assertEqual(updated_record['machine_serials'], [second_serial])
        with self.app.app_context():
            saved = app_module.db.session.get(app_module.PurchaseOrder, record_id)
            self.assertEqual([machine.product_serial for machine in saved.machines], [second_serial])
            self.assertEqual(saved.product_serial, second_serial)

        self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)

    def test_duplicate_machine_link_is_rejected_by_database_constraint(self):
        client = self._client_for(self.po_user_id)
        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serials': [self.product_one_serial],
            'start_date': '2026-08-10',
            'po_number': f'MULTI-DUPLICATE-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(created.status_code, 201)
        record_id = created.get_json()['purchase_order']['id']

        with self.app.app_context():
            app_module.db.session.add(app_module.PurchaseOrderMachine(
                purchase_order_id=record_id,
                product_serial=self.product_one_serial,
                position=1,
            ))
            with self.assertRaises(IntegrityError):
                app_module.db.session.commit()
            app_module.db.session.rollback()
            self.assertEqual(
                app_module.PurchaseOrderMachine.query.filter_by(purchase_order_id=record_id).count(),
                1,
            )

        self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)

    def test_legacy_backfill_is_additive_and_does_not_resurrect_removed_links(self):
        second_serial = f'PO-BACKFILL-SECOND-{self.suffix}'
        missing_serial = f'PO-BACKFILL-MISSING-{self.suffix}'
        with self.app.app_context():
            app_module.db.session.add(app_module.Product(
                serial_number=second_serial,
                name=f'Backfill Second {self.suffix}',
                client_id=self.client_one_id,
            ))
            legacy = app_module.PurchaseOrder(
                client_id=self.client_one_id,
                po_number=f'BACKFILL-LEGACY-{self.suffix}',
                po_date=app_module.date(2026, 8, 10),
                po_type=app_module.PO_TYPE_SINGLE_VISIT,
                product_serial=self.product_one_serial,
            )
            missing = app_module.PurchaseOrder(
                client_id=self.client_one_id,
                po_number=f'BACKFILL-MISSING-{self.suffix}',
                po_date=app_module.date(2026, 8, 10),
                po_type=app_module.PO_TYPE_SINGLE_VISIT,
                product_serial=missing_serial,
            )
            multi = app_module.PurchaseOrder(
                client_id=self.client_one_id,
                po_number=f'BACKFILL-MULTI-{self.suffix}',
                po_date=app_module.date(2026, 8, 10),
                po_type=app_module.PO_TYPE_SINGLE_VISIT,
                product_serial=self.product_one_serial,
            )
            app_module.db.session.add_all([legacy, missing, multi])
            app_module.db.session.flush()
            app_module.apply_purchase_order_machines(
                multi,
                [self.product_one_serial, second_serial],
            )
            app_module.db.session.commit()
            self.created_product_serials.append(second_serial)
            previous_ready = app_module._purchase_order_schema_ready
            try:
                app_module._purchase_order_schema_ready = False
                app_module.ensure_purchase_order_schema()
                self.assertEqual(
                    [machine.product_serial for machine in app_module.db.session.get(app_module.PurchaseOrder, legacy.id).machines],
                    [self.product_one_serial],
                )
                self.assertEqual(
                    [machine.product_serial for machine in app_module.db.session.get(app_module.PurchaseOrder, missing.id).machines],
                    [missing_serial],
                )
                app_module.ensure_purchase_order_schema()
                self.assertEqual(
                    app_module.PurchaseOrderMachine.query.filter_by(purchase_order_id=legacy.id).count(),
                    1,
                )

                multi_saved = app_module.db.session.get(app_module.PurchaseOrder, multi.id)
                app_module.db.session.delete(multi_saved.machines[0])
                multi_saved.product_serial = second_serial
                app_module.db.session.commit()
                app_module._purchase_order_schema_ready = False
                app_module.ensure_purchase_order_schema()
                refreshed = app_module.db.session.get(app_module.PurchaseOrder, multi.id)
                self.assertEqual(
                    [machine.product_serial for machine in refreshed.machines],
                    [second_serial],
                )
            finally:
                app_module._purchase_order_schema_ready = previous_ready

    def test_purchase_order_reads_ignore_a_corrupted_legacy_mirror(self):
        client = self._client_for(self.po_user_id)
        second_serial = f'PO-MIRROR-SECOND-{self.suffix}'
        with self.app.app_context():
            app_module.db.session.add(app_module.Product(
                serial_number=second_serial,
                name=f'Mirror Second {self.suffix}',
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ))
            app_module.db.session.commit()
        self.created_product_serials.append(second_serial)
        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serials': [self.product_one_serial, second_serial],
            'start_date': '2026-08-10',
            'po_number': f'MIRROR-CORRUPT-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(created.status_code, 201)
        record_id = created.get_json()['purchase_order']['id']
        with self.app.app_context():
            saved = app_module.db.session.get(app_module.PurchaseOrder, record_id)
            saved.product_serial = 'CORRUPTED-MIRROR'
            app_module.db.session.commit()

        listing = client.get('/get_purchase_orders')
        self.assertEqual(listing.status_code, 200)
        record = next(item for item in listing.get_json()['purchase_orders'] if item['id'] == record_id)
        self.assertEqual(record['machine_serials'], [self.product_one_serial, second_serial])
        self.assertEqual(record['product_serial'], self.product_one_serial)
        self.assertNotIn('CORRUPTED-MIRROR', record['machine_search'])
        self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)

    def test_machine_filter_matches_any_link_and_export_keeps_one_po_row(self):
        client = self._client_for(self.po_user_id)
        second_serial = f'PO-EXPORT-SECOND-{self.suffix}'
        second_name = f'Export Second {self.suffix}'
        with self.app.app_context():
            app_module.db.session.add(app_module.Product(
                serial_number=second_serial,
                name=second_name,
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ))
            app_module.db.session.commit()
        self.created_product_serials.append(second_serial)
        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serials': [self.product_one_serial, second_serial],
            'start_date': '2026-08-10',
            'po_number': f'EXPORT-MULTI-{self.suffix}',
            'po_type': 'single',
            'amount': '42.00',
        })
        self.assertEqual(created.status_code, 201)
        record_id = created.get_json()['purchase_order']['id']
        response = client.get(f'/export_purchase_orders?machine={quote(second_name)}')
        self.assertEqual(response.status_code, 200)
        worksheet = load_workbook(BytesIO(response.data), data_only=False)['P.O. Register']
        self.assertEqual(worksheet['D2'].value, f'EXPORT-MULTI-{self.suffix}')
        self.assertEqual(worksheet['G2'].value, f'{self.product_one_serial}\n{second_serial}')
        self.assertEqual(worksheet['H2'].value, f'CT One {self.suffix}\n{second_name}')
        self.assertIsNone(worksheet['D3'].value)
        self.assertEqual(worksheet['J4'].value, '=SUM(J2:J2)')
        self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)

    def test_export_machine_columns_stay_aligned_when_a_product_row_is_missing(self):
        """Line k of Machine Serial must always describe line k of Machine Name.

        A link whose Product row is gone contributes an empty name. The row builder emits
        '' for it deliberately; if that is ever "tidied" into a filter that drops empties,
        the two columns slip out of step and the spreadsheet attributes the wrong model to
        a serial. It still opens, still looks entirely normal, and the amount total is
        still right -- which is why only an explicit alignment assertion catches it.

        The product is deleted through the session rather than /delete_product because
        that route refuses with 409 precisely so this state cannot be reached by hand. It
        is still reachable: the backfill deliberately keeps a legacy link whose product
        row was purged before the link existed, rather than dropping the row.
        """
        client = self._client_for(self.po_user_id)
        orphan_serial = f'PO-ORPHAN-{self.suffix}'
        with self.app.app_context():
            app_module.db.session.add(app_module.Product(
                serial_number=orphan_serial,
                name=f'Orphan Model {self.suffix}',
                client_id=self.client_one_id,
                start_warranty_date=date(2026, 8, 1),
                end_warranty_date=date(2027, 8, 1),
                under_contract=True,
            ))
            app_module.db.session.commit()
        self.created_product_serials.append(orphan_serial)

        created = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serials': [self.product_one_serial, orphan_serial],
            'start_date': '2026-08-10',
            'po_number': f'ORPHAN-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(created.status_code, 201, created.get_data(as_text=True))
        record_id = created.get_json()['purchase_order']['id']

        with self.app.app_context():
            orphan = app_module.db.session.get(app_module.Product, orphan_serial)
            app_module.db.session.delete(orphan)
            app_module.db.session.commit()

        response = client.get(f'/export_purchase_orders?number=ORPHAN-{self.suffix}')
        self.assertEqual(response.status_code, 200)
        worksheet = load_workbook(BytesIO(response.data), data_only=False)['P.O. Register']
        serial_lines = (worksheet['G2'].value or '').split('\n')
        name_lines = (worksheet['H2'].value or '').split('\n')
        self.assertEqual(serial_lines, [self.product_one_serial, orphan_serial])
        # The alignment guard: one name line per serial line, empty rather than absent.
        self.assertEqual(len(name_lines), len(serial_lines))
        self.assertEqual(name_lines[0], f'CT One {self.suffix}')
        self.assertEqual(name_lines[1], '')
        self.assertIsNone(worksheet['D3'].value)

        self.assertEqual(client.delete(f'/delete_purchase_order/{record_id}').status_code, 200)

    def test_purchase_order_listing_uses_one_association_query(self):
        statements = []

        def capture(_connection, _cursor, statement, _parameters, _context, _executemany):
            statements.append(statement.lower())

        with self.app.app_context():
            event.listen(app_module.db.engine, 'before_cursor_execute', capture)
            try:
                records = app_module.PurchaseOrder.query.options(
                    app_module.joinedload(app_module.PurchaseOrder.client),
                    app_module.joinedload(app_module.PurchaseOrder.created_by_user),
                    app_module.selectinload(app_module.PurchaseOrder.machines).joinedload(
                        app_module.PurchaseOrderMachine.product
                    ),
                ).all()
                [app_module.purchase_order_to_dict(record) for record in records]
            finally:
                event.remove(app_module.db.engine, 'before_cursor_execute', capture)
        association_queries = [
            statement for statement in statements
            if statement.lstrip().startswith('select') and 'purchase_order_machine' in statement
        ]
        self.assertEqual(len(association_queries), 1)
        self.assertLessEqual(len(statements), 4)

    def test_machine_normalizer_accepts_legacy_strings_dedupes_and_caps(self):
        with self.app.app_context():
            serials, error = app_module.normalize_purchase_order_machines({
                'product_serials': [self.product_one_serial, self.product_one_serial.lower()],
            }, self.client_one_id)
            self.assertIsNone(error)
            self.assertEqual(serials, [self.product_one_serial])

            serials, error = app_module.normalize_purchase_order_machines({
                'product_serial': f'{self.product_one_serial},{self.product_one_serial}',
            }, self.client_one_id)
            self.assertIsNone(error)
            self.assertEqual(serials, [self.product_one_serial])

            serials, error = app_module.normalize_purchase_order_machines({
                'product_serials': [f'NOT-REGISTERED-{index}' for index in range(51)],
            }, self.client_one_id)
            self.assertIsNone(serials)
            self.assertEqual(error, 'A P.O. may reference no more than 50 machines.')

    def test_settings_reports_the_stored_grant_not_the_effective_permission(self):
        """The Settings payload must carry the stored flag, not the effective permission.

        approval_user_to_dict() drives the Settings switches, and saveApprovalUser()
        posts the rendered state straight back. Reporting can_manage_purchase_orders()
        -- which folds in is_admin_authorized() -- made the switch render checked for
        every superadmin and the regional admin, so saving any unrelated change on their
        card wrote po_admin_access=True and an audit line for a grant nobody performed.
        The regional admin reaches the same is_admin_authorized() branch.
        """
        with self.app.app_context():
            superadmin = app_module.db.session.get(app_module.User, self.superadmin_id)
            self.assertFalse(
                superadmin.po_admin_access,
                'fixture precondition: the superadmin holds no explicit grant',
            )
            # Effective access is unchanged -- admins still reach the register.
            self.assertTrue(app_module.can_manage_purchase_orders(superadmin))
            # ...but the Settings switch must show the stored grant, which is False.
            self.assertFalse(
                app_module.approval_user_to_dict(superadmin)['po_admin_access']
            )

            # Positive control: a real grantee still reports True, so the assertion
            # above cannot pass by the serializer simply always returning False.
            granted = app_module.db.session.get(app_module.User, self.po_user_id)
            self.assertTrue(granted.po_admin_access)
            self.assertTrue(
                app_module.approval_user_to_dict(granted)['po_admin_access']
            )

    def test_saving_an_unrelated_permission_does_not_grant_po_access(self):
        """A superadmin's card round-trips without silently persisting the flag."""
        superadmin = self._client_for(self.superadmin_id)
        rendered = superadmin.get('/settings/approval-routing-data')
        self.assertEqual(rendered.status_code, 200)
        row = next(
            user for user in rendered.get_json()['users']
            if user['id'] == self.superadmin_id
        )
        self.assertFalse(row['po_admin_access'], 'the switch must render unchecked')

        # Post back exactly what the UI rendered, as saveApprovalUser() does.
        saved = superadmin.post('/settings/update-approval-user', json={
            'user_id': self.superadmin_id,
            'is_active': True,
            'po_admin_access': row['po_admin_access'],
        })
        self.assertEqual(saved.status_code, 200)
        with self.app.app_context():
            self.assertFalse(
                app_module.db.session.get(app_module.User, self.superadmin_id).po_admin_access
            )

    def test_writes_are_refused_without_the_capability(self):
        owner = self._client_for(self.po_user_id)
        seeded = owner.post('/add_purchase_order', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'po_date': '2026-08-06',
            'end_date': '2026-08-20',
            'po_number': f'GUARD-{self.suffix}',
            'po_type': 'quarterly',
        })
        self.assertEqual(seeded.status_code, 201)
        record_id = seeded.get_json()['purchase_order']['id']

        intruder = self._client_for(self.plain_user_id)
        self.assertEqual(intruder.post('/add_purchase_order', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'po_date': '2026-08-06',
            'end_date': '2026-08-20',
            'po_number': f'INTRUDER-{self.suffix}',
            'po_type': 'quarterly',
        }).status_code, 403)
        self.assertEqual(intruder.put(f'/update_purchase_order/{record_id}', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'po_date': '2026-08-06',
            'po_number': f'INTRUDER-EDIT-{self.suffix}',
            'po_type': 'quarterly',
        }).status_code, 403)
        self.assertEqual(
            intruder.delete(f'/delete_purchase_order/{record_id}').status_code, 403
        )

        with self.app.app_context():
            survivor = app_module.db.session.get(app_module.PurchaseOrder, record_id)
            self.assertIsNotNone(survivor, 'the refused delete must not have landed')
            self.assertEqual(survivor.po_number, f'GUARD-{self.suffix}')

        # Positive control: the same three calls succeed for the capability holder.
        self.assertEqual(owner.put(f'/update_purchase_order/{record_id}', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'po_date': '2026-08-06',
            'po_number': f'GUARD-EDITED-{self.suffix}',
            'po_type': 'single',
        }).status_code, 200)
        self.assertEqual(
            owner.delete(f'/delete_purchase_order/{record_id}').status_code, 200
        )

    def test_po_type_must_be_single_semi_annual_or_quarterly(self):
        client = self._client_for(self.po_user_id)

        def add(po_type, number):
            payload = {
                'client_id': self.client_two_id,
                'product_serial': self.product_two_serial,
                'po_date': '2026-08-06',
                'po_number': number,
                'po_type': po_type,
            }
            return client.post('/add_purchase_order', json=payload)

        for rejected in ('maybe', '', 'Contracts', 'true', '1', 'contract', 'single_visit'):
            response = add(rejected, f'TYPE-BAD-{self.suffix}')
            self.assertEqual(response.status_code, 400, f'accepted po_type={rejected!r}')
            self.assertIn('Single', response.get_json()['error'])

        # Positive control: all current frequency values are accepted and stored as text.
        for accepted, label in (('single', 'Single'), ('semi_annual', 'Semi Annual'), ('quarterly', 'Quarterly')):
            response = add(accepted, f'TYPE-OK-{accepted}-{self.suffix}')
            self.assertEqual(response.status_code, 201)
            body = response.get_json()['purchase_order']
            self.assertEqual(body['po_type'], accepted)
            self.assertEqual(body['po_type_label'], label)

    def test_po_dates_are_product_owned_not_client_supplied(self):
        client = self._client_for(self.po_user_id)
        response = client.post('/add_purchase_order', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'start_date': '1900-01-01',
            'end_date': '1900-01-02',
            'po_number': f'DATE-OWNED-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.get_json()['purchase_order']['start_date'], '2026-08-01')
        self.assertEqual(response.get_json()['purchase_order']['end_date'], '2027-08-01')

    def test_current_frequency_amount_rules(self):
        client = self._client_for(self.po_user_id)

        missing_machine = client.post('/add_purchase_order', json={
            'client_id': self.client_two_id,
            'start_date': '2026-08-06',
            'po_number': f'END-REQUIRED-{self.suffix}',
            'po_type': 'quarterly',
        })
        self.assertEqual(missing_machine.status_code, 400)
        self.assertIn('machine', missing_machine.get_json()['error'].lower())

        for invalid_amount in ('0', '-1', '12.345', 'not-a-number'):
            response = client.post('/add_purchase_order', json={
                'client_id': self.client_two_id,
                'product_serial': self.product_two_serial,
                'start_date': '2026-08-06',
                'po_number': f'AMOUNT-BAD-{invalid_amount}-{self.suffix}',
                'po_type': 'single',
                'amount': invalid_amount,
            })
            self.assertEqual(response.status_code, 400, invalid_amount)
            self.assertIn('amount', response.get_json()['error'].lower())

        current_frequency = client.post('/add_purchase_order', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'start_date': '2026-08-06',
            'po_number': f'SINGLE-CURRENT-{self.suffix}',
            'po_type': 'single',
            'amount': '99.90',
        })
        self.assertEqual(current_frequency.status_code, 201)
        record = current_frequency.get_json()['purchase_order']
        self.assertEqual(record['end_date'], '2027-08-01')
        self.assertEqual(record['amount'], '99.90')

        cleared = client.put(f"/update_purchase_order/{record['id']}", json={
            'start_date': '2026-08-06',
            'end_date': '',
            'po_number': f'SINGLE-CURRENT-{self.suffix}',
            'po_type': 'single',
            'amount': '',
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
        })
        self.assertEqual(cleared.status_code, 200)
        self.assertEqual(cleared.get_json()['purchase_order']['amount'], '')

    def test_legacy_contract_remains_readable_but_edit_requires_machine(self):
        with self.app.app_context():
            legacy = app_module.PurchaseOrder(
                client_id=self.client_two_id,
                po_number=f'LEGACY-CONTRACT-{self.suffix}',
                po_date=app_module.date(2026, 7, 1),
                end_date=None,
                po_type=app_module.PO_TYPE_CONTRACT,
            )
            app_module.db.session.add(legacy)
            app_module.db.session.commit()
            legacy_id = legacy.id

        client = self._client_for(self.po_user_id)
        listing = client.get('/get_purchase_orders')
        self.assertEqual(listing.status_code, 200)
        legacy_payload = next(
            item for item in listing.get_json()['purchase_orders']
            if item['id'] == legacy_id
        )
        self.assertEqual(legacy_payload['end_date'], '')
        self.assertEqual(legacy_payload['po_type'], 'contract')
        self.assertEqual(legacy_payload['product_serial'], '')

        edit_without_machine = client.put(f'/update_purchase_order/{legacy_id}', json={
            'client_id': self.client_two_id,
            'start_date': '2026-07-01',
            'end_date': '2026-07-31',
            'po_number': f'LEGACY-CONTRACT-EDITED-{self.suffix}',
            'po_type': 'contract',
        })
        self.assertEqual(edit_without_machine.status_code, 400)
        self.assertIn('Single', edit_without_machine.get_json()['error'])

        edit_with_machine = client.put(f'/update_purchase_order/{legacy_id}', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'start_date': '1900-01-01',
            'end_date': '1900-01-02',
            'po_number': f'LEGACY-CONTRACT-EDITED-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(edit_with_machine.status_code, 200)
        self.assertEqual(
            edit_with_machine.get_json()['purchase_order']['product_serial'],
            self.product_two_serial,
        )

    def test_filtered_sorted_excel_export_contains_complete_register_details(self):
        client = self._client_for(self.po_user_id)
        first = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': self.product_one_serial,
            'start_date': '2026-08-01',
            'end_date': '2026-08-31',
            'po_number': f'EXPORT-A-{self.suffix}',
            'po_type': 'quarterly',
            'amount': '25.00',
        })
        second = client.post('/add_purchase_order', json={
            'client_id': self.client_one_id,
            'product_serial': self.product_one_serial,
            'start_date': '2026-08-02',
            'po_number': f'EXPORT-B-{self.suffix}',
            'po_type': 'single',
            'amount': '100.00',
        })
        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)

        query = (
            f'/export_purchase_orders?number={quote("EXPORT-")}'
            '&sort=amount&direction=desc'
        )
        response = client.get(query)
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data), data_only=False)
        worksheet = workbook['P.O. Register']
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            [
                'P.O. ID', 'Start Date', 'End Date', 'P.O. Number',
                'Medical Center', 'Complete Address', 'Machine Serial',
                'Machine Name', 'P.O. Type', 'Amount (PHP)', 'Computed Amount (PHP)',
                'Scheduled Dates', 'Created At', 'Created By', 'Updated At',
            ],
        )
        self.assertEqual(worksheet['D2'].value, f'EXPORT-B-{self.suffix}')
        self.assertEqual(worksheet['G2'].value, self.product_one_serial)
        self.assertEqual(worksheet['H2'].value, f'CT One {self.suffix}')
        self.assertEqual(worksheet['J2'].value, 100)
        self.assertEqual(worksheet['J3'].value, 25)
        self.assertEqual(worksheet['K2'].value, 50)
        self.assertEqual(worksheet['K3'].value, 6.25)
        self.assertEqual(worksheet['F2'].value, 'Test address one')
        self.assertEqual(worksheet['J5'].value, '=SUM(J2:J3)')
        self.assertEqual(worksheet['K5'].value, '=SUM(K2:K3)')
        self.assertEqual(worksheet.freeze_panes, 'A2')
        self.assertEqual(worksheet.auto_filter.ref, 'A1:O3')

        unauthorized = self._client_for(self.plain_user_id)
        self.assertEqual(unauthorized.get('/export_purchase_orders').status_code, 403)

    def test_missing_records_report_404_rather_than_a_silent_success(self):
        """delete_client() reports success for an id that never existed; this must not."""
        client = self._client_for(self.po_user_id)
        created = client.post('/add_purchase_order', json={
            'client_id': self.client_two_id,
            'product_serial': self.product_two_serial,
            'po_date': '2026-08-06',
            'po_number': f'GONE-{self.suffix}',
            'po_type': 'single',
        })
        self.assertEqual(created.status_code, 201)
        record_id = created.get_json()['purchase_order']['id']

        # Positive control: the first delete really does succeed.
        self.assertEqual(
            client.delete(f'/delete_purchase_order/{record_id}').status_code, 200
        )

        self.assertEqual(
            client.delete(f'/delete_purchase_order/{record_id}').status_code, 404
        )
        self.assertEqual(client.put(f'/update_purchase_order/{record_id}', json={
            'client_id': self.client_two_id,
            'po_date': '2026-08-06',
            'po_number': f'GONE-EDIT-{self.suffix}',
            'po_type': 'single',
        }).status_code, 404)


class PurchaseOrderRenewalTests(unittest.TestCase):
    """Focused renewal controls with per-test disposable P.O. fixtures."""

    @classmethod
    def setUpClass(cls):
        cls.app = app_module.app
        cls.app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
        cls.suffix = uuid.uuid4().hex[:8]
        with cls.app.app_context():
            app_module.db.create_all()
            app_module.ensure_user_admin_capability_columns()
            app_module.ensure_purchase_order_schema()
            cls.user = app_module.User(
                username=f'po_renewal_{cls.suffix}',
                password=app_module.generate_password_hash('test-password'),
                role='staff',
                is_active=True,
                po_admin_access=True,
            )
            app_module.db.session.add(cls.user)
            app_module.db.session.commit()
            cls.user_id = cls.user.id

    @classmethod
    def tearDownClass(cls):
        with cls.app.app_context():
            user = app_module.db.session.get(app_module.User, cls.user_id)
            if user:
                app_module.db.session.delete(user)
            app_module.db.session.commit()
            app_module.db.session.remove()

    def setUp(self):
        self.client = self.app.test_client()
        response = self.client.post('/login', data={
            'username': f'po_renewal_{self.suffix}',
            'password': 'test-password',
        })
        self.assertIn(response.status_code, (302, 303))
        self.created_serials = []
        with self.app.app_context():
            self.client_record = app_module.Client(
                name=f'P.O. Renewal Client {self.suffix}-{uuid.uuid4().hex[:6]}',
                address='Renewal address',
            )
            app_module.db.session.add(self.client_record)
            app_module.db.session.commit()
            self.client_id = self.client_record.id

    def tearDown(self):
        with self.app.app_context():
            for purchase_order in app_module.PurchaseOrder.query.filter_by(client_id=self.client_id).all():
                app_module.db.session.delete(purchase_order)
            for serial in self.created_serials:
                product = app_module.db.session.get(app_module.Product, serial)
                if product:
                    app_module.db.session.delete(product)
            client = app_module.db.session.get(app_module.Client, self.client_id)
            if client:
                app_module.db.session.delete(client)
            app_module.db.session.commit()
            app_module.db.session.remove()

    def _product(self, label, *, start=None, end=None, under_contract=True):
        serial = f'PO-RENEW-{self.suffix}-{label}-{uuid.uuid4().hex[:6]}'
        with self.app.app_context():
            product = app_module.Product(
                serial_number=serial,
                name=f'Renewal Machine {label}',
                client_id=self.client_id,
                start_warranty_date=start,
                end_warranty_date=end,
                under_contract=under_contract,
            )
            app_module.db.session.add(product)
            app_module.db.session.commit()
        self.created_serials.append(serial)
        return serial

    def _po(self, serials, *, start, end, po_type='single', amount='100.00'):
        number = f'RENEWAL-OLD-{self.suffix}-{uuid.uuid4().hex[:6]}'
        with self.app.app_context():
            purchase_order = app_module.PurchaseOrder(
                client_id=self.client_id,
                po_number=number,
                po_date=start,
                end_date=end,
                po_type=po_type,
                amount=amount,
                created_by=self.user_id,
            )
            app_module.db.session.add(purchase_order)
            app_module.db.session.flush()
            app_module.apply_purchase_order_machines(purchase_order, list(serials))
            app_module.db.session.commit()
            return purchase_order.id

    def _payload(self, serials, *, start, end, date_mode='renewal', po_type='single'):
        return {
            'client_id': self.client_id,
            'product_serials': list(serials),
            'po_number': f'RENEWAL-NEW-{self.suffix}-{uuid.uuid4().hex[:6]}',
            'po_type': po_type,
            'amount': '1000.00',
            'date_mode': date_mode,
            'start_date': start,
            'end_date': end,
        }

    def _count_pos(self):
        with self.app.app_context():
            return app_module.PurchaseOrder.query.filter_by(client_id=self.client_id).count()

    def test_valid_renewal_updates_all_products_and_preserves_old_snapshot(self):
        today = app_module.get_manila_today()
        old_end = today - timedelta(days=30)
        serial = self._product('single', start=today - timedelta(days=400), end=None)
        old_id = self._po(
            [serial],
            start=today - timedelta(days=395),
            end=old_end,
            amount='321.45',
        )
        new_start = old_end + timedelta(days=1)
        new_end = today + timedelta(days=365)
        with self.app.app_context():
            old = app_module.db.session.get(app_module.PurchaseOrder, old_id)
            old_snapshot = (old.po_date, old.end_date, old.amount)
            old_schedule = app_module.purchase_order_schedule(
                old.po_date, old.end_date, old.po_type, old.amount,
            )

        response = self.client.post(
            '/add_purchase_order',
            json=self._payload([serial], start=new_start.isoformat(), end=new_end.isoformat()),
        )
        self.assertEqual(response.status_code, 201, response.get_data(as_text=True))
        body = response.get_json()['purchase_order']
        self.assertEqual(body['start_date'], new_start.isoformat())
        self.assertEqual(body['end_date'], new_end.isoformat())
        self.assertEqual(body['schedule_mode'], 'annual')
        self.assertEqual(body['schedule_mode_label'], 'Annual')

        with self.app.app_context():
            old = app_module.db.session.get(app_module.PurchaseOrder, old_id)
            self.assertEqual((old.po_date, old.end_date, old.amount), old_snapshot)
            self.assertEqual(
                app_module.purchase_order_schedule(old.po_date, old.end_date, old.po_type, old.amount),
                old_schedule,
            )
            product = app_module.db.session.get(app_module.Product, serial)
            self.assertEqual(product.start_warranty_date, new_start)
            self.assertEqual(product.end_warranty_date, new_end)

    def test_renewal_confirmation_required_and_missing_confirmation_is_atomic(self):
        today = app_module.get_manila_today()
        old_end = today - timedelta(days=14)
        original_start = today - timedelta(days=500)
        serial = self._product('confirm', start=original_start, end=None, under_contract=False)
        old_id = self._po([serial], start=today - timedelta(days=500), end=old_end)
        payload = self._payload(
            [serial],
            start=(old_end + timedelta(days=1)).isoformat(),
            end=(today + timedelta(days=180)).isoformat(),
        )
        before_count = self._count_pos()
        pending = self.client.post('/add_purchase_order', json=payload)
        self.assertEqual(pending.status_code, 409)
        self.assertEqual(pending.get_json()['code'], 'under_contract_confirmation_required')
        self.assertEqual(self._count_pos(), before_count)
        with self.app.app_context():
            product = app_module.db.session.get(app_module.Product, serial)
            self.assertEqual(product.start_warranty_date, original_start)
            self.assertIsNone(product.end_warranty_date)
            self.assertFalse(product.under_contract)
            old = app_module.db.session.get(app_module.PurchaseOrder, old_id)
            self.assertEqual(old.end_date, old_end)

        payload['under_contract_confirmed_serials'] = [serial]
        saved = self.client.post('/add_purchase_order', json=payload)
        self.assertEqual(saved.status_code, 201, saved.get_data(as_text=True))
        with self.app.app_context():
            product = app_module.db.session.get(app_module.Product, serial)
            self.assertTrue(product.under_contract)
            self.assertEqual(product.start_warranty_date, old_end + timedelta(days=1))
            self.assertEqual(product.end_warranty_date, today + timedelta(days=180))

    def test_latest_active_no_history_and_null_end_records_block_renewal(self):
        today = app_module.get_manila_today()
        expired_end = today - timedelta(days=20)
        active = self._product('active', start=today - timedelta(days=500), end=None)
        self._po([active], start=today - timedelta(days=500), end=expired_end)
        self._po([active], start=today - timedelta(days=10), end=today)

        no_history = self._product('none', start=today - timedelta(days=100), end=None)
        null_end = self._product('null', start=today - timedelta(days=100), end=None)
        self._po([null_end], start=today - timedelta(days=20), end=None)
        null_end_after_expired = self._product('null-ignored', start=today - timedelta(days=100), end=None)
        ignored_end = today - timedelta(days=15)
        ignored_id = self._po([null_end_after_expired], start=today - timedelta(days=30), end=ignored_end)
        self._po([null_end_after_expired], start=today - timedelta(days=10), end=None)

        for serial in (active, no_history, null_end):
            response = self.client.post(
                '/add_purchase_order',
                json=self._payload(
                    [serial],
                    start=(today + timedelta(days=1)).isoformat(),
                    end=(today + timedelta(days=90)).isoformat(),
                ),
            )
            self.assertEqual(response.status_code, 400, response.get_data(as_text=True))
            self.assertIn('expired', response.get_json()['error'].lower())

        with self.app.app_context():
            context = app_module.purchase_order_renewal_context([null_end_after_expired])
            self.assertTrue(context[null_end_after_expired]['renewal_eligible'])
            self.assertEqual(context[null_end_after_expired]['latest_prior_end_date'], ignored_end)
            self.assertEqual(context[null_end_after_expired]['latest_prior_po_id'], ignored_id)

    def test_renewal_requires_strict_non_overlapping_and_valid_submitted_dates(self):
        today = app_module.get_manila_today()
        old_end = today - timedelta(days=10)
        serial = self._product('boundary', start=today - timedelta(days=300), end=None)
        self._po([serial], start=today - timedelta(days=300), end=old_end)
        before_count = self._count_pos()

        for start, end, expected in (
            (old_end.isoformat(), (today + timedelta(days=90)).isoformat(), 'after'),
            ((old_end - timedelta(days=1)).isoformat(), (today + timedelta(days=90)).isoformat(), 'after'),
            ('not-a-date', (today + timedelta(days=90)).isoformat(), 'valid'),
            ((old_end + timedelta(days=1)).isoformat(), '', 'End Date'),
            ((old_end + timedelta(days=1)).isoformat(), '2026-02-30', 'valid'),
        ):
            response = self.client.post(
                '/add_purchase_order',
                json=self._payload([serial], start=start, end=end),
            )
            self.assertEqual(response.status_code, 400, response.get_data(as_text=True))
            self.assertIn(expected.lower(), response.get_json()['error'].lower())
            self.assertEqual(self._count_pos(), before_count)

        invalid_mode = self.client.post(
            '/add_purchase_order',
            json=self._payload(
                [serial],
                start=(old_end + timedelta(days=1)).isoformat(),
                end=(today + timedelta(days=90)).isoformat(),
                date_mode='client-spoof',
            ),
        )
        self.assertEqual(invalid_mode.status_code, 400)
        self.assertIn('date mode', invalid_mode.get_json()['error'].lower())

    def test_multiple_eligible_machines_use_one_range_and_mixed_selection_is_rejected(self):
        today = app_module.get_manila_today()
        first_end = today - timedelta(days=40)
        second_end = today - timedelta(days=12)
        first = self._product('multi-one', start=today - timedelta(days=500), end=None)
        second = self._product('multi-two', start=today - timedelta(days=500), end=None)
        self._po([first], start=today - timedelta(days=500), end=first_end)
        self._po([second], start=today - timedelta(days=500), end=second_end)
        new_start = second_end + timedelta(days=1)
        new_end = today + timedelta(days=300)
        saved = self.client.post(
            '/add_purchase_order',
            json=self._payload([first, second], start=new_start.isoformat(), end=new_end.isoformat()),
        )
        self.assertEqual(saved.status_code, 201, saved.get_data(as_text=True))
        self.assertEqual(saved.get_json()['purchase_order']['machine_serials'], [first, second])
        with self.app.app_context():
            for serial in (first, second):
                product = app_module.db.session.get(app_module.Product, serial)
                self.assertEqual(product.start_warranty_date, new_start)
                self.assertEqual(product.end_warranty_date, new_end)

        ordinary = self._product('mixed', start=today - timedelta(days=100), end=None)
        mixed = self.client.post(
            '/add_purchase_order',
            json=self._payload(
                [first, ordinary],
                start=(new_end + timedelta(days=1)).isoformat(),
                end=(today + timedelta(days=400)).isoformat(),
            ),
        )
        self.assertEqual(mixed.status_code, 400)
        self.assertIn('eligible', mixed.get_json()['error'].lower())

    def test_edit_rejects_renewal_mode_and_normal_mode_remains_product_owned(self):
        today = app_module.get_manila_today()
        old_end = today - timedelta(days=25)
        product_start = today - timedelta(days=400)
        product_end = today + timedelta(days=200)
        serial = self._product('edit', start=product_start, end=product_end)
        old_id = self._po([serial], start=today - timedelta(days=400), end=old_end)
        with self.app.app_context():
            old = app_module.db.session.get(app_module.PurchaseOrder, old_id)
            original_snapshot = (old.po_date, old.end_date, old.po_number)
        renewal_edit = self.client.put(
            f'/update_purchase_order/{old_id}',
            json=self._payload(
                [serial],
                start=(old_end + timedelta(days=1)).isoformat(),
                end=(today + timedelta(days=200)).isoformat(),
            ),
        )
        self.assertEqual(renewal_edit.status_code, 400)
        self.assertIn('only available', renewal_edit.get_json()['error'].lower())
        with self.app.app_context():
            old = app_module.db.session.get(app_module.PurchaseOrder, old_id)
            product = app_module.db.session.get(app_module.Product, serial)
            self.assertEqual((old.po_date, old.end_date, old.po_number), original_snapshot)
            self.assertEqual((product.start_warranty_date, product.end_warranty_date), (product_start, product_end))

        normal_edit = self.client.put(
            f'/update_purchase_order/{old_id}',
            json={
                'client_id': self.client_id,
                'product_serials': [serial],
                'po_number': f'NORMAL-EDIT-{self.suffix}',
                'po_type': 'single',
                'date_mode': 'product',
                'start_date': '1900-01-01',
                'end_date': '1900-01-02',
            },
        )
        self.assertEqual(normal_edit.status_code, 200, normal_edit.get_data(as_text=True))
        self.assertEqual(normal_edit.get_json()['purchase_order']['start_date'], product_start.isoformat())

    def test_purchase_order_api_exposes_latest_renewal_metadata(self):
        today = app_module.get_manila_today()
        prior_end = today - timedelta(days=7)
        eligible = self._product('api-eligible', start=today - timedelta(days=100), end=None)
        no_history = self._product('api-none', start=today - timedelta(days=100), end=None)
        first_tie_id = self._po([eligible], start=today - timedelta(days=100), end=prior_end)
        second_tie_id = self._po([eligible], start=today - timedelta(days=90), end=prior_end)
        response = self.client.get('/get_purchase_orders')
        self.assertEqual(response.status_code, 200)
        products = {
            item['serial_number']: item
            for item in response.get_json()['products']
        }
        self.assertTrue(products[eligible]['renewal_eligible'])
        self.assertEqual(products[eligible]['latest_prior_end_date'], prior_end.isoformat())
        self.assertEqual(
            products[eligible]['default_renewal_start_date'],
            (prior_end + timedelta(days=1)).isoformat(),
        )
        with self.app.app_context():
            context = app_module.purchase_order_renewal_context([eligible])
            self.assertEqual(context[eligible]['latest_prior_po_id'], max(first_tie_id, second_tie_id))
        self.assertFalse(products[no_history]['renewal_eligible'])
        self.assertEqual(products[no_history]['latest_prior_end_date'], '')
        self.assertEqual(products[no_history]['default_renewal_start_date'], '')

    def test_template_contains_renewal_controls_and_history_notice(self):
        template = (ROOT / 'templates' / 'po_details.html').read_text(encoding='utf-8')
        for expected in (
            'id="po-renewal-notice"',
            'Renewal dates will update the selected Product coverage dates while preserving every existing P.O. snapshot.',
            'date_mode',
            'renewal_eligible',
            'latest_prior_end_date',
            'default_renewal_start_date',
            'Mixed renewal and Product-date selections are not allowed.',
            "payload.date_mode = state.dateMode",
        ):
            self.assertIn(expected, template)


if __name__ == '__main__':
    unittest.main()
